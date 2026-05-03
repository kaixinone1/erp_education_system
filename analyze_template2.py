from openpyxl import load_workbook

filepath = r"D:\erp_thirteen\数据库信息\模板\义务教育学校教职工绩效工资审批表（2）.xlsx"

try:
    wb = load_workbook(filepath)
    ws = wb.active

    print(f"工作表名称: {ws.title}")
    print(f"最大行: {ws.max_row}")
    print(f"最大列: {ws.max_column}")
    print()

    print("=" * 80)
    print("检查所有列（包括L-O列）：")
    print("=" * 80)

    for row in range(1, min(32, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(16, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                col_letter = chr(64 + col) if col <= 26 else f"col{col}"
                row_data.append(f"{col_letter}{row}={repr(str(cell.value))[:30]}")
        if row_data:
            print(f"第{row}行: {', '.join(row_data)}")

    print()
    print("=" * 80)
    print("检查呈报单位意见区域 (E3:E9)：")
    print("=" * 80)
    for row in range(3, 10):
        for col in range(5, 12):  # E到K列
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                col_letter = chr(64 + col)
                print(f"  {col_letter}{row}: {repr(cell.value)[:50]}")

    print()
    print("=" * 80)
    print("检查教育局意见区域 (F10:K18)：")
    print("=" * 80)
    for row in range(10, 19):
        for col in range(6, 12):  # F到K列
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                col_letter = chr(64 + col)
                print(f"  {col_letter}{row}: {repr(cell.value)[:50]}")

    print()
    print("=" * 80)
    print("检查人事部门意见区域 (F19:K30)：")
    print("=" * 80)
    for row in range(19, 31):
        for col in range(6, 12):  # F到K列
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                col_letter = chr(64 + col)
                print(f"  {col_letter}{row}: {repr(cell.value)[:50]}")

except FileNotFoundError:
    print(f"文件不存在: {filepath}")
except Exception as e:
    print(f"错误: {e}")