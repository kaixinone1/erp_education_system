"""诊断绩效工资审批表数据一致性 - 对比Excel和Fill API"""
import psycopg2
import json
import os
import requests
from openpyxl import load_workbook

BASE_URL = "http://127.0.0.1:8000/api/universal-template"

conn = psycopg2.connect(
    host='localhost', port=5432, database='taiping_education',
    user='taiping_user', password='taiping_password'
)
cursor = conn.cursor()

# 1. 查询最近的绩效工资保存记录
cursor.execute("""
    SELECT id, "模板名称", "单位名称", "年月", "查询条件", "统计范围", "填报口径", "excel路径", "保存时间"
    FROM saved_exports
    WHERE "模板名称" LIKE '%绩效%'
    ORDER BY "保存时间" DESC LIMIT 1
""")
record = cursor.fetchone()
if not record:
    print("没有找到绩效工资保存记录")
    exit()

print(f"最新保存记录: ID={record[0]}, 年月={record[3]}, 查询条件={record[4]}")

# 2. 读取Excel文件
saved_dir = r"D:\erp_fifteen\tp_education_system\backend\exports\templates\已保存"
# 找到最新的Excel文件
excel_files = [f for f in os.listdir(saved_dir) if f.endswith('.xlsx') and '绩效' in f]
excel_files.sort(key=lambda x: os.path.getmtime(os.path.join(saved_dir, x)), reverse=True)
latest_excel = os.path.join(saved_dir, excel_files[0]) if excel_files else None
print(f"最新Excel文件: {latest_excel}")

excel_data = {}
if latest_excel and os.path.exists(latest_excel):
    wb = load_workbook(latest_excel)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None:
                excel_data[cell.coordinate] = str(cell.value)

# 3. 调用fill API
query_condition = record[4] if isinstance(record[4], dict) else (json.loads(record[4]) if record[4] else {})
print(f"\n调用fill API, 查询条件={query_condition}")

resp = requests.post(f"{BASE_URL}/fill", json={
    "模板ID": "tpl_15cc984d",
    "查询条件": query_condition,
    "统计范围": json.loads(record[5]) if record[5] and isinstance(record[5], str) and record[5] != '{}' else None,
    "填报口径": json.loads(record[6]) if record[6] and isinstance(record[6], str) and record[6] != '{}' else None
}, timeout=30)

fill_data = {}
if resp.status_code == 200:
    data = resp.json()
    if data.get('成功'):
        filled = data['数据']['配置']
        for cell in filled.get('单元格数据', []):
            key = f"[{cell.get('行号')},{cell.get('列号')}]"
            fill_data[key] = str(cell.get('显示值', '') or '')
    else:
        print(f"Fill API返回失败: {data}")
else:
    print(f"Fill API错误: {resp.status_code}, {resp.text[:300]}")

# 4. 对比
print("\n" + "=" * 70)
print("=== 数据一致性对比 (Excel vs Fill API 预览) ===")
print(f"{'Excel坐标':<8} {'字段名':<22} {'Excel值':<16} {'Fill预览值':<16} {'一致?'}")
print("-" * 70)

key_coords = [
    ('B7', '科员级人数'), ('B10', '高级教师人数'), ('B11', '一级教师人数'),
    ('B12', '二级教师人数'), ('B13', '三级教师人数'), ('B15', '技师人数'),
    ('B18', '初级工人数'), ('B20', '绩效工资合计人数'), ('B21', '乡镇补贴合计人数'),
    ('B24', '岗位遗留问题合计人数'), ('B25', '退休干部'), ('B26', '退休工人'),
    ('B27', '离休干部'), 
    ('D7', '科员级小计'), ('D10', '高级教师小计'), ('D11', '一级教师小计'),
    ('D12', '二级教师小计'), ('D13', '三级教师小计'), ('D15', '技师小计'),
    ('D18', '初级工小计'), ('D20', '绩效工资合计'), ('D21', '乡镇补贴合计'),
    ('D24', '岗位遗留问题合计'),
    ('A28', '备注'),
]

mismatches = []
for coord, name in key_coords:
    col_letter = coord[0]
    row_num = int(coord[1:])
    col_num = ord(col_letter) - ord('A') + 1
    excel_val = excel_data.get(coord, '(空)')
    fill_key = f"[{row_num},{col_num}]"
    fill_val = fill_data.get(fill_key, '(空)')
    match = '✓' if excel_val == fill_val else '✗ 不一致!'
    if match != '✓':
        mismatches.append((coord, name, excel_val, fill_val))
    print(f"{coord:<8} {name:<22} {excel_val[:15]:<16} {fill_val[:15]:<16} {match}")

if mismatches:
    print(f"\n❌ 发现 {len(mismatches)} 处数据不一致:")
    for coord, name, excel, fill in mismatches:
        print(f"  {coord} ({name}): Excel={excel}, Fill预览={fill}")
else:
    print("\n✓ 所有数据一致!")

# 5. 也检查一下备注
print(f"\n=== 备注内容 ===")
print(f"Excel A28: {excel_data.get('A28', '(空)')[:200]}")
fill_remark_key = None
for key in fill_data:
    if '28' in key:
        print(f"Fill {key}: {fill_data[key][:200]}")
        fill_remark_key = key

cursor.close()
conn.close()