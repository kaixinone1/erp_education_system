import xlrd
import openpyxl
import os

# 找到最新的上传文件
upload_dir = r'D:\erp_thirteen\tp_education_system\backend\uploads'
xls_files = [f for f in os.listdir(upload_dir) if f.endswith('.xls') and '最新工资' in f]
if not xls_files:
    print('未找到最新工资包数据文件')
    exit(1)

# 按修改时间排序，取最新的
xls_files.sort(key=lambda f: os.path.getmtime(os.path.join(upload_dir, f)), reverse=True)
src = os.path.join(upload_dir, xls_files[0])
print(f'源文件: {xls_files[0]}')

# 用xlrd读取
wb = xlrd.open_workbook(src)
sheet = wb.sheet_by_index(0)
print(f'读取: {sheet.nrows}行, {sheet.ncols}列')

if sheet.nrows == 0:
    print('错误: 文件读取为空')
    exit(1)

# 用openpyxl写入xlsx
dst = os.path.join(upload_dir, xls_files[0].replace('.xls', '.xlsx'))
owb = openpyxl.Workbook()
ows = owb.active

for row_idx in range(sheet.nrows):
    for col_idx in range(sheet.ncols):
        val = sheet.cell_value(row_idx, col_idx)
        ows.cell(row=row_idx + 1, column=col_idx + 1, value=val)

owb.save(dst)
print(f'已转换: {os.path.basename(dst)} ({sheet.nrows}行, {sheet.ncols}列)')

# 验证转换结果
owb2 = openpyxl.load_workbook(dst)
ows2 = owb2.active
print(f'验证: {ows2.max_row}行, {ows2.max_column}列')

wb.release_resources()