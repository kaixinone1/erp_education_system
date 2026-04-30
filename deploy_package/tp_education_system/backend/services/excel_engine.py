"""
Excel模板引擎 - 处理Excel模板的填充
"""
from openpyxl import load_workbook
from typing import Dict, Any, List
import os
from datetime import datetime


class ExcelEngine:
    """Excel模板填充引擎"""

    @staticmethod
    def fill_excel_template(
        template_path: str,
        output_path: str,
        data: Dict[str, Any],
        field_mappings: List[Dict[str, Any]]
    ) -> str:
        """
        填充Excel模板

        Args:
            template_path: 模板文件路径
            output_path: 输出文件路径
            data: 填充数据
            field_mappings: 字段映射配置

        Returns:
            输出文件路径
        """
        # 加载工作簿
        wb = load_workbook(template_path)
        ws = wb.active

        # 根据字段映射填充数据
        for field in field_mappings:
            position_data = field.get('position_data', {})

            # 获取单元格位置（支持两种格式）
            if 'col_index' in position_data and 'row_index' in position_data:
                # 新格式：使用行列索引（0-based）
                col = position_data['col_index'] + 1  # 转为1-based
                row = position_data['row_index'] + 1
            elif 'col' in position_data and 'row' in position_data:
                # 直接指定
                col = position_data['col']
                row = position_data['row']
            else:
                continue

            # 获取数据值
            data_source = field.get('data_source', '')
            if data_source:
                # 从数据源获取值
                if '.' in data_source:
                    # 格式：table_name.column_name
                    parts = data_source.split('.')
                    column_name = parts[-1]
                    value = data.get(column_name, '')
                else:
                    value = data.get(data_source, '')
            else:
                value = field.get('default_value', '')

            # 填充单元格
            cell = ws.cell(row=row, column=col)
            cell.value = value if value is not None else ''

        # 保存工作簿
        wb.save(output_path)
        return output_path

    @staticmethod
    def preview_excel_template(
        template_path: str,
        data: Dict[str, Any],
        field_mappings: List[Dict[str, Any]]
    ) -> bytes:
        """
        预览填充后的Excel（返回字节数据）
        """
        import io

        wb = load_workbook(template_path)
        ws = wb.active

        # 填充数据
        for field in field_mappings:
            position_data = field.get('position_data', {})

            if 'col_index' in position_data and 'row_index' in position_data:
                col = position_data['col_index'] + 1
                row = position_data['row_index'] + 1
            else:
                continue

            data_source = field.get('data_source', '')
            if data_source:
                if '.' in data_source:
                    column_name = data_source.split('.')[-1]
                    value = data.get(column_name, '')
                else:
                    value = data.get(data_source, '')
            else:
                value = field.get('default_value', '')

            cell = ws.cell(row=row, column=col)
            cell.value = value if value is not None else ''

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


# 便捷函数
def fill_excel_from_template(
    template_id: str,
    business_id: int,
    db_connection_func,
    output_dir: str = None
) -> str:
    """
    从模板填充Excel
    """
    import json
    import psycopg2

    conn = db_connection_func()
    cursor = conn.cursor()

    try:
        # 获取模板信息
        cursor.execute("""
            SELECT template_name, file_path FROM document_templates WHERE template_id = %s
        """, (template_id,))
        template_row = cursor.fetchone()
        if not template_row:
            raise ValueError(f"模板不存在: {template_id}")
        template_name = template_row[0]
        template_path = template_row[1]

        # 获取字段映射
        cursor.execute("""
            SELECT field_name, field_label, data_source, position_data, default_value
            FROM template_field_mappings
            WHERE template_id = %s
            ORDER BY sort_order
        """, (template_id,))

        field_rows = cursor.fetchall()
        field_mappings = []
        for row in field_rows:
            position_data = row[3]
            if isinstance(position_data, str):
                position_data = json.loads(position_data)
            field_mappings.append({
                'field_name': row[0],
                'field_label': row[1],
                'data_source': row[2],
                'position_data': position_data,
                'default_value': row[4]
            })

        # 获取教师数据
        cursor.execute("""
            SELECT id, name, id_card, archive_birth_date, ethnicity,
                   native_place, work_start_date, contact_phone
            FROM teacher_basic_info WHERE id = %s
        """, (business_id,))

        teacher_row = cursor.fetchone()
        if not teacher_row:
            raise ValueError(f"教师不存在: {business_id}")

        data = {
            'id': teacher_row[0],
            'name': teacher_row[1] or '',
            'id_card': teacher_row[2] or '',
            'archive_birth_date': str(teacher_row[3]) if teacher_row[3] else '',
            'ethnicity': teacher_row[4] or '',
            'native_place': teacher_row[5] or '',
            'work_start_date': str(teacher_row[6]) if teacher_row[6] else '',
            'contact_phone': teacher_row[7] or ''
        }

        # 生成输出路径 - 使用"模板名称+教师姓名"格式
        if output_dir is None:
            output_dir = r'd:\erp_thirteen\tp_education_system\backend\uploads\generated'
        os.makedirs(output_dir, exist_ok=True)

        # 清理文件名中的非法字符
        safe_template_name = "".join(c for c in template_name if c.isalnum() or c in (' ', '_', '-'))
        safe_teacher_name = "".join(c for c in data['name'] if c.isalnum() or c in (' ', '_', '-'))
        
        filename = f'{safe_template_name}_{safe_teacher_name}.xlsx'
        output_path = os.path.join(output_dir, filename)

        # 填充Excel
        ExcelEngine.fill_excel_template(template_path, output_path, data, field_mappings)

        return output_path

    finally:
        cursor.close()
        conn.close()
