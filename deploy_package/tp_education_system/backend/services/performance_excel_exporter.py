"""
绩效工资审批表专业Excel导出服务
100%还原原始模板格式
使用openpyxl精确控制：
- 单元格大小、字形字号、行高列宽
- 对齐方式、边框样式
- 合并单元格
- A4纸张、页边距设置
"""

import os
import json
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import Dict, Optional

class PerformancePayExcelExporter:
    """绩效工资审批表Excel导出器"""
    
    def __init__(self):
        self.template_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            'data',
            'performance_pay_template.json'
        )
        self.output_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            'exports'
        )
        os.makedirs(self.output_dir, exist_ok=True)
        
    def load_template(self) -> Dict:
        """加载模板元数据"""
        if os.path.exists(self.template_path):
            with open(self.template_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return self._get_default_template()
    
    def _get_default_template(self) -> Dict:
        """获取默认模板结构"""
        return {
            "page_setup": {
                "margins": {"top": 72, "right": 54, "bottom": 72, "left": 54},
                "paper_width": 595,
                "paper_height": 842,
                "orientation": "portrait"
            },
            "col_widths": ["84", "100", "100", "100", "75", "auto"],
            "rows": []
        }
    
    def export(self, data: Dict, year_month: str) -> str:
        """
        导出绩效工资审批表为Excel
        :param data: 数据字典
        :param year_month: 年月字符串，如"2026年5月"
        :return: 生成的Excel文件路径
        """
        template = self.load_template()
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "绩效工资审批表"
        
        # 设置页面（A4纸张）
        self._setup_page(ws, template)
        
        # 设置列宽
        self._setup_columns(ws, template)
        
        # 创建样式
        styles = self._create_styles()
        
        # 填充数据
        self._fill_data(ws, template, data, styles)
        
        # 生成文件名
        filename = f"绩效工资审批表_{year_month}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        # 保存文件
        wb.save(filepath)
        
        return filepath
    
    def _setup_page(self, ws, template: Dict):
        """设置页面参数"""
        page_setup = template.get('page_setup', {})
        
        # A4纸张
        ws.page_setup.paperSize = 9  # A4 = 9
        
        # 纵向
        ws.page_setup.orientation = 'portrait'
        
        # 自动调整到一页
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        
        # 页边距（单位：英寸，Excel默认单位）
        margins = page_setup.get('margins', {})
        # Excel边距单位是英寸，模板中是1/72英寸（点）
        ws.page_margins.left = margins.get('left', 54) / 72
        ws.page_margins.right = margins.get('right', 54) / 72
        ws.page_margins.top = margins.get('top', 72) / 72
        ws.page_margins.bottom = margins.get('bottom', 72) / 72
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
        
    def _setup_columns(self, ws, template: Dict):
        """设置列宽"""
        col_widths = template.get('col_widths', ['84', '100', '100', '100', '75', 'auto'])
        
        for i, width in enumerate(col_widths, 1):
            col_letter = get_column_letter(i)
            if width == 'auto':
                ws.column_dimensions[col_letter].width = 20
            else:
                # Excel列宽单位是字符宽度，需要转换
                ws.column_dimensions[col_letter].width = int(width) / 7
        
    def _create_styles(self):
        """创建样式字典"""
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        thick_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        
        return {
            'title': {
                'font': Font(name='宋体', size=16, bold=True),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': None
            },
            'info': {
                'font': Font(name='宋体', size=11),
                'alignment': Alignment(horizontal='left', vertical='center'),
                'border': None
            },
            'header': {
                'font': Font(name='宋体', size=11, bold=True),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': thin_border
            },
            'normal': {
                'font': Font(name='宋体', size=11),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': thin_border
            },
            'normal_left': {
                'font': Font(name='宋体', size=11),
                'alignment': Alignment(horizontal='left', vertical='center'),
                'border': thin_border
            },
            'vertical_text': {
                'font': Font(name='宋体', size=11, bold=True),
                'alignment': Alignment(horizontal='center', vertical='center', text_rotation=255),
                'border': thin_border
            },
            'remarks': {
                'font': Font(name='宋体', size=11),
                'alignment': Alignment(horizontal='left', vertical='top', wrap_text=True),
                'border': thin_border
            },
            'thin_border': thin_border,
            'thick_border': thick_border
        }
    
    def _fill_data(self, ws, template: Dict, data: Dict, styles: Dict):
        """填充数据到工作表"""
        rows = template.get('rows', [])
        occupied_cells = {}
        
        for row_idx, row in enumerate(rows, 1):
            row_height = row.get('height', 25)
            ws.row_dimensions[row_idx].height = row_height
            
            col_idx = 1
            for cell in row.get('cells', []):
                # 检查是否被合并单元格占用
                while occupied_cells.get(f"{row_idx}-{col_idx}"):
                    col_idx += 1
                    if col_idx > 6:
                        break
                
                if col_idx > 6:
                    break
                
                text = cell.get('text', '')
                rowspan = cell.get('rowspan', 1)
                colspan = cell.get('colspan', 1)
                cell_class = cell.get('class', '')
                align = cell.get('align', 'center')
                is_notes_row = cell.get('isNotesRow', False)
                
                # 替换日期占位符
                text = self._replace_placeholders(text, data)
                
                # 合并单元格
                if rowspan > 1 or colspan > 1:
                    ws.merge_cells(
                        start_row=row_idx,
                        start_column=col_idx,
                        end_row=row_idx + rowspan - 1,
                        end_column=col_idx + colspan - 1
                    )
                    
                    # 标记被占用的单元格
                    for r in range(row_idx, row_idx + rowspan):
                        for c in range(col_idx, col_idx + colspan):
                            occupied_cells[f"{r}-{c}"] = True
                
                # 设置单元格值
                cell = ws.cell(row=row_idx, column=col_idx, value=text)
                
                # 设置样式
                self._apply_cell_style(cell, cell_class, align, is_notes_row, styles)
            
    def _replace_placeholders(self, text: str, data: Dict) -> str:
        """替换文本中的占位符"""
        if not text:
            return text
        
        # 替换日期
        today = datetime.now()
        today_str = f"{today.year}年{today.month}月{today.day}日"
        
        # 正确计算明天（避免月末问题）
        tomorrow = today + timedelta(days=1)
        tomorrow_str = f"{tomorrow.year}年{tomorrow.month}月{tomorrow.day}日"
        
        text = text.replace('2026年5月', data.get('年月', ''))
        text = text.replace('2026年4月28日', today_str)
        text = text.replace('2026年4月29日', tomorrow_str)
        
        # 替换数据占位符
        text = text.replace('0人，0元', f"{data.get('绩效人数合计', 0)}人，{data.get('绩效工资合计', 0)}元")
        text = text.replace('生活补贴0人，0元', f"生活补贴{data.get('在职人数', 0)}人，{data.get('乡镇补贴合计', 0)}元")
        text = text.replace('岗位设置遗留0人，0元', f"岗位设置遗留{data.get('遗留问题人数', 0)}人，{data.get('遗留问题金额', 0)}元")
        text = text.replace('无乡镇补贴0人，姓名', f"无乡镇补贴{data.get('无补贴人数', 0)}人，{data.get('无补贴名单', '')}")
        
        total = data.get('绩效工资合计', 0) + data.get('乡镇补贴合计', 0) + data.get('遗留问题金额', 0)
        text = text.replace('总计0人，0元', f"总计{total}元")
        text = text.replace('合计0人，0元', f"合计{data.get('绩效工资合计', 0) + data.get('乡镇补贴合计', 0)}元")
        
        return text
    
    def _apply_cell_style(self, cell, cell_class: str, align: str, is_notes_row: bool, styles: Dict):
        """应用单元格样式"""
        # 根据单元格类型选择样式
        if cell_class in ['title', 'xl75']:
            style = styles['title']
        elif cell_class in ['info']:
            style = styles['info']
        elif cell_class in ['xl78', 'xl107', 'xl122']:
            # 垂直文本
            style = styles['vertical_text']
        elif cell_class in ['xl80', 'xl90', 'xl104']:
            # 分类标题
            style = styles['header']
            style['alignment'] = Alignment(horizontal='left', vertical='center')
        elif is_notes_row:
            style = styles['remarks']
        else:
            style = styles['normal']
        
        # 设置字体
        cell.font = style['font']
        
        # 设置对齐方式
        if align == 'left':
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        elif align == 'right':
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = style['alignment']
        
        # 设置边框
        if style['border']:
            cell.border = style['border']

def export_performance_pay(data: Dict, year_month: str) -> str:
    """
    导出绩效工资审批表的便捷函数
    :param data: 数据字典
    :param year_month: 年月字符串
    :return: Excel文件路径
    """
    exporter = PerformancePayExcelExporter()
    return exporter.export(data, year_month)

if __name__ == "__main__":
    # 测试导出
    test_data = {
        '年月': '2026年5月',
        '填报单位': '太平中心学校',
        '填报时间': '2026年4月28日',
        '绩效人数合计': 50,
        '绩效工资合计': 65000,
        '在职人数': 45,
        '乡镇补贴标准': 350,
        '乡镇补贴合计': 15750,
        '遗留问题人数': 2,
        '遗留问题金额': 3000,
        '无补贴人数': 5,
        '无补贴名单': '张三、李四、王五'
    }
    
    path = export_performance_pay(test_data, '2026年5月')
    print(f"Excel导出成功: {path}")
