from openpyxl import load_workbook
import os

filepath = r"D:\erp_thirteen\数据库信息\模板\义务教育学校教职工绩效工资审批表（2）.xlsx"

print(f"文件路径: {filepath}")
print(f"文件是否存在: {os.path.exists(filepath)}")
print(f"文件大小: {os.path.getsize(filepath) if os.path.exists(filepath) else 'N/A'} bytes")
print()

wb = load_workbook(filepath)
ws = wb.active

print(f"工作表名称: {ws.title}")
print(f"最大行: {ws.max_row}")
print(f"最大列: {ws.max_column}")
print()

print("=" * 80)
print("所有非空单元格内容：")
print("=" * 80)

for row in range(1, ws.max_row + 1):
    row_data = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            col_letter = chr(64 + col) if col <= 26 else f"col{col}"
            row_data.append(f"{col_letter}{row}={repr(str(cell.value))[:60]}")
    if row_data:
        print(f"第{row}行: {', '.join(row_data)}")

print()
print("=" * 80)
print("合并单元格：")
print("=" * 80)
for merge in ws.merged_cells.ranges:
    print(f"  {merge}")