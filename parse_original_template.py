"""
逆向解析原始绩效审批表Excel模板
提取所有格式信息
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

filepath = r"D:\erp_thirteen\数据库信息\模板\义务教育学校教职工绩效工资审批表.xlsx"

print("=" * 80)
print("逆向解析原始Excel模板")
print("=" * 80)

wb = load_workbook(filepath)
ws = wb.active

print(f"\n工作表名称: {ws.title}")
print(f"最大行: {ws.max_row}")
print(f"最大列: {ws.max_column}")

# 页面设置
print("\n【页面设置】")
print(f"  纸张大小: {ws.page_setup.paperSize} (9=A4)")
print(f"  打印方向: {ws.page_setup.orientation}")
print(f"  页边距 - 左: {ws.page_margins.left:.2f}, 右: {ws.page_margins.right:.2f}")
print(f"  页边距 - 上: {ws.page_margins.top:.2f}, 下: {ws.page_margins.bottom:.2f}")

# 列宽
print("\n【列宽设置】")
for i in range(1, ws.max_column + 1):
    col_letter = get_column_letter(i)
    width = ws.column_dimensions[col_letter].width
    print(f"  {col_letter}列: {width:.2f}")

# 行高
print("\n【行高设置】")
for i in range(1, ws.max_row + 1):
    height = ws.row_dimensions[i].height
    if height:
        print(f"  第{i}行: {height:.2f} pt")

# 合并单元格
print("\n【合并单元格】")
for merge in ws.merged_cells.ranges:
    print(f"  {merge}")

# 单元格内容（前20行，前15列）
print("\n【单元格内容】")
for row in range(1, min(33, ws.max_row + 1)):
    row_data = []
    for col in range(1, min(15, ws.max_column + 1)):
        cell = ws.cell(row=row, column=col)
        if cell.value:
            row_data.append(f"{get_column_letter(col)}{row}='{cell.value}'")
    if row_data:
        print(f"  第{row}行: {', '.join(row_data)}")

print("\n解析完成！")