#!/usr/bin/env python3
"""
数据导出引擎 - 将JSON配置导出为Excel文件，保持原始格式
核心功能：JSON数据转Excel、样式还原、合并单元格处理
"""
import os
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

class DataExportEngine:
    """
    数据导出引擎 - 负责将JSON配置转换为Excel文件
    """
    
    def __init__(self):
        pass
    
    def export_to_excel(self, config: Dict[str, Any], output_path: str) -> bool:
        """
        导出为Excel文件，保持格式
        
        Args:
            config: 填充后的JSON数据
            output_path: 输出文件路径
            
        Returns:
            是否成功
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            
            # 应用页面设置
            self._apply_page_setup(ws, config.get("page_setup", {}))
            
            # 应用列宽
            self._apply_column_widths(ws, config.get("column_widths", []))
            
            # 写入数据和样式
            self._write_data(ws, config.get("rows", []))
            
            # 处理合并单元格
            self._apply_merged_cells(ws, config.get("merged_cells", []))
            
            # 保存文件
            wb.save(output_path)
            return True
            
        except Exception as e:
            print(f"导出Excel失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _apply_page_setup(self, ws, page_setup: Dict[str, Any]):
        """应用页面设置"""
        # 设置纸张大小
        if page_setup.get("paper_size"):
            ws.page_setup.paper_size = page_setup["paper_size"]
        
        # 设置方向
        if page_setup.get("orientation") == "landscape":
            ws.page_setup.orientation = "landscape"
        else:
            ws.page_setup.orientation = "portrait"
        
        # 设置边距
        margins = page_setup.get("margins", {})
        page_margins = PageMargins()
        page_margins.top = margins.get("top", 72) / 72  # 转换为英寸
        page_margins.right = margins.get("right", 54) / 72
        page_margins.bottom = margins.get("bottom", 72) / 72
        page_margins.left = margins.get("left", 54) / 72
        page_margins.header = margins.get("header", 0) / 72
        page_margins.footer = margins.get("footer", 0) / 72
        ws.page_margins = page_margins
        
        # 设置打印区域
        if page_setup.get("print_area"):
            ws.print_area = page_setup["print_area"]
        
        # 设置缩放
        if page_setup.get("fit_to_page"):
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = page_setup.get("fit_to_width", 1)
            ws.page_setup.fitToHeight = page_setup.get("fit_to_height", 0)
    
    def _apply_column_widths(self, ws, column_widths: List[float]):
        """应用列宽"""
        for idx, width in enumerate(column_widths):
            col_letter = get_column_letter(idx + 1)
            # 转换为字符宽度（Excel的列宽单位）
            ws.column_dimensions[col_letter].width = width / 7.2
    
    def _write_data(self, ws, rows: List[Dict[str, Any]]):
        """写入数据和样式"""
        for row in rows:
            row_num = row["row_number"]
            height = row.get("height", 15)
            
            # 设置行高
            ws.row_dimensions[row_num].height = height
            
            # 写入单元格
            for cell in row.get("cells", []):
                col_num = cell["column_number"]
                text = cell["text"]
                
                # 获取单元格
                ws_cell = ws.cell(row=row_num, column=col_num, value=text)
                
                # 应用样式
                style = cell.get("style", {})
                self._apply_cell_style(ws_cell, style)
    
    def _apply_cell_style(self, ws_cell, style: Dict[str, Any]):
        """应用单元格样式"""
        # 字体样式
        font_info = style.get("font", {})
        underline_val = font_info.get("underline", False)
        if underline_val is True:
            underline_val = "single"
        elif underline_val not in ['single', 'double', 'singleAccounting', 'doubleAccounting']:
            underline_val = None
        
        font = Font(
            name=font_info.get("name", "宋体"),
            size=font_info.get("size", 11),
            bold=font_info.get("bold", False),
            italic=font_info.get("italic", False),
            underline=underline_val,
            color=font_info.get("color")
        )
        ws_cell.font = font
        
        # 填充样式
        fill_info = style.get("fill", {})
        if fill_info.get("fg_color"):
            fill = PatternFill(
                fill_type=fill_info.get("type", "solid"),
                fgColor=fill_info["fg_color"]
            )
            ws_cell.fill = fill
        
        # 对齐样式
        align_info = style.get("alignment", {})
        alignment = Alignment(
            horizontal=align_info.get("horizontal", "general"),
            vertical=align_info.get("vertical", "bottom"),
            wrap_text=align_info.get("wrap_text", False),
            text_rotation=align_info.get("text_rotation", 0),
            indent=align_info.get("indent", 0)
        )
        ws_cell.alignment = alignment
        
        # 边框样式
        border_info = style.get("border", {})
        border = Border(
            top=self._create_border_side(border_info.get("top", {})),
            bottom=self._create_border_side(border_info.get("bottom", {})),
            left=self._create_border_side(border_info.get("left", {})),
            right=self._create_border_side(border_info.get("right", {}))
        )
        ws_cell.border = border
        
        # 数字格式
        number_format = style.get("number_format")
        if number_format:
            ws_cell.number_format = number_format
    
    def _create_border_side(self, side_info: Dict[str, Any]) -> Side:
        """创建边框边"""
        style = side_info.get("style", "thin")
        color = side_info.get("color")
        
        if not style or style == "none":
            return Side(style=None)
        
        return Side(style=style, color=color)
    
    def _apply_merged_cells(self, ws, merged_cells: List[Dict[str, Any]]):
        """应用合并单元格"""
        for merged in merged_cells:
            start_row = merged["start_row"]
            start_col = merged["start_col"]
            end_row = merged["end_row"]
            end_col = merged["end_col"]
            
            start_cell = f"{get_column_letter(start_col)}{start_row}"
            end_cell = f"{get_column_letter(end_col)}{end_row}"
            
            ws.merge_cells(f"{start_cell}:{end_cell}")
    
    def batch_export(self, configs: List[Dict[str, Any]], output_dir: str) -> List[str]:
        """
        批量导出
        
        Args:
            configs: JSON配置列表
            output_dir: 输出目录
            
        Returns:
            导出的文件路径列表
        """
        exported_files = []
        
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        for i, config in enumerate(configs):
            template_name = config.get("template_name", f"template_{i+1}")
            template_name = os.path.splitext(template_name)[0]
            output_path = os.path.join(output_dir, f"{template_name}.xlsx")
            
            if self.export_to_excel(config, output_path):
                exported_files.append(output_path)
        
        return exported_files
    
    def export_to_pdf(self, config: Dict[str, Any], output_path: str) -> bool:
        """
        导出为PDF（需要安装pywin32或其他PDF库）
        
        Args:
            config: JSON配置
            output_path: 输出文件路径
            
        Returns:
            是否成功
        """
        try:
            # 先导出为Excel，再转换为PDF
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name
            
            if not self.export_to_excel(config, tmp_path):
                return False
            
            # 尝试使用win32com转换为PDF
            try:
                import win32com.client
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                
                wb = excel.Workbooks.Open(tmp_path)
                wb.ExportAsFixedFormat(0, output_path)  # 0 = xlTypePDF
                wb.Close()
                excel.Quit()
                
                os.unlink(tmp_path)
                return True
            except ImportError:
                print("需要安装pywin32才能导出PDF")
                os.unlink(tmp_path)
                return False
            except Exception as e:
                print(f"转换PDF失败: {e}")
                os.unlink(tmp_path)
                return False
                
        except Exception as e:
            print(f"导出PDF失败: {e}")
            return False

# 单例导出
data_export_engine = DataExportEngine()