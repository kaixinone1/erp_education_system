#!/usr/bin/env python3
"""
模板导入引擎 - 将Excel模板转换为JSON配置
核心功能：提取页面设置、列宽、行高、单元格样式、合并单元格
"""
import json
import os
from typing import Dict, List, Any
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class RGBEncoder(json.JSONEncoder):
    """自定义JSON编码器，处理RGB对象"""
    def default(self, obj):
        # 处理RGB对象
        if hasattr(obj, '__class__') and obj.__class__.__name__ == 'RGB':
            return str(obj)
        # 处理Color对象
        if hasattr(obj, 'rgb'):
            return obj.rgb
        # 处理其他不可序列化的对象
        try:
            return str(obj)
        except:
            return None
        return json.JSONEncoder.default(self, obj)

class TemplateImportEngine:
    """
    模板导入引擎 - 负责将Excel模板转换为JSON配置
    """
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xlsm']
    
    def import_template(self, excel_path: str) -> Dict[str, Any]:
        """
        导入Excel模板，生成JSON配置
        
        Args:
            excel_path: Excel文件路径
        
        Returns:
            JSON配置字典
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"文件不存在: {excel_path}")
        
        ext = os.path.splitext(excel_path)[1].lower()
        if ext not in self.supported_formats:
            raise ValueError(f"不支持的文件格式: {ext}，支持格式: {self.supported_formats}")
        
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        config = {
            "template_id": self._generate_template_id(excel_path),
            "template_name": os.path.basename(excel_path),
            "page_setup": self._extract_page_setup(ws),
            "column_widths": self._extract_column_widths(ws),
            "rows": self._extract_rows_data(ws),
            "merged_cells": self._extract_merged_cells(ws),
            "field_mappings": {}
        }
        
        return config
    
    def _generate_template_id(self, excel_path: str) -> str:
        """生成模板ID"""
        name = os.path.splitext(os.path.basename(excel_path))[0]
        # 转换为小写，去除特殊字符，替换空格为下划线
        import re
        return re.sub(r'[^a-z0-9_]', '', name.lower().replace(' ', '_'))
    
    def _extract_page_setup(self, ws) -> Dict[str, Any]:
        """提取页面设置"""
        # 处理不同版本openpyxl的兼容性
        paper_size = "A4"
        try:
            # openpyxl >= 3.1
            paper_size = ws.page_setup.paperSize if ws.page_setup.paperSize else "A4"
        except AttributeError:
            try:
                # openpyxl < 3.1
                paper_size = ws.page_setup.paper_size if ws.page_setup.paper_size else "A4"
            except:
                paper_size = "A4"
        
        orientation = "portrait"
        try:
            orientation = "landscape" if ws.page_setup.orientation == "landscape" else "portrait"
        except:
            orientation = "portrait"
        
        fit_to_page = False
        try:
            fit_to_page = ws.page_setup.fitToPage
        except:
            fit_to_page = False
        
        margins = {
            "top": 72,
            "right": 54,
            "bottom": 72,
            "left": 54,
            "header": 0,
            "footer": 0
        }
        try:
            margins = {
                "top": round(ws.page_margins.top * 72, 1),
                "right": round(ws.page_margins.right * 72, 1),
                "bottom": round(ws.page_margins.bottom * 72, 1),
                "left": round(ws.page_margins.left * 72, 1),
                "header": round(ws.page_margins.header * 72, 1),
                "footer": round(ws.page_margins.footer * 72, 1)
            }
        except:
            pass
        
        print_area = ""
        try:
            print_area = ws.print_area or ""
        except:
            print_area = ""
        
        fit_to_width = 1
        fit_to_height = 0
        try:
            fit_to_width = ws.page_setup.fitToWidth
            fit_to_height = ws.page_setup.fitToHeight
        except:
            pass
        
        return {
            "paper_size": paper_size,
            "orientation": orientation,
            "margins": margins,
            "print_area": print_area,
            "fit_to_page": fit_to_page,
            "fit_to_width": fit_to_width,
            "fit_to_height": fit_to_height
        }
    
    def _extract_column_widths(self, ws) -> List[float]:
        """提取列宽"""
        widths = []
        max_col = ws.max_column
        
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            width = ws.column_dimensions[col_letter].width
            # 默认宽度为8.43字符
            widths.append(round(width * 7.2, 1) if width else 60)  # 转换为磅
        
        return widths
    
    def _extract_rows_data(self, ws) -> List[Dict[str, Any]]:
        """提取行数据"""
        rows = []
        max_row = ws.max_row
        max_col = ws.max_column
        
        for row_num in range(1, max_row + 1):
            row_data = {
                "row_number": row_num,
                "height": round(ws.row_dimensions[row_num].height, 1) if ws.row_dimensions[row_num].height else 15,
                "cells": []
            }
            
            for col_num in range(1, max_col + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell_info = self._extract_cell_info(cell)
                row_data["cells"].append(cell_info)
            
            rows.append(row_data)
        
        return rows
    
    def _extract_cell_info(self, cell) -> Dict[str, Any]:
        """提取单元格信息"""
        cell_info = {
            "column_number": cell.column,
            "text": self._get_cell_value(cell),
            "style": self._extract_cell_style(cell)
        }
        
        return cell_info
    
    def _get_cell_value(self, cell) -> str:
        """获取单元格值"""
        value = cell.value
        
        if value is None:
            return ""
        
        if isinstance(value, (int, float)):
            return str(value)
        
        return str(value)
    
    def _rgb_to_str(self, rgb_obj):
        """将RGB对象转换为字符串"""
        if rgb_obj is None:
            return None
        # 处理不同类型的RGB对象
        if hasattr(rgb_obj, 'rgb'):
            return rgb_obj.rgb
        if isinstance(rgb_obj, str):
            return rgb_obj
        # 处理直接的RGB类型对象
        try:
            # 尝试获取hex属性（某些版本的RGB对象有hex属性）
            if hasattr(rgb_obj, 'hex'):
                return rgb_obj.hex
            # 尝试获取tuple属性
            if hasattr(rgb_obj, '__iter__'):
                rgb_tuple = tuple(rgb_obj)
                return '{:02X}{:02X}{:02X}'.format(*rgb_tuple)
        except:
            pass
        return None
    
    def _extract_cell_style(self, cell) -> Dict[str, Any]:
        """提取单元格样式"""
        style = cell.style
        font = cell.font
        fill = cell.fill
        alignment = cell.alignment
        border = cell.border
        
        style_dict = {
            "font": {
                "name": font.name or "宋体",
                "size": font.size or 11,
                "bold": font.bold or False,
                "italic": font.italic or False,
                "underline": font.underline or False,
                "color": self._rgb_to_str(font.color)
            },
            "fill": {
                "type": fill.fill_type,
                "fg_color": self._rgb_to_str(fill.fgColor),
                "bg_color": self._rgb_to_str(fill.bgColor)
            },
            "alignment": {
                "horizontal": alignment.horizontal or "general",
                "vertical": alignment.vertical or "bottom",
                "wrap_text": alignment.wrapText or False,
                "text_rotation": alignment.textRotation or 0,
                "indent": alignment.indent or 0
            },
            "border": {
                "top": self._border_side_to_dict(border.top),
                "bottom": self._border_side_to_dict(border.bottom),
                "left": self._border_side_to_dict(border.left),
                "right": self._border_side_to_dict(border.right)
            },
            "number_format": cell.number_format or "General"
        }
        
        return style_dict
    
    def _border_side_to_dict(self, side: Side) -> Dict[str, Any]:
        """将边框边转换为字典"""
        if side is None:
            return {"style": None, "color": None}
        
        return {
            "style": side.style,
            "color": self._rgb_to_str(side.color)
        }
    
    def _extract_merged_cells(self, ws) -> List[Dict[str, Any]]:
        """提取合并单元格"""
        merged_cells = []
        
        for merged in ws.merged_cells.ranges:
            merged_info = {
                "start_row": merged.min_row,
                "start_col": merged.min_col,
                "end_row": merged.max_row,
                "end_col": merged.max_col,
                "coordinate": str(merged)
            }
            merged_cells.append(merged_info)
        
        return merged_cells
    
    def save_config(self, config: Dict[str, Any], output_path: str) -> bool:
        """保存配置到JSON文件"""
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, ensure_ascii=False, indent=2, cls=RGBEncoder)
            return True
        except Exception as e:
            print(f"保存配置失败: {e}")
            return False
    
    def load_config(self, config_path: str) -> Dict[str, Any]:
        """加载JSON配置文件"""
        if not os.path.exists(config_path):
            raise FileNotFoundError(f"配置文件不存在: {config_path}")
        
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)

# 单例导出
template_import_engine = TemplateImportEngine()