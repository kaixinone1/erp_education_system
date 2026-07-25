"""
Excel完整元数据提取器
适用于任意Excel模板，100%还原
"""
import os
import json
from typing import Dict, List, Any, Optional
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# TODO: 数据自动填报功能时启用
# from sqlalchemy import create_engine
# from sqlalchemy.orm import sessionmaker
# 
# 
# DATABASE_URL = "postgresql://taiping_user:taiping_password@localhost:5432/taiping_education"
# engine = create_engine(DATABASE_URL)
# SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


class ExcelMetadataExtractor:
    """Excel完整元数据提取器"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def load(self, file_path: str) -> bool:
        """加载Excel文件"""
        try:
            if not os.path.exists(file_path):
                return False
            self.workbook = load_workbook(file_path, data_only=True)
            if self.workbook is None or len(self.workbook.worksheets) == 0:
                return False
            self.worksheet = self.workbook.active
            return True
        except Exception as e:
            print(f"加载文件失败: {e}")
            return False
    
    def _rgb_to_hex(self, rgb) -> Optional[str]:
        """
        将RGB转换为十六进制颜色
        
        Excel使用ARGB格式：AARRGGBB
        CSS使用：#RRGGBB 或 rgba(R, G, B, A)
        """
        if rgb is None:
            return None
        if isinstance(rgb, str):
            if rgb.startswith('#'):
                return rgb
            if rgb == '00000000':
                return None
            # Excel ARGB格式：AARRGGBB
            if len(rgb) == 8:
                alpha = rgb[0:2]
                red = rgb[2:4]
                green = rgb[4:6]
                blue = rgb[6:8]
                # 如果是完全不透明（alpha=FF），返回6位hex
                if alpha == 'FF':
                    return f"#{red}{green}{blue}"
                # 如果是完全透明，返回None
                if alpha == '00':
                    return None
                # 否则返回rgba格式
                alpha_val = int(alpha, 16) / 255
                return f"rgba({int(red, 16)}, {int(green, 16)}, {int(blue, 16)}, {alpha_val:.2f})"
            return f"#{rgb}"
        return None
    
    def _get_font_info(self, cell) -> Dict[str, Any]:
        """获取字体信息"""
        font_info = {
            "name": "宋体",
            "size": 11.0,
            "bold": False,
            "italic": False,
            "underline": False,
            "strike": False,
            "color": "#000000"  # 默认黑色
        }
        
        try:
            if cell.font:
                if cell.font.name:
                    font_info["name"] = cell.font.name
                if cell.font.size:
                    font_info["size"] = float(cell.font.size)
                if cell.font.bold:
                    font_info["bold"] = True
                if cell.font.italic:
                    font_info["italic"] = True
                if cell.font.underline:
                    font_info["underline"] = True
                if cell.font.strike:
                    font_info["strike"] = True
                if cell.font.color and cell.font.color.rgb:
                    color = self._rgb_to_hex(cell.font.color.rgb)
                    if color:
                        font_info["color"] = color
        except Exception:
            pass
        
        return font_info
    
    def _get_alignment_info(self, cell) -> Dict[str, Any]:
        """获取对齐信息"""
        alignment_info = {
            "horizontal": "general",
            "vertical": "bottom",
            "wrap_text": False,
            "text_rotation": 0
        }
        
        try:
            if cell.alignment:
                if cell.alignment.horizontal:
                    alignment_info["horizontal"] = cell.alignment.horizontal
                if cell.alignment.vertical:
                    alignment_info["vertical"] = cell.alignment.vertical
                if cell.alignment.wrap_text:
                    alignment_info["wrap_text"] = True
                if cell.alignment.textRotation:
                    alignment_info["text_rotation"] = int(cell.alignment.textRotation)
        except Exception:
            pass
        
        return alignment_info
    
    def _get_border_info(self, cell, is_merged=False, is_master=False, rowspan=1, colspan=1) -> Dict[str, Any]:
        """
        获取边框信息
        
        对于合并单元格：
        - 主单元格：提取整个合并区域四周边框
        - 合并区域内部：无边框
        """
        border_info = {
            "top": None,
            "bottom": None,
            "left": None,
            "right": None
        }
        
        try:
            if is_merged and is_master:
                start_row = cell.row - 1
                start_col = cell.column - 1
                end_row = start_row + rowspan - 1
                end_col = start_col + colspan - 1
                
                top_border = None
                for col_idx in range(start_col, end_col + 1):
                    border_cell = self.worksheet.cell(row=start_row + 1, column=col_idx + 1)
                    if border_cell.border and border_cell.border.top and border_cell.border.top.style:
                        top_border = {
                            "style": border_cell.border.top.style,
                            "color": self._rgb_to_hex(border_cell.border.top.color.rgb) if border_cell.border.top.color else "#000000"
                        }
                        break
                
                bottom_border = None
                for col_idx in range(start_col, end_col + 1):
                    border_cell = self.worksheet.cell(row=end_row + 1, column=col_idx + 1)
                    if border_cell.border and border_cell.border.bottom and border_cell.border.bottom.style:
                        bottom_border = {
                            "style": border_cell.border.bottom.style,
                            "color": self._rgb_to_hex(border_cell.border.bottom.color.rgb) if border_cell.border.bottom.color else "#000000"
                        }
                        break
                
                left_border = None
                for row_idx in range(start_row, end_row + 1):
                    border_cell = self.worksheet.cell(row=row_idx + 1, column=start_col + 1)
                    if border_cell.border and border_cell.border.left and border_cell.border.left.style:
                        left_border = {
                            "style": border_cell.border.left.style,
                            "color": self._rgb_to_hex(border_cell.border.left.color.rgb) if border_cell.border.left.color else "#000000"
                        }
                        break
                
                right_border = None
                for row_idx in range(start_row, end_row + 1):
                    border_cell = self.worksheet.cell(row=row_idx + 1, column=end_col + 1)
                    if border_cell.border and border_cell.border.right and border_cell.border.right.style:
                        right_border = {
                            "style": border_cell.border.right.style,
                            "color": self._rgb_to_hex(border_cell.border.right.color.rgb) if border_cell.border.right.color else "#000000"
                        }
                        break
                
                border_info["top"] = top_border
                border_info["bottom"] = bottom_border
                border_info["left"] = left_border
                border_info["right"] = right_border
                
            elif not is_merged:
                if cell.border:
                    if cell.border.top and cell.border.top.style:
                        border_info["top"] = {
                            "style": cell.border.top.style,
                            "color": self._rgb_to_hex(cell.border.top.color.rgb) if cell.border.top.color else "#000000"
                        }
                    if cell.border.bottom and cell.border.bottom.style:
                        border_info["bottom"] = {
                            "style": cell.border.bottom.style,
                            "color": self._rgb_to_hex(cell.border.bottom.color.rgb) if cell.border.bottom.color else "#000000"
                        }
                    if cell.border.left and cell.border.left.style:
                        border_info["left"] = {
                            "style": cell.border.left.style,
                            "color": self._rgb_to_hex(cell.border.left.color.rgb) if cell.border.left.color else "#000000"
                        }
                    if cell.border.right and cell.border.right.style:
                        border_info["right"] = {
                            "style": cell.border.right.style,
                            "color": self._rgb_to_hex(cell.border.right.color.rgb) if cell.border.right.color else "#000000"
                        }
        except Exception:
            pass
        
        return border_info
    
    def _get_fill_info(self, cell) -> Optional[Dict[str, Any]]:
        """获取填充信息"""
        try:
            if cell.fill and cell.fill.fgColor:
                color = self._rgb_to_hex(cell.fill.fgColor.rgb)
                if color:
                    return {
                        "color": color,
                        "pattern": cell.fill.patternType or "solid"
                    }
        except Exception:
            pass
        
        return None
    
    def _get_merge_info(self) -> Dict[tuple, Dict[str, Any]]:
        """获取所有合并单元格信息"""
        merge_map = {}
        
        try:
            for merge_range in self.worksheet.merged_cells.ranges:
                min_row = merge_range.min_row - 1
                max_row = merge_range.max_row - 1
                min_col = merge_range.min_col - 1
                max_col = merge_range.max_col - 1
                
                rowspan = max_row - min_row + 1
                colspan = max_col - min_col + 1
                
                merge_type = "independent"
                if rowspan == 1 and colspan > 1:
                    merge_type = "horizontal"
                elif rowspan > 1 and colspan == 1:
                    merge_type = "vertical"
                elif rowspan > 1 and colspan > 1:
                    merge_type = "region"
                
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        is_master = (r == min_row and c == min_col)
                        merge_map[(r, c)] = {
                            "type": merge_type,
                            "rowspan": rowspan,
                            "colspan": colspan,
                            "is_master": is_master
                        }
        except Exception:
            pass
        
        return merge_map
    
    # TODO: 数据自动填报功能时启用以下三个方法
    # def _load_auto_fill_rules(self) -> List[Dict[str, Any]]:
    #     """从数据库加载启用的自动填充规则"""
    #     rules = []
    #     try:
    #         session = SessionLocal()
    #         try:
    #             from sqlalchemy import text
    #             result = session.execute(
    #                 text("SELECT * FROM template_auto_fill_rules WHERE enabled = true")
    #             )
    #             for row in result:
    #                 rules.append({
    #                     "id": row.id,
    #                     "label_pattern": row.label_pattern,
    #                     "fill_type": row.fill_type,
    #                     "date_format": row.date_format,
    #                     "position": row.position
    #                 })
    #         finally:
    #             session.close()
    #     except Exception as e:
    #         print(f"加载自动填充规则失败: {e}")
    #     
    #     return rules
    # 
    # def _apply_auto_fill_rules(self, cells: List[Dict[str, Any]], rules: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    #     """
    #     应用自动填充规则
    #     
    #     Args:
    #         cells: 单元格列表
    #         rules: 自动填充规则列表
    #     
    #     Returns:
    #         更新后的单元格列表
    #     """
    #     if not rules:
    #         return cells
    #     
    #     cell_map = {(c['row'], c['col']): c for c in cells}
    #     
    #     for rule in rules:
    #         label_pattern = rule.get('label_pattern')
    #         fill_type = rule.get('fill_type', 'current_date')
    #         date_format = rule.get('date_format', 'YYYY年MM月DD日')
    #         position = rule.get('position', 'same_row_next_cell')
    #         
    #         for cell in cells:
    #             cell_value = cell.get('value', '')
    #             
    #             if label_pattern in cell_value:
    #                 current_row = cell['row']
    #                 current_col = cell['col']
    #                 
    #                 if position == 'same_row_next_cell':
    #                     next_col = current_col + 1
    #                     next_cell_key = (current_row, next_col)
    #                     
    #                     if next_cell_key in cell_map:
    #                         next_cell = cell_map[next_cell_key]
    #                         next_value = next_cell.get('value', '').strip()
    #                         
    #                         if not next_value:
    #                             fill_value = self._generate_fill_value(fill_type, date_format)
    #                             next_cell['auto_fill'] = {
    #                                 "type": fill_type,
    #                                 "format": date_format,
    #                                 "value": fill_value,
    #                                 "rule_id": rule.get('id')
    #                             }
    #                             next_cell['value'] = fill_value
    #     
    #     return cells
    # 
    # def _generate_fill_value(self, fill_type: str, date_format: str) -> str:
    #     """
    #     生成填充值
    #     
    #     Args:
    #         fill_type: 填充类型
    #         date_format: 日期格式
    #     
    #     Returns:
    #         填充值
    #     """
    #     now = datetime.now()
    #     
    #     if fill_type == 'current_date':
    #         format_map = {
    #             'YYYY年MM月DD日': f'{now.year}年{now.month:02d}月{now.day:02d}日',
    #             'YYYY年M月D日': f'{now.year}年{now.month}月{now.day}日',
    #             'YYYY-MM-DD': f'{now.year}-{now.month:02d}-{now.day:02d}',
    #             'YYYY/MM/DD': f'{now.year}/{now.month:02d}/{now.day:02d}',
    #             'MM月DD日': f'{now.month:02d}月{now.day:02d}日'
    #         }
    #         return format_map.get(date_format, f'{now.year}年{now.month:02d}月{now.day:02d}日')
    #     
    #     return ''

    
    def extract_all(self) -> Dict[str, Any]:
        """提取完整元数据"""
        if not self.worksheet:
            return {}
        
        try:
            total_rows = int(self.worksheet.max_row)
            total_cols = int(self.worksheet.max_column)
            
            sheet_info = {
                "sheet_name": self.worksheet.title or "Sheet1",
                "total_rows": total_rows,
                "total_cols": total_cols
            }
            
            page_setup = {
                "paper_size": 9,
                "orientation": "portrait",
                "margins": {
                    "top": 1.0,
                    "right": 0.75,
                    "bottom": 1.0,
                    "left": 0.75
                },
                "fit_to_page": False,
                "fit_to_width": 1,
                "fit_to_height": 1,
                "scale": 100,
                "print_area": None,
                "print_title_rows": None,
                "print_title_cols": None,
                "print_grid_lines": False,
                "print_row_col_headers": False,
                "horizontal_centered": False,
                "vertical_centered": False
            }
            
            try:
                if self.worksheet.page_setup:
                    if self.worksheet.page_setup.paperSize:
                        page_setup["paper_size"] = int(self.worksheet.page_setup.paperSize)
                    if self.worksheet.page_setup.orientation:
                        page_setup["orientation"] = self.worksheet.page_setup.orientation
                    if hasattr(self.worksheet.page_setup, 'fitToPage'):
                        page_setup["fit_to_page"] = bool(self.worksheet.page_setup.fitToPage)
                    if hasattr(self.worksheet.page_setup, 'fitToWidth'):
                        page_setup["fit_to_width"] = int(self.worksheet.page_setup.fitToWidth or 1)
                    if hasattr(self.worksheet.page_setup, 'fitToHeight'):
                        page_setup["fit_to_height"] = int(self.worksheet.page_setup.fitToHeight or 1)
                    if hasattr(self.worksheet.page_setup, 'scale'):
                        page_setup["scale"] = int(self.worksheet.page_setup.scale or 100)
                
                if self.worksheet.page_margins:
                    page_setup["margins"] = {
                        "top": float(self.worksheet.page_margins.top or 1.0),
                        "right": float(self.worksheet.page_margins.right or 0.75),
                        "bottom": float(self.worksheet.page_margins.bottom or 1.0),
                        "left": float(self.worksheet.page_margins.left or 0.75)
                    }
                
                if self.worksheet.print_area:
                    page_setup["print_area"] = str(self.worksheet.print_area)
                
                if self.worksheet.print_title_rows:
                    page_setup["print_title_rows"] = str(self.worksheet.print_title_rows)
                
                if self.worksheet.print_title_cols:
                    page_setup["print_title_cols"] = str(self.worksheet.print_title_cols)
                
                if hasattr(self.worksheet, 'print_options'):
                    if hasattr(self.worksheet.print_options, 'gridLines'):
                        page_setup["print_grid_lines"] = bool(self.worksheet.print_options.gridLines)
                    if hasattr(self.worksheet.print_options, 'horizontalCentered'):
                        page_setup["horizontal_centered"] = bool(self.worksheet.print_options.horizontalCentered)
                    if hasattr(self.worksheet.print_options, 'verticalCentered'):
                        page_setup["vertical_centered"] = bool(self.worksheet.print_options.verticalCentered)
                
            except Exception as e:
                print(f"提取页面设置失败: {e}")
            
            row_heights = {}
            for row_idx in range(1, total_rows + 1):
                try:
                    height = self.worksheet.row_dimensions[row_idx].height
                    if height:
                        row_heights[str(row_idx - 1)] = float(height)
                except Exception:
                    pass
            
            col_widths = {}
            for col_idx in range(1, total_cols + 1):
                try:
                    col_letter = get_column_letter(col_idx)
                    width = self.worksheet.column_dimensions[col_letter].width
                    if width:
                        col_widths[str(col_idx - 1)] = float(width * 7)
                except Exception:
                    pass
            
            dimensions = {
                "row_heights": row_heights,
                "col_widths": col_widths
            }
            
            merge_map = self._get_merge_info()
            
            merge_info = []
            for (r, c), info in merge_map.items():
                if info["is_master"]:
                    merge_info.append({
                        "row": r,
                        "col": c,
                        "type": info["type"],
                        "rowspan": info["rowspan"],
                        "colspan": info["colspan"]
                    })
            
            cells = []
            for row_idx in range(total_rows):
                for col_idx in range(total_cols):
                    try:
                        cell = self.worksheet.cell(row=row_idx + 1, column=col_idx + 1)
                        
                        merge_data = merge_map.get((row_idx, col_idx), {
                            "type": "independent",
                            "rowspan": 1,
                            "colspan": 1,
                            "is_master": True
                        })
                        
                        if not merge_data["is_master"]:
                            continue
                        
                        value = ""
                        if cell.value is not None:
                            value = str(cell.value)
                        
                        cell_data = {
                            "row": row_idx,
                            "col": col_idx,
                            "value": value,
                            "is_merged": merge_data["type"] != "independent",
                            "is_master": merge_data["is_master"],
                            "rowspan": merge_data["rowspan"],
                            "colspan": merge_data["colspan"],
                            "font": self._get_font_info(cell),
                            "alignment": self._get_alignment_info(cell),
                            "border": self._get_border_info(
                                cell, 
                                is_merged=merge_data["type"] != "independent",
                                is_master=merge_data["is_master"],
                                rowspan=merge_data["rowspan"],
                                colspan=merge_data["colspan"]
                            ),
                            "fill": self._get_fill_info(cell)
                        }
                        
                        cells.append(cell_data)
                    except Exception as e:
                        print(f"提取单元格({row_idx},{col_idx})失败: {e}")
            
            # TODO: 数据自动填报功能时启用
            # auto_fill_rules = self._load_auto_fill_rules()
            # cells = self._apply_auto_fill_rules(cells, auto_fill_rules)
            
            return {
                "sheet_info": sheet_info,
                "page_setup": page_setup,
                "dimensions": dimensions,
                "merge_info": merge_info,
                "cells": cells,
                "extract_time": datetime.now().isoformat()
            }
            
        except Exception as e:
            print(f"提取元数据失败: {e}")
            import traceback
            traceback.print_exc()
            return {}


def extract_excel_metadata(file_path: str) -> Dict[str, Any]:
    """提取Excel文件完整元数据的便捷函数"""
    extractor = ExcelMetadataExtractor()
    if extractor.load(file_path):
        return extractor.extract_all()
    return {}
