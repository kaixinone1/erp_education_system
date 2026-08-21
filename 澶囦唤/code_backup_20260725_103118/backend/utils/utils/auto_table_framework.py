#!/usr/bin/env python3
"""
自动表管理框架 - 零配置方案
直接从数据库表结构读取字段信息，无需配置文件
"""
import psycopg2
from psycopg2.extras import RealDictCursor
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from typing import Dict, List, Any, Optional
from datetime import datetime, date
import hashlib
import json
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 数据库配置
DATABASE_CONFIG = {
    "host": "localhost",
    "port": "5432",
    "database": "taiping_education",
    "user": "taiping_user",
    "password": "taiping_password"
}


def get_db_connection():
    """获取数据库连接"""
    return psycopg2.connect(**DATABASE_CONFIG)


class AutoTableManager:
    """自动表管理器 - 零配置"""
    
    def __init__(self, table_name: str):
        self.table_name = table_name
        self._schema = None
    
    def get_schema(self) -> List[Dict[str, Any]]:
        """从数据库读取表结构"""
        if self._schema is not None:
            return self._schema
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # 查询表结构
            cursor.execute("""
                SELECT 
                    column_name,
                    data_type,
                    is_nullable,
                    character_maximum_length,
                    column_default
                FROM information_schema.columns
                WHERE table_name = %s
                ORDER BY ordinal_position
            """, (self.table_name,))
            
            rows = cursor.fetchall()
            
            schema = []
            for row in rows:
                col_name = row[0]
                data_type = row[1]
                is_nullable = row[2] == 'YES'
                max_length = row[3]
                
                # 跳过系统字段
                if col_name in ['id', 'created_at', 'updated_at']:
                    continue
                
                # 映射字段类型
                field_type = self._map_data_type(data_type)
                
                schema.append({
                    'name': col_name,
                    'type': field_type,
                    'length': max_length,
                    'nullable': is_nullable,
                    'db_type': data_type
                })
            
            self._schema = schema
            return schema
            
        finally:
            cursor.close()
            conn.close()
    
    def _map_data_type(self, db_type: str) -> str:
        """映射数据库类型到前端类型"""
        type_mapping = {
            'character varying': 'VARCHAR',
            'text': 'TEXT',
            'integer': 'INTEGER',
            'bigint': 'INTEGER',
            'numeric': 'DECIMAL',
            'decimal': 'DECIMAL',
            'double precision': 'DECIMAL',
            'date': 'DATE',
            'timestamp without time zone': 'DATETIME',
            'timestamp with time zone': 'DATETIME',
            'boolean': 'BOOLEAN'
        }
        return type_mapping.get(db_type, 'VARCHAR')
    
    def get_data(self, filters: Dict = None, page: int = 1, page_size: int = 20) -> Dict:
        """获取数据"""
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        try:
            # 动态获取表中实际存在的列名
            cursor.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = %s
                ORDER BY ordinal_position
            """, (self.table_name,))
            db_columns = [row['column_name'] for row in cursor.fetchall()]
            
            # 构建字段列表（只选择表中实际存在的列）
            field_names = [f'"{col}"' for col in db_columns]
            
            # 构建查询
            sql = f"SELECT {', '.join(field_names)} FROM {self.table_name} WHERE 1=1"
            params = []
            
            if filters:
                for key, value in filters.items():
                    if key in db_columns:
                        sql += f" AND \"{key}\" = %s"
                        params.append(value)
            
            # 获取总数
            count_sql = f"SELECT COUNT(*) FROM {self.table_name} WHERE 1=1"
            if filters:
                for key, value in filters.items():
                    if key in db_columns:
                        count_sql += f" AND \"{key}\" = %s"
            
            cursor.execute(count_sql, params)
            count_row = cursor.fetchone()
            if count_row:
                total = list(count_row.values())[0]
            else:
                total = 0
            
            # 分页 - 使用 id 排序（所有表都有 id 列）
            order_col = 'updated_at' if 'updated_at' in db_columns else 'id'
            sql += f" ORDER BY \"{order_col}\" DESC"
            sql += f" LIMIT {page_size} OFFSET {(page - 1) * page_size}"
            
            cursor.execute(sql, params)
            rows = cursor.fetchall()
            
            return {
                'data': [dict(row) for row in rows],
                'total': total,
                'page': page,
                'page_size': page_size
            }
            
        finally:
            cursor.close()
            conn.close()
    
    def get_by_teacher_id(self, record_id: int) -> Optional[Dict]:
        """根据ID获取单条数据（record_id 可以是 teacher_id 或 id）"""
        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        try:
            # 动态获取表中实际存在的列名
            cursor.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = %s
            """, (self.table_name,))
            db_columns = [row['column_name'] for row in cursor.fetchall()]
            field_names = [f'"{col}"' for col in db_columns]
            
            where_col = 'teacher_id' if 'teacher_id' in db_columns else 'id'
            
            cursor.execute(f"""
                SELECT {', '.join(field_names)}
                FROM {self.table_name}
                WHERE "{where_col}" = %s
            """, (record_id,))
            
            row = cursor.fetchone()
            return dict(row) if row else None
            
        finally:
            cursor.close()
            conn.close()
    
    def update_data(self, record_id: int, data: Dict) -> bool:
        """更新数据（record_id 可以是 teacher_id 或 id）"""
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # 动态获取表中实际存在的列名
            cursor.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = %s
            """, (self.table_name,))
            db_columns = {row[0] for row in cursor.fetchall()}
            
            updates = []
            values = []
            
            for key, value in data.items():
                if key in db_columns and key not in ['id', 'teacher_id']:
                    updates.append(f'"{key}" = %s')
                    values.append(value)
            
            # 如果表有 updated_at 列，自动更新
            if 'updated_at' in db_columns:
                updates.append('"updated_at" = NOW()')
            
            if not updates:
                return True
            
            values.append(record_id)
            
            # 使用合适的 WHERE 条件
            where_col = 'teacher_id' if 'teacher_id' in db_columns else 'id'
            
            sql = f"""
                UPDATE {self.table_name}
                SET {', '.join(updates)}
                WHERE "{where_col}" = %s
            """
            
            cursor.execute(sql, values)
            conn.commit()
            return cursor.rowcount > 0
            
        except Exception as e:
            conn.rollback()
            print(f"更新失败: {e}")
            return False
        finally:
            cursor.close()
            conn.close()
    
    def create_data(self, data: Dict) -> Optional[int]:
        """创建新数据，返回新记录的ID"""
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # 获取表结构，只插入存在的字段
            schema = self.get_schema()
            valid_fields = {f['name'] for f in schema}
            
            columns = []
            placeholders = []
            values = []
            
            for key, value in data.items():
                if key in valid_fields and key not in ['id', 'teacher_id']:
                    columns.append(f'"{key}"')
                    placeholders.append('%s')
                    values.append(value)
            
            if not columns:
                return None
            
            sql = f"""
                INSERT INTO {self.table_name} ({', '.join(columns)})
                VALUES ({', '.join(placeholders)})
                RETURNING id
            """
            
            cursor.execute(sql, values)
            new_id = cursor.fetchone()[0]
            conn.commit()
            return new_id
            
        except Exception as e:
            conn.rollback()
            print(f"创建失败: {e}")
            return None
        finally:
            cursor.close()
            conn.close()
    
    def delete_data(self, record_id: int) -> bool:
        """删除数据（record_id 可以是 teacher_id 或 id）"""
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # 动态获取表中实际存在的列名
            cursor.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = %s
            """, (self.table_name,))
            db_columns = {row[0] for row in cursor.fetchall()}
            
            where_col = 'teacher_id' if 'teacher_id' in db_columns else 'id'
            
            cursor.execute(f"""
                DELETE FROM {self.table_name}
                WHERE "{where_col}" = %s
            """, (record_id,))
            
            conn.commit()
            return cursor.rowcount > 0
            
        except Exception as e:
            conn.rollback()
            print(f"删除失败: {e}")
            return False
        finally:
            cursor.close()
            conn.close()
    
    def get_table_fingerprint(self) -> str:
        """生成表指纹：MD5(表名 + 排序后的字段名列表 + 字段数量)"""
        schema = self.get_schema()
        field_names = sorted([f['name'] for f in schema if f['name'] not in ['id', 'created_at', 'updated_at']])
        fingerprint_str = f"{self.table_name}|{','.join(field_names)}|{len(field_names)}"
        return hashlib.md5(fingerprint_str.encode('utf-8')).hexdigest()
    
    def export_template(self) -> io.BytesIO:
        """导出Excel模板，包含表指纹"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "数据模板"
        
        schema = self.get_schema()
        # 排除系统字段
        business_fields = [f for f in schema if f['name'] not in ['id', 'created_at', 'updated_at']]
        
        if not business_fields:
            ws['A1'] = '无业务字段'
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        
        # 样式定义
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头（中文名）
        for col_idx, field in enumerate(business_fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field['name'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 根据数据类型设置列宽
        for col_idx, field in enumerate(business_fields, 1):
            col_letter = get_column_letter(col_idx)
            data_type = field.get('data_type', '').upper()
            if 'VARCHAR' in data_type or 'TEXT' in data_type:
                ws.column_dimensions[col_letter].width = 20
            elif 'INT' in data_type or 'NUMERIC' in data_type or 'DECIMAL' in data_type:
                ws.column_dimensions[col_letter].width = 15
            elif 'DATE' in data_type or 'TIMESTAMP' in data_type:
                ws.column_dimensions[col_letter].width = 18
            else:
                ws.column_dimensions[col_letter].width = 18
        
        # 添加数据验证行（示例数据行）
        for col_idx, field in enumerate(business_fields, 1):
            cell = ws.cell(row=2, column=col_idx)
            data_type = field.get('data_type', '').upper()
            if 'INT' in data_type or 'NUMERIC' in data_type or 'DECIMAL' in data_type:
                cell.value = 0
            elif 'DATE' in data_type:
                cell.value = '2024-01-01'
            else:
                cell.value = ''
            cell.border = thin_border
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 隐藏Sheet：存储表指纹
        ws_fingerprint = wb.create_sheet("__表指纹__")
        ws_fingerprint.sheet_state = 'hidden'
        ws_fingerprint['A1'] = 'table_name'
        ws_fingerprint['B1'] = self.table_name
        ws_fingerprint['A2'] = 'fingerprint'
        ws_fingerprint['B2'] = self.get_table_fingerprint()
        ws_fingerprint['A3'] = 'field_count'
        ws_fingerprint['B3'] = len(business_fields)
        ws_fingerprint['A4'] = 'fields'
        ws_fingerprint['B4'] = json.dumps([f['name'] for f in business_fields], ensure_ascii=False)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def import_data_with_validation(
        self, 
        file_content: bytes,
        duplicate_strategy: str = 'skip'  # 'skip', 'overwrite', 'cancel'
    ) -> Dict[str, Any]:
        """
        导入数据并进行四层验证+重复检测
        返回: {status, message, inserted_count, skipped_count, errors, duplicate_records}
        """
        from openpyxl import load_workbook
        
        try:
            wb = load_workbook(io.BytesIO(file_content), data_only=True)
            
            # ========== 第四层：表指纹验证 ==========
            if '__表指纹__' in wb.sheetnames:
                fp_ws = wb['__表指纹__']
                stored_table_name = fp_ws['B1'].value
                stored_fingerprint = fp_ws['B2'].value
                current_fingerprint = self.get_table_fingerprint()
                
                if stored_table_name != self.table_name:
                    return {
                        "status": "error",
                        "message": f"表名不匹配！模板表名：【{stored_table_name}】，当前表名：【{self.table_name}】。请确认选择了正确的文件。",
                        "inserted_count": 0,
                        "skipped_count": 0,
                        "errors": [],
                        "duplicate_records": []
                    }
                
                if stored_fingerprint != current_fingerprint:
                    # 显示详细差异
                    stored_fields_str = fp_ws['B4'].value or ''
                    try:
                        stored_fields = json.loads(stored_fields_str) if stored_fields_str else []
                    except:
                        stored_fields = []
                    
                    current_fields = [f['name'] for f in self.get_schema() if f['name'] not in ['id', 'created_at', 'updated_at']]
                    
                    missing_in_file = set(current_fields) - set(stored_fields)
                    extra_in_file = set(stored_fields) - set(current_fields)
                    
                    diff_msg = []
                    if missing_in_file:
                        diff_msg.append(f"文件中缺少字段：{', '.join(missing_in_file)}")
                    if extra_in_file:
                        diff_msg.append(f"文件中多出字段：{', '.join(extra_in_file)}")
                    
                    return {
                        "status": "error",
                        "message": f"表指纹不匹配！文件不属于当前表。{'；'.join(diff_msg)}",
                        "inserted_count": 0,
                        "skipped_count": 0,
                        "errors": [],
                        "duplicate_records": []
                    }
            else:
                return {
                    "status": "error",
                    "message": "文件不是有效的模板文件（缺少表指纹）。请使用本系统导出的模板文件。",
                    "inserted_count": 0,
                    "skipped_count": 0,
                    "errors": [],
                    "duplicate_records": []
                }
            
            # 读取数据Sheet
            ws = wb['数据模板']
            
            # ========== 第一层：字段名匹配验证 ==========
            # 读取表头行
            file_headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    file_headers.append(str(header).strip())
            
            schema = self.get_schema()
            business_fields = [f for f in schema if f['name'] not in ['id', 'created_at', 'updated_at']]
            current_field_names = [f['name'] for f in business_fields]
            
            # 验证字段名匹配
            if file_headers != current_field_names:
                missing = set(current_field_names) - set(file_headers)
                extra = set(file_headers) - set(current_field_names)
                error_parts = []
                if missing:
                    error_parts.append(f"缺少字段：{', '.join(missing)}")
                if extra:
                    error_parts.append(f"多余字段：{', '.join(extra)}")
                return {
                    "status": "error",
                    "message": f"字段名不匹配！{'；'.join(error_parts)}",
                    "inserted_count": 0,
                    "skipped_count": 0,
                    "errors": [],
                    "duplicate_records": []
                }
            
            # ========== 第二层：字段数量验证 ==========
            if len(file_headers) != len(current_field_names):
                return {
                    "status": "error",
                    "message": f"字段数量不匹配！文件：{len(file_headers)}列，当前表：{len(current_field_names)}列",
                    "inserted_count": 0,
                    "skipped_count": 0,
                    "errors": [],
                    "duplicate_records": []
                }
            
            # 构建字段名到数据类型的映射
            field_type_map = {f['name']: f.get('data_type', 'VARCHAR').upper() for f in business_fields}
            
            # 查找业务唯一键（用于重复检测）
            unique_key_candidates = ['身份证号码', '身份证号', '工号', '编号', '学号', '档案号', '证件号码']
            unique_key = None
            for candidate in unique_key_candidates:
                if candidate in current_field_names:
                    unique_key = candidate
                    break
            
            # ========== 第三层：数据类型验证 + 读取数据 ==========
            errors = []
            valid_rows = []
            
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                row_has_error = False
                
                for col_idx, field_name in enumerate(file_headers, 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    
                    # 跳过完全空行
                    if cell_value is None:
                        row_data[field_name] = None
                        continue
                    
                    data_type = field_type_map.get(field_name, 'VARCHAR')
                    
                    # 数据类型验证
                    try:
                        if 'INT' in data_type or 'NUMERIC' in data_type or 'DECIMAL' in data_type:
                            if cell_value == '' or cell_value is None:
                                row_data[field_name] = None
                            else:
                                float_val = float(str(cell_value).replace(',', ''))
                                if 'INT' in data_type:
                                    row_data[field_name] = int(float_val)
                                else:
                                    row_data[field_name] = float_val
                        elif 'DATE' in data_type or 'TIMESTAMP' in data_type:
                            if isinstance(cell_value, datetime):
                                row_data[field_name] = cell_value.strftime('%Y-%m-%d')
                            elif isinstance(cell_value, date):
                                row_data[field_name] = cell_value.strftime('%Y-%m-%d')
                            elif str(cell_value).strip():
                                # 尝试解析日期字符串
                                date_str = str(cell_value).strip()
                                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日', '%Y.%m.%d']:
                                    try:
                                        parsed = datetime.strptime(date_str, fmt)
                                        row_data[field_name] = parsed.strftime('%Y-%m-%d')
                                        break
                                    except:
                                        continue
                                if field_name not in row_data:
                                    errors.append(f"第{row_idx}行，字段【{field_name}】日期格式不正确：{cell_value}")
                                    row_has_error = True
                            else:
                                row_data[field_name] = None
                        else:
                            row_data[field_name] = str(cell_value).strip()
                    except Exception as e:
                        errors.append(f"第{row_idx}行，字段【{field_name}】数据类型错误：{cell_value}（{str(e)}）")
                        row_has_error = True
                
                if not row_has_error:
                    # 检查是否全为空行
                    if any(v is not None and str(v).strip() != '' for v in row_data.values()):
                        valid_rows.append(row_data)
            
            if not valid_rows:
                return {
                    "status": "error",
                    "message": "文件中没有有效数据行",
                    "inserted_count": 0,
                    "skipped_count": 0,
                    "errors": errors,
                    "duplicate_records": []
                }
            
            # ========== 重复检测 ==========
            duplicate_records = []
            if unique_key:
                conn = get_db_connection()
                cursor = conn.cursor(cursor_factory=RealDictCursor)
                try:
                    for row_data in valid_rows:
                        key_value = row_data.get(unique_key)
                        if key_value:
                            cursor.execute(
                                f'SELECT * FROM {self.table_name} WHERE "{unique_key}" = %s',
                                (str(key_value),)
                            )
                            existing = cursor.fetchone()
                            if existing:
                                duplicate_records.append({
                                    "unique_key": unique_key,
                                    "key_value": str(key_value),
                                    "existing_id": existing.get('id'),
                                    "row_data": {k: str(v) if v else '' for k, v in row_data.items()}
                                })
                finally:
                    cursor.close()
                    conn.close()
                
                if duplicate_records:
                    if duplicate_strategy == 'cancel':
                        return {
                            "status": "duplicate_found",
                            "message": f"发现 {len(duplicate_records)} 条重复记录（{unique_key}重复）",
                            "inserted_count": 0,
                            "skipped_count": 0,
                            "errors": errors,
                            "duplicate_records": duplicate_records,
                            "unique_key": unique_key
                        }
            
            # ========== 插入数据 ==========
            conn = get_db_connection()
            cursor = conn.cursor()
            inserted_count = 0
            skipped_count = 0
            
            try:
                # 获取实际列名
                cursor.execute("""
                    SELECT column_name FROM information_schema.columns
                    WHERE table_name = %s
                """, (self.table_name,))
                db_columns = {row[0] for row in cursor.fetchall()}
                
                for row_data in valid_rows:
                    # 检查是否重复
                    if unique_key and duplicate_strategy == 'skip':
                        key_value = row_data.get(unique_key)
                        if key_value:
                            cursor.execute(
                                f'SELECT id FROM {self.table_name} WHERE "{unique_key}" = %s',
                                (str(key_value),)
                            )
                            if cursor.fetchone():
                                skipped_count += 1
                                continue
                    
                    if unique_key and duplicate_strategy == 'overwrite':
                        key_value = row_data.get(unique_key)
                        if key_value:
                            # 先删除旧记录
                            cursor.execute(
                                f'DELETE FROM {self.table_name} WHERE "{unique_key}" = %s',
                                (str(key_value),)
                            )
                    
                    # 构建插入语句
                    columns = []
                    placeholders = []
                    values = []
                    
                    for key, value in row_data.items():
                        if key in db_columns and key not in ['id']:
                            columns.append(f'"{key}"')
                            placeholders.append('%s')
                            values.append(value)
                    
                    if columns:
                        sql = f"""
                            INSERT INTO {self.table_name} ({', '.join(columns)})
                            VALUES ({', '.join(placeholders)})
                        """
                        cursor.execute(sql, values)
                        inserted_count += 1
                
                conn.commit()
                
                return {
                    "status": "success",
                    "message": f"成功导入 {inserted_count} 条数据，跳过 {skipped_count} 条重复",
                    "inserted_count": inserted_count,
                    "skipped_count": skipped_count,
                    "errors": errors,
                    "duplicate_records": []
                }
                
            except Exception as e:
                conn.rollback()
                return {
                    "status": "error",
                    "message": f"数据插入失败: {str(e)}",
                    "inserted_count": 0,
                    "skipped_count": 0,
                    "errors": errors,
                    "duplicate_records": []
                }
            finally:
                cursor.close()
                conn.close()
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {
                "status": "error",
                "message": f"导入处理失败: {str(e)}",
                "inserted_count": 0,
                "skipped_count": 0,
                "errors": [],
                "duplicate_records": []
            }


# 全局表管理器缓存
_table_managers: Dict[str, AutoTableManager] = {}


def get_table_manager(table_name: str) -> AutoTableManager:
    """获取表管理器（带缓存）"""
    if table_name not in _table_managers:
        _table_managers[table_name] = AutoTableManager(table_name)
    return _table_managers[table_name]


def load_table_translation(table_name: str) -> dict:
    """加载表的中文翻译"""
    import json
    import os
    
    config_path = os.path.join(os.path.dirname(__file__), '..', 'config', 'table_translations.json')
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                translations = json.load(f)
                return translations.get('tables', {}).get(table_name, {})
    except Exception as e:
        print(f"加载翻译文件失败: {e}")
    
    return {}


def create_auto_table_routes(table_name: str) -> APIRouter:
    """为表创建自动API路由"""
    router = APIRouter(prefix=f"/api/auto-table/{table_name}")
    manager = get_table_manager(table_name)
    translation = load_table_translation(table_name)
    
    # 获取表结构
    @router.get("/schema")
    async def get_schema():
        """获取表结构"""
        schema = manager.get_schema()
        field_translations = translation.get('fields', {})
        
        # 创建字段映射
        schema_map = {field['name']: field for field in schema}
        
        # 按照翻译文件中定义的顺序排列字段
        ordered_schema = []
        for field_name in field_translations.keys():
            if field_name in schema_map:
                field = schema_map[field_name]
                field['label'] = field_translations[field_name]
                ordered_schema.append(field)
        
        # 添加翻译文件中没有的字段（如果有的话）
        for field in schema:
            if field['name'] not in field_translations:
                field['label'] = field['name']
                ordered_schema.append(field)
        
        return {
            "status": "success",
            "data": {
                "table_name": table_name,
                "chinese_name": translation.get('chinese_name', table_name),
                "description": translation.get('description', ''),
                "fields": ordered_schema
            }
        }
    
    # 列表查询
    @router.get("/list")
    async def list_data(
        page: int = 1,
        page_size: int = 20,
        teacher_id: int = None
    ):
        """获取数据列表"""
        filters = {}
        if teacher_id:
            filters['teacher_id'] = teacher_id
        
        result = manager.get_data(filters=filters if filters else None, page=page, page_size=page_size)
        
        return {
            "status": "success",
            **result
        }
    
    # 获取单条
    @router.get("/detail/{teacher_id}")
    async def get_detail(teacher_id: int):
        """获取单条数据详情"""
        try:
            data = manager.get_by_teacher_id(teacher_id)
            if not data:
                # 返回空数据而不是404错误，让前端知道数据不存在但API调用成功
                return {
                    "status": "success",
                    "data": {},
                    "message": "数据不存在，请先汇集数据"
                }
            
            return {
                "status": "success",
                "data": data
            }
        except Exception as e:
            print(f"获取详情失败: {e}")
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"获取详情失败: {str(e)}")
    
    # 更新
    @router.put("/update/{teacher_id}")
    async def update_data(teacher_id: int, data: Dict[str, Any]):
        """更新数据"""
        success = manager.update_data(teacher_id, data)
        if success:
            return {"status": "success", "message": "更新成功"}
        else:
            raise HTTPException(status_code=500, detail="更新失败")
    
    # 删除
    @router.delete("/delete/{teacher_id}")
    async def delete_data(teacher_id: int):
        """删除数据"""
        success = manager.delete_data(teacher_id)
        if success:
            return {"status": "success", "message": "删除成功"}
        else:
            raise HTTPException(status_code=500, detail="删除失败")
    
    # 计算工作年限（退休表专用）
    if table_name == 'retirement_report_data':
        @router.post("/calculate/{teacher_id}")
        async def calculate(teacher_id: int):
            """计算退休信息"""
            from utils.retirement_calculator import calculate_retirement_info
            
            data = manager.get_by_teacher_id(teacher_id)
            if not data:
                raise HTTPException(status_code=404, detail="数据不存在")
            
            birth_date = data.get('出生日期')
            gender = data.get('性别')
            personal_identity = data.get('个人身份', '干部')
            work_start_date = data.get('参加工作时间')
            
            if not birth_date or not gender or not work_start_date:
                raise HTTPException(status_code=400, detail="缺少必要的计算参数")
            
            # 转换日期
            if isinstance(birth_date, str):
                birth_date = datetime.strptime(birth_date, '%Y-%m-%d').date()
            if isinstance(work_start_date, str):
                work_start_date = datetime.strptime(work_start_date, '%Y-%m-%d').date()
            
            result = calculate_retirement_info(
                birth_date=birth_date,
                gender=gender,
                personal_identity=personal_identity,
                work_start_date=work_start_date
            )
            
            return {
                "status": "success",
                "data": {
                    "original_retirement_date": str(result['original_retirement_date']),
                    "delay_months": result['delay_months'],
                    "calculated_retirement_date": str(result['calculated_retirement_date']),
                    "actual_retirement_date": str(result['actual_retirement_date']),
                    "work_years": result['work_years']
                }
            }
        
        @router.post("/save-calculation/{teacher_id}")
        async def save_calculation(teacher_id: int, data: Dict[str, Any]):
            """保存计算结果"""
            retirement_date = data.get('retirement_date')
            work_years = data.get('work_years')
            
            if not retirement_date or work_years is None:
                raise HTTPException(status_code=400, detail="缺少退休日期或工作年限")
            
            success = manager.update_data(teacher_id, {
                '退休时间': retirement_date,
                '工作年限': work_years
            })
            
            if success:
                return {"status": "success", "message": "保存成功"}
            else:
                raise HTTPException(status_code=500, detail="保存失败")
    
    return router


# 创建通用动态路由
from fastapi import Request

def create_dynamic_auto_table_router() -> APIRouter:
    """创建通用动态路由处理器 - 支持任意表名"""
    router = APIRouter(prefix="/api/auto-table")
    
    @router.get("/{table_name}/schema")
    async def dynamic_schema(table_name: str):
        """动态获取任意表的结构"""
        try:
            manager = get_table_manager(table_name)
            schema = manager.get_schema()
            translation = load_table_translation(table_name)
            field_translations = translation.get('fields', {})
            
            # 创建字段映射
            schema_map = {field['name']: field for field in schema}
            
            # 按照翻译文件中定义的顺序排列字段
            ordered_schema = []
            for field_name in field_translations.keys():
                if field_name in schema_map:
                    field = schema_map[field_name]
                    field['label'] = field_translations[field_name]
                    ordered_schema.append(field)
            
            # 添加翻译文件中没有的字段
            for field in schema:
                if field['name'] not in field_translations:
                    field['label'] = field['name']
                    ordered_schema.append(field)
            
            return {
                "status": "success",
                "data": {
                    "table_name": table_name,
                    "chinese_name": translation.get('chinese_name', table_name),
                    "description": translation.get('description', ''),
                    "fields": ordered_schema
                }
            }
        except Exception as e:
            raise HTTPException(status_code=404, detail=f"表不存在或无法访问: {e}")
    
    @router.get("/{table_name}/list")
    async def dynamic_list(
        table_name: str,
        page: int = 1,
        page_size: int = 20,
        teacher_id: int = None
    ):
        """动态获取任意表的数据列表"""
        try:
            manager = get_table_manager(table_name)
            filters = {}
            if teacher_id:
                filters['teacher_id'] = teacher_id
            
            result = manager.get_data(filters=filters if filters else None, page=page, page_size=page_size)
            
            return {
                "status": "success",
                **result
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"获取数据失败: {e}")
    
    @router.get("/{table_name}/detail/{teacher_id}")
    async def dynamic_detail(table_name: str, teacher_id: int):
        """动态获取单条数据详情"""
        try:
            manager = get_table_manager(table_name)
            data = manager.get_by_teacher_id(teacher_id)
            if not data:
                raise HTTPException(status_code=404, detail="数据不存在")
            
            return {
                "status": "success",
                "data": data
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"获取详情失败: {e}")
    
    @router.post("/{table_name}/create")
    async def dynamic_create(table_name: str, data: Dict[str, Any]):
        """动态创建数据"""
        try:
            manager = get_table_manager(table_name)
            new_id = manager.create_data(data)
            if new_id:
                return {"status": "success", "message": "创建成功", "id": new_id}
            else:
                raise HTTPException(status_code=500, detail="创建失败")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"创建失败: {e}")
    
    @router.put("/{table_name}/update/{teacher_id}")
    async def dynamic_update(table_name: str, teacher_id: int, data: Dict[str, Any]):
        """动态更新数据"""
        try:
            manager = get_table_manager(table_name)
            success = manager.update_data(teacher_id, data)
            if success:
                return {"status": "success", "message": "更新成功"}
            else:
                raise HTTPException(status_code=500, detail="更新失败")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"更新失败: {e}")
    
    @router.delete("/{table_name}/delete/{teacher_id}")
    async def dynamic_delete(table_name: str, teacher_id: int):
        """动态删除数据"""
        try:
            manager = get_table_manager(table_name)
            success = manager.delete_data(teacher_id)
            if success:
                return {"status": "success", "message": "删除成功"}
            else:
                raise HTTPException(status_code=500, detail="删除失败")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"删除失败: {e}")
    
    # ========== 导入导出功能 ==========
    
    @router.get("/{table_name}/export-template")
    async def dynamic_export_template(table_name: str):
        """导出Excel模板（含表指纹）"""
        try:
            from datetime import datetime as dt
            
            manager = get_table_manager(table_name)
            template_bytes = manager.export_template()
            
            # 文件名：表名+模板+日期
            today = dt.now().strftime('%Y年%m月%d日')
            filename = f"{table_name}_模板_{today}.xlsx"
            # URL编码中文文件名
            from urllib.parse import quote
            encoded_filename = quote(filename)
            
            return StreamingResponse(
                template_bytes,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
                }
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"导出模板失败: {e}")
    
    @router.post("/{table_name}/import-data")
    async def dynamic_import_data(table_name: str, request: Request):
        """导入数据（含四层验证+重复检测）"""
        try:
            # 读取上传的文件
            form = await request.form()
            file = form.get('file')
            duplicate_strategy = form.get('duplicate_strategy', 'skip')
            
            if not file:
                raise HTTPException(status_code=400, detail="未上传文件")
            
            file_content = await file.read()
            
            manager = get_table_manager(table_name)
            result = manager.import_data_with_validation(
                file_content=file_content,
                duplicate_strategy=duplicate_strategy
            )
            
            return result
            
        except HTTPException:
            raise
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=f"导入数据失败: {e}")
    
    return router
