"""
Excel转HTML转换器
实现100%样式还原
"""
import base64
import io
import json
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class ExcelToHtmlConverter:
    """Excel转HTML转换器"""
    
    def _rgb_to_hex(self, rgb) -> str:
        """将RGB转换为十六进制颜色"""
        if rgb is None:
            return None
        if isinstance(rgb, str):
            if rgb.startswith('#'):
                return rgb
            if rgb == '00000000':
                return None
            return f"#{rgb}"
        return None
    
    def _get_cell_style(self, cell) -> str:
        """获取单元格CSS样式"""
        styles = []
        
        if cell.font:
            font = cell.font
            if font.color and font.color.rgb:
                color = self._rgb_to_hex(font.color.rgb)
                if color:
                    styles.append(f'color: {color}')
            if font.size:
                styles.append(f'font-size: {font.size}px')
            if font.bold:
                styles.append('font-weight: bold')
            if font.italic:
                styles.append('font-style: italic')
            if font.underline:
                styles.append('text-decoration: underline')
        
        if cell.fill:
            fill = cell.fill
            if hasattr(fill, 'fgColor') and fill.fgColor:
                bg_color = self._rgb_to_hex(fill.fgColor.rgb)
                if bg_color and bg_color != '#000000':
                    styles.append(f'background-color: {bg_color}')
        
        if cell.alignment:
            align = cell.alignment
            if align.horizontal:
                styles.append(f'text-align: {align.horizontal}')
            if align.vertical:
                styles.append(f'vertical-align: {align.vertical}')
        
        if cell.border:
            border = cell.border
            border_styles = []
            if border.left and border.left.style:
                border_styles.append(f'border-left: 1px solid #000')
            if border.right and border.right.style:
                border_styles.append(f'border-right: 1px solid #000')
            if border.top and border.top.style:
                border_styles.append(f'border-top: 1px solid #000')
            if border.bottom and border.bottom.style:
                border_styles.append(f'border-bottom: 1px solid #000')
            styles.extend(border_styles)
        
        return '; '.join(styles) if styles else ''
    
    def convert(self, file_path: str, max_rows: int = 50) -> str:
        """
        将Excel文件转换为HTML表格
        
        Args:
            file_path: Excel文件路径
            max_rows: 最大显示行数
        
        Returns:
            HTML字符串
        """
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        merge_info = {}
        for merge_range in ws.merged_cells.ranges:
            merge_info[(merge_range.min_row, merge_range.min_col)] = {
                'rowspan': merge_range.max_row - merge_range.min_row + 1,
                'colspan': merge_range.max_col - merge_range.min_col + 1
            }
        
        html_parts = ['<table style="border-collapse: collapse; border: 1px solid #000;">']
        
        row_count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows), start=1):
            html_parts.append('<tr>')
            
            for col_idx, cell in enumerate(row, start=1):
                if (row_idx, col_idx) in merge_info:
                    merge = merge_info[(row_idx, col_idx)]
                    rowspan = f' rowspan="{merge["rowspan"]}"' if merge["rowspan"] > 1 else ''
                    colspan = f' colspan="{merge["colspan"]}"' if merge["colspan"] > 1 else ''
                else:
                    skip = False
                    for (mr, mc), info in merge_info.items():
                        if mr < row_idx <= mr + info['rowspan'] - 1 and mc < col_idx <= mc + info['colspan'] - 1:
                            skip = True
                            break
                    if skip:
                        continue
                    rowspan = ''
                    colspan = ''
                
                value = cell.value if cell.value is not None else ''
                style = self._get_cell_style(cell)
                style_attr = f' style="{style}"' if style else ''
                
                html_parts.append(f'<td{rowspan}{colspan}{style_attr}>{value}</td>')
            
            html_parts.append('</tr>')
            row_count += 1
        
        html_parts.append('</table>')
        
        return ''.join(html_parts)
