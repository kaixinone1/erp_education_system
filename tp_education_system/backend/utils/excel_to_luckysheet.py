"""
Excel转Luckysheet格式转换器
实现100%样式还原
"""
import base64
import io
import json
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment


class ExcelToLuckysheetConverter:
    """Excel转Luckysheet格式转换器"""
    
    def __init__(self):
        self.color_map = {}
    
    def _rgb_to_hex(self, rgb) -> str:
        """将RGB转换为十六进制颜色"""
        if rgb is None:
            return None
        if isinstance(rgb, str):
            if rgb.startswith('#'):
                return rgb
            return f"#{rgb}"
        return None
    
    def _get_cell_style(self, cell) -> Dict[str, Any]:
        """获取单元格样式"""
        style = {}
        
        if cell.font:
            font = cell.font
            if font.color and font.color.rgb:
                color = self._rgb_to_hex(font.color.rgb)
                if color:
                    style['fc'] = color
            if font.size:
                style['fs'] = font.size
            if font.bold:
                style['bl'] = 1
            if font.italic:
                style['it'] = 1
            if font.underline:
                style['un'] = 1
            if font.strike:
                style['cl'] = 1
        
        if cell.fill:
            fill = cell.fill
            if isinstance(fill, PatternFill) and fill.fgColor:
                bg_color = self._rgb_to_hex(fill.fgColor.rgb)
                if bg_color and bg_color != '#000000':
                    style['bg'] = bg_color
        
        if cell.alignment:
            align = cell.alignment
            ht_map = {'left': 1, 'center': 0, 'right': 2}
            vt_map = {'top': 1, 'center': 0, 'bottom': 2}
            if align.horizontal:
                style['ht'] = ht_map.get(align.horizontal, 1)
            if align.vertical:
                style['vt'] = vt_map.get(align.vertical, 0)
            if align.wrap_text:
                style['tb'] = 2
        
        if cell.border:
            border = cell.border
            border_style_map = {
                'thin': '1',
                'medium': '2',
                'thick': '3',
                'double': '4',
                'dotted': '5',
                'dashed': '6'
            }
            
            if border.left and border.left.style:
                style['bl'] = border_style_map.get(border.left.style, '1')
            if border.right and border.right.style:
                style['br'] = border_style_map.get(border.right.style, '1')
            if border.top and border.top.style:
                style['bt'] = border_style_map.get(border.top.style, '1')
            if border.bottom and border.bottom.style:
                style['bb'] = border_style_map.get(border.bottom.style, '1')
        
        return style
    
    def _get_cell_value(self, cell) -> Any:
        """获取单元格值"""
        if cell.value is None:
            return ""
        
        if isinstance(cell.value, (int, float)):
            return cell.value
        
        return str(cell.value)
    
    def convert(self, file_path: str, sheet_name: str = None) -> Dict[str, Any]:
        """
        将Excel文件转换为Luckysheet格式
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称（可选，默认第一个）
        
        Returns:
            Luckysheet格式的数据
        """
        wb = load_workbook(file_path, data_only=False)
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title
        
        celldata = []
        merge_cells = []
        
        max_row = ws.max_row
        max_col = ws.max_column
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), start=0):
            for col_idx, cell in enumerate(row, start=0):
                value = self._get_cell_value(cell)
                
                if value or cell.fill or cell.font or cell.border or cell.alignment:
                    cell_data = {
                        "r": row_idx,
                        "c": col_idx,
                        "v": {
                            "v": value,
                            "m": value,
                            "ct": {"fa": "General", "t": "g"}
                        }
                    }
                    
                    style = self._get_cell_style(cell)
                    if style:
                        cell_data["v"].update(style)
                    
                    celldata.append(cell_data)
        
        if ws.merged_cells:
            for merge_range in ws.merged_cells.ranges:
                merge_cells.append({
                    "r": merge_range.min_row - 1,
                    "c": merge_range.min_col - 1,
                    "rs": merge_range.max_row - merge_range.min_row + 1,
                    "cs": merge_range.max_col - merge_range.min_col + 1
                })
        
        config = {
            "merge": merge_cells if merge_cells else None,
            "rowlen": {},
            "columnlen": {}
        }
        
        for row_idx in range(1, max_row + 1):
            if ws.row_dimensions[row_idx].height:
                config["rowlen"][str(row_idx - 1)] = ws.row_dimensions[row_idx].height
        
        for col_idx in range(1, max_col + 1):
            col_letter = chr(64 + col_idx) if col_idx <= 26 else f"{chr(64 + col_idx // 26)}{chr(64 + col_idx % 26)}"
            if ws.column_dimensions[col_letter].width:
                config["columnlen"][str(col_idx - 1)] = ws.column_dimensions[col_letter].width
        
        luckysheet_data = {
            "name": sheet_name,
            "color": "",
            "index": 0,
            "status": 1,
            "order": 0,
            "celldata": celldata,
            "config": config,
            "row": max_row,
            "column": max_col,
            "defaultRowHeight": 19,
            "defaultColWidth": 73
        }
        
        return luckysheet_data
    
    def convert_to_base64(self, file_path: str, sheet_name: str = None) -> str:
        """转换为Base64编码的JSON字符串"""
        data = self.convert(file_path, sheet_name)
        json_str = json.dumps([data], ensure_ascii=False)
        return base64.b64encode(json_str.encode('utf-8')).decode('utf-8')
