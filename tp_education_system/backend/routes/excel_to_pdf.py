"""
Excel 转 PDF 工具
支持多种转换方式，按优先级尝试
"""
import os
import subprocess
import tempfile


def convert_excel_to_pdf(excel_path: str, pdf_path: str, data: dict = None) -> bool:
    """
    将 Excel 转换为 PDF
    如果提供了 data，则在转换前填充数据
    """
    # 方法1: 使用 Microsoft Excel (COM接口，Windows最佳格式保持)
    if _convert_with_ms_excel(excel_path, pdf_path, data):
        return True
    
    # 方法2: 使用 LibreOffice
    if _convert_with_libreoffice(excel_path, pdf_path):
        return True
    
    # 方法3: 使用 Python reportlab (备用方案)
    if _convert_with_python(excel_path, pdf_path):
        return True
    
    return False


def _convert_with_ms_excel(excel_path: str, pdf_path: str, data: dict = None) -> bool:
    """使用 Microsoft Excel COM 接口转换，支持数据填充"""
    try:
        import win32com.client
        import pythoncom
        
        pythoncom.CoInitialize()
        
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        # 打开工作簿
        wb = excel.Workbooks.Open(os.path.abspath(excel_path))
        
        # 如果有数据，填充到单元格
        if data:
            ws = wb.Worksheets(1)  # 第一个工作表
            
            # 遍历所有单元格查找占位符
            for row in range(1, ws.UsedRange.Rows.Count + 1):
                for col in range(1, ws.UsedRange.Columns.Count + 1):
                    cell = ws.Cells(row, col)
                    if cell.Value and isinstance(cell.Value, str):
                        cell_text = str(cell.Value)
                        # 查找并替换占位符
                        for placeholder, value in data.items():
                            placeholder_pattern = '{{' + placeholder + '}}'
                            if placeholder_pattern in cell_text:
                                cell.Value = cell_text.replace(placeholder_pattern, str(value) if value is not None else '')
        
        # 导出为 PDF (0 = xlQualityStandard)
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
        
        # 关闭（不保存修改）
        wb.Close(False)
        excel.Quit()
        
        pythoncom.CoUninitialize()
        
        if os.path.exists(pdf_path):
            print(f"使用 Microsoft Excel 成功转换: {pdf_path}")
            return True
    except Exception as e:
        print(f"Microsoft Excel 转换失败: {e}")
        import traceback
        traceback.print_exc()
    
    return False


def _convert_with_libreoffice(excel_path: str, pdf_path: str) -> bool:
    """使用 LibreOffice 转换"""
    soffice_paths = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        r"C:\Program Files\LibreOffice 7\program\soffice.exe",
        r"C:\Program Files\LibreOffice 26.2\program\soffice.exe",
    ]
    
    soffice_path = None
    for path in soffice_paths:
        if os.path.exists(path):
            soffice_path = path
            break
    
    if not soffice_path:
        print("未找到 LibreOffice")
        return False
    
    try:
        # 创建临时输出目录
        output_dir = os.path.dirname(pdf_path)
        
        result = subprocess.run(
            [soffice_path, '--headless', '--convert-to', 'pdf', '--outdir', output_dir, excel_path],
            capture_output=True,
            text=True,
            timeout=120
        )
        
        if result.returncode == 0:
            # 重命名生成的PDF文件
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            generated_pdf = os.path.join(output_dir, base_name + '.pdf')
            
            if os.path.exists(generated_pdf):
                if generated_pdf != pdf_path:
                    if os.path.exists(pdf_path):
                        os.remove(pdf_path)
                    os.rename(generated_pdf, pdf_path)
                print(f"使用 LibreOffice 成功转换: {pdf_path}")
                return True
        else:
            print(f"LibreOffice 转换失败: {result.stderr}")
    except Exception as e:
        print(f"LibreOffice 转换异常: {e}")
    
    return False


def _convert_with_python(excel_path: str, pdf_path: str) -> bool:
    """使用 Python reportlab 转换（备用方案）"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        
        # 注册中文字体
        font_name = 'Helvetica'
        font_paths = [
            'C:/Windows/Fonts/simsun.ttc',
            'C:/Windows/Fonts/simhei.ttf',
            'C:/Windows/Fonts/msyh.ttc',
        ]
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    font_name = 'ChineseFont'
                    pdfmetrics.registerFont(TTFont(font_name, font_path))
                    break
                except:
                    continue
        
        # 读取Excel数据
        file_ext = excel_path.lower().split('.')[-1]
        data = []
        
        if file_ext == 'xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            for row in ws.iter_rows():
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                data.append(row_data)
        elif file_ext == 'xls':
            import xlrd
            wb = xlrd.open_workbook(excel_path)
            ws = wb.sheet_by_index(0)
            for row_idx in range(ws.nrows):
                row_data = [str(ws.cell_value(row_idx, col_idx)) for col_idx in range(ws.ncols)]
                data.append(row_data)
        
        if not data:
            data = [["No data"]]
        
        # 创建PDF
        doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4))
        table = Table(data)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        doc.build([table])
        
        print(f"使用 Python 成功转换: {pdf_path}")
        return True
    except Exception as e:
        print(f"Python 转换失败: {e}")
    
    return False
