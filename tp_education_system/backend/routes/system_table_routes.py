"""
全系统表字段导出与导入API
"""
from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Dict, Any
import psycopg2
import json
import io
import os
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

router = APIRouter(prefix="/api/system-tables", tags=["系统表导出导入"])

DATABASE_CONFIG = {
    "host": "localhost",
    "port": "5432",
    "database": "taiping_education",
    "user": "taiping_user",
    "password": "taiping_password"
}

EXCLUDED_TABLES = {
    'auto_fill_history', 'fill_records', 'navigation_backups',
    'pending_triggers', 'trigger_conditions', 'trigger_logs',
    'todo_templates', 'todo_history', 'todo_items', 'user_custom_todos',
    'custom_todo_stats', 'system_messages', 'saved_exports',
    'template_usage_logs', 'template_usage_records', 'template_fill_records',
    'report_generation_logs', 'template_data_records',
    'business_checklist', 'business_checklist_instances',
    'business_checklist_item_records', 'business_checklist_items',
    'business_checklist_templates', 'checklist_instances',
    'checklist_item_records', 'checklist_template_items',
    'checklist_templates', 'checklist_trigger_conditions',
    'checklist_trigger_type_defs', 'checklist_trigger_types',
    'unified_todos', 'query_configs', 'filter_condition_templates',
    'template_configs', 'template_field_mapping', 'template_field_mappings',
    'template_field_values', 'template_page_settings', 'template_placeholders',
    'template_regions', 'template_tags', 'templates', 'report_templates',
    'report_template_fields', 'report_calculate_fields', 'report_field_sources',
    'report_tag_filters', 'report_data_remarks', 'document_templates',
    'universal_templates', 'universal_field_mapping', 'template_auto_fill_rules',
    'data_filling_field_mappings', 'field_mapping_cache', 'field_mapping_history',
    'id_card', 'info', 'personal_statistics', 'table_field_relations',
    'statistics', 'death_registration_stats', 'octogenarian_subsidy_stats',
    'retirement_approval_stats', 'retirement_reminder_stats',
    'intermediate_tables', 'navigation_modules', 'unit_hierarchy',
    'school_information_table', 'town_subsidy_standards',
    'shi_ye_dan_wei_tui_xiu_ren_yuan_tui_xiu_fei_bian_dong_ming_xi_b',
    'test_wage', 'test_wage_fix',
}


def get_db_connection():
    return psycopg2.connect(**DATABASE_CONFIG)


def get_user_tables():
    """获取需要导出的用户数据表"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT table_name FROM information_schema.tables
        WHERE table_schema = 'public' AND table_type = 'BASE TABLE'
        ORDER BY table_name
    """)
    all_tables = [r[0] for r in cursor.fetchall()]
    cursor.close()
    conn.close()
    return [t for t in all_tables if t not in EXCLUDED_TABLES]


def load_table_name_mappings():
    """
    从 table_name_mappings.json 加载表名映射
    返回: {english_table_name: chinese_table_name}
    """
    try:
        config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
        mappings_file = os.path.join(config_dir, 'table_name_mappings.json')
        if os.path.exists(mappings_file):
            with open(mappings_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return data.get('reverse_mappings', {})
    except Exception as e:
        print(f'加载表名映射失败: {e}')
    return {}


def get_chinese_table_name(english_name: str, table_mappings: dict) -> str:
    """获取表的英文名对应的中文显示名称"""
    return table_mappings.get(english_name, english_name)


def get_english_table_name(chinese_name: str, table_mappings: dict) -> str:
    """将中文表名转换回英文表名"""
    for eng, chn in table_mappings.items():
        if chn == chinese_name:
            return eng
    return chinese_name


def load_field_mappings():
    """
    加载字段映射配置，构建英文字段名→中文字段名对照表
    返回: {table_name: {english_db_col: chinese_display_name}}
    """
    mappings = {}
    try:
        config_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
        field_mappings_file = os.path.join(config_dir, 'field_mappings.json')
        if os.path.exists(field_mappings_file):
            with open(field_mappings_file, 'r', encoding='utf-8') as f:
                field_configs = json.load(f)
            for config in field_configs.get('configs', []):
                table_name = config.get('table_name', '')
                if not table_name:
                    continue
                if table_name not in mappings:
                    mappings[table_name] = {}
                for fm in config.get('field_mappings', []):
                    source = fm.get('sourceField', '')  # 中文显示名
                    target = fm.get('targetField', source)  # 数据库列名（英文或中文）
                    if target and source:
                        # key=数据库列名, value=中文显示名
                        mappings[table_name][target] = source
    except Exception as e:
        print(f'加载字段映射失败: {e}')
    return mappings


def get_chinese_field_name(table_name: str, column_name: str, field_mappings: dict) -> str:
    """获取字段的中文显示名称（column_name=数据库列名）"""
    if table_name in field_mappings and column_name in field_mappings[table_name]:
        return field_mappings[table_name][column_name]
    return column_name


def get_english_field_name(table_name: str, chinese_name: str, field_mappings: dict) -> str:
    """将中文字段名转换回英文数据库列名"""
    if table_name in field_mappings:
        for db_col, chn_display in field_mappings[table_name].items():
            if chn_display == chinese_name:
                return db_col
    return chinese_name


def get_table_columns(table_name: str):
    """获取表的列信息"""
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT column_name, data_type, is_nullable, character_maximum_length
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s
        ORDER BY ordinal_position
    """, (table_name,))
    columns = [{'column_name': r[0], 'data_type': r[1], 'is_nullable': r[2], 'max_length': r[3]} for r in cursor.fetchall()]
    cursor.close()
    conn.close()
    return columns


@router.get("/export-fields")
async def export_all_table_fields():
    """导出全系统表中字段到Excel工作簿"""
    try:
        tables = get_user_tables()
        wb = Workbook()
        wb.remove(wb.active)

        # 加载字段映射和表名映射
        field_mappings = load_field_mappings()
        table_mappings = load_table_name_mappings()

        header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        table_count = 0
        for table_name in tables:
            columns = get_table_columns(table_name)
            if not columns:
                continue

            # 获取中文表名
            chinese_name = get_chinese_table_name(table_name, table_mappings)
            # Excel sheet name max 31 chars, 优先使用中文名，超长截断
            sheet_name = chinese_name[:31]
            ws = wb.create_sheet(title=sheet_name)

            # Title row - Chinese table name
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            title_cell = ws.cell(row=1, column=1, value=chinese_name)
            title_cell.font = Font(name='宋体', size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Headers
            headers = ['字段名', '数据类型', '最大长度', '填写内容']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Column data
            for row_idx, col in enumerate(columns, 3):
                # Column name - use Chinese display name
                chinese_col_name = get_chinese_field_name(table_name, col['column_name'], field_mappings)
                cell_name = ws.cell(row=row_idx, column=1, value=chinese_col_name)
                cell_name.font = Font(name='宋体', size=10)
                cell_name.alignment = Alignment(horizontal='left', vertical='center')
                cell_name.border = thin_border

                # Data type
                cell_type = ws.cell(row=row_idx, column=2, value=col['data_type'])
                cell_type.font = Font(name='宋体', size=10)
                cell_type.alignment = Alignment(horizontal='center', vertical='center')
                cell_type.border = thin_border

                # Max length
                max_len = col['max_length'] if col['max_length'] else ''
                cell_len = ws.cell(row=row_idx, column=3, value=max_len)
                cell_len.font = Font(name='宋体', size=10)
                cell_len.alignment = Alignment(horizontal='center', vertical='center')
                cell_len.border = thin_border

                # Fill content (empty for user to fill)
                cell_fill = ws.cell(row=row_idx, column=4, value='')
                cell_fill.font = Font(name='宋体', size=10)
                cell_fill.alignment = Alignment(horizontal='left', vertical='center')
                cell_fill.border = thin_border

                # Highlight required fields
                if col['is_nullable'] == 'NO':
                    cell_name.font = Font(name='宋体', size=10, bold=True, color='CC0000')

            # Column widths
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 40

            table_count += 1

        # Add index sheet
        ws_index = wb.create_sheet(title='目录', index=0)
        ws_index.cell(row=1, column=1, value='系统数据表字段目录').font = Font(name='宋体', size=16, bold=True)
        ws_index.cell(row=2, column=1, value=f'生成时间：{datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")}').font = Font(name='宋体', size=10)
        ws_index.cell(row=4, column=1, value='说明：').font = Font(name='宋体', size=10, bold=True)
        ws_index.cell(row=5, column=1, value='1. 红色加粗字段为必填字段').font = Font(name='宋体', size=10)
        ws_index.cell(row=6, column=1, value='2. 每个工作表对应一个数据表，请在各工作表中填写数据').font = Font(name='宋体', size=10)
        ws_index.cell(row=7, column=1, value='3. 填写完成后，通过导入功能上传此文件，系统将自动分拣到对应表中').font = Font(name='宋体', size=10)
        ws_index.cell(row=8, column=1, value='4. 请勿修改工作表名称和列名').font = Font(name='宋体', size=10)

        ws_index.cell(row=10, column=1, value='工作表目录：').font = Font(name='宋体', size=10, bold=True)
        ws_index.column_dimensions['A'].width = 50
        ws_index.column_dimensions['B'].width = 30

        for idx, table_name in enumerate(tables):
            chinese_name = get_chinese_table_name(table_name, table_mappings)
            sheet_name = chinese_name[:31]
            display_text = f'{chinese_name}({table_name})'
            ws_index.cell(row=11 + idx, column=1, value=sheet_name).font = Font(name='宋体', size=10)
            ws_index.cell(row=11 + idx, column=2, value=display_text).font = Font(name='宋体', size=10)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'系统数据表字段模板_{datetime.now().strftime("%Y%m%d")}.xlsx'
        encoded_filename = urllib.parse.quote(filename)

        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename*=UTF-8\'\'{encoded_filename}'}
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/import-from-workbook")
async def import_from_workbook(file: UploadFile = File(...)):
    """从工作簿导入数据，自动分拣到各表"""
    try:
        contents = await file.read()
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

        user_tables = set(get_user_tables())
        results = []
        total_rows = 0

        # 加载字段映射和表名映射
        field_mappings = load_field_mappings()
        table_mappings = load_table_name_mappings()

        conn = get_db_connection()
        cursor = conn.cursor()

        for sheet_name in wb.sheetnames:
            if sheet_name == '目录':
                continue

            # 查找实际表名：先尝试中文表名→英文表名，再尝试直接匹配
            table_name = None
            # 1. 中文表名反查英文表名
            english_from_cn = get_english_table_name(sheet_name, table_mappings)
            if english_from_cn and english_from_cn in user_tables:
                table_name = english_from_cn
            # 2. 直接匹配英文表名
            elif sheet_name in user_tables:
                table_name = sheet_name
            # 3. 前缀匹配（兼容截断的sheet名）
            else:
                for ut in user_tables:
                    if ut.startswith(sheet_name):
                        table_name = ut
                        break

            if not table_name:
                continue

            ws = wb[sheet_name]

            # Read headers from row 2
            headers = []
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=2, column=col_idx).value
                if header:
                    headers.append(header)
                else:
                    break

            # Find the "填写内容" column (column D)
            fill_col_idx = None
            for idx, h in enumerate(headers):
                if h == '填写内容':
                    fill_col_idx = idx + 1
                    break

            if fill_col_idx is None:
                continue

            # Find the "字段名" column (column A)
            field_col_idx = None
            for idx, h in enumerate(headers):
                if h == '字段名':
                    field_col_idx = idx + 1
                    break

            if field_col_idx is None:
                continue

            # Read data rows (starting from row 3)
            field_values = {}
            for row_idx in range(3, ws.max_row + 1):
                field_name = ws.cell(row=row_idx, column=field_col_idx).value
                field_value = ws.cell(row=row_idx, column=fill_col_idx).value

                if field_name and field_value is not None and str(field_value).strip():
                    field_values[field_name] = str(field_value).strip()

            if not field_values:
                continue

            # Get actual column info for this table
            columns_info = get_table_columns(table_name)
            valid_columns = {c['column_name'] for c in columns_info}

            # Filter to only valid columns, mapping Chinese names back to English
            valid_fields = {}
            for k, v in field_values.items():
                # 尝试将中文字段名映射回英文列名
                english_name = get_english_field_name(table_name, k, field_mappings)
                if english_name in valid_columns:
                    valid_fields[english_name] = v
                elif k in valid_columns:
                    valid_fields[k] = v
            if not valid_fields:
                continue

            # Build INSERT statement
            columns = list(valid_fields.keys())
            values = list(valid_fields.values())
            placeholders = ', '.join(['%s'] * len(values))
            columns_str = ', '.join([f'"{c}"' for c in columns])

            sql = f'INSERT INTO \"{table_name}\" ({columns_str}) VALUES ({placeholders})'

            try:
                cursor.execute(sql, values)
                conn.commit()
                results.append(f'{table_name}: 导入成功')
                total_rows += 1
            except Exception as e:
                conn.rollback()
                results.append(f'{table_name}: 导入失败 - {str(e)}')

        cursor.close()
        conn.close()

        return {
            'status': 'success',
            'message': f'导入完成，共处理 {total_rows} 条记录',
            'details': results
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))