"""
通用模板自动填报系统API路由
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, Dict, Any, List
import os
import json
import re
import copy
import shutil
import subprocess
import traceback
import asyncio
import threading
from datetime import datetime

from services.universal_template_service import template_engine, get_db_connection
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

router = APIRouter(prefix="/api/universal-template", tags=["通用模板自动填报"])

CONFIG_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
FIELD_CONFIGS_DIR = os.path.join(CONFIG_DIR, 'field_configs')
MAPPINGS_FILE = os.path.join(CONFIG_DIR, 'table_name_mappings.json')


def _load_table_mappings() -> Dict[str, str]:
    en_to_cn = {}
    if not os.path.exists(MAPPINGS_FILE):
        return en_to_cn
    with open(MAPPINGS_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    for chinese_name, info in data.get('mappings', {}).items():
        english_name = info.get('english_name', '')
        if english_name:
            en_to_cn[english_name] = chinese_name
    return en_to_cn


def _load_field_configs_index() -> Dict[str, Dict[str, Any]]:
    index = {}
    if not os.path.exists(FIELD_CONFIGS_DIR):
        return index
    for filename in os.listdir(FIELD_CONFIGS_DIR):
        if not filename.endswith('.json'):
            continue
        filepath = os.path.join(FIELD_CONFIGS_DIR, filename)
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            table_name = data.get('table_name', '')
            if table_name:
                index[table_name] = {
                    'config_name': data.get('config_name', ''),
                    'table_type': data.get('table_type', 'master'),
                    'field_configs': data.get('field_configs', [])
                }
        except Exception:
            continue
    return index


def _resolve_chinese_datasource(datasource: str) -> str:
    if not datasource or datasource.startswith('公式:'):
        return datasource
    if '.' not in datasource:
        return datasource
    parts = datasource.split('.', 1)
    if len(parts) != 2:
        return datasource
    english_table, english_field = parts
    en_to_cn = _load_table_mappings()
    field_index = _load_field_configs_index()
    chinese_table = en_to_cn.get(english_table, english_table)
    config = field_index.get(english_table)
    chinese_field = english_field
    if config:
        for fc in config.get('field_configs', []):
            if fc.get('targetField', '') == english_field:
                chinese_field = fc.get('sourceField', english_field)
                break
    return f"{chinese_table}.{chinese_field}"

TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads', 'templates')
EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'exports', 'templates')
SAVED_DIR = os.path.join(EXPORT_DIR, '已保存')
os.makedirs(TEMPLATE_DIR, exist_ok=True)
os.makedirs(EXPORT_DIR, exist_ok=True)
os.makedirs(SAVED_DIR, exist_ok=True)


def _init_saved_exports_table():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS saved_exports (
                id SERIAL PRIMARY KEY,
                模板ID VARCHAR(100) NOT NULL,
                模板名称 VARCHAR(200),
                单位名称 VARCHAR(200),
                年月 VARCHAR(20),
                查询条件 JSONB DEFAULT '{}',
                统计范围 JSONB DEFAULT '{}',
                填报口径 JSONB DEFAULT '{}',
                Excel路径 VARCHAR(500),
                PDF路径 VARCHAR(500),
                保存时间 TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                备注 TEXT
            )
        """)
        conn.commit()
        cursor.close()
        conn.close()
        print("[OK] saved_exports表初始化成功")
    except Exception as e:
        print(f"[WARNING] saved_exports表初始化失败: {e}")

_init_saved_exports_table()

def _build_export_filename(config, request, include_seconds=False):
    now = datetime.now()
    # 优先使用用户选择的年月，没有则用当前系统年月
    user_ym = request.查询条件.get('年月', '') if request.查询条件 else ''
    if user_ym:
        ym_match = re.match(r'^(\d{4})-(\d{1,2})$', user_ym)
        if ym_match:
            year = int(ym_match.group(1))
            month = int(ym_match.group(2)) + 1
            if month > 12:
                month = 1
                year += 1
            month_str = f"{year}年{month}月"
        else:
            month_str = f"{now.year}年{now.month}月"
    else:
        month_str = f"{now.year}年{now.month}月"
    date_str = now.strftime("%Y%m%d_%H%M%S") if include_seconds else now.strftime("%Y%m%d")
    name = config.get('模板名称', '')
    unit_name = _get_fill_unit_name(request)
    template_category = config.get('模板分类', '')
    
    # 个人表：使用教师姓名
    if template_category == '个人表' and unit_name:
        return f"{unit_name}_{name}({date_str})"
    # 单位表：使用单位名称+年月
    if unit_name:
        return f"{unit_name}{month_str}{name}({date_str})"
    return f"{month_str}{name}({date_str})"


def _get_fill_unit_name(request):
    """获取导出文件名中的单位/个人名称
    
    优先级：
    1. 封面单位（登录时选择的单位，确保每次一致）
    2. 统计范围中的单位名称
    3. 个人表：通过身份证号或职工ID查询教师姓名
    """
    # 优先使用登录时选择的封面单位（保证每次导出文件名中单位名称一致）
    if getattr(request, '封面单位', None):
        return request.封面单位
    
    if request.统计范围:
        scope = request.统计范围.get('单位范围', {})
        for level in ['学校', '镇', '县', '地区', '省']:
            if level in scope and scope[level].get('unit_name'):
                return scope[level]['unit_name']
    if request.查询条件:
        # 个人表：通过身份证号查询教师姓名
        id_card = request.查询条件.get('身份证号', '')
        if id_card:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT \"姓名\" FROM teacher_basic_info WHERE \"身份证号码\" = %s LIMIT 1",
                    (id_card.strip(),)
                )
                row = cursor.fetchone()
                cursor.close()
                conn.close()
                if row and row[0]:
                    return row[0]  # 返回教师姓名
            except Exception:
                pass
        
        # 个人表：通过职工ID查询教师姓名
        employee_id = request.查询条件.get('职工ID')
        if employee_id:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT \"姓名\" FROM teacher_basic_info WHERE id = %s",
                    (int(employee_id),)
                )
                row = cursor.fetchone()
                cursor.close()
                conn.close()
                if row and row[0]:
                    return row[0]  # 返回教师姓名
            except Exception:
                pass
    return None


SOFFICE_PATHS = [
    shutil.which('soffice.exe'),
    shutil.which('soffice'),
    r'C:\Program Files\LibreOffice\program\soffice.exe',
    r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
]
SOFFICE_PATH = None
for p in SOFFICE_PATHS:
    if p and os.path.exists(p):
        SOFFICE_PATH = p
        break

LIBREOFFICE_PROFILE = os.path.join(EXPORT_DIR, '.libreoffice_profile')

if SOFFICE_PATH:
    print(f"[OK] LibreOffice检测成功: {SOFFICE_PATH}")
    os.makedirs(LIBREOFFICE_PROFILE, exist_ok=True)
    try:
        subprocess.run(
            [SOFFICE_PATH, '--headless',
             f'-env:UserInstallation=file:///{LIBREOFFICE_PROFILE.replace(os.sep, "/")}',
             '--terminate_after_init'],
            capture_output=True, text=True, timeout=30
        )
        print(f"[OK] LibreOffice配置文件初始化成功: {LIBREOFFICE_PROFILE}")
    except Exception as e:
        print(f"[WARNING] LibreOffice配置文件初始化失败: {e}")
else:
    print("[WARNING] 未检测到LibreOffice，PDF导出功能不可用")


class ImportTemplateRequest(BaseModel):
    模板名称: str
    模板类型: str


class FieldMappingRequest(BaseModel):
    模板ID: str
    字段名称: str
    行号: int
    列号: int
    数据源: str
    转换函数: Optional[str] = None
    默认值: Optional[str] = None
    字典值选择: Optional[List[Any]] = None


class FillTemplateRequest(BaseModel):
    模板ID: str
    查询条件: Dict[str, Any]
    统计范围: Optional[Dict[str, Any]] = None
    填报口径: Optional[Dict[str, Any]] = None
    备注: Optional[str] = None
    填报配置: Optional[Dict[str, Any]] = None  # 前端传来的已填充配置，用于保存时避免重新填充
    封面单位: Optional[str] = None  # 登录时选择的单位，用于封面显示
    saved_export_id: Optional[int] = None  # 已保存记录ID，用于导出时直接下载已保存文件
    导出格式: Optional[str] = 'Excel'  # 导出格式：Excel/Word/PDF


@router.get("/check-filename/{filename}")
async def check_filename(filename: str):
    """
    检测文件名是否已被其他模板使用
    
    返回:
        存在: bool, 引用模板列表: list
    """
    try:
        file_path = os.path.join(TEMPLATE_DIR, filename)
        disk_exists = os.path.exists(file_path)

        referenced_templates = []
        if disk_exists:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute(
                "SELECT 模板ID, 模板名称 FROM template_configs WHERE 原始文件路径 = %s",
                (filename,)
            )
            for row in cursor.fetchall():
                referenced_templates.append({"模板ID": row[0], "模板名称": row[1]})
            cursor.close()
            conn.close()

        return {
            "成功": True,
            "磁盘存在": disk_exists,
            "被引用模板": referenced_templates,
            "有冲突": len(referenced_templates) > 0
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/import")
async def import_template(
    file: UploadFile = File(...),
    模板名称: str = Form(...),
    模板类型: str = Form(...),
    模板分类: str = Form(None),
    保存文件名: str = Form(None)
):
    try:
        save_filename = 保存文件名 if 保存文件名 else file.filename
        temp_path = os.path.join(TEMPLATE_DIR, save_filename)
        with open(temp_path, 'wb') as f:
            f.write(file.file.read())
        
        config = template_engine.import_template(temp_path, 模板名称, 模板类型)
        if 模板分类:
            config['模板分类'] = 模板分类
        
        template_engine.save_template_config(config)
        
        return {
            "成功": True,
            "消息": "模板导入成功",
            "数据": config
        }
    except Exception as e:
        import traceback
        print(f"导入模板失败: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/download/{template_id}")
async def download_template_file(template_id: str):
    """
    下载原始模板文件（供Luckysheet直接加载）
    
    参数:
        template_id: 模板ID
    
    返回:
        Excel文件流
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT 原始文件路径 FROM template_configs WHERE 模板ID = %s
        """, (template_id,))
        
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not row:
            raise HTTPException(status_code=404, detail="模板不存在")
        
        file_name = row[0]
        if not file_name:
            raise HTTPException(status_code=404, detail="原始文件不存在")
        
        original_file_path = os.path.join(TEMPLATE_DIR, file_name)
        if not os.path.exists(original_file_path):
            raise HTTPException(status_code=404, detail="原始文件不存在")
        
        return FileResponse(
            original_file_path,
            filename=os.path.basename(original_file_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class RemarksUpdateRequest(BaseModel):
    备注: Optional[str] = None
    模板名称: Optional[str] = None
    单位名称: Optional[str] = None
    年月: Optional[str] = None


@router.get("/remarks")
async def get_all_remarks():
    """获取所有已保存的导出记录（备注编辑用）"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, 模板ID, 模板名称, 单位名称, 年月, 备注, 保存时间,
                   CASE WHEN "excel路径" IS NOT NULL AND "excel路径" != '' THEN true ELSE false END as 有Excel,
                   CASE WHEN "pdf路径" IS NOT NULL AND "pdf路径" != '' THEN true ELSE false END as 有PDF
            FROM saved_exports
            ORDER BY 保存时间 DESC
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        records = []
        for row in rows:
            records.append({
                "ID": row[0],
                "模板ID": row[1],
                "模板名称": row[2],
                "单位名称": row[3] or '',
                "年月": row[4] or '',
                "备注": row[5] or '',
                "保存时间": row[6].strftime("%Y-%m-%d %H:%M:%S") if row[6] else '',
                "有Excel": row[7],
                "有PDF": row[8],
                "有HTML": False  # HTML不再保存到磁盘
            })

        return {"成功": True, "数据": records}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/remarks")
async def create_remarks(request: RemarksUpdateRequest):
    """手动创建一条备注记录"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        now = datetime.now()
        cursor.execute(
            """INSERT INTO saved_exports (模板ID, 模板名称, 单位名称, 年月, 查询条件, 统计范围, 填报口径, 备注, 保存时间)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (
                '',
                request.模板名称 or '',
                request.单位名称 or '',
                request.年月 or '',
                '{}',
                '{}',
                '{}',
                request.备注 or '',
                now
            )
        )
        conn.commit()
        cursor.close()
        conn.close()
        return {"成功": True, "消息": "备注创建成功"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/remarks/{record_id}")
async def update_remarks(record_id: int, request: RemarksUpdateRequest):
    """更新导出记录的备注信息"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        updates = []
        values = []
        if request.备注 is not None:
            updates.append("备注 = %s")
            values.append(request.备注)
        if request.模板名称 is not None:
            updates.append("模板名称 = %s")
            values.append(request.模板名称)
        if request.单位名称 is not None:
            updates.append("单位名称 = %s")
            values.append(request.单位名称)
        if request.年月 is not None:
            updates.append("年月 = %s")
            values.append(request.年月)

        if not updates:
            cursor.close()
            conn.close()
            return {"成功": True, "消息": "无更新内容"}

        values.append(record_id)
        sql = f"UPDATE saved_exports SET {', '.join(updates)} WHERE id = %s"
        cursor.execute(sql, values)
        conn.commit()
        cursor.close()
        conn.close()

        return {"成功": True, "消息": "更新成功"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/remarks/{record_id}")
async def delete_remarks(record_id: int):
    """删除导出记录"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM saved_exports WHERE id = %s", (record_id,))
        conn.commit()
        cursor.close()
        conn.close()

        return {"成功": True, "消息": "删除成功"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/list")
async def list_templates():
    """
    列出所有模板
    
    返回:
        模板列表
    """
    try:
        templates = template_engine.list_templates()
        return {
            "成功": True,
            "数据": templates
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/config/{template_id}")
async def get_template_config(template_id: str):
    """
    获取模板配置
    
    参数:
        template_id: 模板ID
    
    返回:
        模板配置JSON
    """
    try:
        config = template_engine.load_template_config(template_id)
        if not config:
            raise HTTPException(status_code=404, detail="模板不存在")
        
        return {
            "成功": True,
            "数据": config
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/preview/{template_id}")
async def preview_template(template_id: str, teacher_id: Optional[int] = None, id_card: Optional[str] = None, 年月: Optional[str] = None):
    """
    预览模板（返回完整元数据，用于Luckysheet渲染）
    
    参数:
        template_id: 模板ID
        teacher_id: 教师ID（可选，用于提取备注）
        id_card: 身份证号（可选，个人模板优先使用）
        年月: 用户选择的年月（如：2026-05），用于解析{{年月+1}}等日期占位符
    
    返回:
        完整元数据（包含单元格、样式、合并单元格、行高列宽等）
    """
    try:
        config = template_engine.load_template_config(template_id)
        if not config:
            raise HTTPException(status_code=404, detail="模板不存在")
        
        original_file_path = config.get('原始文件路径')
        original_filename = config.get('原始文件', '')
        
        if not original_file_path or not os.path.exists(original_file_path):
            original_file_path = os.path.join(TEMPLATE_DIR, original_filename)
        
        # 检查是否是Word模板（.docx）
        is_word_template = original_file_path and original_file_path.lower().endswith('.docx')
        if is_word_template:
            # Word模板无法在Excel编辑器预览，返回空元数据，前端直接进入填报
            return {
                "成功": True,
                "数据": {
                    "metadata": None,
                    "HTML": "",
                    "模板名称": config.get('模板名称', ''),
                    "模板ID": template_id,
                    "备注": "",
                    "模板类型": "word"
                }
            }
        
        if os.path.exists(original_file_path):
            metadata = template_engine.extract_full_metadata(original_file_path)
        else:
            metadata = {
                'filename': config.get('原始文件', ''),
                'sheets': ['Sheet1'],
                'active_sheet': 0,
                'cells': [],
                'styles': {},
                'dimensions': {
                    'rows': {},
                    'columns': {},
                    'max_row': len(config.get('行数据', [])),
                    'max_column': len(config.get('列宽', []))
                },
                'merged_cells': [],
                'page_setup': config.get('页面设置', {}),
                'page_margins': config.get('页边距', {})
            }
            
            for row_data in config.get('行数据', []):
                for cell in row_data.get('单元格', []):
                    if cell.get('文本'):
                        metadata['cells'].append({
                            'r': row_data['行号'] - 1,
                            'c': cell['列号'] - 1,
                            'v': {
                                'v': cell['文本'],
                                'm': cell['文本'],
                            }
                        })
            
            for merged in config.get('合并单元格', []):
                metadata['merged_cells'].append({
                    'range': f"{merged['起始']}:{merged['结束']}",
                    'r': merged['起始行'] - 1,
                    'c': merged['起始列'] - 1,
                    'rs': merged['结束行'] - merged['起始行'] + 1,
                    'cs': merged['结束列'] - merged['起始列'] + 1
                })
        
        # 构建查询参数：优先使用id_card，其次使用teacher_id
        query_params = {}
        if 年月:
            query_params['年月'] = 年月
        if id_card:
            query_params['身份证号'] = id_card
        elif teacher_id is not None and teacher_id > 0:
            query_params['职工ID'] = str(teacher_id)
        
        # 如果有查询参数（身份证号或教师ID），先填充数据再渲染HTML
        filled_config = config
        备注 = ''
        if query_params and (id_card or (teacher_id is not None and teacher_id > 0)):
            try:
                # 加载字段映射
                mappings = template_engine.load_field_mappings(template_id)
                filled_config = copy.deepcopy(config)
                filled_config['字段映射'] = mappings
                filled_config = template_engine.fill_template_data(filled_config, query_params)
                
                # 提取备注信息
                if filled_config and '单元格数据' in filled_config:
                    for cell in filled_config['单元格数据']:
                        显示值 = str(cell.get('显示值', '') or '')
                        if 显示值.startswith('备注：'):
                            备注 = 显示值.replace('备注：', '').strip()
                            break
            except Exception as e:
                print(f"填充数据失败: {e}")
                import traceback
                traceback.print_exc()
        
        return {
            "成功": True,
            "数据": {
                "metadata": metadata,
                "HTML": template_engine.preview_template(filled_config, query_params, excel_path=original_file_path),
                "模板名称": config['模板名称'],
                "模板ID": template_id,
                "备注": 备注
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"预览错误: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/find-teacher-by-idcard")
async def find_teacher_by_idcard(id_card: str = Query(..., description="身份证号码")):
    """
    根据身份证号查找教师信息
    
    参数:
        id_card: 身份证号码（18位）
    
    返回:
        教师基本信息
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute(
                "SELECT id, \"姓名\", \"身份证号码\" FROM teacher_basic_info WHERE \"身份证号码\" = %s LIMIT 1",
                (id_card.strip(),)
            )
            row = cursor.fetchone()
            if row:
                return {
                    "成功": True,
                    "数据": {
                        "id": row[0],
                        "姓名": row[1],
                        "身份证号": row[2]
                    }
                }
            else:
                return {
                    "成功": False,
                    "消息": "未找到该身份证号对应的教师"
                }
        finally:
            cursor.close()
            conn.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/field-mapping")
async def save_field_mapping(request: FieldMappingRequest):
    """
    保存字段映射
    
    参数:
        模板ID: 模板ID
        字段名称: 字段名称
        行号: 单元格行号
        列号: 单元格列号
        数据源: 数据源（表名.字段名）
    
    返回:
        成功/失败消息
    """
    try:
        template_engine.save_field_mapping(
            request.模板ID,
            request.字段名称,
            request.行号,
            request.列号,
            request.数据源,
            request.转换函数,
            request.默认值,
            request.字典值选择
        )
        
        return {
            "成功": True,
            "消息": "字段映射保存成功"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/field-mappings/{template_id}")
async def get_field_mappings(template_id: str):
    """
    获取字段映射
    
    参数:
        template_id: 模板ID
    
    返回:
        字段映射字典（包含数据源_中文字段，直接从表名映射和字段配置文件解析）
    """
    try:
        mappings = template_engine.load_field_mappings(template_id)
        for field_name, mapping in mappings.items():
            datasource = mapping.get('数据源', '')
            if datasource:
                mapping['数据源_中文'] = _resolve_chinese_datasource(datasource)
        return {
            "成功": True,
            "数据": mappings
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/field-mapping/{template_id}/{field_name}")
async def delete_field_mapping(template_id: str, field_name: str):
    try:
        deleted = template_engine.delete_field_mapping(template_id, field_name)
        if deleted:
            return {"成功": True, "消息": "字段映射已删除"}
        else:
            return {"成功": False, "消息": "未找到该字段映射"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _auto_fill_retirement_report(config: dict, query_params: dict, template_engine, 封面单位: str = '') -> dict:
    """
    退休呈报表代码自动填报
    从多个数据源（教师基础信息、岗位聘任信息、最新工资包数据、退休补充信息）自动汇集数据
    
    参数:
        config: 模板配置
        query_params: 查询参数 {"职工ID": "xxx", "年月": "2026-07"}
        template_engine: 模板引擎实例
        封面单位: 登录时选择的单位名称（用于封面，与"发给退休费的单位"分开）
    
    返回:
        填充后的配置
    """
    import copy
    from services.universal_template_service import get_db_connection
    
    filled_config = copy.deepcopy(config)
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        teacher_id = int(query_params.get('职工ID', 0))
        if teacher_id <= 0:
            return filled_config
        
        # 1. 查询教师基础信息
        cursor.execute("""
            SELECT id, "姓名", "身份证号码", "档案出生日期", "民族", "籍贯", "参加工作日期"
            FROM teacher_basic_info WHERE id = %s
        """, (teacher_id,))
        teacher_row = cursor.fetchone()
        if not teacher_row:
            return filled_config
        
        teacher_id_val = teacher_row[0]
        teacher_name = teacher_row[1]
        id_card = teacher_row[2]
        
        # 导入自动填报函数
        from routes.status_change_routes import _build_retirement_report_data
        import sys
        import os
        
        # 查询学历
        cursor.execute("""
            SELECT education, graduate_date FROM teacher_education_record
            WHERE teacher_id = %s ORDER BY graduate_date DESC LIMIT 1
        """, (teacher_id_val,))
        education_row = cursor.fetchone()
        
        # 调用自动填报函数
        report_data = _build_retirement_report_data(cursor, teacher_id_val, teacher_row, id_card, education_row)
        
        if not report_data:
            return filled_config
        
        # 2. 构建模板合并单元格映射，用于确定正确的值填充位置
        # 关键：标签和值在分开的合并单元格中，必须填充到值合并单元格的主单元格
        merges = config.get('合并单元格', [])
        merge_master = {}  # (行,列) → (起始行,起始列,结束行,结束列)
        for mc in merges:
            sr, sc = mc.get('起始行'), mc.get('起始列')
            er, ec = mc.get('结束行'), mc.get('结束列')
            if sr and sc and er and ec:
                for r in range(sr, er + 1):
                    for c in range(sc, ec + 1):
                        merge_master[(r, c)] = (sr, sc, er, ec)
        
        cells = filled_config.get('单元格数据', [])
        cell_index = {}
        for cell in cells:
            cell_index[(cell.get('行号'), cell.get('列号'))] = cell
        
        # 3. 基本信息区域：字段名 → 值单元格主单元格(行,列)
        # 这些映射基于模板实际合并单元格结构分析得出
        basic_info_value_cells = {
            '姓名': (22, 2),
            '性别': (22, 7),
            '出生日期': (22, 12),
            '民族': (23, 2),
            '文化程度': (23, 7),
            '是否独生子女': (23, 12),
            '入党年月': (24, 2),
            '职务': (24, 7),
            '技术职称': (24, 12),
            '参加工作时间': (25, 3),
            '工作年限': (25, 9),
            '籍贯': (26, 2),
            '现住址': (26, 8),
            '退休原因': (33, 4),
            '退休后居住地址': (38, 2),
        }
        
        def set_cell_value(row, col, value):
            """设置指定单元格的值，自动处理合并单元格（填充到主单元格）"""
            master = merge_master.get((row, col))
            if master:
                row, col = master[0], master[1]  # 使用主单元格
            cell = cell_index.get((row, col))
            if cell:
                cell['值'] = str(value)
                cell['显示值'] = str(value)
        
        # 填充基本信息
        for field_name, (row, col) in basic_info_value_cells.items():
            val = report_data.get(field_name)
            if val is not None and val != '':
                set_cell_value(row, col, val)
        
        # 4. 封面区域：单位名称、姓名
        # 封面单位（行18列17）：使用登录时选择的单位，没有则留空
        unit_name = 封面单位 or ''
        if unit_name:
            set_cell_value(18, 17, unit_name)
        # 封面姓名（行19列17）
        if report_data.get('姓名'):
            set_cell_value(19, 17, report_data['姓名'])
        
        # 5. 工作简历区域（表头在行28，数据填在行29）
        # 自何年何月 → 行29列2
        # 至何年何月 → 行29列6
        # 在何单位任何职 → 行29列9
        # 证明人及其住址 → 行29列12
        work_exp = report_data.get('工作经历')
        if work_exp and len(work_exp) > 0:
            first_exp = work_exp[0]
            if isinstance(first_exp, dict):
                set_cell_value(29, 2, first_exp.get('自何年何月', ''))
                set_cell_value(29, 6, first_exp.get('至何年何月', ''))
                set_cell_value(29, 9, first_exp.get('所在单位及职务', ''))
                set_cell_value(29, 12, first_exp.get('证明人及其住址', ''))
            elif isinstance(first_exp, str):
                # 字符串类型：直接从 report_data 获取各字段
                set_cell_value(29, 2, report_data.get('自何年何月', ''))
                set_cell_value(29, 6, report_data.get('至何年何月', ''))
                set_cell_value(29, 9, report_data.get('所在单位及职务', ''))
                set_cell_value(29, 12, report_data.get('证明人及其住址', ''))
        
        # 6. 供养直系亲属（从退休补充信息表查询，非硬编码）
        support_info = report_data.get('直系亲属信息') or report_data.get('直系亲属供养情况')
        set_cell_value(35, 4, support_info if support_info else '')
        
        # 7. 最后一次职务升降时间：模板中已有标签"最后一次职务（技术职称）升降时间"，此处只填日期值
        last_promotion = report_data.get('最后一次职务升降时间')
        if last_promotion:
            # 格式化日期为中文格式
            try:
                from datetime import datetime
                if isinstance(last_promotion, datetime):
                    formatted_date = f"{last_promotion.year}年{last_promotion.month:02d}月{last_promotion.day:02d}日"
                else:
                    dt = datetime.strptime(str(last_promotion)[:10], '%Y-%m-%d')
                    formatted_date = f"{dt.year}年{dt.month:02d}月{dt.day:02d}日"
            except:
                formatted_date = str(last_promotion)
            set_cell_value(28, 15, formatted_date)
        
        # 8. 工资信息区域（三个时间点：2014年9月30日、最后升降、退休时）
        # 数据来源于最新工资包数据表，按人员分类填充到对应行
        person_category = report_data.get('人员分类', '')
        
        # 辅助函数：从岗位名称中提取等级数字和纯等级名称
        # 如"12级专技"→(12, "十二级")，"八级义教"→(8, "八级")
        CN_NUM_MAP = {'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'七':7,'八':8,'九':9,'十':10,'十一':11,'十二':12,'十三':13}
        NUM_CN_MAP = {1:'一',2:'二',3:'三',4:'四',5:'五',6:'六',7:'七',8:'八',9:'九',10:'十',11:'十一',12:'十二',13:'十三'}
        
        def extract_grade_level(job_title_str):
            """从岗位名称中提取等级数字和中文等级，如'12级专技'→(12,'十二级')"""
            import re
            if not job_title_str:
                return None, None
            # 匹配阿拉伯数字+级
            m = re.match(r'(\d+)级', str(job_title_str))
            if m:
                num = int(m.group(1))
                cn = NUM_CN_MAP.get(num, str(num))
                return num, f'{cn}级'
            # 匹配中文数字+级
            for cn_num in sorted(CN_NUM_MAP.keys(), key=len, reverse=True):
                if str(job_title_str).startswith(cn_num + '级'):
                    num = CN_NUM_MAP[cn_num]
                    return num, f'{cn_num}级'
            return None, str(job_title_str)
        
        def get_duty_by_grade(grade_num):
            """根据岗位等级数字映射职务级别"""
            if grade_num is None:
                return None
            if 5 <= grade_num <= 7:
                return '副高级'
            elif 8 <= grade_num <= 10:
                return '中级'
            elif 11 <= grade_num <= 13:
                return '初级'
            return None
        
        def extract_salary_level_num(level_str):
            """从级别薪级字符串中提取数字，如'专技22级'→'22'，'义教32级'→'32'"""
            import re
            if not level_str:
                return None
            m = re.search(r'(\d+)', str(level_str))
            if m:
                return m.group(1)
            return str(level_str)
        
        # 确定人员分类对应的行号和列映射
        # 格式: { 时间点: { 人员分类: (行号, 岗位/技术等级(列18), 级别薪级/对应原职务(列20), 薪级(列22)) } }
        time_points = {
            # 2014年9月30日
            '2014': {
                'other':      (22, 18, None, 22),  # 机关工人: 技术等级(18), 级别薪级(22)
                'management': (23, 18, 20, 22),     # 事业管理: 岗位(18), 对应原职务(20), 薪级(22)
                'technical':  (25, 18, 20, 22),     # 事业专技: 岗位(18), 对应原职务(20), 薪级(22)
                'worker':     (27, 18, 20, 22),     # 事业工勤: 岗位(18), 对应技术等级(20), 薪级(22)
            },
            # 最后一次职务升降
            '最新': {
                'other':      (28, 18, None, 22),
                'management': (30, 18, 20, 22),
                'technical':  (32, 18, 20, 22),
                'worker':     (34, 18, 20, 22),
            },
            # 退休时
            '退休': {
                'other':      (35, 18, None, 22),
                'management': (36, 18, 20, 22),
                'technical':  (38, 18, 20, 22),
                'worker':     (39, 18, 20, 22),
            },
        }
        
        # 从 report_data 提取三个时间点的数据
        # 字段名映射: (岗位/技术等级字段, 级别薪级/对应原职务字段, 薪级字段)
        time_field_names = {
            '2014': ('2014年9月30日职务岗位', '2014年9月30日级别薪级', '2014年9月30日薪级工资'),
            '最新': ('最后一次职务升降岗位', '最后一次职务升降薪级', '最后一次职务升降薪级工资'),
            '退休': ('退休时职务岗位', '退休时级别薪级', '退休时薪级工资'),
        }
        
        for tp_name, (job_field, level_field, salary_field) in time_field_names.items():
            rows = time_points[tp_name].get(person_category)
            if not rows:
                continue
            row, job_col, level_col, salary_col = rows
            
            # 岗位：提取纯等级（如"12级专技"→"十二级"，"八级义教"→"八级"）
            job_raw = report_data.get(job_field)
            if job_raw is not None:
                grade_num, grade_name = extract_grade_level(job_raw)
                if grade_name:
                    set_cell_value(row, job_col, grade_name)
            
            # 对应原职务：根据岗位等级映射（5-7副高/8-10中级/11-13初级）
            if level_col is not None:
                # 从岗位中提取等级，映射到职务
                job_raw_for_duty = report_data.get(job_field)
                if job_raw_for_duty:
                    grade_num_for_duty, _ = extract_grade_level(job_raw_for_duty)
                    duty = get_duty_by_grade(grade_num_for_duty)
                    if duty:
                        set_cell_value(row, level_col, duty)
            
            # 薪级：从级别薪级字段提取数字（如"专技22级"→"22"）
            level_raw = report_data.get(level_field)
            if level_raw is not None and salary_col is not None:
                salary_num = extract_salary_level_num(level_raw)
                if salary_num:
                    set_cell_value(row, salary_col, salary_num)
        
        # 9. "同意...同志"之间的姓名填充（审批信息页三处）
        # 扫描所有单元格，找"同意"和"同志"之间的空单元格填充姓名
        # 关键：跳过与"同意"单元格在同一合并区域内的空单元格，避免覆盖"同意"文字
        for row in range(1, 21):
            row_cells = [c for c in cells if c.get('行号') == row]
            for i, cell in enumerate(row_cells):
                text = str(cell.get('显示值', '') or '')
                if '同意' in text:
                    # 获取"同意"所在单元格的合并主单元格
                    agree_row = cell.get('行号')
                    agree_col = cell.get('列号')
                    agree_master = merge_master.get((agree_row, agree_col))
                    agree_master_key = (agree_master[0], agree_master[1]) if agree_master else (agree_row, agree_col)
                    
                    # 向右查找"同志"
                    for j in range(i + 1, len(row_cells)):
                        next_text = str(row_cells[j].get('显示值', '') or '')
                        if '同志' in next_text:
                            # 在"同意"和"同志"之间填充姓名
                            for k in range(i + 1, j):
                                mid_text = str(row_cells[k].get('显示值', '') or '')
                                if not mid_text.strip() or mid_text == 'None':
                                    # 检查该空单元格是否与"同意"在同一合并区域
                                    mid_row = row_cells[k].get('行号')
                                    mid_col = row_cells[k].get('列号')
                                    mid_master = merge_master.get((mid_row, mid_col))
                                    mid_master_key = (mid_master[0], mid_master[1]) if mid_master else (mid_row, mid_col)
                                    if mid_master_key == agree_master_key:
                                        continue  # 跳过，避免覆盖"同意"文字
                                    set_cell_value(mid_row, mid_col, teacher_name)
                            break
        
        return filled_config
        
    finally:
        cursor.close()
        conn.close()


@router.post("/fill")
async def fill_template(request: FillTemplateRequest):
    """
    自动填报数据
    
    参数:
        模板ID: 模板ID
        查询条件: 查询条件（如：{"职工ID": "xxx", "年月": "2026-05"}）
        统计范围: 五级单位层级
        填报口径: 选中的标签筛选条件
    
    返回:
        填充后的模板配置JSON
    """
    try:
        # #region debug-point A:fill-entry
        try:
            import json as _dj, time as _dt
            _logp = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'debug_fill_log.txt')
            with open(_logp, 'a', encoding='utf-8') as _df:
                _df.write(f"[{_dt.time()}] A:fill_entry 模板ID={request.模板ID} 查询条件={request.查询条件} 统计范围={request.统计范围} 填报口径={request.填报口径}\n")
        except Exception:
            pass
        # #endregion
        config = template_engine.load_template_config(request.模板ID)
        if not config:
            raise HTTPException(status_code=404, detail="模板不存在")
        
        if request.统计范围:
            config['统计范围'] = request.统计范围
        if request.填报口径:
            config['填报口径'] = request.填报口径
        
        mappings = template_engine.load_field_mappings(request.模板ID)
        config['字段映射'] = mappings
        
        original_file_path = config.get('原始文件路径')
        is_word_template = original_file_path and original_file_path.lower().endswith('.docx')
        
        # ========== 退休呈报表：代码自动填报 ==========
        if config.get('模板名称') in ['职工退休呈报表', '职工退休呈报表（Word版）', '枣阳市机关事业单位养老保险改革过渡期内职务升降退休人员信息申报表']:
            from services.universal_template_service import get_db_connection
            filled_config = _auto_fill_retirement_report(config, request.查询条件, template_engine, request.封面单位 or '')
        else:
            filled_config = template_engine.fill_template_data(config, request.查询条件)
        
        # Word模板不需要生成HTML预览
        if is_word_template:
            html = ''
        else:
            html = template_engine.generate_print_html(filled_config, request.查询条件, excel_path=config.get('原始文件路径'))
        
        # 提取备注信息（从填充后的单元格数据中）
        备注 = ''
        if filled_config and '单元格数据' in filled_config:
            for cell in filled_config['单元格数据']:
                显示值 = str(cell.get('显示值', '') or '')
                if 显示值.startswith('备注：'):
                    备注 = 显示值.replace('备注：', '').strip()
                    break
        
        return {
            "成功": True,
            "数据": {
                "配置": filled_config,
                "HTML": html,
                "备注": 备注,
                "模板类型": "word" if is_word_template else "excel"
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"填报失败: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/performance-remarks")
def get_performance_remarks(year: int = datetime.now().year, month: int = datetime.now().month):
    """
    获取绩效工资审批表的备注信息
    
    参数:
        year: 年份
        month: 月份
    
    返回:
        格式化的备注信息
    """
    try:
        remarks = template_engine.get_performance_remarks(year, month)
        return {"成功": True, "数据": {"备注": remarks}}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _find_latest_saved_file(template_id, unit_name, year_month):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            """SELECT "excel路径", "pdf路径" FROM saved_exports
               WHERE 模板ID = %s AND 单位名称 = %s AND 年月 = %s
               ORDER BY 保存时间 DESC LIMIT 1""",
            (template_id, unit_name, year_month)
        )
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if row and row[0] and os.path.exists(row[0]):
            return {'Excel路径': row[0], 'PDF路径': row[1]}
        return None
    except Exception as e:
        print(f"[WARNING] 查找保存文件失败: {e}")
        return None


def _get_system_ym():
    now = datetime.now()
    return f"{now.year}年{now.month}月"


WORD_TEMPLATE_PATH = r"D:\erp_thirteen\数据库信息\模板\职工退休呈报表.docx"
JOB_PROMOTION_TEMPLATE_PATH = r"D:\erp_thirteen\数据库信息\模板\枣阳市机关事业单位养老保险改革过渡期内职务升降退休人员信息申报表.docx"


def _build_word_fill_data(report_data: dict, 封面单位: str = '') -> dict:
    """
    将退休呈报表数据转换为Word模板占位符映射
    
    _build_retirement_report_data 返回的是扁平dict，所有字段直接在顶层。
    
    Args:
        report_data: 由 _build_retirement_report_data 返回的扁平数据字典
        封面单位: 登录时选择的单位名称（用于封面，与"发给退休费的单位"分开）
    
    Returns:
        dict: {占位符名: 值}
    """
    data = {}
    
    def _g(key, default=''):
        """从扁平dict取值，None转为默认值"""
        v = report_data.get(key, None)
        return default if v is None else v
    
    def _g_str(key, default=''):
        """取值并转为字符串"""
        v = _g(key, default)
        return str(v) if v != default else default
    
    # === 基础信息 ===
    def _fmt_date_cn(date_val, precision='month'):
        """将日期值转为中文格式：YYYY年M月 或 YYYY年M月D日
        
        precision: 'month' → 只到月（YYYY年M月），'day' → 到日（YYYY年M月D日）
        """
        if not date_val:
            return ''
        s = str(date_val).strip()
        if not s:
            return ''
        s = s.replace('-', '.').replace('/', '.').replace('年', '.').replace('月', '.').replace('日', '')
        parts = [p.strip() for p in s.split('.') if p.strip()]
        if precision == 'day' and len(parts) >= 3:
            return f"{parts[0]}年{int(parts[1])}月{int(parts[2])}日"
        elif len(parts) >= 2:
            return f"{parts[0]}年{int(parts[1])}月"
        elif len(parts) == 1:
            return f"{parts[0]}年"
        return s
    
    def _fmt_bool_cn(value):
        """将布尔值转为中文"是"/"否"（兼容字符串和数字）"""
        if value is None:
            return ''
        if isinstance(value, bool):
            return '是' if value else '否'
        if isinstance(value, (int, float)):
            return '是' if value else '否'
        v = str(value).strip().lower()
        if v in ('true', '1', 'yes', '是'):
            return '是'
        if v in ('false', '0', 'no', '否', ''):
            return '否'
        return v
    
    data['姓名'] = _g('姓名')
    data['性别'] = _g('性别')
    data['出生年月'] = _fmt_date_cn(_g('出生日期'), 'month')
    data['民族'] = _g('民族')
    data['文化程度'] = _g('文化程度')
    data['是否独生子女'] = _fmt_bool_cn(_g('是否独生子女'))
    data['入党年月'] = _fmt_date_cn(_g('入党年月'), 'month')
    data['职务'] = _g('职务')
    data['技术职称'] = _g('技术职称')
    data['参加工作时间'] = _fmt_date_cn(_g('参加工作时间'), 'month')
    data['工作年限'] = _g_str('工作年限')
    data['籍贯'] = _g('籍贯')
    data['现住址'] = _g('现住址')
    data['退休原因'] = _g('退休原因')
    
    # 直系亲属信息
    data['直系亲属信息'] = _g('直系亲属信息') or _g('直系亲属供养情况')
    
    # 退休后居住地址 & 发给退休费的单位（从退休补充信息表获取）
    data['退休后居住地址'] = _g('退休后居住地址')
    data['发给退休费的单位'] = _g('发给退休费的单位')
    # 封面单位名称：取自登录时选择的单位，没有则留空
    # 注意：发给退休费的单位 ≠ 登录单位，两者是不同的概念，不可混淆
    data['单位名称'] = 封面单位 or ''
    
    # === 工作简历 ===
    data['自何年何月'] = _g('自何年何月')
    data['至何年何月'] = _g('至何年何月')
    data['在何单位任何职'] = _g('所在单位及职务') or _g('工作经历')
    data['证明人及其住址'] = _g('证明人及其住址')
    
    # === 工资信息 ===
    # 从扁平dict获取各时间点的岗位/薪级/职务信息
    # 兼容分类字段名（如 事业专技岗位2）和通用字段名（如 2014年9月30日职务岗位）
    
    def _extract_post_level(job_title_str):
        """从职务岗位字符串提取岗位等级（保留"级"后缀）
        如 '12级专技'→'12级', '十级专技'→'十级', '五级专技'→'五级'
        """
        if not job_title_str:
            return ''
        s = str(job_title_str)
        # 先匹配阿拉伯数字+级：如 '12级专技' → '12级'
        m = re.search(r'(\d+级)', s)
        if m:
            return m.group(1)
        # 再匹配中文数字+级：如 '十级专技' → '十级'
        m = re.search(r'([一二三四五六七八九十]+级)', s)
        if m:
            return m.group(1)
        # 最后尝试只匹配中文数字，加上'级'
        m = re.search(r'[一二三四五六七八九十]+', s)
        if m:
            return m.group() + '级'
        return s
    
    def _extract_salary_level(salary_level_str):
        """从薪级字符串提取薪级（保留"级"后缀）
        如 '专技13级'→'13级', '专技22级'→'22级', '22'→'22级'
        """
        if not salary_level_str:
            return ''
        s = str(salary_level_str)
        # 匹配数字+级：如 '专技13级' → '13级'
        m = re.search(r'(\d+级)', s)
        if m:
            return m.group(1)
        # 匹配纯数字，加上'级'：如 '22' → '22级'
        m = re.search(r'(\d+)', s)
        if m:
            return m.group(1) + '级'
        return s
    
    # 岗位等级→职务映射（用于对应原职务）
    # 等级数字越小级别越高：1-4正高级, 5-7副高级, 8-10中级, 11-13初级
    CHINESE_NUM_MAP = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '十': 10, '十一': 11, '十二': 12, '十三': 13}
    
    def _parse_level_num(job_title_str):
        """解析岗位等级数字"""
        if not job_title_str:
            return None
        s = str(job_title_str)
        m = re.search(r'(\d+)', s)
        if m:
            return int(m.group(1))
        for ch_name, num in sorted(CHINESE_NUM_MAP.items(), key=lambda x: -len(x[0])):
            if ch_name in s:
                return num
        return None
    
    def _map_level_to_title(level_num):
        """根据岗位等级数字映射职务名称"""
        if level_num is None:
            return ''
        if 1 <= level_num <= 4:
            return '正高级'
        if 5 <= level_num <= 7:
            return '副高级'
        if 8 <= level_num <= 10:
            return '中级'
        if 11 <= level_num <= 13:
            return '初级'
        return ''
    
    def _get_post_and_title(job_title_val, duty_val):
        """从岗位原始值获取岗位等级和对应原职务
        - 岗位：保留'xx级'格式（如'12级'、'八级'）
        - 对应原职务：根据等级映射为 正高级/副高级/中级/初级
        """
        post = _extract_post_level(job_title_val)
        # 对应原职务：根据岗位等级数字映射
        # 如果duty_val已经是映射后的职务名称（副高级/中级/初级/正高级），直接使用
        if duty_val and str(duty_val).strip() in ('正高级', '副高级', '中级', '初级'):
            title = str(duty_val)
        else:
            level_num = _parse_level_num(job_title_val)
            title = _map_level_to_title(level_num)
        return post, title
    
    # 退休时（行26: 事业专技岗位8）
    post_ret_raw = _g('退休时事业专技岗位8') or _g('退休时事业管理岗位7') or _g('退休时事业工勤岗位9') or _g('退休时职务岗位')
    duty_ret_raw = _g('对应原职务8') or _g('对应原职务7') or _g('对应技术等级9')
    data['岗位8'], data['职务8'] = _get_post_and_title(post_ret_raw, duty_ret_raw)
    sl_ret = _g('薪级8') or _g('薪级7') or _g('薪级9') or _g('退休时级别薪级')
    data['薪级8'] = _extract_salary_level(sl_ret)
    
    # 2014年9月30日（行18: 事业专技岗位2）
    post_2014_raw = _g('事业专技岗位2') or _g('事业管理岗位1') or _g('事业工勤岗位3') or _g('2014年9月30日职务岗位')
    duty_2014_raw = _g('对应原职务2') or _g('对应原职务1') or _g('对应技术等级3')
    data['岗位2'], data['职务2'] = _get_post_and_title(post_2014_raw, duty_2014_raw)
    sl_2014 = _g('薪级2') or _g('薪级1') or _g('薪级3') or _g('2014年9月30日级别薪级')
    data['薪级2'] = _extract_salary_level(sl_2014)
    
    # 最后一次职务升降（行22: 事业专技岗位5）
    post_prom_raw = _g('事业专技岗位5') or _g('事业管理岗位4') or _g('事业工勤岗位6') or _g('最后一次职务升降岗位')
    duty_prom_raw = _g('对应原职务5') or _g('对应原职务4') or _g('对应技术等级6')
    data['岗位5'], data['职务5'] = _get_post_and_title(post_prom_raw, duty_prom_raw)
    sl_prom = _g('薪级5') or _g('薪级4') or _g('薪级6') or _g('最后一次职务升降薪级')
    data['薪级5'] = _extract_salary_level(sl_prom)
    
    # 最后一次职务升降时间：模板中已有标签，只填格式化后的日期值
    promotion_time = _g('最后一次职务升降时间')
    if promotion_time:
        try:
            from datetime import datetime
            if isinstance(promotion_time, datetime):
                formatted = f"{promotion_time.year}年{promotion_time.month:02d}月{promotion_time.day:02d}日"
            else:
                dt = datetime.strptime(str(promotion_time)[:10], '%Y-%m-%d')
                formatted = f"{dt.year}年{dt.month:02d}月{dt.day:02d}日"
        except:
            formatted = str(promotion_time)
        data['最后一次职务（技术职称升降时间'] = formatted
    else:
        data['最后一次职务（技术职称升降时间'] = ''
    
    # === 审批信息（从退休补充信息中获取） ===
    # 退休方式序号：默认填"二"（法定退休年龄退休）
    data['退休方式序号'] = _g('退休方式序号', '二')
    data['审批退休方式序号'] = _g('审批退休方式序号')
    
    # 审批日期：当值为空时保留"    年  月"格式（5空格+年+2空格+月，用于手写填入）
    exec_ym = _g('退休执行年月')
    data['退休执行年月'] = exec_ym if exec_ym else '    年  月'
    exec_ym2 = _g('审批退休执行年月')
    data['审批退休执行年月'] = exec_ym2 if exec_ym2 else '    年  月'
    
    # 补贴信息：金额为空时预留空格（用于手写填入）
    data['独生子女费金额'] = _g('独生子女费金额') if _g('独生子女费金额') else '       '
    data['特殊贡献奖金额'] = _g('特殊贡献奖金额') if _g('特殊贡献奖金额') else '       '
    subsidy_ym = _g('补贴执行年月')
    data['补贴执行年月'] = subsidy_ym if subsidy_ym else '    年  月'
    
    # 空的工资行 - 填空
    for empty_field in ['岗位1', '岗位3', '岗位4', '岗位6', '岗位7', '岗位9',
                        '职务1', '职务3', '职务4', '职务6', '职务7', '职务9',
                        '薪级1', '薪级3', '薪级4', '薪级6', '薪级7', '薪级9']:
        if empty_field not in data:
            data[empty_field] = ''
    
    return data


def _fill_word_retirement_report(report_data: dict, output_path: str, 封面单位: str = '', template_path: str = None) -> str:
    """
    使用Word模板填充退休呈报表
    
    Args:
        report_data: 退休呈报表数据
        output_path: 输出路径
        封面单位: 登录时选择的单位名称
        template_path: Word模板路径（不传则使用默认路径）
    
    Returns:
        输出文件路径
    """
    from services.word_template_filler import fill_word_template
    
    if template_path is None:
        template_path = WORD_TEMPLATE_PATH
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Word模板不存在: {template_path}")
    
    fill_data = _build_word_fill_data(report_data, 封面单位)
    return fill_word_template(template_path, output_path, fill_data)


def _build_job_promotion_fill_data(report_data: dict, 封面单位: str = '', personal_id: str = '') -> dict:
    """
    构建职务升降退休人员信息申报表的填充数据
    
    上半部分：2014年9月30日职务（职级）、岗位工资信息
    下半部分：视为2014年9月职务（职级）、岗位升降工资信息（=最后一次职务升降数据）
    
    Args:
        report_data: 退休呈报表数据（来自 _build_retirement_report_data）
        封面单位: 登录时选择的单位名称
        personal_id: 个人编号（从退休补充信息表查询）
    
    Returns:
        Word模板占位符映射
    """
    # 复用 _build_word_fill_data 获取岗位/薪级数据
    base_data = _build_word_fill_data(report_data, 封面单位)
    
    # 新模板使用具体的占位符名称（如 事业专技岗位2），需要从 base_data 的通用键名映射
    # _build_word_fill_data 生成: 岗位2, 岗位5, 岗位8, 薪级2, 薪级5, 薪级8
    # 新模板占位符: 事业管理岗位1/4, 事业专技岗位2/5, 事业工勤岗位3/6, 薪级1-6
    data = {}
    
    # 映射岗位数据到具体占位符名称
    # 2014年9月30日（上半部分）
    data['事业管理岗位1'] = base_data.get('岗位1', '')
    data['事业专技岗位2'] = base_data.get('岗位2', '')
    data['事业工勤岗位3'] = base_data.get('岗位3', '')
    data['薪级1'] = base_data.get('薪级1', '')
    data['薪级2'] = base_data.get('薪级2', '')
    data['薪级3'] = base_data.get('薪级3', '')
    
    # 视为2014年9月 = 最后一次职务升降（下半部分）
    data['事业管理岗位4'] = base_data.get('岗位4', '')
    data['事业专技岗位5'] = base_data.get('岗位5', '')
    data['事业工勤岗位6'] = base_data.get('岗位6', '')
    data['薪级4'] = base_data.get('薪级4', '')
    data['薪级5'] = base_data.get('薪级5', '')
    data['薪级6'] = base_data.get('薪级6', '')
    
    # 姓名、性别、个人编号
    data['姓名'] = report_data.get('姓名', '')
    data['性别'] = report_data.get('性别', '')
    data['个人编号'] = personal_id or report_data.get('个人编号', '')
    
    return data


def _fill_job_promotion_report(report_data: dict, output_path: str, 封面单位: str = '', personal_id: str = '', template_path: str = None) -> str:
    """
    使用Word模板填充职务升降退休人员信息申报表
    
    Args:
        report_data: 退休呈报表数据
        output_path: 输出路径
        封面单位: 登录时选择的单位名称
        personal_id: 个人编号
        template_path: Word模板路径（不传则使用默认路径）
    
    Returns:
        输出文件路径
    """
    from services.word_template_filler import fill_word_template
    
    if template_path is None:
        template_path = JOB_PROMOTION_TEMPLATE_PATH
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"职务升降申报表模板不存在: {template_path}")
    
    fill_data = _build_job_promotion_fill_data(report_data, 封面单位, personal_id)
    return fill_word_template(template_path, output_path, fill_data)


def _write_filled_to_excel(config, filled_config, output_path):
    original_path = config.get('原始文件路径')
    wb = load_workbook(original_path)
    ws = wb.active

    for cell_data in filled_config.get('单元格数据', []):
        row_num = cell_data.get('行号')
        col_num = cell_data.get('列号')
        display_val = cell_data.get('显示值', '')
        if row_num and col_num and display_val != '':
            try:
                s = str(display_val)
                if s.replace('-', '').replace('.', '').isdigit():
                    val = float(s) if '.' in s else int(s)
                else:
                    val = display_val
            except (ValueError, AttributeError):
                val = display_val
            cell = ws.cell(row=row_num, column=col_num, value=val)
            if isinstance(val, str) and ('\n' in val or '\r' in val):
                # 设置自动换行
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                # 第28行备注：根据内容行数自动增加行高
                if row_num == 28:
                    行数 = val.count('\n') + 1
                    # 每行约15磅，最小不低于原始行高
                    原始行高 = ws.row_dimensions[28].height or 15
                    需要行高 = max(原始行高, 行数 * 15)
                    ws.row_dimensions[28].height = 需要行高

    for merge_cell in filled_config.get('合并单元格', []):
        start = merge_cell.get('起始', '')
        end = merge_cell.get('结束', '')
        if start and end:
            ws.merge_cells(f'{start}:{end}')

    for merge_cell in filled_config.get('合并单元格', []):
        start = merge_cell.get('起始', '')
        if start:
            try:
                col_letter, row_num = start[0], int(start[1:])
                cell = ws.cell(row=row_num, column=ord(col_letter) - ord('A') + 1)
                if cell.value and isinstance(cell.value, str) and ('\n' in cell.value or '\r' in cell.value):
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            except Exception:
                pass

    # 三级教师下划线加粗：教师系列最后一行，底部边框应与其他列一致为2px
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value and '三级教师' in str(cell.value):
                medium_side = Side(style='medium', color='000000')
                cell.border = Border(
                    top=cell.border.top if cell.border else Side(),
                    bottom=medium_side,
                    left=cell.border.left if cell.border else Side(),
                    right=cell.border.right if cell.border else Side()
                )

    wb.save(output_path)
    return output_path


def _do_full_fill_to_excel(config, request, output_path):
    if request.统计范围:
        config['统计范围'] = request.统计范围
    if request.填报口径:
        config['填报口径'] = request.填报口径
    mappings = template_engine.load_field_mappings(request.模板ID)
    config['字段映射'] = mappings
    filled_config = template_engine.fill_template_data(config, request.查询条件)
    _write_filled_to_excel(config, filled_config, output_path)
    return filled_config


@router.post("/export")
async def export_template(request: FillTemplateRequest):
    """
    导出已保存的模板文件（不再重新填充数据，保证数据一致性）
    
    优先使用 saved_export_id 下载已保存文件；
    如果未提供 saved_export_id，则返回错误提示用户先保存。
    """
    try:
        # 新流程：通过保存记录ID下载已保存文件
        saved_id = getattr(request, 'saved_export_id', None)
        export_format = getattr(request, '导出格式', 'Excel')
        
        if saved_id:
            return await download_history_file(saved_id, export_format)
        
        # 旧流程兼容：提示用户先保存
        config = template_engine.load_template_config(request.模板ID)
        if not config:
            raise HTTPException(status_code=404, detail="模板不存在")
        
        # 查找最近一次保存记录
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'SELECT id FROM saved_exports WHERE "模板ID" = %s ORDER BY "保存时间" DESC LIMIT 1',
            (request.模板ID,)
        )
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if row:
            return await download_history_file(row[0], export_format)
        
        raise HTTPException(status_code=404, detail="没有找到已保存的文件，请先点击「填报」保存后再导出")
    except HTTPException:
        raise
    except Exception as e:
        print(f"导出失败: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


def _write_compatibility_json(request, filled_config, template_config):
    """
    方案A兼容写入：将通用模板填充结果写入旧系统JSON格式
    确保 performance_pay_history_routes.py 等旧模块能继续读取数据
    """
    # 旧系统数据目录
    OLD_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'performance_pay_approval')
    os.makedirs(OLD_DATA_DIR, exist_ok=True)

    # 解析年月
    user_ym = request.查询条件.get('年月', '') if request.查询条件 else ''
    year = datetime.now().year
    month = datetime.now().month
    if user_ym:
        ym_match = re.match(r'^(\d{4})-(\d{1,2})$', user_ym)
        if ym_match:
            year = int(ym_match.group(1))
            month = int(ym_match.group(2))

    # 从 filled_config 的单元格数据中提取关键统计字段
    cells = filled_config.get('单元格数据', [])
    cell_by_field = {}
    for cell in cells:
        field_name = cell.get('字段名称', '')
        if field_name:
            cell_by_field[field_name] = cell

    def _get_cell_value(field_name, default=0):
        """从填充配置中获取指定字段的值"""
        cell = cell_by_field.get(field_name)
        if cell:
            val = cell.get('值')
            if val is not None and val != '':
                try:
                    return float(val) if '.' in str(val) else int(val)
                except (ValueError, TypeError):
                    return str(val)
        return default

    def _get_cell_display(field_name, default=''):
        """从填充配置中获取指定字段的显示值"""
        cell = cell_by_field.get(field_name)
        if cell:
            return str(cell.get('显示值', '') or '')
        return default

    # 构建兼容JSON
    unit_name = ''
    if request.统计范围 and request.统计范围.get('单位范围'):
        scope = request.统计范围['单位范围']
        for level_key in ['学校', '镇', '县', '地区', '省']:
            if scope.get(level_key) and scope[level_key].get('unit_name'):
                unit_name = scope[level_key]['unit_name']
                break
    if not unit_name:
        unit_name = request.封面单位 or '太平中心学校'

    compatibility_data = {
        '填报单位': unit_name,
        '年月': f"{year}年{month}月",
        '填报时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        # 行政管理人员
        '副处级人数': _get_cell_value('副处级人数', 0),
        '副处级标准': _get_cell_value('副处级标准', 0),
        '正科级人数': _get_cell_value('正科级人数', 0),
        '正科级标准': _get_cell_value('正科级标准', 0),
        '副科级人数': _get_cell_value('副科级人数', 0),
        '副科级标准': _get_cell_value('副科级标准', 0),
        '科员级人数': _get_cell_value('科员级人数', 0),
        '科员级标准': _get_cell_value('科员级标准', 1185),
        '办事员级人数': _get_cell_value('办事员级人数', 0),
        '办事员级标准': _get_cell_value('办事员级标准', 0),
        # 专业技术人员
        '正高级教师人数': _get_cell_value('正高级教师人数', 0),
        '正高级教师标准': _get_cell_value('正高级教师标准', 1862),
        '高级教师人数': _get_cell_value('高级教师人数', 0),
        '高级教师标准': _get_cell_value('高级教师标准', 1523),
        '一级教师人数': _get_cell_value('一级教师人数', 0),
        '一级教师标准': _get_cell_value('一级教师标准', 1309),
        '二级教师人数': _get_cell_value('二级教师人数', 0),
        '二级教师标准': _get_cell_value('二级教师标准', 1241),
        '三级教师人数': _get_cell_value('三级教师人数', 0),
        '三级教师标准': _get_cell_value('三级教师标准', 1128),
        # 工人
        '高级技师人数': _get_cell_value('高级技师人数', 0),
        '高级技师标准': _get_cell_value('高级技师标准', 0),
        '技师人数': _get_cell_value('技师人数', 0),
        '技师标准': _get_cell_value('技师标准', 1331),
        '高级工人数': _get_cell_value('高级工人数', 0),
        '高级工标准': _get_cell_value('高级工标准', 1219),
        '中级工人数': _get_cell_value('中级工人数', 0),
        '中级工标准': _get_cell_value('中级工标准', 1185),
        '初级工人数': _get_cell_value('初级工人数', 0),
        '初级工标准': _get_cell_value('初级工标准', 1106),
        '普工人数': _get_cell_value('普工人数', 0),
        '普工标准': _get_cell_value('普工标准', 1106),
        # 汇总
        '绩效人数合计': _get_cell_value('绩效人数合计', 0),
        '绩效工资合计': _get_cell_value('绩效工资合计', 0),
        # 乡镇补贴
        '在职人数': _get_cell_value('在职人数', 0),
        '乡镇补贴标准': _get_cell_value('乡镇补贴标准', 350),
        '乡镇补贴合计': _get_cell_value('乡镇补贴合计', 0),
        # 退休人员
        '退休干部': _get_cell_value('退休干部', 0),
        '退休职工': _get_cell_value('退休职工', 0),
        '离休干部人数': _get_cell_value('离休干部人数', 0),
        # 遗留问题
        '遗留问题详情': _get_cell_display('遗留问题详情', ''),
        '遗留问题人数': _get_cell_value('遗留问题人数', 0),
        '遗留问题金额': _get_cell_value('遗留问题金额', 0),
        '无补贴人数': _get_cell_value('无补贴人数', 0),
        '无补贴名单': _get_cell_display('无补贴名单', ''),
        # 备注
        '备注': request.备注 or _get_cell_display('备注', ''),
    }

    # 写入旧系统JSON文件
    filename = f"performance_pay_{year}_{month:02d}.json"
    filepath = os.path.join(OLD_DATA_DIR, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(compatibility_data, f, ensure_ascii=False, indent=2)
    print(f"[兼容写入] 绩效工资JSON已写入: {filepath}")


@router.post("/save")
async def save_filled_template(request: FillTemplateRequest):
    import traceback
    try:
        config = template_engine.load_template_config(request.模板ID)
        if not config:
            raise HTTPException(status_code=404, detail="模板不存在")

        original_file_path = config.get('原始文件路径')
        if not original_file_path or not os.path.exists(original_file_path):
            raise HTTPException(status_code=404, detail="原始文件不存在")

        base_name = _build_export_filename(config, request, include_seconds=True)
        is_word_template = original_file_path.lower().endswith('.docx')

        # Word模板：跳过Excel生成，直接生成Word+PDF
        if is_word_template:
            xlsx_filename = None
            xlsx_path = None
            filled_config = {}
            filled_html = ''
            html_path = None
        else:
            xlsx_filename = f"{base_name}.xlsx"
            xlsx_path = os.path.join(SAVED_DIR, xlsx_filename)

            # 使用前端传来的已填充配置（避免重复填充，保证数据一致性）
            HAS_FILLED_CONFIG = request.填报配置 is not None
            print(f"[SAVE DEBUG] 模板={config.get('模板名称','')}, 收到填报配置={HAS_FILLED_CONFIG}, 收到备注={'YES' if request.备注 else 'NO'}")
            
            if request.填报配置:
                filled_config = request.填报配置
                # 记录关键数据用于调试
                key_cells_log = {}
                for cell in filled_config.get('单元格数据', []):
                    r, c = cell.get('行号'), cell.get('列号')
                    if r in [7, 10, 11, 12, 13, 15, 18, 20, 21, 24, 25, 26, 27, 28]:
                        key_cells_log[f"({r},{c})"] = str(cell.get('显示值',''))[:60]
                print(f"[SAVE DEBUG] 填报配置关键数据: {key_cells_log}")
                
                # 如果有编辑后的备注，也要更新到填报配置中
                # 注意：只在备注确实被修改过且内容不包含完整表格文本时才更新
                if request.备注 is not None and request.备注 != '':
                    # 检测备注内容是否异常（包含整个表格文本的异常情况）
                    备注内容 = str(request.备注)
                    # 异常检测：包含表格结构关键词，或长度超过正常备注范围
                    is_abnormal_remark = (
                        ('项目' in 备注内容 and '人数' in 备注内容 and '标准' in 备注内容 and '小计' in 备注内容) or
                        len(备注内容) > 800
                    )
                    if is_abnormal_remark:
                        print(f"[SAVE WARNING] 备注内容异常！包含完整表格文本或过长，拒绝使用！备注长度={len(备注内容)}")
                        # 不使用异常备注，保持filled_config中的原始备注
                    else:
                        print(f"[SAVE DEBUG] 备注内容正常，长度={len(备注内容)}，将更新到配置中")
                        updated = False
                        # 第一优先级：精确行号+列号匹配（A28 = 行28,列1）
                        for cell in filled_config.get('单元格数据', []):
                            if cell.get('行号') == 28 and cell.get('列号') == 1:
                                cell['显示值'] = f"备注：\n{request.备注}"
                                cell['值'] = cell['显示值']
                                print(f"[SAVE DEBUG] 备注已精确更新到单元格 (28,1)")
                                updated = True
                                break
                        # 第二优先级：回退到文本匹配
                        if not updated:
                            for cell in filled_config.get('单元格数据', []):
                                显示值 = str(cell.get('显示值', '') or '')
                                if 显示值.strip().startswith('备注') or 显示值.strip() == '备注':
                                    cell['显示值'] = f"备注：\n{request.备注}"
                                    cell['值'] = cell['显示值']
                                    print(f"[SAVE DEBUG] 备注已文本匹配更新到单元格 ({cell.get('行号')},{cell.get('列号')})")
                                    updated = True
                                    break
                        if not updated:
                            print(f"[SAVE WARNING] 未找到备注单元格，无法更新备注！")
                _write_filled_to_excel(config, filled_config, xlsx_path)
            else:
                # 兜底：没有填报配置时，重新填充
                print(f"[SAVE WARNING] 未收到填报配置，走兜底路径重新填充！")
                filled_config = _do_full_fill_to_excel(config, request, xlsx_path)
                if request.备注:
                    for cell in filled_config.get('单元格数据', []):
                        显示值 = str(cell.get('显示值', '') or '')
                        if 显示值.startswith('备注'):
                            cell['显示值'] = request.备注 if not 显示值.startswith('备注：') else f"备注：{request.备注}"
                            _write_filled_to_excel(config, filled_config, xlsx_path)
                            break
                    else:
                        _write_filled_to_excel(config, filled_config, xlsx_path)

            # 保存HTML文件用于打印（与Excel同名，仅扩展名不同）
            html_filename = f"{base_name}.html"
            html_path = os.path.join(SAVED_DIR, html_filename)
            try:
                html_content = template_engine.generate_print_html(filled_config, request.查询条件, excel_path=config.get('原始文件路径'))
                with open(html_path, 'w', encoding='utf-8') as f:
                    f.write(html_content)
                print(f"[SAVE DEBUG] HTML已保存: {html_path}")
            except Exception as html_e:
                print(f"[SAVE WARNING] HTML保存失败: {html_e}")
                html_path = None

        # Word模板填充：根据源文件类型确定模板路径和填充方式
        docx_path = None
        template_name = config.get('模板名称', '')
        is_retirement_template = template_name in ['职工退休呈报表', '职工退休呈报表（Word版）']
        is_job_promotion_template = template_name == '枣阳市机关事业单位养老保险改革过渡期内职务升降退休人员信息申报表'
        
        # 确定Word模板路径：优先使用源文件（.docx），其次使用硬编码路径（兼容旧配置）
        word_template_path = None
        if is_word_template:
            word_template_path = original_file_path
        elif is_retirement_template and os.path.exists(WORD_TEMPLATE_PATH):
            word_template_path = WORD_TEMPLATE_PATH
        elif is_job_promotion_template and os.path.exists(JOB_PROMOTION_TEMPLATE_PATH):
            word_template_path = JOB_PROMOTION_TEMPLATE_PATH
        
        if word_template_path:
            try:
                docx_filename = f"{base_name}.docx"
                docx_path = os.path.join(SAVED_DIR, docx_filename)
                # 查询教师信息用于构建填报数据
                from routes.status_change_routes import _build_retirement_report_data
                conn2 = get_db_connection()
                cursor2 = conn2.cursor()
                try:
                    teacher_id_val = int(request.查询条件.get('职工ID', 0))
                    cursor2.execute("""
                        SELECT id, "姓名", "身份证号码", "档案出生日期", "民族", "籍贯", "参加工作日期"
                        FROM teacher_basic_info WHERE id = %s
                    """, (teacher_id_val,))
                    teacher_row = cursor2.fetchone()
                    if not teacher_row:
                        raise ValueError(f"未找到教师ID={teacher_id_val}")
                    id_card = teacher_row[2]
                    cursor2.execute("""
                        SELECT education, graduate_date FROM teacher_education_record
                        WHERE teacher_id = %s ORDER BY graduate_date DESC LIMIT 1
                    """, (teacher_id_val,))
                    education_row = cursor2.fetchone()
                    report_data = _build_retirement_report_data(cursor2, teacher_id_val, teacher_row, id_card, education_row)
                    if is_job_promotion_template:
                        # 职务升降申报表：额外查询个人编号
                        personal_id = ''
                        try:
                            cursor2.execute('SELECT "个人编号" FROM tui_xiu_bu_chong_xin_xi WHERE name = %s', (teacher_row[1],))
                            pi_row = cursor2.fetchone()
                            if pi_row and pi_row[0]:
                                personal_id = str(pi_row[0])
                        except Exception:
                            pass
                        _fill_job_promotion_report(report_data, docx_path, request.封面单位 or '', personal_id, template_path=word_template_path)
                        print(f"[Word] 职务升降申报表已生成: {docx_path}")
                    else:
                        _fill_word_retirement_report(report_data, docx_path, request.封面单位 or '', template_path=word_template_path)
                        print(f"[Word] 退休呈报表已生成: {docx_path}")
                finally:
                    cursor2.close()
                    conn2.close()
            except Exception as e:
                print(f"[Word] 生成失败: {e}")
                import traceback
                traceback.print_exc()

        # PDF转换：同步执行，确保保存后立即可用
        # 先确定需要转换的源文件
        pdf_source_path = None
        pdf_source_filename = None
        pdf_candidate = None
        if SOFFICE_PATH:
            if docx_path and os.path.exists(docx_path):
                pdf_source_path = docx_path
                pdf_source_filename = os.path.basename(docx_path)
            elif xlsx_path and os.path.exists(xlsx_path):
                pdf_source_path = xlsx_path
                pdf_source_filename = xlsx_filename

        now = datetime.now()
        unit_name = _get_fill_unit_name(request)
        # 使用用户选择的年月，月份+1（报表月份 = 填报月份 + 1）
        user_ym = request.查询条件.get('年月', '') if request.查询条件 else ''
        if user_ym:
            ym_match = re.match(r'^(\d{4})-(\d{1,2})$', user_ym)
            if ym_match:
                year = int(ym_match.group(1))
                month = int(ym_match.group(2)) + 1
                if month > 12:
                    month = 1
                    year += 1
                user_ym = f"{year}年{month}月"
        system_ym = user_ym if user_ym else _get_system_ym()

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            """INSERT INTO saved_exports (模板ID, 模板名称, 单位名称, 年月, 查询条件, 统计范围, 填报口径, Excel路径, PDF路径, 保存时间)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id""",
            (
                request.模板ID,
                config.get('模板名称', ''),
                unit_name or '',
                system_ym,
                json.dumps(request.查询条件, ensure_ascii=False),
                json.dumps(request.统计范围, ensure_ascii=False) if request.统计范围 else '{}',
                json.dumps(request.填报口径, ensure_ascii=False) if request.填报口径 else '{}',
                # Word模板：docx路径存入Excel路径列（Excel路径对Word模板无意义，复用此列存储Word路径）
                docx_path or xlsx_path,
                '',  # PDF路径先为空，转换后更新
                now
            )
        )
        saved_id = cursor.fetchone()[0]
        conn.commit()

        # PDF转换（同步执行，确保保存后立即可用）
        if pdf_source_path and SOFFICE_PATH:
            pdf_name = f"{base_name}.pdf"
            pdf_candidate = os.path.join(SAVED_DIR, pdf_name)

            # 修复：退休呈报表Word模板的Section 1为CONTINUOUS，
            # LibreOffice转PDF时不会自动分页，导致封面内容溢出到表格区域
            # 创建临时副本，将Section 1改为NEW_PAGE后再转换，原docx不受影响
            convert_source_path = pdf_source_path
            convert_source_filename = pdf_source_filename
            temp_docx_path = None
            if pdf_source_path.lower().endswith('.docx') and (is_retirement_template or is_job_promotion_template):
                try:
                    from docx import Document as DocxDocument
                    temp_docx_path = pdf_source_path + '.pdf_fix.tmp.docx'
                    shutil.copy2(pdf_source_path, temp_docx_path)
                    tmp_doc = DocxDocument(temp_docx_path)
                    if len(tmp_doc.sections) > 1:
                        for si in range(1, len(tmp_doc.sections)):
                            if tmp_doc.sections[si].start_type == 0:  # CONTINUOUS
                                tmp_doc.sections[si].start_type = 2  # NEW_PAGE
                                print(f"[PDF] 已修复Section {si}分页: CONTINUOUS → NEW_PAGE")
                        tmp_doc.save(temp_docx_path)
                        convert_source_path = temp_docx_path
                        convert_source_filename = os.path.basename(temp_docx_path)
                except Exception as fix_e:
                    print(f"[PDF] Section修复失败(使用原文件): {fix_e}")

            try:
                result = subprocess.run(
                    [SOFFICE_PATH, '--headless',
                     f'-env:UserInstallation=file:///{LIBREOFFICE_PROFILE.replace(os.sep, "/")}',
                     '--convert-to', 'pdf', convert_source_filename, '--outdir', SAVED_DIR],
                    capture_output=True, text=True, timeout=120,
                    cwd=SAVED_DIR
                )
                # 如果用的是临时文件转换，PDF输出名基于临时文件名，需重命名为正确名称
                actual_pdf = None
                if temp_docx_path:
                    tmp_pdf_name = os.path.basename(temp_docx_path).replace('.docx', '.pdf')
                    actual_pdf = os.path.join(SAVED_DIR, tmp_pdf_name)
                    if os.path.exists(actual_pdf):
                        if os.path.exists(pdf_candidate):
                            os.remove(pdf_candidate)
                        os.rename(actual_pdf, pdf_candidate)

                if result.returncode == 0 and os.path.exists(pdf_candidate):
                    cursor.execute(
                        'UPDATE saved_exports SET pdf路径 = %s WHERE id = %s',
                        (pdf_candidate, saved_id)
                    )
                    conn.commit()
                    print(f"[PDF] 转换成功: {pdf_candidate}")
                else:
                    print(f"[PDF] 转换失败: {result.stderr}")
            except Exception as e:
                print(f"[PDF] 转换异常: {e}")
            finally:
                # 清理临时文件
                if temp_docx_path and os.path.exists(temp_docx_path):
                    try:
                        os.remove(temp_docx_path)
                    except Exception:
                        pass

        cursor.close()
        conn.close()

        # ============================================================
        # 方案A兼容写入：当保存的是绩效工资审批表时，额外写一份JSON到旧目录
        # 确保绩效工资历史、统计、上传模块能继续读取数据
        # ============================================================
        if request.模板ID == 'tpl_15cc984d' and filled_config and not is_word_template:
            try:
                _write_compatibility_json(request, filled_config, config)
            except Exception as compat_e:
                print(f"[兼容写入] 绩效工资审批表兼容JSON写入失败: {compat_e}")

        return {
            "成功": True,
            "消息": "保存成功",
            "数据": {
                "Excel文件": xlsx_filename,
                "Word文件": os.path.basename(docx_path) if docx_path else None,
                "PDF文件": os.path.basename(pdf_candidate) if (pdf_source_path and SOFFICE_PATH and os.path.exists(pdf_candidate)) else None,
                "保存时间": now.strftime("%Y-%m-%d %H:%M:%S"),
                "记录ID": saved_id
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"保存失败: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export-preview/{template_id}")
async def export_preview(template_id: str):
    """
    导出预览（直接返回原始模板文件，保证100%格式一致）
    
    参数:
        template_id: 模板ID
    
    返回:
        Excel文件下载
    """
    try:
        config = template_engine.load_template_config(template_id)
        if not config:
            raise HTTPException(status_code=404, detail="模板不存在")
        
        original_file_path = config.get('原始文件路径')
        original_filename = config.get('原始文件', '')
        
        if not original_file_path or not os.path.exists(original_file_path):
            original_file_path = os.path.join(TEMPLATE_DIR, original_filename)
        
        if not os.path.exists(original_file_path):
            raise HTTPException(status_code=404, detail="原始文件不存在")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        is_docx = original_file_path.lower().endswith('.docx')
        ext = '.docx' if is_docx else '.xlsx'
        filename = f"{config['模板名称']}_{timestamp}{ext}"
        output_path = os.path.join(EXPORT_DIR, filename)
        
        shutil.copy2(original_file_path, output_path)
        
        if is_docx:
            media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        else:
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        return FileResponse(
            output_path,
            filename=filename,
            media_type=media_type
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/download-template/{template_id}")
async def download_template(template_id: str):
    """
    下载原始模板文件（空表）

    参数:
        template_id: 模板ID
    返回:
        原始Excel文件下载
    """
    try:
        config = template_engine.load_template_config(template_id)
        if not config:
            raise HTTPException(status_code=404, detail="模板不存在")

        original_file_path = config.get('原始文件路径', '')
        original_filename = config.get('原始文件', '')

        if not original_file_path or not os.path.exists(original_file_path):
            original_file_path = os.path.join(TEMPLATE_DIR, original_filename)

        if not os.path.exists(original_file_path):
            raise HTTPException(status_code=404, detail="原始模板文件不存在")

        is_docx = original_file_path.lower().endswith('.docx')
        ext = '.docx' if is_docx else '.xlsx'
        download_filename = f"{config['模板名称']}_模板{ext}"
        if is_docx:
            media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        else:
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return FileResponse(
            original_file_path,
            filename=download_filename,
            media_type=media_type
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/check-libreoffice")
async def check_libreoffice():
    """检查服务器是否安装了LibreOffice"""
    return {
        "成功": True,
        "可用": SOFFICE_PATH is not None,
        "路径": SOFFICE_PATH or ""
    }


@router.post("/export-pdf")
async def export_pdf(request: FillTemplateRequest):
    """
    导出PDF（优先从已保存文件下载，不再重新填充数据）
    """
    if not SOFFICE_PATH:
        raise HTTPException(status_code=503, detail="服务器未安装LibreOffice，PDF导出不可用")

    try:
        # 新流程：通过保存记录ID下载已保存PDF文件
        saved_id = getattr(request, 'saved_export_id', None)
        
        if saved_id:
            return await download_history_file(saved_id, 'PDF')
        
        # 旧流程兼容：查找最近一次保存记录的PDF
        config = template_engine.load_template_config(request.模板ID)
        if not config:
            raise HTTPException(status_code=404, detail="模板不存在")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'SELECT id FROM saved_exports WHERE "模板ID" = %s ORDER BY "保存时间" DESC LIMIT 1',
            (request.模板ID,)
        )
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if row:
            return await download_history_file(row[0], 'PDF')
        
        raise HTTPException(status_code=404, detail="没有找到已保存的PDF文件，请先点击「填报」保存后再导出")
    except HTTPException:
        raise
    except Exception as e:
        print(f"PDF导出失败: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/available-tables")
async def get_available_tables():
    """
    获取可用的数据表列表（用于字段映射配置的数据源表选择）
    从数据库中动态读取所有表，结合 table_name_mappings.json 提供中文表名
    """
    try:
        import psycopg2
        from psycopg2.extras import RealDictCursor
        
        conn = get_db_connection()
        if not conn:
            raise HTTPException(status_code=500, detail="无法连接数据库")
        
        cur = conn.cursor()
        
        cur.execute("""
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema = 'public'
            AND table_type IN ('BASE TABLE', 'VIEW')
            AND table_name NOT LIKE '_prisma_%'
            AND table_name NOT LIKE 'pg_%'
            ORDER BY table_name
        """)
        db_tables = [row[0] for row in cur.fetchall()]
        cur.close()
        conn.close()
        
        MAPPINGS_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'table_name_mappings.json')
        chinese_map = {}
        if os.path.exists(MAPPINGS_FILE):
            with open(MAPPINGS_FILE, 'r', encoding='utf-8') as f:
                mappings_data = json.load(f)
            chinese_map = mappings_data.get('reverse_mappings', {})
        
        tables = []
        for eng_name in db_tables:
            chinese_name = chinese_map.get(eng_name, None)
            tables.append({
                '英文表名': eng_name,
                '中文表名': chinese_name,
                '显示名称': chinese_name if chinese_name else eng_name
            })
        
        return {'成功': True, '数据': tables}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# 全局缓存和翻译器
_field_configs_cache = {}
_en_to_cn_global = {}
_cn_to_en = {}

def _build_global_en_to_cn_map():
    global _en_to_cn_global, _cn_to_en
    if _en_to_cn_global:
        return
    import os as _os
    configs_dir = _os.path.join(_os.path.dirname(_os.path.dirname(__file__)), 'config', 'field_configs')
    if not _os.path.exists(configs_dir):
        return
    for filename in _os.listdir(configs_dir):
        if not filename.endswith('.json'):
            continue
        filepath = _os.path.join(configs_dir, filename)
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            for fc in data.get('field_configs', []):
                sf = fc.get('sourceField', '')
                tf = fc.get('targetField', sf)
                rdf = fc.get('relation_display_field', '')
                rt = fc.get('relation_type', '')
                if tf and tf != sf and all(c.isascii() and (c.isalnum() or c == '_') for c in tf):
                    if tf not in _en_to_cn_global:
                        _en_to_cn_global[tf] = sf
                        _cn_to_en[sf] = tf
                if rt in ('to_master', 'to_dict') and rdf and rdf != sf and all(c.isascii() and (c.isalnum() or c == '_') for c in rdf):
                    if rdf not in _en_to_cn_global:
                        _en_to_cn_global[rdf] = sf
        except Exception:
            pass

_COLUMN_WORD_MAP = {
    'id': 'ID', 'name': '名称', 'code': '编码', 'type': '类型', 'date': '日期',
    'time': '时间', 'year': '年份', 'month': '月份', 'status': '状态',
    'title': '标题', 'description': '描述', 'remark': '备注', 'remarks': '备注',
    'amount': '金额', 'count': '数量', 'people': '人数', 'person': '人',
    'standard': '标准', 'total': '合计', 'subtotal': '小计',
    'unit': '单位', 'level': '级别', 'category': '类别', 'tag': '标签',
    'school': '学校', 'teacher': '教师', 'student': '学生', 'employee': '职工',
    'phone': '电话', 'address': '地址', 'email': '邮箱', 'gender': '性别',
    'birth': '出生', 'age': '年龄', 'work': '工作', 'start': '开始', 'end': '结束',
    'entry': '进入', 'retire': '退休', 'retired': '退休', 'death': '死亡',
    'salary': '工资', 'pay': '工资', 'wage': '工资', 'subsidy': '补贴',
    'performance': '绩效', 'approval': '审批', 'report': '报表', 'export': '导出',
    'template': '模板', 'config': '配置', 'setting': '设置', 'rule': '规则',
    'detail': '明细', 'record': '记录', 'log': '日志', 'history': '历史',
    'backup': '备份', 'data': '数据', 'info': '信息', 'file': '文件',
    'path': '路径', 'source': '来源', 'target': '目标', 'note': '说明',
    'notes': '说明', 'content': '内容', 'value': '值', 'key': '键',
    'order': '顺序', 'sort': '排序', 'active': '激活', 'enabled': '启用',
    'effective': '生效', 'expiry': '到期', 'cancel': '取消', 'change': '变动',
    'registration': '登记', 'certificate': '证书', 'cert': '证书',
    'major': '专业', 'degree': '学位', 'education': '教育', 'training': '培训',
    'professional': '专业', 'technical': '技术', 'position': '岗位', 'post': '岗位',
    'job': '职务', 'duty': '职务', 'appointment': '聘任', 'hire': '雇用',
    'native': '籍贯', 'place': '地点', 'household': '户籍', 'ethnicity': '民族',
    'cadre': '干部', 'worker': '工人', 'official': '干部',
    'current': '当前', 'new': '新', 'old': '旧', 'original': '原始',
    'monthly': '月度', 'yearly': '年度', 'daily': '每日',
    'created': '创建', 'updated': '更新', 'creator': '创建人', 'updater': '更新人',
    'handler': '处理人', 'assignee': '负责人', 'operator': '操作人',
    'parent': '父级', 'child': '子级', 'root': '根',
    'prefix': '前缀', 'suffix': '后缀', 'display': '显示', 'hidden': '隐藏',
    'required': '必填', 'optional': '可选', 'default': '默认',
    'format': '格式', 'length': '长度', 'width': '宽度', 'height': '高度',
    'page': '页面', 'margin': '边距', 'orientation': '方向', 'landscape': '横向',
    'portrait': '纵向', 'paper': '纸张', 'size': '尺寸',
    'header': '表头', 'footer': '表尾', 'body': '正文', 'row': '行', 'col': '列',
    'merge': '合并', 'cell': '单元格', 'region': '区域',
    'x0': '左', 'y0': '上', 'x1': '右', 'y1': '下',
    'legacy': '遗留', 'issue': '问题', 'issues': '问题',
    'town': '乡镇', 'village': '村', 'city': '城市', 'province': '省份',
    'nation': '国家', 'country': '国家',
    'bank': '银行', 'account': '账户', 'card': '卡',
    'contact': '联系', 'relation': '关系', 'family': '家庭', 'member': '成员',
    'party': '党派', 'join': '加入', 'only': '独生', 'child_only': '独生子女',
    'pension': '退休金', 'insurance': '保险',
    'census': '户籍', 'residence': '居住', 'current_residence': '现居住地',
    'todo': '待办', 'task': '任务', 'trigger': '触发', 'condition': '条件',
    'message': '消息', 'notification': '通知',
    'filled': '已填报', 'saved': '已保存', 'exported': '已导出',
    'uploaded': '已上传', 'downloaded': '已下载', 'scanned': '已扫描',
    'generate': '生成', 'generated': '已生成', 'generation': '生成',
    'fill': '填报', 'filling': '填报',
    'step': '步骤', 'flow': '流程', 'process': '过程', 'progress': '进度',
    'completed': '已完成', 'returned': '已退回', 'return': '退回',
    'archive': '档案', 'archived': '已归档',
    'admin': '管理', 'manage': '管理', 'management': '管理',
    'system': '系统', 'user': '用户', 'role': '角色', 'permission': '权限',
    'password': '密码', 'token': '令牌', 'session': '会话',
    'api': '接口', 'endpoint': '端点', 'request': '请求', 'response': '响应',
    'error': '错误', 'warning': '警告', 'success': '成功', 'failed': '失败',
    'pending': '待处理', 'processing': '处理中', 'reviewing': '审核中',
    'approved': '已批准', 'rejected': '已拒绝',
    'pdf': 'PDF', 'excel': 'Excel', 'word': 'Word', 'xml': 'XML',
    'json': 'JSON', 'html': 'HTML', 'csv': 'CSV', 'image': '图片',
    'text': '文本', 'number': '数字', 'string': '字符串', 'boolean': '布尔',
    'integer': '整数', 'float': '浮点', 'decimal': '小数',
    'select': '选择', 'input': '输入', 'output': '输出', 'search': '搜索',
    'filter': '过滤', 'query': '查询', 'sort_by': '排序依据',
    'group': '分组', 'aggregate': '聚合', 'sum': '求和', 'avg': '平均',
    'max': '最大', 'min': '最小', 'count': '计数', 'product': '乘积',
    'calculation': '计算', 'calculate': '计算', 'formula': '公式',
    'field': '字段', 'table': '表', 'column': '列', 'row_num': '行号',
    'col_num': '列号', 'cell_num': '单元格号',
    'mapping': '映射', 'relation': '关系', 'reference': '引用',
    'foreign': '外键', 'primary': '主键', 'unique_key': '唯一键',
    'index': '索引', 'constraint': '约束', 'schema': '模式',
    'auto': '自动', 'manual': '手动', 'batch': '批量', 'single': '单个',
    'multi': '多个', 'all': '全部', 'none': '无', 'any': '任意',
    'every': '每个', 'each': '每', 'per': '每',
    'before': '之前', 'after': '之后', 'during': '期间',
    'from': '从', 'to': '到', 'by': '由', 'via': '通过',
    'online': '在线', 'offline': '离线', 'local': '本地', 'remote': '远程',
    'internal': '内部', 'external': '外部', 'public': '公开', 'private': '私有',
    'result': '结果', 'summary': '汇总', 'statistics': '统计', 'stats': '统计',
    'total_people': '总人数', 'total_amount': '总金额',
    'people_count': '人数', 'subsidy_amount': '补贴金额',
    'subsidy_standard': '补贴标准', 'subsidy_people': '补贴人数',
    'report_unit': '填报单位', 'report_date': '填报日期', 'report_period': '填报期间',
    'level_name': '级别名称', 'level_code': '级别编码',
    'monthly_standard': '月度标准', 'effective_date': '生效日期',
    'retired_cadre_count': '退休干部人数', 'retired_worker_count': '退休工人人数',
    'legacy_total_amount': '遗留总金额', 'legacy_total_people': '遗留总人数',
    'no_subsidy_count': '无补贴人数', 'excel_file_path': 'Excel文件路径',
    'pdf_file_path': 'PDF文件路径', 'scanned_file_path': '扫描文件路径',
    'teacher_id': '教师ID', 'teacher_name': '教师姓名',
    'id_card': '身份证号码', 'id_card_1': '身份证号码1', 'id_card_2': '身份证号码2',
    'birth_date': '出生日期', 'archive_birth_date': '档案出生日期',
    'work_start_date': '参加工作日期', 'retirement_date': '退休日期',
    'contact_phone': '联系电话', 'contact_phone_1': '联系电话1',
    'contact_phone_2': '联系电话2',
    'created_at': '创建时间', 'updated_at': '更新时间',
    'created_by': '创建人', 'updated_by': '更新人',
    'is_active': '是否启用', 'is_enabled': '是否启用',
    'is_cadre': '是否干部', 'is_only_child': '是否独生子女',
    'is_landscape': '是否横向', 'is_required': '是否必填',
    'is_visible': '是否可见',
    'is_read': '是否已读', 'is_admin': '是否管理员',
    'employment_status': '任职状态', 'work_years': '工作年限',
    'job_title': '职务', 'job_title_post': '职务岗位',
    'post_level': '岗位等级', 'post_level_1': '岗位等级1',
    'post_time': '岗位时间',
    'professional_title': '职称', 'professional_title_1': '职称1',
    'professional_title_2': '职称2', 'professional_title_3': '职称3',
    'professional_title_time': '职称取得时间',
    'graduation_school': '毕业学校', 'graduation_date': '毕业日期',
    'entry_date': '进入本单位日期', 'pension_unit': '退休金单位',
    'retirement_reason': '退休原因', 'retirement_address': '退休后居住地址',
    'retirement_certificate_number': '退休证编号',
    'birth_date_display': '出生日期显示', 'work_start_date_display': '参加工作日期显示',
    'work_years_display': '工作年限显示',
    'receive_date': '签收日期', 'recipient_name': '签收人',
    'sort_order': '排序号', 'task_items': '任务项',
    'due_date': '到期日期', 'plan_date': '计划日期',
    'remind_days': '提醒天数', 'advance_notice': '提前通知',
    'completed_at': '完成时间', 'started_at': '开始时间',
    'returned_at': '退回时间', 'return_reason': '退回原因',
    'return_count': '退回次数',
    'handle_time': '处理时间', 'handle_note': '处理备注',
    'trigger_time': '触发时间', 'trigger_reason': '触发原因',
    'trigger_type': '触发类型', 'trigger_value': '触发值',
    'listen_table': '监听表', 'listen_field': '监听字段',
    'old_value': '旧值', 'new_value': '新值',
    'template_id': '模板ID', 'template_code': '模板编码',
    'template_name': '模板名称', 'template_type': '模板类型',
    'file_name': '文件名', 'file_path': '文件路径', 'file_size': '文件大小',
    'file_type': '文件类型', 'file_format': '文件格式',
    'business_type': '业务类型', 'business_id': '业务ID',
    'related_type': '关联类型', 'related_id': '关联ID',
    'action': '操作', 'action_taken': '已执行操作',
    'fill_type': '填报类型', 'fill_target': '填报目标',
    'fill_date': '填报日期', 'filled_by': '填报人',
    'original_filename': '原始文件名', 'original_file_path': '原始文件路径',
    'original_status': '原始状态', 'new_status': '新状态',
    'original_post': '原始岗位', 'new_post': '新岗位',
    'original_retirement_date': '原始退休日期', 'new_retirement_date': '新退休日期',
    'delay_months': '延迟月数', 'estimate_type': '估算类型',
    'change_type': '变动类型', 'change_date': '变动日期',
    'change_content': '变动内容', 'change_reason': '变动原因',
    'change_category': '变动类别', 'change_detail': '变动详情',
    'effective_month': '生效月份', 'expiry_month': '到期月份',
    'cancel_performance_month': '取消绩效月份',
    'generation_params': '生成参数', 'generated_file_path': '生成文件路径',
    'generated_file_name': '生成文件名', 'generated_at': '生成时间',
    'generated_by': '生成人', 'error_message': '错误信息',
    'template_file_path': '模板文件路径',
    'source_type': '来源类型', 'source_config': '来源配置',
    'intermediate_table': '中间表', 'intermediate_table_cn': '中间表中文名',
    'intermediate_field': '中间字段', 'intermediate_field_cn': '中间字段中文名',
    'placeholder_name': '占位符名', 'placeholder_config': '占位符配置',
    'placeholder': '占位符', 'placeholders': '占位符',
    'activation_type': '激活类型', 'activation_config': '激活配置',
    'label_pattern': '标签模式', 'date_format': '日期格式',
    'format_type': '格式类型', 'table_index': '表序号',
    'row_index': '行序号', 'cell_index': '单元格序号',
    'page_num': '页码', 'x_pos': 'X位置', 'y_pos': 'Y位置',
    'css_selector': 'CSS选择器', 'page_width': '页面宽度',
    'page_height': '页面高度', 'paper_size': '纸张尺寸',
    'margin_top': '上边距', 'margin_bottom': '下边距',
    'margin_left': '左边距', 'margin_right': '右边距',
    'width_cm': '宽度厘米', 'height_cm': '高度厘米',
    'margin_left_cm': '左边距厘米', 'margin_right_cm': '右边距厘米',
    'margin_top_cm': '上边距厘米', 'margin_bottom_cm': '下边距厘米',
    'full_path': '完整路径', 'parent_id': '父级ID',
    'school_dict_id': '学校字典ID', 'unit_hierarchy': '单位层级',
    'unit_level': '单位级别', 'unit_name': '单位名称',
    'module_id': '模块ID', 'modules_data': '模块数据',
    'backup_name': '备份名称',
    'record_id': '记录ID', 'field_id': '字段ID', 'field_name': '字段名',
    'field_label': '字段标签', 'field_type': '字段类型',
    'field_value': '字段值', 'value_type': '值类型',
    'calculation_formula': '计算公式', 'depends_on_fields': '依赖字段',
    'logic_type': '逻辑类型', 'logic_operator': '逻辑运算符',
    'aggregate_func': '聚合函数', 'filter_condition': '过滤条件',
    'tag_name': '标签名', 'tag_name_cn': '标签中文名',
    'tag_id': '标签ID', 'tag_color': '标签颜色',
    'condition_name': '条件名称', 'condition_id': '条件ID',
    'user_id': '用户ID', 'user_name': '用户名',
    'assignee_id': '负责人ID', 'assignee_name': '负责人姓名',
    'related_teacher_id': '关联教师ID', 'related_teacher_name': '关联教师姓名',
    'log_id': '日志ID', 'usage_count': '使用次数',
    'last_used_at': '最后使用时间',
    'row_count': '行数', 'col_count': '列数',
    'cell_count': '单元格数', 'merge_count': '合并单元格数',
    'metadata_path': '元数据路径',
    'config_json': '配置JSON', 'data_json': '数据JSON',
    'progress_update_time': '进度更新时间', 'start_process_time': '开始处理时间',
    'complete_time': '完成时间', 'create_time': '创建时间',
}

def _generate_chinese_column_name(col_name):
    _build_global_en_to_cn_map()
    
    if col_name in _COLUMN_WORD_MAP:
        return _COLUMN_WORD_MAP[col_name]
    
    if col_name in _en_to_cn_global:
        return _en_to_cn_global[col_name]
    
    parts = col_name.split('_')
    if len(parts) == 1:
        return col_name
    
    translated = []
    for part in parts:
        if part in _COLUMN_WORD_MAP:
            translated.append(_COLUMN_WORD_MAP[part])
        else:
            translated.append(part)
    
    if translated[0] == 'is' and len(translated) >= 2:
        rest = ''.join(translated[1:])
        return '是否' + rest
    
    result = ''.join(translated)
    
    if result == col_name:
        return col_name
    
    return result

def _resolve_chinese_name(col_name, field_relations, en_to_cn, source_fields, biz_col_names, field_name_map=None):
    rel = field_relations.get(col_name, {})
    chinese_name = rel.get('中文字段名', '')
    
    if not chinese_name:
        chinese_name = en_to_cn.get(col_name, '')
    
    if not chinese_name and field_name_map:
        chinese_name = field_name_map.get(col_name.lower(), '')
    
    if not chinese_name:
        try:
            idx = biz_col_names.index(col_name) if col_name in biz_col_names else -1
            if 0 <= idx < len(source_fields):
                chinese_name = source_fields[idx]
        except ValueError:
            pass
    
    if not chinese_name:
        chinese_name = _generate_chinese_column_name(col_name)
    
    return chinese_name

def _load_dict_relations_from_configs(table_name):
    """从 field_configs 目录加载: dict_relations(字典关联), source_fields(中文字段名顺序列表), en_to_cn(英文→中文映射), field_name_map(英文字段名→中文字段名映射)"""
    import os as _os
    configs_dir = _os.path.join(_os.path.dirname(_os.path.dirname(__file__)), 'config', 'field_configs')
    if not _os.path.exists(configs_dir):
        return [], [], {}, {}
    
    dict_relations = []
    source_fields = []
    en_to_cn = {}
    field_name_map = {}
    
    for filename in _os.listdir(configs_dir):
        if not filename.endswith('.json'):
            continue
        filepath = _os.path.join(configs_dir, filename)
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            t_name = data.get('table_name', '')
            
            for fc in data.get('field_configs', []):
                sf = fc.get('sourceField', '')
                tf = fc.get('targetField', sf)
                rdf = fc.get('relation_display_field', '')
                rt = fc.get('relation_type', '')
                
                if t_name == table_name:
                    if tf.lower() not in field_name_map:
                        field_name_map[tf.lower()] = sf
                    if rt == 'to_dict':
                        dict_relations.append({
                            '中文字段名': sf,
                            'targetField': tf,
                            '字典表': fc.get('relation_table', ''),
                            '字典值字段': rdf
                        })
                    source_fields.append(sf)
                
                if rt == 'to_master' and sf and rdf and sf != rdf:
                    if rdf not in en_to_cn:
                        en_to_cn[rdf] = sf
                elif rt == 'to_dict' and sf and rdf and sf != rdf:
                    if rdf not in en_to_cn:
                        en_to_cn[rdf] = sf
                    
        except Exception:
            pass
    
    return dict_relations, source_fields, en_to_cn, field_name_map

def _query_dict_values(dict_table, value_field):
    """查询字典表的所有可选值"""
    try:
        conn = get_db_connection()
        if not conn:
            return []
        cur = conn.cursor()
        cur.execute(f"SELECT DISTINCT \"{value_field}\" FROM \"{dict_table}\" WHERE \"{value_field}\" IS NOT NULL AND \"{value_field}\" != '' ORDER BY \"{value_field}\"")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return [row[0] for row in rows]
    except Exception:
        return []

@router.get("/table-columns/{table_name}")
async def get_table_columns(table_name: str):
    """获取指定表的所有列信息（用于字段映射下拉选择）"""
    try:
        import psycopg2
        
        conn = get_db_connection()
        if not conn:
            raise HTTPException(status_code=500, detail="无法连接数据库")
        
        cur = conn.cursor()
        
        cur.execute("""
            SELECT column_name, data_type
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = %s
            ORDER BY ordinal_position
        """, (table_name,))
        columns = [{'字段名': row[0], '数据类型': row[1]} for row in cur.fetchall()]
        
        cur.execute("SELECT field_name, 关联类型, 关联表, 关联显示字段, 字段名 FROM table_field_relations WHERE table_name = %s", (table_name,))
        field_relations = {}
        for row in cur.fetchall():
            field_relations[row[0]] = {
                '关联类型': row[1], 
                '字典表': row[2], 
                '字典值字段': row[3],
                '中文字段名': row[4]
            }
        
        cur.close()
        conn.close()
        
        dict_relations_list, source_fields, en_to_cn, field_name_map = _load_dict_relations_from_configs(table_name)
        
        skip_cols = {'id', 'created_at', 'updated_at'}
        biz_col_names = [c['字段名'] for c in columns if c['字段名'] not in skip_cols]
        
        for col in columns:
            col_name = col['字段名']
            
            chinese_name = _resolve_chinese_name(col_name, field_relations, en_to_cn, source_fields, biz_col_names, field_name_map)
            
            col['中文字段名'] = chinese_name
            col['显示名称'] = chinese_name
        
        return {'成功': True, '数据': columns}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/dict-values/{table_name}")
async def get_dict_values(table_name: str):
    """
    获取字典表的所有可选值列表
    """
    try:
        DICT_CONFIGS_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config', 'dict_tables_config.json')
        value_field = None
        label_field = None
        
        if os.path.exists(DICT_CONFIGS_FILE):
            with open(DICT_CONFIGS_FILE, 'r', encoding='utf-8') as f:
                dict_configs = json.load(f)
            for dt in dict_configs.get('dict_tables', []):
                if dt['table_name'] == table_name:
                    value_field = dt.get('value_field', '')
                    label_field = dt.get('label_field', '')
                    break
        
        if not value_field:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("SELECT column_name FROM information_schema.columns WHERE table_schema = 'public' AND table_name = %s ORDER BY ordinal_position LIMIT 1", (table_name,))
            first_col = cur.fetchone()
            cur.close()
            conn.close()
            value_field = first_col[0] if first_col else 'id'
            label_field = value_field
        
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(f"SELECT DISTINCT \"{label_field}\" FROM \"{table_name}\" WHERE \"{label_field}\" IS NOT NULL AND \"{label_field}\" != '' ORDER BY \"{label_field}\"")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        
        values = [{'值': row[0], '显示名称': str(row[0])} for row in rows]
        return {'成功': True, '数据': values, '字典表': table_name, '值字段': value_field}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/table-distinct-values/{table_name}/{column_name}")
async def get_table_distinct_values(table_name: str, column_name: str):
    """查询指定表指定字段的可选值列表（用于前端下拉筛选）。
    
    原则：数据源表决定数据源字段，数据源字段决定可选值来源。
    
    - 如果字段有关联字典表 → 从字典表查询全部值（全集），保证通用性。
      字典表是标准全集，不同单位可能只用到其中一部分。
      如果只取当前表的子集，未来新出现的值将无法被筛选，导致数据失真。
    - 如果字段无字典关联 → 从当前表查 distinct 值（保持原有逻辑）。
    """
    try:
        conn = get_db_connection()
        if not conn:
            raise HTTPException(status_code=500, detail="无法连接数据库")
        cur = conn.cursor()

        dict_table = None
        dict_id_field = 'id'
        dict_label_field = None

        cur.execute("""
            SELECT "关联表", "关联字段", "关联显示字段"
            FROM table_field_relations
            WHERE "表名" = %s AND "字段名" = %s AND "关联类型" = 'dict' AND "关联表" IS NOT NULL
            LIMIT 1
        """, (table_name, column_name))
        rel = cur.fetchone()

        if rel:
            dict_table = rel[0]
            dict_id_field = rel[1] or 'id'
            dict_label_field = rel[2]
        else:
            dict_relations_list, _, _, _ = _load_dict_relations_from_configs(table_name)
            for dr in dict_relations_list:
                tf = dr.get('targetField', '')
                if tf.lower() == column_name.lower():
                    dict_table = dr.get('字典表', '')
                    dict_label_field = dr.get('字典值字段', '')
                    dict_id_field = 'id'
                    break

        if dict_table and dict_label_field:
            if not dict_label_field:
                cur.execute("""
                    SELECT column_name FROM information_schema.columns
                    WHERE table_schema = 'public' AND table_name = %s
                    ORDER BY ordinal_position LIMIT 1
                """, (dict_table,))
                lc = cur.fetchone()
                dict_label_field = lc[0] if lc else 'id'

            try:
                query = f'SELECT "{dict_id_field}", "{dict_label_field}" FROM "{dict_table}" ORDER BY "{dict_id_field}"'
                cur.execute(query)
                rows = cur.fetchall()
                resolved = [{'值': row[0], '标签': str(row[1]) if row[1] is not None else str(row[0])} for row in rows]
                cur.close()
                conn.close()
                return {'成功': True, '数据': resolved, '表名': table_name, '字段名': column_name, '数量': len(resolved), '已解析': True}
            except Exception:
                pass

        cur.execute("""
            SELECT data_type FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = %s AND column_name = %s
        """, (table_name, column_name))
        col_info = cur.fetchone()
        data_type = (col_info[0] or '').lower() if col_info else ''

        numeric_types = {'integer', 'bigint', 'smallint', 'numeric', 'decimal', 'real', 'double precision', 'serial', 'bigserial', 'smallserial'}
        is_numeric = data_type in numeric_types

        if is_numeric:
            query = f'SELECT DISTINCT "{column_name}" FROM "{table_name}" WHERE "{column_name}" IS NOT NULL ORDER BY "{column_name}"'
        else:
            query = f'SELECT DISTINCT "{column_name}" FROM "{table_name}" WHERE "{column_name}" IS NOT NULL AND "{column_name}" != \'\' ORDER BY "{column_name}"'

        cur.execute(query)
        rows = cur.fetchall()
        raw_values = [row[0] for row in rows]

        cur.close()
        conn.close()

        if not raw_values:
            return {'成功': True, '数据': [], '表名': table_name, '字段名': column_name, '数量': 0}

        return {'成功': True, '数据': raw_values, '表名': table_name, '字段名': column_name, '数量': len(raw_values)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tags")
async def get_all_tags():
    """
    获取所有标签列表
    从 personal_dict_dictionary 表动态读取
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT id, biao_qian FROM personal_dict_dictionary ORDER BY id")
        rows = cur.fetchall()
        cur.close()
        conn.close()

        tags = [{'id': r[0], '标签名称': r[1]} for r in rows]

        return {
            '成功': True,
            '数据': tags
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/search-employee")
async def search_employee(
    keyword: str = Query(..., description="搜索关键词（身份证号/姓名/ID）")
):
    """
    根据身份证号、姓名或ID搜索职工
    """
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        if keyword.isdigit() and len(keyword) < 10:
            cur.execute(
                """SELECT id, \"姓名\", \"身份证号码\" FROM teacher_basic_info WHERE id = %s LIMIT 5""",
                (int(keyword),)
            )
        else:
            cur.execute(
                """SELECT id, \"姓名\", \"身份证号码\" FROM teacher_basic_info
                   WHERE \"姓名\" LIKE %s OR \"身份证号码\" LIKE %s LIMIT 10""",
                (f"%{keyword}%", f"%{keyword}%")
            )

        rows = cur.fetchall()
        cur.close()
        conn.close()

        results = [{'职工ID': r[0], '姓名': r[1], '身份证号': r[2] or ''} for r in rows]

        return {'成功': True, '数据': results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/saved-files/{template_id}")
async def get_saved_files(template_id: str, 年月: Optional[str] = None):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        formatted_ym = 年月
        if 年月:
            ym_match = re.match(r'^(\d{4})-(\d{1,2})$', 年月)
            if ym_match:
                y, m = ym_match.group(1), ym_match.group(2)
                formatted_ym = f"{y}年{int(m)}月"

        if formatted_ym:
            cursor.execute(
                """SELECT id, 模板名称, 单位名称, 年月, Excel路径, PDF路径, 保存时间
                   FROM saved_exports WHERE 模板ID = %s AND 年月 = %s
                   ORDER BY 保存时间 DESC""",
                (template_id, formatted_ym)
            )
            rows = cursor.fetchall()
        else:
            cursor.execute(
                """SELECT id, 模板名称, 单位名称, 年月, Excel路径, PDF路径, 保存时间
                   FROM saved_exports WHERE 模板ID = %s
                   ORDER BY 保存时间 DESC LIMIT 50""",
                (template_id,)
            )
            rows = cursor.fetchall()
        cursor.close()
        conn.close()
        results = []
        for row in rows:
            excel_path = row[4]
            pdf_path = row[5]
            # Word文件检测：优先从Excel路径检测（Word模板时将docx路径存入Excel路径列）
            # 其次从PDF路径推导（兼容旧数据，PDF路径存储后Word与PDF同名）
            has_word = False
            if excel_path and excel_path.endswith('.docx') and os.path.exists(excel_path):
                has_word = True
            elif pdf_path and os.path.exists(pdf_path):
                word_candidate = pdf_path.replace('.pdf', '.docx')
                has_word = os.path.exists(word_candidate)
            # Excel文件检测：Excel路径非docx且存在
            has_excel = False
            if excel_path and not excel_path.endswith('.docx') and os.path.exists(excel_path):
                has_excel = True
            # HTML文件检测：与Excel/Word同名，仅扩展名不同
            has_html = False
            if excel_path:
                html_candidate = os.path.splitext(excel_path)[0] + '.html'
                has_html = os.path.exists(html_candidate)
            results.append({
                "ID": row[0],
                "模板名称": row[1],
                "单位名称": row[2],
                "年月": row[3],
                "有Excel": has_excel,
                "有PDF": bool(pdf_path and os.path.exists(pdf_path)),
                "有HTML": has_html,
                "有Word": has_word,
                "保存时间": row[6].strftime("%Y-%m-%d %H:%M:%S") if hasattr(row[6], 'strftime') else str(row[6])
            })
        return {"成功": True, "数据": results}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{template_id}")
async def delete_template(template_id: str):
    try:
        result = template_engine.delete_template(template_id)
        if not result.get("成功"):
            raise HTTPException(status_code=404, detail=result.get("消息", "删除失败"))
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class HistoryQueryRequest(BaseModel):
    模板ID: str
    单位名称: Optional[str] = None
    起始日期: Optional[str] = None
    截止日期: Optional[str] = None


@router.post("/history")
async def query_history(request: HistoryQueryRequest):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        conditions = ["模板ID = %s"]
        params = [request.模板ID]
        if request.单位名称:
            conditions.append("单位名称 = %s")
            params.append(request.单位名称)
        if request.起始日期:
            conditions.append("保存时间 >= %s::date")
            params.append(request.起始日期)
        if request.截止日期:
            conditions.append("保存时间 <= %s::date + interval '1 day'")
            params.append(request.截止日期)
        where = " AND ".join(conditions)
        sql = f'SELECT id, 模板名称, 单位名称, 年月, 查询条件, Excel路径, PDF路径, 保存时间 FROM saved_exports WHERE {where} ORDER BY 保存时间 DESC'
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        results = []
        for row in rows:
            pdf_path = row[6]
            # Word路径：从PDF路径推导
            word_path = None
            if pdf_path and os.path.exists(pdf_path):
                word_candidate = pdf_path.replace('.pdf', '.docx')
                if os.path.exists(word_candidate):
                    word_path = word_candidate
            entry = {
                "ID": row[0],
                "模板名称": row[1],
                "单位名称": row[2],
                "年月": row[3],
                "查询条件": row[4] if isinstance(row[4], dict) else json.loads(row[4]) if row[4] else {},
                "Excel路径": row[5],
                "PDF路径": pdf_path,
                "Word路径": word_path,
                "保存时间": row[7].strftime("%Y-%m-%d %H:%M:%S") if hasattr(row[7], 'strftime') else str(row[7])
            }
            results.append(entry)
        cursor.close()
        conn.close()
        return {"成功": True, "数据": results}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/history-file/{record_id}")
async def download_history_file(record_id: int, format: str = "Excel", inline: int = 0):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT "excel路径", "pdf路径" FROM saved_exports WHERE id = %s', (record_id,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if not row:
            raise HTTPException(status_code=404, detail="记录不存在")
        excel_path, pdf_path = row[0], row[1]
        # Word格式：优先从Excel路径检测（Word模板时将docx路径存入Excel路径列）
        # 其次从PDF路径推导（兼容旧数据）
        word_path = None
        if excel_path and excel_path.endswith('.docx') and os.path.exists(excel_path):
            word_path = excel_path
        elif pdf_path and os.path.exists(pdf_path):
            word_candidate = pdf_path.replace('.pdf', '.docx')
            if os.path.exists(word_candidate):
                word_path = word_candidate
        if format == "Word" and word_path:
            return FileResponse(
                word_path,
                filename=os.path.basename(word_path),
                media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            )
        if format == "PDF" and pdf_path and os.path.exists(pdf_path):
            response = FileResponse(
                pdf_path,
                filename=os.path.basename(pdf_path),
                media_type="application/pdf"
            )
            if inline:
                response.headers["Content-Disposition"] = "inline"
            return response
        if format == "HTML":
            # 查找与Excel/Word同名的HTML文件
            html_path = None
            if excel_path and os.path.exists(excel_path):
                html_candidate = os.path.splitext(excel_path)[0] + '.html'
                if os.path.exists(html_candidate):
                    html_path = html_candidate
            if html_path:
                return FileResponse(
                    html_path,
                    filename=os.path.basename(html_path),
                    media_type="text/html; charset=utf-8"
                )
            raise HTTPException(status_code=404, detail="HTML文件不存在，请先保存后重试")
        if excel_path and os.path.exists(excel_path):
            return FileResponse(
                excel_path,
                filename=os.path.basename(excel_path),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        raise HTTPException(status_code=404, detail="文件不存在")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
