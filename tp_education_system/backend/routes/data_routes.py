from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse
from typing import List, Dict, Any, Optional
import json
import os
import glob
from sqlalchemy import create_engine, text

router = APIRouter(prefix="/api/data", tags=["data"])

# 数据库连接
DATABASE_URL = "postgresql://taiping_user:taiping_password@localhost:5432/taiping_education"
engine = create_engine(DATABASE_URL)

# 数据库连接配置
DATABASE_CONFIG = {
    'host': 'localhost',
    'port': '5432',
    'database': 'taiping_education',
    'user': 'taiping_user',
    'password': 'taiping_password'
}

def get_db_connection():
    import psycopg2
    return psycopg2.connect(**DATABASE_CONFIG)

# 配置文件路径
CONFIG_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'config')
SCHEMA_FILE = os.path.join(CONFIG_DIR, 'merged_schema_mappings.json')


def convert_chinese_to_english_fields(table_name: str, data: Dict[str, Any]) -> Dict[str, Any]:
    """将中文字段名转换为英文字段名"""
    try:
        with open(SCHEMA_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)
        
        tables_config = config.get('tables', {})
        table_config = tables_config.get(table_name, {})
        fields_config = table_config.get('fields', [])
        
        # 构建中文到英文的映射
        chinese_to_english = {}
        for field in fields_config:
            source = field.get('sourceField', '')  # 中文名
            target = field.get('targetField', '')   # 英文名
            if source and target:
                chinese_to_english[source] = target
        
        # 转换数据
        converted = {}
        for key, value in data.items():
            # 如果是中文字段名，转换为英文
            if key in chinese_to_english:
                converted[chinese_to_english[key]] = value
            else:
                # 保留原字段名
                converted[key] = value
        
        return converted
    except Exception as e:
        print(f"字段名转换失败: {e}")
        return data

# 默认字典关联配置
# 注意：只配置实际存在的字典表，code_field 必须是字典表的实际主键字段
DEFAULT_DICT_MAPPINGS = {
    '学历类型': {'table': 'education_dictionary', 'code_field': '学历类型', 'name_field': '类型名称', 'alias': '学历类型名称'},
    '学历': {'table': 'dict_education_dictionary', 'code_field': 'code', 'name_field': '学历', 'alias': '学历_name'},
    'position_level': {'table': 'dict_position', 'code_field': 'code', 'name_field': 'name', 'alias': 'position_level_name'},
    'personal_identity': {'table': 'dict_data_personal_identity', 'code_field': 'code', 'name_field': 'name', 'alias': 'personal_identity_name', 'link_field': 'personal_identity_code'},
}

# 检查字典表是否存在的缓存
dict_table_cache = {}

def check_dict_table_exists(table_name: str) -> bool:
    """检查字典表是否存在"""
    if table_name in dict_table_cache:
        return dict_table_cache[table_name]
    
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_name = :table_name
                )
            """), {"table_name": table_name})
            exists = result.scalar()
            dict_table_cache[table_name] = exists
            return exists
    except Exception as e:
        print(f"检查字典表存在性失败: {e}")
        return False


def read_json_file(file_path: str) -> Optional[Dict[str, Any]]:
    """读取JSON文件"""
    try:
        if os.path.exists(file_path):
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        return None
    except Exception as e:
        print(f"读取文件失败 {file_path}: {e}")
        return None


def get_table_schema_config(table_name: str) -> Dict[str, Any]:
    """获取表结构配置"""
    config = read_json_file(SCHEMA_FILE)
    if config and "tables" in config:
        return config["tables"].get(table_name, {})
    return {}


def get_dict_mappings_for_table(table_name: str, columns: List[str]) -> List[Dict[str, Any]]:
    """获取表的字典关联配置 - 从 merged_schema_mappings.json 动态读取"""
    mappings = []

    # 1. 从 merged_schema_mappings.json 读取表结构配置
    schema_config = read_json_file(SCHEMA_FILE)
    table_schema = schema_config.get("tables", {}).get(table_name, {})
    fields_config = table_schema.get("fields", [])

    # 2. 遍历字段配置，查找关联字段
    for field_config in fields_config:
        target_field = field_config.get("targetField") or field_config.get("english_name")
        if not target_field or target_field not in columns:
            continue

        relation_type = field_config.get("relation_type")
        relation_table = field_config.get("relation_table")

        # 处理字典关联 (to_dict)
        if relation_type == "to_dict" and relation_table:
            # 检查字典表是否存在
            if check_dict_table_exists(relation_table):
                # 获取字典表的字段（code_field用于关联，name_field用于显示）
                code_field, name_field = get_dict_fields(relation_table)
                mapping = {
                    'field': target_field,
                    'table': relation_table,
                    'code_field': code_field,
                    'name_field': name_field,
                    'alias': f"{target_field}_name"
                }
                mappings.append(mapping)
                print(f"字典关联: {table_name}.{target_field} -> {relation_table}.{code_field} (显示: {name_field})")

        # 处理主表关联 (to_master)
        elif relation_type == "to_master" and relation_table:
            # 检查主表是否存在
            if check_dict_table_exists(relation_table):
                # 动态查找表中的身份证字段（id_card, id_card_1, id_card_2, 身份证号码等）
                id_card_field = None
                for col in columns:
                    if col.startswith('id_card') or col == '身份证号码':
                        id_card_field = col
                        break
                
                # 动态查找表中的姓名字段（name, 姓名等）
                name_field_in_table = None
                for col in columns:
                    if col in ('name', '姓名'):
                        name_field_in_table = col
                        break
                
                # 如果找到身份证字段且当前是name字段，创建关联
                if id_card_field and name_field_in_table and target_field in ('name', '姓名'):
                    # 动态获取主表的主键关联字段
                    master_code_field, master_name_field = get_dict_fields(relation_table)
                    # code_field 应该是主表中的身份证字段（如 身份证号码）
                    # 如果 get_dict_fields 返回的第一个字段不是身份证字段，则查找主表中的身份证字段
                    master_id_card_col = None
                    for mc in [master_code_field, '身份证号码', 'id_card']:
                        if mc and mc != 'id':  # 排除自增id
                            master_id_card_col = mc
                            break
                    if not master_id_card_col:
                        master_id_card_col = master_code_field
                    
                    mapping = {
                        'field': id_card_field,
                        'table': relation_table,
                        'code_field': master_id_card_col,
                        'name_field': master_name_field if master_name_field != 'id' else name_field_in_table,
                        'alias': f'{name_field_in_table}_display',
                        'is_master_relation': True
                    }
                    mappings.append(mapping)
                    print(f"主表关联(姓名): {table_name}.{id_card_field} -> {relation_table}.{master_id_card_col} (显示: {master_name_field})")

    # 3. 兼容旧的硬编码配置（作为后备）
    for field_name in columns:
        if field_name in DEFAULT_DICT_MAPPINGS:
            # 检查是否已经在动态配置中
            already_configured = any(m['field'] == field_name for m in mappings)
            if already_configured:
                continue

            dict_config = DEFAULT_DICT_MAPPINGS[field_name]
            if check_dict_table_exists(dict_config['table']):
                mapping = {
                    'field': field_name,
                    'table': dict_config['table'],
                    'code_field': dict_config['code_field'],
                    'name_field': dict_config['name_field'],
                    'alias': dict_config['alias']
                }
                if 'link_field' in dict_config:
                    mapping['link_field'] = dict_config['link_field']
                mappings.append(mapping)

    # 4. 回退：从 field_configs/ 目录读取（配置驱动，无需手动同步）
    if not mappings:
        field_config_dir = os.path.join(CONFIG_DIR, 'field_configs')
        if os.path.isdir(field_config_dir):
            for cfg_file in glob.glob(os.path.join(field_config_dir, '*.json')):
                try:
                    cfg = read_json_file(cfg_file)
                    if not cfg:
                        continue
                    cfg_table = cfg.get('table_name', '')
                    if cfg_table != table_name:
                        continue
                    for fc in cfg.get('field_configs', []):
                        tf = fc.get('targetField', '')
                        if not tf or tf not in columns:
                            continue
                        rt = fc.get('relation_type', '')
                        rel_table = fc.get('relation_table', '')
                        if rt == 'to_dict' and rel_table and check_dict_table_exists(rel_table):
                            code_field, name_field = get_dict_fields(rel_table)
                            rel_display = fc.get('relation_display_field', '')
                            if rel_display:
                                name_field = rel_display
                            mappings.append({
                                'field': tf,
                                'table': rel_table,
                                'code_field': code_field,
                                'name_field': name_field,
                                'alias': f'{tf}_name'
                            })
                            print(f"字典关联(field_configs): {table_name}.{tf} -> {rel_table}.{code_field} (显示: {name_field})")
                except Exception:
                    pass

    return mappings


def get_dict_fields(table_name: str) -> tuple:
    """获取字典表的字段信息，返回 (code_field, name_field)
    字典表结构为: (id, value_field, created_at, updated_at)
    - code_field: 用于关联的字段（id）
    - name_field: 用于显示的字段（value_field，通常是第二个字段）
    """
    try:
        with engine.connect() as conn:
            # 获取表的所有字段
            result = conn.execute(text("""
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = :table_name
                ORDER BY ordinal_position
            """), {"table_name": table_name})
            columns = [row.column_name for row in result]

            # 过滤掉系统字段
            system_fields = {'created_at', 'updated_at'}
            value_fields = [c for c in columns if c not in system_fields]

            if len(value_fields) >= 2:
                # 第一个字段是id（用于关联），第二个字段是值（用于显示）
                return value_fields[0], value_fields[1]  # (id, value_field)
            elif len(value_fields) == 1:
                # 只有一个字段，既用于关联也用于显示
                return value_fields[0], value_fields[0]
    except Exception as e:
        print(f"获取字典表字段失败: {e}")
    return "id", "name"  # 默认返回


@router.get("/schema/{table_name}")
async def get_table_schema(table_name: str):
    """获取表结构定义 - 从配置文件读取，name使用数据库实际列名"""
    try:
        # 从配置文件读取表结构
        config = read_json_file(SCHEMA_FILE)
        if config and "tables" in config:
            table_schema = config["tables"].get(table_name)
            if table_schema and "fields" in table_schema:
                # 检查是否有有效的字段映射（支持 targetField/sourceField 或 name/chinese_name 两种格式）
                has_valid_mapping = any(
                    f.get("targetField") or f.get("sourceField") or f.get("name")
                    for f in table_schema["fields"]
                )
                
                if has_valid_mapping:
                    # 获取数据库实际列名
                    db_columns = []
                    try:
                        with engine.connect() as conn:
                            cols_result = conn.execute(text("""
                                SELECT column_name FROM information_schema.columns
                                WHERE table_name = :tbl ORDER BY ordinal_position
                            """), {"tbl": table_name})
                            db_columns = [row[0] for row in cols_result]
                    except Exception:
                        pass
                    
                    # 构建 targetField → sourceField 映射
                    field_map = {}
                    for field in table_schema["fields"]:
                        tf = field.get("targetField", "")
                        sf = field.get("sourceField", "")
                        if tf and sf:
                            field_map[tf] = sf
                    
                    # 转换为前端期望的格式
                    fields = []
                    for field in table_schema["fields"]:
                        # 兼容两种字段格式：targetField/sourceField 和 name/chinese_name
                        tf = field.get("targetField", "") or field.get("name", "")
                        sf = field.get("sourceField", "") or field.get("chinese_name", "")
                        
                        # 跳过id字段
                        if tf == "id":
                            continue
                        
                        # 确定 name：优先使用数据库实际列名，回退到 targetField
                        db_name = tf
                        if db_columns:
                            if tf in db_columns:
                                db_name = tf
                            elif sf in db_columns:
                                db_name = sf
                            elif field.get("english_name", "") in db_columns:
                                db_name = field.get("english_name", "")
                        
                        converted_field = {
                            "name": db_name,
                            "label": sf or tf,
                            "source_name": sf or tf,
                            "type": field.get("dataType") or field.get("data_type", "VARCHAR"),
                            "required": field.get("required", False),
                            "unique": field.get("unique", False),
                            "length": field.get("length", 255),
                        }
                        fields.append(converted_field)
                    
                    return {
                        "name": table_name,
                        "chinese_name": table_schema.get("chinese_name", ""),
                        "fields": fields
                    }
                # 配置文件中有表定义但字段映射无效，回退到数据库读取
        
        # 如果配置文件中没有，从数据库读取
        with engine.connect() as conn:
            # 获取字段基本信息
            result = conn.execute(text("""
                SELECT column_name, data_type, is_nullable, character_maximum_length
                FROM information_schema.columns
                WHERE table_name = :table_name
                ORDER BY ordinal_position
            """), {"table_name": table_name})
            
            # 获取字段注释
            comments_result = conn.execute(text("""
                SELECT a.attname as column_name, d.description
                FROM pg_class c
                JOIN pg_attribute a ON a.attrelid = c.oid
                LEFT JOIN pg_description d ON d.objoid = c.oid AND d.objsubid = a.attnum
                WHERE c.relname = :table_name AND a.attnum > 0 AND NOT a.attisdropped
            """), {"table_name": table_name})
            
            # 构建注释字典
            comments = {}
            for row in comments_result:
                comments[row.column_name] = row.description
            
            fields = []
            for row in result:
                if row.column_name == "id":
                    continue
                field_type = map_postgres_type_to_generic(row.data_type)
                # 使用注释作为标签，如果没有注释则使用字段名
                label = comments.get(row.column_name) or row.column_name
                field = {
                    "name": row.column_name,
                    "label": label,
                    "source_name": label,
                    "type": field_type,
                    "required": row.is_nullable == "NO"
                }
                if field_type == "VARCHAR" and row.character_maximum_length:
                    field["length"] = row.character_maximum_length
                fields.append(field)
            
            # 获取表的中文名称（从表注释）
            table_comment_result = conn.execute(text("""
                SELECT obj_description(c.oid, 'pg_class') as comment
                FROM pg_class c
                WHERE c.relname = :table_name AND c.relkind = 'r'
            """), {"table_name": table_name})
            table_comment_row = table_comment_result.fetchone()
            table_chinese_name = table_comment_row.comment if table_comment_row and table_comment_row.comment else table_name
            
            if fields:
                return {
                    "name": table_name,
                    "chinese_name": table_chinese_name,
                    "fields": fields
                }
        
        # 返回默认结构
        return {
            "name": table_name,
            "chinese_name": table_name,
            "fields": [
                {"name": "name", "label": "名称", "source_name": "名称", "type": "VARCHAR", "length": 50},
                {"name": "created_at", "label": "创建时间", "source_name": "创建时间", "type": "DATETIME"}
            ]
        }
    except Exception as e:
        print(f"获取表结构失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取表结构失败: {str(e)}")


@router.get("/config/schema")
async def get_schema_mappings():
    """获取整个 schema mappings，包括所有表和字典表"""
    try:
        config = read_json_file(SCHEMA_FILE)
        if config:
            return config
        
        # 返回默认结构
        return {
            "tables": {},
            "dictionaries": {}
        }
    except Exception as e:
        print(f"获取 schema mappings 失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取 schema mappings 失败: {str(e)}")


def map_postgres_type_to_generic(pg_type: str) -> str:
    """将PostgreSQL类型映射为通用类型"""
    type_mapping = {
        "integer": "INTEGER",
        "bigint": "INTEGER",
        "smallint": "INTEGER",
        "numeric": "DECIMAL",
        "decimal": "DECIMAL",
        "real": "DECIMAL",
        "double precision": "DECIMAL",
        "character varying": "VARCHAR",
        "varchar": "VARCHAR",
        "character": "VARCHAR",
        "char": "VARCHAR",
        "text": "TEXT",
        "date": "DATE",
        "timestamp": "DATETIME",
        "timestamp without time zone": "DATETIME",
        "timestamp with time zone": "DATETIME",
        "boolean": "BOOLEAN",
        "bool": "BOOLEAN"
    }
    return type_mapping.get(pg_type.lower(), "VARCHAR")


@router.get("/{table_name}")
async def get_table_data(
    table_name: str,
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=10000),
    filter: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None)
):
    """获取表数据，支持搜索和筛选，自动关联字典表"""
    try:
        with engine.connect() as conn:
            # 获取表的所有列名
            columns_result = conn.execute(text("""
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = :table_name
                ORDER BY ordinal_position
            """), {"table_name": table_name})
            columns = [row.column_name for row in columns_result]

            # 获取字典关联配置
            dict_mappings = get_dict_mappings_for_table(table_name, columns)
            use_table_alias = len(dict_mappings) > 0

            # 构建搜索条件
            where_conditions = []
            params = {}

            # 处理关键词搜索
            if keyword:
                search_conditions = []
                for col in columns:
                    # 如果使用表别名，列名需要加 t. 前缀
                    col_ref = f"t.{col}" if use_table_alias else col
                    search_conditions.append(f"CAST({col_ref} AS TEXT) ILIKE :keyword")
                
                # 同时搜索字典中文名称
                for mapping in dict_mappings:
                    dict_alias = f"dict_{mapping['field']}"
                    search_conditions.append(f"CAST({dict_alias}.{mapping['name_field']} AS TEXT) ILIKE :keyword")
                
                where_conditions.append("(" + " OR ".join(search_conditions) + ")")
                params["keyword"] = f"%{keyword}%"

            # 处理筛选条件
            if filter:
                try:
                    filter_data = json.loads(filter)
                    # 新格式：条件数组 [{field, operator, value}, ...]
                    if isinstance(filter_data, list):
                        for cond in filter_data:
                            field = cond.get('field', '')
                            operator = cond.get('operator', 'contains')
                            value = cond.get('value', '')
                            
                            if not field:
                                continue
                            if field not in columns:
                                continue
                            
                            param_name = f"filter_{field}_{len(params)}"
                            col_ref = f"t.{field}" if use_table_alias else field
                            
                            if operator == 'is_null':
                                where_conditions.append(f"{col_ref} IS NULL")
                            elif operator == 'is_not_null':
                                where_conditions.append(f"{col_ref} IS NOT NULL")
                            elif operator == 'eq':
                                where_conditions.append(f"CAST({col_ref} AS TEXT) = :{param_name}")
                                params[param_name] = str(value)
                            elif operator == 'neq':
                                where_conditions.append(f"CAST({col_ref} AS TEXT) != :{param_name}")
                                params[param_name] = str(value)
                            elif operator == 'gt':
                                where_conditions.append(f"CAST({col_ref} AS TEXT) > :{param_name}")
                                params[param_name] = str(value)
                            elif operator == 'gte':
                                where_conditions.append(f"CAST({col_ref} AS TEXT) >= :{param_name}")
                                params[param_name] = str(value)
                            elif operator == 'lt':
                                where_conditions.append(f"CAST({col_ref} AS TEXT) < :{param_name}")
                                params[param_name] = str(value)
                            elif operator == 'lte':
                                where_conditions.append(f"CAST({col_ref} AS TEXT) <= :{param_name}")
                                params[param_name] = str(value)
                            elif operator == 'not_contains':
                                where_conditions.append(f"CAST({col_ref} AS TEXT) NOT ILIKE :{param_name}")
                                params[param_name] = f"%{value}%"
                            elif operator == 'in':
                                # 支持逗号分隔的多值，如 "退休,离休"
                                values = [v.strip() for v in str(value).split(',') if v.strip()]
                                if values:
                                    in_params = []
                                    for i, v in enumerate(values):
                                        pname = f"{param_name}_{i}"
                                        in_params.append(f":{pname}")
                                        params[pname] = v
                                    where_conditions.append(f"CAST({col_ref} AS TEXT) IN ({', '.join(in_params)})")
                            elif operator == 'starts_with':
                                where_conditions.append(f"CAST({col_ref} AS TEXT) ILIKE :{param_name}")
                                params[param_name] = f"{value}%"
                            elif operator == 'ends_with':
                                where_conditions.append(f"CAST({col_ref} AS TEXT) ILIKE :{param_name}")
                                params[param_name] = f"%{value}"
                            else:  # contains (默认)
                                if value:
                                    where_conditions.append(f"CAST({col_ref} AS TEXT) ILIKE :{param_name}")
                                    params[param_name] = f"%{value}%"
                    # 旧格式兼容：字典 {field: value, ...}
                    elif isinstance(filter_data, dict):
                        for field, value in filter_data.items():
                            if field in columns and value:
                                param_name = f"filter_{field}"
                                col_ref = f"t.{field}" if use_table_alias else field
                                where_conditions.append(f"CAST({col_ref} AS TEXT) ILIKE :{param_name}")
                                params[param_name] = f"%{value}%"
                except json.JSONDecodeError:
                    pass

            # 构建WHERE子句
            where_clause = ""
            if where_conditions:
                where_clause = "WHERE " + " AND ".join(where_conditions)

            # 获取总记录数
            # 如果使用表别名，COUNT查询也需要使用相同的表别名和JOIN
            if use_table_alias:
                count_join_clauses = []
                for mapping in dict_mappings:
                    dict_alias = f"dict_{mapping['field']}"
                    count_join_clauses.append(
                        f"LEFT JOIN {mapping['table']} {dict_alias} "
                        f"ON CAST(t.{mapping['field']} AS TEXT) = CAST({dict_alias}.{mapping['code_field']} AS TEXT)"
                    )
                count_sql = f"SELECT COUNT(*) FROM {table_name} t {' '.join(count_join_clauses)} {where_clause}"
            else:
                count_sql = f"SELECT COUNT(*) FROM {table_name} {where_clause}"
            count_result = conn.execute(text(count_sql), params)
            total = count_result.scalar()

            # 获取分页数据
            offset = (page - 1) * size
            params["size"] = size
            params["offset"] = offset

            # 构建查询SQL，关联字典表
            if use_table_alias:
                select_fields = [f"t.{col}" for col in columns]
                join_clauses = []

                for idx, mapping in enumerate(dict_mappings):
                    dict_alias = f"dict_{mapping['field']}"
                    select_fields.append(f"{dict_alias}.{mapping['name_field']} as {mapping['alias']}")
                    # 使用指定的关联字段，如果没有则使用原字段
                    link_field = mapping.get('link_field', mapping['field'])
                    join_clauses.append(
                        f"LEFT JOIN {mapping['table']} {dict_alias} "
                        f"ON CAST(t.{link_field} AS TEXT) = CAST({dict_alias}.{mapping['code_field']} AS TEXT)"
                    )

                data_sql = f"""
                    SELECT {', '.join(select_fields)}
                    FROM {table_name} t
                    {' '.join(join_clauses)}
                    {where_clause}
                    ORDER BY t.id ASC
                    LIMIT :size OFFSET :offset
                """
            else:
                data_sql = f"SELECT * FROM {table_name} {where_clause} ORDER BY id ASC LIMIT :size OFFSET :offset"

            result = conn.execute(text(data_sql), params)

            # 转换结果为字典列表
            data = []
            for row in result:
                row_dict = {}
                for key, value in row._mapping.items():
                    # 处理日期时间类型
                    if hasattr(value, 'isoformat'):
                        row_dict[key] = value.isoformat()
                    else:
                        row_dict[key] = value
                data.append(row_dict)

            return {
                "data": data,
                "total": total,
                "page": page,
                "size": size
            }
    except Exception as e:
        print(f"获取数据失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取数据失败: {str(e)}")


@router.post("/export")
async def export_data(data: Dict[str, Any]):
    """导出数据到文件（直接下载）"""
    try:
        table_name = data.get('table_name', '')
        export_data_list = data.get('data', [])
        format_type = data.get('format', 'excel')
        filename = data.get('filename', f'export_{table_name}')
        
        # 使用临时目录
        import tempfile
        temp_dir = tempfile.gettempdir()
        
        if format_type == 'excel':
            file_path = os.path.join(temp_dir, f"{filename}.xlsx")
            
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            df = pd.DataFrame(export_data_list)
            df.to_excel(file_path, index=False, engine='openpyxl')

            wb = load_workbook(file_path)
            ws = wb.active

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 15), 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            
            return FileResponse(
                file_path, 
                filename=f"{filename}.xlsx",
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        elif format_type == 'pdf':
            file_path = os.path.join(temp_dir, f"{filename}.pdf")
            
            # 使用reportlab生成PDF
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
            
            doc = SimpleDocTemplate(file_path, pagesize=landscape(letter))
            elements = []
            
            # 添加标题
            styles = getSampleStyleSheet()
            title = Paragraph(f"<b>{filename}</b>", styles['Heading1'])
            elements.append(title)
            
            # 准备表格数据
            if export_data_list:
                headers = list(export_data_list[0].keys())
                table_data = [headers]
                for row in export_data_list:
                    table_data.append([str(row.get(h, '')) for h in headers])
                
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table)
            
            doc.build(elements)
            
            return FileResponse(
                file_path, 
                filename=f"{filename}.pdf",
                media_type='application/pdf'
            )
            
        elif format_type == 'sql':
            file_path = os.path.join(temp_dir, f"{filename}.sql")
            
            with open(file_path, 'w', encoding='utf-8') as f:
                # 写入表结构（简化版）
                f.write(f"-- 导出表: {table_name}\n")
                f.write(f"-- 导出时间: {pd.Timestamp.now()}\n\n")
                
                # 生成INSERT语句
                for row in export_data_list:
                    columns = ', '.join([f'"{k}"' for k in row.keys()])
                    value_list = []
                    for v in row.values():
                        if v is None:
                            value_list.append('NULL')
                        else:
                            escaped = str(v).replace("'", "''")
                            value_list.append(f"'{escaped}'")
                    values = ', '.join(value_list)
                    f.write(f"INSERT INTO \"{table_name}\" ({columns}) VALUES ({values});\n")
            
            return FileResponse(
                file_path, 
                filename=f"{filename}.sql",
                media_type='application/sql'
            )
        else:
            raise HTTPException(status_code=400, detail=f"不支持的导出格式: {format_type}")
            
    except Exception as e:
        print(f"导出失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.post("/{table_name}")
async def create_record(table_name: str, data: Dict[str, Any]):
    """创建记录"""
    try:
        # 保存原始数据用于备注写入
        original_data = data.copy()
        
        # ⚠️ teacher_basic_info 表的数据库列名是中文，不需要转换
        if table_name != 'teacher_basic_info':
            data = convert_chinese_to_english_fields(table_name, data)
        
        with engine.connect() as conn:
            # 构建 INSERT 语句
            columns = []
            placeholders = []
            params = {}
            
            for key, value in data.items():
                columns.append(key)
                placeholders.append(f":{key}")
                params[key] = value
            
            if not columns:
                return {
                    "status": "success",
                    "message": "没有数据",
                    "id": None
                }
            
            sql = f"""
                INSERT INTO {table_name} ({', '.join(columns)})
                VALUES ({', '.join(placeholders)})
                RETURNING id
            """
            
            result = conn.execute(text(sql), params)
            new_id = result.fetchone()[0]
            conn.commit()
            
            # 如果是岗位聘任信息表(information)，写入新增人员备注
            if table_name == 'information':
                await handle_new_employee_remarks(original_data)
            
            # 如果是教师基础信息表(teacher_basic_info)，写入新增人员备注
            if table_name == 'teacher_basic_info':
                await handle_new_teacher_remarks(original_data, new_id)
            
            return {
                "status": "success",
                "message": "创建成功",
                "id": new_id
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"创建失败: {str(e)}")


async def handle_post_change_remarks(record_id: int, new_post_id: int):
    """处理岗位变更，写入备注信息表"""
    conn = None
    cursor = None
    try:
        from datetime import datetime
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 读取字典映射
        cursor.execute("SELECT id, post FROM dict_dictionary_personal")
        dict_mapping = {}
        for dict_row in cursor.fetchall():
            dict_mapping[dict_row[0]] = dict_row[1]
        
        # 获取原岗位和教师信息
        cursor.execute("""
            SELECT p.post_1, p.id_card, p.name
            FROM post_appointment_info p
            WHERE p.id = %s
        """, (record_id,))
        row = cursor.fetchone()
        
        if not row:
            return
        
        original_post_id = row[0]
        id_card = row[1]
        teacher_name = row[2]
        
        # 获取教师ID
        cursor.execute("SELECT id FROM teacher_basic_info WHERE \"身份证号码\" = %s", (id_card,))
        teacher_row = cursor.fetchone()
        teacher_id = teacher_row[0] if teacher_row else None
        
        # 转换岗位名称
        original_post = dict_mapping.get(int(original_post_id), f'未知岗位{original_post_id}') if original_post_id else None
        new_post = dict_mapping.get(int(new_post_id), f'未知岗位{new_post_id}') if new_post_id else None
        
        # 如果岗位没有变化，不写入
        if original_post == new_post or not new_post:
            return
        
        # 判断是否是晋升
        promotion_mapping = {
            '一级教师': {'高级教师': '一级教师晋升高级教师'},
            '二级教师': {'一级教师': '二级教师晋升一级教师'},
            '三级教师': {'二级教师': '三级教师晋升二级教师'},
        }
        
        current_month = datetime.now().strftime('%Y-%m')
        change_category = 'position_change'
        change_detail = ''
        remark_label = ''
        
        if original_post in promotion_mapping and new_post in promotion_mapping[original_post]:
            remark_label = promotion_mapping[original_post][new_post]
        else:
            # 其他岗位变化
            remark_label = f'{original_post}变更为{new_post}'
        
        # 写入备注信息表
        cursor.execute("""
            INSERT INTO performance_pay_remarks (
                report_period, remark_type, teacher_id, teacher_name,
                original_status, new_status, original_post, new_post,
                change_category, change_detail, created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        """, (
            current_month,
            'position_change',
            teacher_id,
            teacher_name,
            None,  # original_status
            None,  # new_status
            original_post,
            new_post,
            change_category,
            remark_label
        ))
        conn.commit()
        print(f"[备注] 已写入岗位变更: {teacher_name} {original_post}->{new_post}")
        
    except Exception as e:
        print(f"[备注] 写入岗位变更失败: {e}")
        if conn:
            try:
                conn.rollback()
            except:
                pass
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


async def handle_status_change_remarks(record_id: int, new_status: str):
    """处理教师状态变更，写入备注信息表"""
    conn = None
    cursor = None
    try:
        from datetime import datetime
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取教师原有信息（包括身份证号）
        cursor.execute("""
            SELECT id, "姓名", "任职状态", "身份证号码" 
            FROM teacher_basic_info 
            WHERE id = %s
        """, (record_id,))
        row = cursor.fetchone()
        
        if not row:
            return
        
        teacher_id = row[0]
        teacher_name = row[1]
        old_status = row[2]
        id_card = row[3]
        
        # 如果状态没有变化，不写入
        if old_status == new_status or not new_status:
            return
        
        # 获取原岗位信息（从岗位聘任信息表查询，并通过字典转换）
        original_post = None
        if id_card:
            # 读取字典映射
            cursor.execute("SELECT id, post FROM dict_dictionary_personal")
            dict_mapping = {}
            for dict_row in cursor.fetchall():
                dict_mapping[dict_row[0]] = dict_row[1]
            
            # 从岗位聘任信息表查询post_1字段
            cursor.execute("""
                SELECT post_1 FROM post_appointment_info WHERE id_card = %s
            """, (id_card,))
            post_row = cursor.fetchone()
            if post_row and post_row[0]:
                post_id = int(post_row[0])
                original_post = dict_mapping.get(post_id, f'未知岗位{post_id}')
        
        current_month = datetime.now().strftime('%Y-%m')
        
        # 写入备注信息表
        cursor.execute("""
            INSERT INTO performance_pay_remarks (
                report_period, remark_type, teacher_id, teacher_name,
                original_status, new_status, original_post, new_post,
                change_category, change_detail, created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        """, (
            current_month,
            'status_change',
            teacher_id,
            teacher_name,
            old_status,
            new_status,
            original_post,
            None,
            'status_change',
            f'{old_status}->{new_status}'
        ))
        conn.commit()
        print(f"[备注] 已写入状态变更: {teacher_name} {old_status}->{new_status}")
        
    except Exception as e:
        print(f"[备注] 写入状态变更失败: {e}")
        if conn:
            try:
                conn.rollback()
            except:
                pass
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


async def handle_new_employee_remarks(data: dict):
    """处理新增人员，写入备注信息表"""
    conn = None
    cursor = None
    try:
        from datetime import datetime
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取教师ID和姓名
        teacher_id = data.get('教师ID') or data.get('teacher_id')
        teacher_name = data.get('姓名') or data.get('name')
        post_2 = data.get('post_2') or data.get('现受聘岗位名称')
        remark = data.get('备注') or data.get('remark') or ''
        
        if not teacher_id or not teacher_name:
            return
        
        current_month = datetime.now().strftime('%Y-%m')
        
        # 检查当前月份是否已存在该教师的新增记录，避免重复
        cursor.execute("""
            SELECT COUNT(*) FROM performance_pay_remarks 
            WHERE report_period = %s AND teacher_id = %s AND remark_type = 'new_add'
        """, (current_month, teacher_id))
        count = cursor.fetchone()[0]
        if count > 0:
            print(f"[备注] 教师 {teacher_name} 本月已有新增记录，跳过写入")
            return
        change_category = 'new_add'
        change_detail = 'transfer_in'
        
        # 根据备注字段判断新增类型
        if '新录聘' in remark or '招聘' in remark:
            change_detail = 'new_hire'
            remark_label = f'{post_2}新录聘' if post_2 else '教师新录聘'
        elif '应届' in remark or '毕业生' in remark:
            change_detail = 'graduate'
            remark_label = f'{post_2}应届毕业生' if post_2 else '教师应届毕业生'
        elif '引进' in remark or '高层次' in remark:
            change_detail = 'talent'
            remark_label = f'{post_2}人才引进' if post_2 else '教师人才引进'
        elif '见习' in remark or '实习' in remark:
            change_detail = 'intern'
            remark_label = f'{post_2}见习期' if post_2 else '教师见习期'
        elif '管理' in remark or '九级管理' in (post_2 or ''):
            change_detail = 'management'
            remark_label = f'{post_2}新增' if post_2 else '九级管理新增'
        else:
            change_detail = 'transfer_in'
            remark_label = f'{post_2}调入' if post_2 else '教师调入'
        
        # 写入备注信息表
        cursor.execute("""
            INSERT INTO performance_pay_remarks (
                report_period, remark_type, teacher_id, teacher_name,
                original_status, new_status, original_post, new_post,
                change_category, change_detail, created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        """, (
            current_month,
            'new_add',
            teacher_id,
            teacher_name,
            None,  # original_status
            '在职',  # new_status
            None,  # original_post
            post_2,  # new_post
            change_category,
            remark_label
        ))
        conn.commit()
        print(f"[备注] 已写入新增人员: {teacher_name}")
        
    except Exception as e:
        print(f"[备注] 写入新增人员失败: {e}")
        if conn:
            try:
                conn.rollback()
            except:
                pass
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


async def handle_new_teacher_remarks(data: dict, new_id: int):
    """处理教师基础信息表新增人员，写入备注信息表"""
    conn = None
    cursor = None
    try:
        from datetime import datetime
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 获取教师ID和姓名
        teacher_id = new_id
        teacher_name = data.get('姓名') or data.get('name') or data.get('teacher_name')
        remark = data.get('备注') or data.get('remark') or ''
        
        if not teacher_id or not teacher_name:
            return
        
        # 检查当前月份是否已存在该教师的新增记录，避免重复
        current_month = datetime.now().strftime('%Y-%m')
        cursor.execute("""
            SELECT COUNT(*) FROM performance_pay_remarks 
            WHERE report_period = %s AND teacher_id = %s AND remark_type = 'new_add'
        """, (current_month, teacher_id))
        count = cursor.fetchone()[0]
        if count > 0:
            print(f"[备注] 教师 {teacher_name} 本月已有新增记录，跳过写入")
            return
        
        # 从岗位聘任信息表获取岗位信息
        post_2 = None
        # 先获取身份证号码
        cursor.execute("SELECT \"身份证号码\" FROM teacher_basic_info WHERE id = %s", (teacher_id,))
        id_card_row = cursor.fetchone()
        if id_card_row:
            id_card = id_card_row[0]
            # 读取字典映射
            cursor.execute("SELECT id, post FROM dict_dictionary_personal")
            dict_mapping = {}
            for dict_row in cursor.fetchall():
                dict_mapping[dict_row[0]] = dict_row[1]
            # 从岗位聘任信息表查询post_1字段
            cursor.execute("SELECT post_1 FROM post_appointment_info WHERE id_card = %s", (id_card,))
            post_row = cursor.fetchone()
            if post_row and post_row[0]:
                post_id = int(post_row[0])
                post_2 = dict_mapping.get(post_id, f'未知岗位{post_id}')
        
        current_month = datetime.now().strftime('%Y-%m')
        change_category = 'new_add'
        
        # 根据备注字段判断新增类型
        if '新录聘' in remark or '招聘' in remark:
            change_detail = 'new_hire'
            remark_label = f'{post_2}新录聘' if post_2 else '教师新录聘'
        elif '应届' in remark or '毕业生' in remark:
            change_detail = 'graduate'
            remark_label = f'{post_2}应届毕业生' if post_2 else '教师应届毕业生'
        elif '引进' in remark or '高层次' in remark:
            change_detail = 'talent'
            remark_label = f'{post_2}人才引进' if post_2 else '教师人才引进'
        elif '见习' in remark or '实习' in remark:
            change_detail = 'intern'
            remark_label = f'{post_2}见习期' if post_2 else '教师见习期'
        elif '管理' in remark or '九级管理' in (post_2 or ''):
            change_detail = 'management'
            remark_label = f'{post_2}新增' if post_2 else '九级管理新增'
        else:
            change_detail = 'transfer_in'
            remark_label = f'{post_2}调入' if post_2 else '教师调入'
        
        # 写入备注信息表
        cursor.execute("""
            INSERT INTO performance_pay_remarks (
                report_period, remark_type, teacher_id, teacher_name,
                original_status, new_status, original_post, new_post,
                change_category, change_detail, created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        """, (
            current_month,
            'new_add',
            teacher_id,
            teacher_name,
            None,  # original_status
            '在职',  # new_status
            None,  # original_post
            post_2,  # new_post
            change_category,
            remark_label
        ))
        conn.commit()
        print(f"[备注] 已写入新增教师: {teacher_name}")
        
    except Exception as e:
        print(f"[备注] 写入新增教师失败: {e}")
        if conn:
            try:
                conn.rollback()
            except:
                pass
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.put("/{table_name}/{record_id}")
async def update_record(table_name: str, record_id: int, data: Dict[str, Any]):
    """更新记录"""
    try:
        # ⚠️ teacher_basic_info 表的数据库列名是中文，不需要转换
        if table_name != 'teacher_basic_info':
            data = convert_chinese_to_english_fields(table_name, data)
        
        # 如果是岗位聘任信息表(post_appointment_info)，检查岗位变更并写入备注表
        if table_name == 'post_appointment_info' and 'post_1' in data:
            await handle_post_change_remarks(record_id, data.get('post_1'))
        
        # 如果是教师基础信息表(teacher_basic_info)，检查状态变更并写入备注表
        if table_name == 'teacher_basic_info' and '任职状态' in data:
            await handle_status_change_remarks(record_id, data.get('任职状态'))
        
        with engine.connect() as conn:
            # 构建 UPDATE 语句
            set_clauses = []
            params = {"record_id": record_id}
            
            for key, value in data.items():
                if key != "id":  # 不更新主键
                    set_clauses.append(f"{key} = :{key}")
                    params[key] = value
            
            if not set_clauses:
                return {
                    "status": "success",
                    "message": "没有需要更新的字段"
                }
            
            update_sql = f"UPDATE {table_name} SET {', '.join(set_clauses)} WHERE id = :record_id"
            result = conn.execute(text(update_sql), params)
            conn.commit()
            
            if result.rowcount > 0:
                return {
                    "status": "success",
                    "message": "更新成功",
                    "updated_rows": result.rowcount
                }
            else:
                raise HTTPException(status_code=404, detail="记录不存在")
    except Exception as e:
        print(f"更新失败: {e}")
        raise HTTPException(status_code=500, detail=f"更新失败: {str(e)}")


@router.delete("/{table_name}/{record_id}")
async def delete_record(table_name: str, record_id: int):
    """删除记录"""
    try:
        with engine.connect() as conn:
            # 执行删除操作
            delete_sql = f"DELETE FROM {table_name} WHERE id = :record_id"
            result = conn.execute(text(delete_sql), {"record_id": record_id})
            conn.commit()
            
            if result.rowcount > 0:
                return {
                    "status": "success",
                    "message": "删除成功",
                    "deleted_count": result.rowcount
                }
            else:
                raise HTTPException(status_code=404, detail="记录不存在")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")


@router.get("/ui-components/{table_name}")
async def get_ui_components(table_name: str):
    """获取UI组件配置"""
    try:
        # 这里应该从配置文件读取
        return {
            "actions": [
                {
                    "key": "smartFill",
                    "label": "智能填充",
                    "icon": "MagicStick",
                    "type": "primary",
                    "requireSelection": False
                },
                {
                    "key": "batchTag",
                    "label": "批量打标签",
                    "icon": "CollectionTag",
                    "type": "default",
                    "requireSelection": True
                },
                {
                    "key": "generateReport",
                    "label": "生成报表",
                    "icon": "Document",
                    "type": "success",
                    "requireSelection": False
                }
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取UI配置失败: {str(e)}")
