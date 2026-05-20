"""
查看Excel模板的详细信息
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

excel_path = r"d:\erp_thirteen\tp_education_system\backend\uploads\templates\义务教育学校教职工绩效工资审批表.xlsx"

wb = load_workbook(excel_path)
ws = wb.active

print("=" * 80)
print("Excel模板详细信息")
print("=" * 80)
print(f"文件: {excel_path}")
print(f"工作表: {ws.title}")
print(f"行数: {ws.max_row}")
print(f"列数: {ws.max_column}")
print()

print("列宽信息:")
print("-" * 80)
for col_idx in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    col_dim = ws.column_dimensions.get(col_letter)
    if col_dim and hasattr(col_dim, 'width') and col_dim.width:
        width = col_dim.width
        print(f"列{col_letter}: {width} 字符宽度 = {int(width * 7)} 像素")
    else:
        print(f"列{col_letter}: 默认宽度 = 8.43 字符宽度 = 59 像素")
print()

print("行高信息:")
print("-" * 80)
for row_idx in range(1, min(ws.max_row + 1, 20)):
    row_dim = ws.row_dimensions[row_idx]
    if row_dim and hasattr(row_dim, 'height') and row_dim.height:
        height = row_dim.height
        print(f"行{row_idx}: {height} 磅")
    else:
        print(f"行{row_idx}: 默认高度 = 15 磅")
print()

print("合并单元格:")
print("-" * 80)
for merged_range in ws.merged_cells.ranges:
    print(f"{merged_range.start_cell.coordinate} 到 {merged_range.end_cell.coordinate}")
print()

print("单元格内容（前10行）:")
print("-" * 80)
for row_idx in range(1, min(ws.max_row + 1, 10)):
    row_content = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        value = str(cell.value) if cell.value is not None else ""
        if value:
            row_content.append(f"{get_column_letter(col_idx)}{row_idx}:{value[:20]}")
    if row_content:
        print(f"行{row_idx}: {', '.join(row_content)}")
print()

print("单元格样式示例（第1行）:")
print("-" * 80)
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(1, col_idx)
    if cell.value:
        print(f"单元格 {get_column_letter(col_idx)}1:")
        print(f"  内容: {cell.value}")
        if cell.font:
            print(f"  字体: {cell.font.name}, 大小: {cell.font.size}pt, 粗体: {cell.font.bold}")
        if cell.alignment:
            print(f"  对齐: 水平={cell.alignment.horizontal}, 垂直={cell.alignment.vertical}")
        if cell.fill and cell.fill.start_color:
            print(f"  填充色: {cell.fill.start_color.rgb}")
        print()

wb.close()
