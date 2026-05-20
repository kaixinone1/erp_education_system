import requests
from openpyxl import load_workbook
import json
import re

print("=" * 150)
print("从原模板文件和网页预览中读取数据对比")
print("=" * 150)

# 1. 从API获取模板信息
response = requests.get("http://127.0.0.1:8000/api/template/list")
data = response.json()

if not data["success"] or not data["templates"]:
    print("获取模板列表失败")
    exit(1)

template = data["templates"][0]
file_path = template["file_path"]
print(f"\n模板文件: {file_path}")

# 2. 从原Excel文件读取内容
wb = load_workbook(file_path, data_only=True)
ws = wb.active

excel_contents = {}
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value:
            excel_contents[(row, col)] = str(cell.value)

print(f"原Excel文件有 {len(excel_contents)} 个有内容的单元格")

# 3. 从网页读取预览内容
# 访问前端页面
frontend_url = "http://localhost:5173"
try:
    page_response = requests.get(frontend_url, timeout=5)
    print(f"\n前端页面访问成功: {frontend_url}")
    print(f"页面状态码: {page_response.status_code}")
except Exception as e:
    print(f"\n访问前端页面失败: {e}")
    print("尝试访问模板设计页面...")
    
# 4. 从API获取预览数据
preview_response = requests.get(f"http://127.0.0.1:8000/api/template/preview?file_path={file_path}")
preview_data = preview_response.json()

if preview_data.get("success"):
    metadata = preview_data["metadata"]
    
    # 模拟前端渲染逻辑
    cellMap = {}
    for cell in metadata['cells']:
        cellMap[f"{cell['row']}-{cell['col']}"] = cell
    
    occupiedCells = set()
    preview_contents = {}
    
    for row in range(metadata['sheet_info']['total_rows']):
        for col in range(metadata['sheet_info']['total_cols']):
            key = f"{row}-{col}"
            
            if key in occupiedCells:
                continue
            
            cell = cellMap.get(key)
            
            if not cell:
                preview_contents[(row + 1, col + 1)] = ""
                continue
            
            rowspan = cell.get('rowspan', 1)
            colspan = cell.get('colspan', 1)
            
            for r in range(rowspan):
                for c in range(colspan):
                    occupiedCells.add(f"{row + r}-{col + c}")
            
            value = cell.get('value', '')
            preview_contents[(row + 1, col + 1)] = value
    
    print(f"预览API返回 {len([v for v in preview_contents.values() if v])} 个有内容的单元格")
    
    # 5. 打印对比表
    print("\n" + "=" * 150)
    print("单元格序号".ljust(15) + "原模板内容".ljust(60) + "预览模板内容")
    print("=" * 150)
    
    all_cells = set(excel_contents.keys()) | set(preview_contents.keys())
    
    for (row, col) in sorted(all_cells):
        excel_val = excel_contents.get((row, col), "")
        preview_val = preview_contents.get((row, col), "")
        
        if excel_val or preview_val:
            cell_ref = f"({row},{col})"
            print(f"{cell_ref.ljust(15)}{excel_val.ljust(60)}{preview_val}")
    
    print("=" * 150)
    
    # 6. 统计不一致
    mismatches = []
    for (row, col) in all_cells:
        excel_val = excel_contents.get((row, col), "")
        preview_val = preview_contents.get((row, col), "")
        
        if excel_val != preview_val:
            mismatches.append((row, col, excel_val, preview_val))
    
    if mismatches:
        print(f"\n发现 {len(mismatches)} 个不一致:")
        for row, col, excel_val, preview_val in mismatches:
            print(f"  ({row},{col}): 原模板='{excel_val}', 预览='{preview_val}'")
    else:
        print("\n所有内容完全一致！")
else:
    print(f"预览API返回错误: {preview_data}")
