from openpyxl import load_workbook
import json

file_path = r'd:\erp_thirteen\tp_education_system\backend\data\templates\义务教育学校教职工绩效工资审批表_20260505_221848.xlsx'
metadata_path = file_path + '.metadata.json'

wb = load_workbook(file_path, data_only=True)
ws = wb.active

with open(metadata_path, 'r', encoding='utf-8') as f:
    metadata = json.load(f)

totalRows = metadata['sheet_info']['total_rows']
totalCols = metadata['sheet_info']['total_cols']
cells = metadata['cells']

cellMap = {}
for cell in cells:
    cellMap[f"{cell['row']}-{cell['col']}"] = cell

occupiedCells = set()

preview_contents = {}

for row in range(totalRows):
    for col in range(totalCols):
        key = f"{row}-{col}"
        
        if key in occupiedCells:
            continue
        
        cell = cellMap.get(key)
        
        if not cell:
            preview_contents[(row + 1, col + 1)] = ""
            continue
        
        rowspan = cell.get('rowspan', 1)
        colspan = cell.get('colspan', 1)
        
        for r in range(rowspan):
            for c in range(colspan):
                occupiedCells.add(f"{row + r}-{col + c}")
        
        value = cell.get('value', '')
        preview_contents[(row + 1, col + 1)] = value

print("=" * 150)
print("单元格序号 | 原模板内容 | 预览模板内容")
print("=" * 150)

for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell_ref = f"({row},{col})"
        
        excel_cell = ws.cell(row=row, column=col)
        excel_value = str(excel_cell.value) if excel_cell.value is not None else ''
        
        preview_value = preview_contents.get((row, col), '')
        
        if excel_value or preview_value:
            print(f"{cell_ref:<12} | {excel_value:<60} | {preview_value:<60}")

print("=" * 150)
