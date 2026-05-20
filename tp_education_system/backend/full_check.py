"""
完整检查：原模板 vs 元数据 vs 前端渲染
找出所有差异
"""
import json
from openpyxl import load_workbook

# 读取原模板
original_file = r"D:\erp_thirteen\数据库信息\模板\义务教育学校教职工绩效工资审批表.xlsx"
wb = load_workbook(original_file, data_only=True)
ws = wb.active

# 读取元数据
metadata_file = r"d:\erp_thirteen\tp_education_system\backend\data\templates\义务教育学校教职工绩效工资审批表_20260505_154157.xlsx.metadata.json"
with open(metadata_file, 'r', encoding='utf-8') as f:
    metadata = json.load(f)

print("=" * 120)
print("完整检查报告")
print("=" * 120)

# 1. 检查单元格数量
print("\n1. 单元格数量检查：")
print(f"   原模板总单元格数: 31行 × 6列 = 186个")
print(f"   元数据中单元格数: {len(metadata['cells'])}个")

# 2. 检查元数据中缺失的单元格
print("\n2. 元数据中缺失的单元格：")
cell_map = {}
for cell in metadata['cells']:
    cell_map[f"{cell['row']}-{cell['col']}"] = cell

missing_cells = []
for row in range(31):
    for col in range(6):
        key = f"{row}-{col}"
        if key not in cell_map:
            missing_cells.append((row, col))
            original_value = ws.cell(row=row + 1, column=col + 1).value
            if original_value is not None and str(original_value).strip():
                print(f"   单元格({row+1}, {col+1}): 原模板有值'{original_value}'，但元数据缺失")

if not missing_cells:
    print("   无缺失")

# 3. 检查合并单元格
print("\n3. 合并单元格检查：")
print(f"   原模板合并区域数: {len(list(ws.merged_cells.ranges))}")
print(f"   元数据合并区域数: {len(metadata.get('merge_info', []))}")

print("\n   原模板合并区域：")
for merge in ws.merged_cells.ranges:
    print(f"      {merge}")

print("\n   元数据合并区域：")
for merge in metadata.get('merge_info', []):
    print(f"      行{merge['row']+1}列{merge['col']+1}: {merge['rowspan']}行×{merge['colspan']}列")

# 4. 检查列宽
print("\n4. 列宽检查：")
for col in range(6):
    col_letter = chr(65 + col)
    original_width = ws.column_dimensions[col_letter].width
    metadata_width = metadata['dimensions']['col_widths'].get(str(col))
    if metadata_width:
        metadata_width_orig = metadata_width / 7  # 转换回Excel单位
    else:
        metadata_width_orig = None
    
    print(f"   列{col_letter}: 原模板={original_width}, 元数据={metadata_width_orig}")

# 5. 检查行高
print("\n5. 行高检查（只显示有差异的）：")
for row in range(31):
    original_height = ws.row_dimensions[row + 1].height
    metadata_height = metadata['dimensions']['row_heights'].get(str(row))
    
    if original_height != metadata_height:
        print(f"   行{row+1}: 原模板={original_height}, 元数据={metadata_height}")

# 6. 检查单元格值
print("\n6. 单元格值检查（只显示有差异的）：")
for row in range(31):
    for col in range(6):
        original_value = ws.cell(row=row + 1, column=col + 1).value
        key = f"{row}-{col}"
        metadata_cell = cell_map.get(key)
        metadata_value = metadata_cell['value'] if metadata_cell else None
        
        # 比较
        orig_str = str(original_value) if original_value is not None else ""
        meta_str = str(metadata_value) if metadata_value is not None else ""
        
        if orig_str != meta_str:
            print(f"   单元格({row+1}, {col+1}):")
            print(f"      原模板: {repr(original_value)}")
            print(f"      元数据: {repr(metadata_value)}")

# 7. 检查单元格样式（抽样）
print("\n7. 单元格样式检查（抽样前10个有值的单元格）：")
count = 0
for row in range(31):
    for col in range(6):
        original_cell = ws.cell(row=row + 1, column=col + 1)
        if original_cell.value is not None and str(original_cell.value).strip():
            key = f"{row}-{col}"
            metadata_cell = cell_map.get(key)
            
            if metadata_cell:
                print(f"\n   单元格({row+1}, {col+1}): {original_cell.value}")
                
                # 字体
                orig_font = original_cell.font
                meta_font = metadata_cell.get('font', {})
                print(f"      字体名称: 原模板={orig_font.name}, 元数据={meta_font.get('name')}")
                print(f"      字体大小: 原模板={orig_font.size}, 元数据={meta_font.get('size')}")
                print(f"      字体加粗: 原模板={orig_font.bold}, 元数据={meta_font.get('bold')}")
                
                # 对齐
                orig_align = original_cell.alignment
                meta_align = metadata_cell.get('alignment', {})
                print(f"      水平对齐: 原模板={orig_align.horizontal}, 元数据={meta_align.get('horizontal')}")
                print(f"      垂直对齐: 原模板={orig_align.vertical}, 元数据={meta_align.get('vertical')}")
                
                count += 1
                if count >= 10:
                    break
    if count >= 10:
        break

wb.close()

print("\n" + "=" * 120)
print("检查完成")
print("=" * 120)
