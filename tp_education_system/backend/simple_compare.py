"""
直接输出：原模板所有单元格内容 vs 元数据所有单元格内容
不进行任何判断，直接显示
"""
import json
from openpyxl import load_workbook

# 读取原模板
original_file = r"D:\erp_thirteen\数据库信息\模板\义务教育学校教职工绩效工资审批表.xlsx"
wb = load_workbook(original_file, data_only=True)
ws = wb.active

# 读取元数据
metadata_file = r"d:\erp_thirteen\tp_education_system\backend\data\templates\义务教育学校教职工绩效工资审批表_20260505_154157.xlsx.metadata.json"
with open(metadata_file, 'r', encoding='utf-8') as f:
    metadata = json.load(f)

print("=" * 100)
print("原模板所有单元格内容（有值的）")
print("=" * 100)

original_values = {}
for row in range(1, 32):
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None and str(cell.value).strip():
            original_values[f"({row},{col})"] = str(cell.value)
            print(f"({row},{col}): {cell.value}")

print("\n" + "=" * 100)
print("元数据所有单元格内容（有值的）")
print("=" * 100)

metadata_values = {}
for cell in metadata['cells']:
    if cell['value'] and str(cell['value']).strip():
        row = cell['row'] + 1
        col = cell['col'] + 1
        metadata_values[f"({row},{col})"] = cell['value']
        print(f"({row},{col}): {cell['value']}")

print("\n" + "=" * 100)
print("原模板有但元数据没有的")
print("=" * 100)

for key, value in original_values.items():
    if key not in metadata_values:
        print(f"{key}: {value}")

print("\n" + "=" * 100)
print("元数据有但原模板没有的")
print("=" * 100)

for key, value in metadata_values.items():
    if key not in original_values:
        print(f"{key}: {value}")

print("\n" + "=" * 100)
print("值不一致的")
print("=" * 100)

for key in original_values:
    if key in metadata_values and original_values[key] != metadata_values[key]:
        print(f"{key}:")
        print(f"  原模板: {original_values[key]}")
        print(f"  元数据: {metadata_values[key]}")

wb.close()
