"""
按单元格逐一比对：原模板 vs 预览页面
完全按照用户要求的方式输出
"""
import json
from openpyxl import load_workbook

# 读取原模板
original_file = r"D:\erp_thirteen\数据库信息\模板\义务教育学校教职工绩效工资审批表.xlsx"
wb = load_workbook(original_file, data_only=True)
ws = wb.active

# 读取元数据
metadata_file = r"d:\erp_thirteen\tp_education_system\backend\data\templates\义务教育学校教职工绩效工资审批表_20260505_154157.xlsx.metadata.json"
with open(metadata_file, 'r', encoding='utf-8') as f:
    metadata = json.load(f)

# 创建单元格映射
cell_map = {}
for cell in metadata['cells']:
    cell_map[f"{cell['row']}-{cell['col']}"] = cell

total_rows = metadata['sheet_info']['total_rows']
total_cols = metadata['sheet_info']['total_cols']

# 输出文件
output_file = r"d:\erp_thirteen\tp_education_system\backend\单元格对比结果.txt"

with open(output_file, 'w', encoding='utf-8') as f:
    f.write("=" * 120 + "\n")
    f.write("单元格逐一比对：原模板 vs 预览页面\n")
    f.write("=" * 120 + "\n\n")
    
    f.write("格式说明：\n")
    f.write("  单元格坐标 | 原模板内容 | 预览内容 | 是否一致\n")
    f.write("-" * 120 + "\n\n")
    
    # 模拟前端渲染逻辑
    occupied_cells = set()
    
    for row in range(total_rows):
        for col in range(total_cols):
            # 原模板内容
            original_cell = ws.cell(row=row + 1, column=col + 1)
            original_value = original_cell.value
            if original_value is None:
                original_str = "(空)"
            else:
                original_str = str(original_value)
            
            # 预览内容（从元数据）
            key = f"{row}-{col}"
            
            # 检查是否被合并单元格占用
            if key in occupied_cells:
                preview_str = "(被合并)"
                match = "✅"
            else:
                metadata_cell = cell_map.get(key)
                if metadata_cell:
                    preview_value = metadata_cell.get('value')
                    if preview_value:
                        preview_str = preview_value
                    else:
                        preview_str = "(空)"
                    
                    # 标记合并单元格占用的区域
                    rowspan = metadata_cell.get('rowspan', 1)
                    colspan = metadata_cell.get('colspan', 1)
                    for r in range(rowspan):
                        for c in range(colspan):
                            occupied_cells.add(f"{row + r}-{col + c}")
                else:
                    preview_str = "(无数据)"
            
            # 比较
            if original_str == preview_str or (original_str == "(空)" and preview_str == "(空)"):
                match = "✅"
            else:
                match = "❌"
            
            # 输出
            cell_name = f"({row + 1}, {col + 1})"
            f.write(f"{cell_name:12} | {original_str:40} | {preview_str:40} | {match}\n")
    
    f.write("\n" + "=" * 120 + "\n")
    f.write("比对完成\n")

wb.close()

print(f"对比结果已保存到: {output_file}")
print("\n请打开文件查看详细对比结果")
