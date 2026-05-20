import requests
import json
from openpyxl import load_workbook

response = requests.get("http://127.0.0.1:8000/api/template/list")
data = response.json()

if data["success"] and data["templates"]:
    template = data["templates"][0]
    file_path = template["file_path"]
    
    print("=" * 150)
    print("从实际 API 获取的模板信息")
    print("=" * 150)
    print(f"模板ID: {template['template_id']}")
    print(f"模板名称: {template['template_name']}")
    print(f"文件路径: {file_path}")
    print(f"行数: {template['row_count']}")
    print(f"列数: {template['col_count']}")
    print(f"单元格数: {template['cell_count']}")
    print(f"合并单元格数: {template['merge_count']}")
    
    print("\n" + "=" * 150)
    print("访问预览 API")
    print("=" * 150)
    
    preview_response = requests.get(f"http://127.0.0.1:8000/api/template/preview?file_path={file_path}")
    preview_data = preview_response.json()
    
    if preview_data.get("success"):
        metadata = preview_data["metadata"]
        
        print(f"元数据提取成功")
        print(f"工作表: {metadata['sheet_info']['sheet_name']}")
        print(f"总行数: {metadata['sheet_info']['total_rows']}")
        print(f"总列数: {metadata['sheet_info']['total_cols']}")
        print(f"单元格数: {len(metadata['cells'])}")
        print(f"合并单元格数: {len(metadata['merge_info'])}")
        
        print("\n" + "=" * 150)
        print("对比原Excel和预览API返回的数据")
        print("=" * 150)
        
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        cellMap = {}
        for cell in metadata['cells']:
            cellMap[f"{cell['row']}-{cell['col']}"] = cell
        
        occupiedCells = set()
        preview_contents = {}
        
        for row in range(metadata['sheet_info']['total_rows']):
            for col in range(metadata['sheet_info']['total_cols']):
                key = f"{row}-{col}"
                
                if key in occupiedCells:
                    continue
                
                cell = cellMap.get(key)
                
                if not cell:
                    preview_contents[(row + 1, col + 1)] = ""
                    continue
                
                rowspan = cell.get('rowspan', 1)
                colspan = cell.get('colspan', 1)
                
                for r in range(rowspan):
                    for c in range(colspan):
                        occupiedCells.add(f"{row + r}-{col + c}")
                
                value = cell.get('value', '')
                preview_contents[(row + 1, col + 1)] = value
        
        print(f"\n{'单元格序号':<15} {'原Excel内容':<60} {'预览API内容':<60}")
        print("-" * 150)
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                excel_cell = ws.cell(row=row, column=col)
                excel_value = str(excel_cell.value) if excel_cell.value is not None else ''
                
                preview_value = preview_contents.get((row, col), '')
                
                if excel_value or preview_value:
                    match = "OK" if excel_value == preview_value else "ERROR"
                    print(f"({row},{col})".ljust(15), f"'{excel_value}'".ljust(60), f"'{preview_value}'".ljust(60), match)
        
        print("=" * 150)
    else:
        print(f"预览API返回错误: {preview_data}")
else:
    print("获取模板列表失败")
