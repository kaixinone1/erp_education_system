"""
通用模板处理服务 - 一个引擎处理所有模板
实现100%格式复制：原模板 == 预览 == 导出
"""
import os
import json
import copy
import re
import hashlib
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import psycopg2

DATABASE_CONFIG = {
    "host": "localhost",
    "port": "5432",
    "database": "taiping_education",
    "user": "taiping_user",
    "password": "taiping_password"
}


def get_db_connection():
    return psycopg2.connect(**DATABASE_CONFIG)


def _auto_spread_field_values(filled_config, field_values, config):
    """
    字段值自动扩散：已填充的字段值自动匹配模板中所有同名标签并填充
    
    工作原理：
    1. 遍历模板中所有单元格，识别标签单元格（如"姓名："、"姓名"）
    2. 如果标签匹配已填充的字段名，查找其右侧或下方的空值单元格
    3. 自动填充值，无需为每个位置单独配置字段映射
    
    例如：字段"姓名"已填充为"张三"，
    模板中所有"姓名："标签旁边的空单元格都会被自动填充为"张三"
    """
    if not field_values:
        return
    
    # 构建字段名到值的映射（去除尾随标点）
    field_map = {}
    for fname, fval in field_values.items():
        if fval is None:
            continue
        fval_str = str(fval)
        if not fval_str.strip():
            continue
        # 键：原始字段名
        field_map[fname] = fval_str
        # 键：去除尾随的中文冒号
        clean = fname.rstrip('：:')
        if clean != fname:
            field_map[clean] = fval_str
    
    cells = filled_config.get('单元格数据', [])
    if not cells:
        return
    
    # 构建行列索引，快速查找单元格
    cell_index = {}
    for cell in cells:
        key = (cell.get('行号'), cell.get('列号'))
        cell_index[key] = cell
    
    # 已填充的位置集合（避免重复填充）
    already_filled = set()
    for cell in cells:
        val = cell.get('显示值')
        if val is not None and str(val).strip():
            already_filled.add((cell.get('行号'), cell.get('列号')))
    
    # 遍历所有单元格，识别标签并填充
    filled_count = 0
    for cell in cells:
        cell_val = str(cell.get('显示值', '') or '')
        cell_val_clean = cell_val.strip().rstrip('：:')
        
        if not cell_val_clean:
            continue
        
        # 检查是否匹配已填充字段
        matched_field = None
        for fname, fval in field_map.items():
            # 标签等于字段名，或标签以字段名+冒号开头
            if cell_val_clean == fname:
                matched_field = fname
                break
            if cell_val.startswith(fname) and cell_val[len(fname):].strip().startswith('：'):
                matched_field = fname
                break
        
        if not matched_field:
            continue
        
        row = cell.get('行号')
        col = cell.get('列号')
        target_val = field_map[matched_field]
        
        # 尝试填充右侧单元格
        right_key = (row, col + 1)
        if right_key in cell_index and right_key not in already_filled:
            right_cell = cell_index[right_key]
            right_val = str(right_cell.get('显示值', '') or '').strip()
            if not right_val or right_val == 'None':
                right_cell['值'] = target_val
                right_cell['显示值'] = target_val
                already_filled.add(right_key)
                filled_count += 1
        
        # 尝试填充下方单元格
        below_key = (row + 1, col)
        if below_key in cell_index and below_key not in already_filled:
            below_cell = cell_index[below_key]
            below_val = str(below_cell.get('显示值', '') or '').strip()
            if not below_val or below_val == 'None':
                below_cell['值'] = target_val
                below_cell['显示值'] = target_val
                already_filled.add(below_key)
                filled_count += 1
    
    # ========== 第二步：处理"同意...同志"模式 ==========
    # 模板中"同意XXX同志"格式，XXX位置需要填入姓名
    # 例如：B11"根据...同意" + J11"同志发放" → 中间H11填入姓名
    #       C16"根据...同意" + L16"同志" → 中间I16填入姓名
    name_value = field_map.get('姓名', None)
    if name_value:
        # 构建合并单元格映射：坐标 → 合并范围的结束行列
        merged_map = {}
        for mc in filled_config.get('合并单元格', []) or []:
            sr, sc = mc.get('起始行'), mc.get('起始列')
            er, ec = mc.get('结束行'), mc.get('结束列')
            if sr and sc and er and ec:
                for r in range(sr, er + 1):
                    for c in range(sc, ec + 1):
                        merged_map[(r, c)] = (er, ec)
        
        # 获取单元格实际占用的列范围（考虑合并）
        def get_cell_span(cell):
            row, col = cell.get('行号'), cell.get('列号')
            key = (row, col)
            if key in merged_map:
                er, ec = merged_map[key]
                return col, ec  # 返回起始列和结束列
            return col, col
        
        # 收集所有包含"同意"的单元格
        agree_cells = []
        comrade_cells = []
        for cell in cells:
            cv = str(cell.get('显示值', '') or '')
            if '同意' in cv:
                agree_cells.append(cell)
            if '同志' in cv:
                comrade_cells.append(cell)
        
        # 在同一行中，匹配"同意"和"同志"配对，填充中间空单元格
        for agree_cell in agree_cells:
            agree_row = agree_cell.get('行号')
            agree_col = agree_cell.get('列号')
            _, agree_end_col = get_cell_span(agree_cell)
            
            # 找到同一行中右侧最近的"同志"单元格
            best_comrade = None
            best_dist = 999
            for comrade_cell in comrade_cells:
                if comrade_cell.get('行号') == agree_row and comrade_cell.get('列号') > agree_end_col:
                    dist = comrade_cell.get('列号') - agree_end_col
                    if dist < best_dist:
                        best_dist = dist
                        best_comrade = comrade_cell
            
            if best_comrade:
                comrade_col = best_comrade.get('列号')
                # 填充"同意"结束列和"同志"起始列之间的所有空单元格
                for mid_col in range(agree_end_col + 1, comrade_col):
                    mid_key = (agree_row, mid_col)
                    if mid_key in cell_index and mid_key not in already_filled:
                        mid_cell = cell_index[mid_key]
                        mid_val = str(mid_cell.get('显示值', '') or '').strip()
                        # 跳过属于合并单元格的（已在merged_map中）
                        if mid_key in merged_map:
                            # 只填充合并单元格的左上角
                            sr_check = merged_map[mid_key][0]
                            if mid_key[0] != sr_check:
                                continue
                        if not mid_val or mid_val == 'None':
                            mid_cell['值'] = name_value
                            mid_cell['显示值'] = name_value
                            already_filled.add(mid_key)
                            filled_count += 1
    
    if filled_count > 0:
        print(f"[自动扩散] 已为 {len(field_values)} 个字段自动填充 {filled_count} 个位置")


class UniversalTemplateEngine:
    """通用模板处理引擎 - 实现100%格式复制"""
    
    def import_template(self, excel_path, template_name, template_type):
        """
        导入Excel模板，生成完整JSON配置（包含所有元数据）
        
        参数:
            excel_path: Excel文件路径
            template_name: 模板名称
            template_type: 模板类型（如：呈报表、审批表、公文）
        
        返回:
            JSON配置字典
        """
        wb = load_workbook(excel_path, read_only=False, keep_links=False)
        ws = wb.active
        
        config = {
            "模板ID": self._generate_template_id(template_name),
            "模板名称": template_name,
            "模板类型": template_type,
            "原始文件": os.path.basename(excel_path),
            "原始文件路径": excel_path,
            "导入时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "页面设置": self._extract_page_setup(ws),
            "页边距": self._extract_page_margins(ws),
            "列宽": self._extract_column_widths(ws),
            "行高": self._extract_row_heights(ws),
            "单元格数据": self._extract_all_cells(ws),
            "样式数据": self._extract_all_styles(ws),
            "行数据": self._extract_rows_data(ws),
            "合并单元格": self._extract_merged_cells(ws),
            "字段映射": {}
        }
        
        wb.close()
        return config
    
    def extract_full_metadata(self, excel_path):
        """
        提取Excel文件的完整元数据（用于Luckysheet预览）
        
        参数:
            excel_path: Excel文件路径
        
        返回:
            完整元数据字典
        """
        wb = load_workbook(excel_path, read_only=False, keep_links=False)
        ws = wb.active
        
        metadata = {
            'filename': os.path.basename(excel_path),
            'sheets': wb.sheetnames,
            'active_sheet': wb.sheetnames.index(ws.title),
            'cells': [],
            'styles': {},
            'dimensions': {
                'rows': {},
                'columns': {},
                'max_row': ws.max_row,
                'max_column': ws.max_column
            },
            'merged_cells': [],
            'page_setup': {},
            'page_margins': {}
        }
        
        for row in ws.iter_rows():
            for cell in row:
                if hasattr(cell, 'column_letter'):
                    cell_key = f"{cell.column_letter}{cell.row}"
                else:
                    cell_key = f"{get_column_letter(cell.column)}{cell.row}"
                
                if cell.value is not None:
                    metadata['cells'].append({
                        'r': cell.row - 1,
                        'c': cell.column - 1,
                        'v': {
                            'v': cell.value,
                            'm': str(cell.value),
                        }
                    })
                
                style = {}
                
                if cell.font:
                    font_color = str(cell.font.color.rgb) if cell.font.color and hasattr(cell.font.color, 'rgb') else None
                    style['font'] = {
                        'name': cell.font.name,
                        'size': cell.font.size,
                        'bold': cell.font.bold,
                        'italic': cell.font.italic,
                        'color': font_color,
                        'underline': cell.font.underline,
                        'strike': cell.font.strike
                    }
                
                if cell.fill and cell.fill.patternType:
                    fg_color = str(cell.fill.fgColor.rgb) if cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') else None
                    bg_color = str(cell.fill.bgColor.rgb) if cell.fill.bgColor and hasattr(cell.fill.bgColor, 'rgb') else None
                    style['fill'] = {
                        'patternType': cell.fill.patternType,
                        'fgColor': fg_color,
                        'bgColor': bg_color
                    }
                
                if cell.border:
                    style['border'] = {}
                    for side in ['left', 'right', 'top', 'bottom']:
                        side_obj = getattr(cell.border, side, None)
                        if side_obj and side_obj.style:
                            border_color = str(side_obj.color.rgb) if side_obj.color and hasattr(side_obj.color, 'rgb') and side_obj.color.rgb else None
                            style['border'][side] = {
                                'style': side_obj.style,
                                'color': border_color
                            }
                
                if cell.alignment:
                    style['alignment'] = {
                        'horizontal': cell.alignment.horizontal,
                        'vertical': cell.alignment.vertical,
                        'wrapText': cell.alignment.wrapText,
                        'textRotation': cell.alignment.textRotation
                    }
                
                if cell.number_format and cell.number_format != 'General':
                    style['numberFormat'] = cell.number_format
                
                if style:
                    metadata['styles'][cell_key] = style
        
        for row_idx in range(1, ws.max_row + 1):
            row_dim = ws.row_dimensions[row_idx]
            if row_dim and row_dim.height:
                metadata['dimensions']['rows'][row_idx] = row_dim.height
        
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            col_dim = ws.column_dimensions.get(col_letter)
            if col_dim and col_dim.width:
                metadata['dimensions']['columns'][col_letter] = col_dim.width
        
        if hasattr(ws, 'merged_cells') and ws.merged_cells:
            for merged_range in ws.merged_cells.ranges:
                metadata['merged_cells'].append({
                    'range': str(merged_range),
                    'r': merged_range.min_row - 1,
                    'c': merged_range.min_col - 1,
                    'rs': merged_range.max_row - merged_range.min_row + 1,
                    'cs': merged_range.max_col - merged_range.min_col + 1
                })
        
        if hasattr(ws, 'page_setup') and ws.page_setup:
            ps = ws.page_setup
            page_setup_data = {
                'paperSize': ps.paperSize,
                'orientation': ps.orientation
            }
            try:
                page_setup_data['fitToPage'] = ps.fitToPage
            except:
                pass
            try:
                page_setup_data['fitToWidth'] = ps.fitToWidth
            except:
                pass
            try:
                page_setup_data['fitToHeight'] = ps.fitToHeight
            except:
                pass
            metadata['page_setup'] = page_setup_data
        
        if hasattr(ws, 'page_margins') and ws.page_margins:
            pm = ws.page_margins
            metadata['page_margins'] = {
                'left': pm.left,
                'right': pm.right,
                'top': pm.top,
                'bottom': pm.bottom,
                'header': pm.header,
                'footer': pm.footer
            }
        
        wb.close()
        return metadata
    
    def _generate_template_id(self, template_name):
        """生成模板ID"""
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        hash_obj = hashlib.md5(f"{template_name}{timestamp}".encode())
        return f"tpl_{hash_obj.hexdigest()[:8]}"
    
    def _extract_all_cells(self, ws):
        """提取所有单元格数据（包含完整值和显示值）"""
        cells = []
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                cells.append({
                    '行号': cell.row,
                    '列号': cell.column,
                    '值': val,
                    '显示值': str(val) if val is not None else '',
                    '坐标': cell.coordinate
                })
        return cells
    
    def _extract_all_styles(self, ws):
        """提取所有单元格样式（完整样式信息）"""
        styles = {}
        for row in ws.iter_rows():
            for cell in row:
                cell_key = cell.coordinate
                style = {}
                
                if cell.font:
                    font_color = self._safe_extract_color(cell.font.color)
                    style['字体'] = {
                        '名称': cell.font.name,
                        '大小': cell.font.size,
                        '粗体': cell.font.bold,
                        '斜体': cell.font.italic,
                        '颜色': font_color,
                        '下划线': cell.font.underline,
                        '删除线': cell.font.strike
                    }
                
                if cell.fill:
                    fg_color = self._safe_extract_color(cell.fill.fgColor)
                    bg_color = self._safe_extract_color(cell.fill.bgColor)
                    if fg_color or bg_color:
                        style['填充'] = {
                            '图案类型': cell.fill.patternType,
                            '前景色': fg_color,
                            '背景色': bg_color
                        }
                
                if cell.border:
                    style['边框'] = {}
                    for side_name in ['left', 'right', 'top', 'bottom']:
                        side_obj = getattr(cell.border, side_name, None)
                        if side_obj and side_obj.style:
                            border_color = self._safe_extract_color(side_obj.color)
                            style['边框'][side_name] = {
                                '样式': side_obj.style,
                                '颜色': border_color
                            }
                
                if cell.alignment:
                    style['对齐'] = {
                        '水平': cell.alignment.horizontal,
                        '垂直': cell.alignment.vertical,
                        '自动换行': cell.alignment.wrapText,
                        '旋转角度': cell.alignment.textRotation
                    }
                
                if cell.number_format and cell.number_format != 'General':
                    style['数字格式'] = cell.number_format
                
                if style:
                    styles[cell_key] = style
        return styles
    
    def _safe_extract_color(self, color_obj):
        if not color_obj:
            return None
        try:
            rgb_val = None
            try:
                rgb_val = color_obj.rgb
            except (TypeError, AttributeError):
                pass

            if rgb_val and isinstance(rgb_val, str) and len(rgb_val) >= 6 and not str(rgb_val).startswith('Values must be'):
                if rgb_val == '00000000':
                    rgb_val = None
                else:
                    if len(rgb_val) == 8:
                        rgb_val = rgb_val[2:]
                    return str(rgb_val)

            if hasattr(color_obj, 'theme') and color_obj.theme is not None:
                theme_colors = {
                    0: 'FFFFFF', 1: '000000', 2: 'EEECE1', 3: '1F497D',
                    4: '4F81BD', 5: 'C0504D', 6: '9BBB59', 7: '8064A2',
                    8: '4BACC6', 9: 'F79646', 10: 'FF0000', 11: 'FFFF00',
                    12: '0000FF', 13: '800080', 14: '008000'
                }
                return theme_colors.get(color_obj.theme)

            if hasattr(color_obj, 'indexed') and color_obj.indexed is not None:
                indexed_colors = {
                    0: '000000', 1: 'FFFFFF', 2: 'FF0000', 3: '00FF00',
                    4: '0000FF', 5: 'FFFF00', 6: 'FF00FF', 7: '00FFFF',
                    8: '000000', 9: 'FFFFFF', 10: 'FF0000', 11: '00FF00',
                    12: '0000FF', 13: 'FFFF00', 14: 'FF00FF', 15: '00FFFF'
                }
                return indexed_colors.get(color_obj.indexed)

            return None
        except Exception:
            return None
    
    def _extract_row_heights(self, ws):
        """提取所有行高"""
        row_heights = {}
        for row_idx in range(1, ws.max_row + 1):
            row_dim = ws.row_dimensions[row_idx]
            if row_dim and row_dim.height:
                row_heights[row_idx] = row_dim.height
        return row_heights
    
    def _extract_page_margins(self, ws):
        """提取页边距"""
        if hasattr(ws, 'page_margins') and ws.page_margins:
            pm = ws.page_margins
            return {
                '左': pm.left,
                '右': pm.right,
                '上': pm.top,
                '下': pm.bottom,
                '页眉': pm.header,
                '页脚': pm.footer
            }
        return {
            '左': 0.71,
            '右': 0.71,
            '上': 0.98,
            '下': 0.98,
            '页眉': 0.5,
            '页脚': 0.5
        }
    
    def _extract_page_setup(self, ws):
        """提取页面设置"""
        paper_size = None
        try:
            paper_size = ws.page_setup.paperSize
        except:
            pass
        
        paper_width = None
        paper_height = None
        try:
            paper_width = ws.page_setup.paperWidth
        except:
            pass
        try:
            paper_height = ws.page_setup.paperHeight
        except:
            pass
        
        orientation = "纵向"
        try:
            if ws.page_setup.orientation == 'landscape':
                orientation = "横向"
        except:
            pass
        
        scale = None
        try:
            scale = ws.page_setup.scale
        except:
            pass
        
        margins = {"上": 0.75, "右": 0.7, "下": 0.75, "左": 0.7}
        try:
            if ws.page_margins:
                margins["上"] = ws.page_margins.top
                margins["右"] = ws.page_margins.right
                margins["下"] = ws.page_margins.bottom
                margins["左"] = ws.page_margins.left
        except:
            pass
        
        return {
            "纸张类型": paper_size,
            "纸张宽度": paper_width,
            "纸张高度": paper_height,
            "方向": orientation,
            "缩放比例": scale,
            "边距": margins
        }
    
    def _extract_column_widths(self, ws):
        col_widths = []
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            col_dim = ws.column_dimensions.get(col_letter)
            if col_dim and col_dim.width:
                width = col_dim.width
            else:
                width = 8.43
            col_widths.append(width)
        return col_widths
    
    def _extract_rows_data(self, ws):
        """提取行数据（修复：正确保留合并单元格的主单元格）"""
        rows_data = []
        
        for row_idx in range(1, ws.max_row + 1):
            row_data = {
                "行号": row_idx,
                "高度": ws.row_dimensions[row_idx].height if ws.row_dimensions[row_idx] and ws.row_dimensions[row_idx].height else 25.0,
                "单元格": []
            }
            
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, col_idx)
                
                # 检查是否是合并单元格的从属单元格（非主单元格）
                is_slave = False
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= row_idx <= merged_range.max_row and 
                        merged_range.min_col <= col_idx <= merged_range.max_col and
                        not (merged_range.min_row == row_idx and merged_range.min_col == col_idx)):
                        is_slave = True
                        break
                
                if is_slave:
                    continue
                
                cell_data = {
                    "列号": col_idx,
                    "文本": str(cell.value) if cell.value is not None else '',
                    "跨行": 1,
                    "跨列": 1,
                    "样式": self._extract_cell_style(cell),
                    "对齐": self._extract_cell_alignment(cell)
                }
                
                for merged_range in ws.merged_cells.ranges:
                    if merged_range.min_row == row_idx and merged_range.min_col == col_idx:
                        cell_data["跨行"] = merged_range.max_row - merged_range.min_row + 1
                        cell_data["跨列"] = merged_range.max_col - merged_range.min_col + 1
                        break
                
                row_data["单元格"].append(cell_data)
            
            rows_data.append(row_data)
        
        return rows_data
    
    def _extract_cell_style(self, cell):
        """提取单元格样式"""
        styles = []
        
        if cell.font:
            if cell.font.size:
                styles.append(f"font-size:{int(cell.font.size)}pt")
            if cell.font.bold:
                styles.append("font-weight:bold")
            if cell.font.name:
                styles.append(f"font-family:{cell.font.name}")
        
        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
            if cell.fill.start_color.rgb != '00000000':
                styles.append(f"background-color:#{cell.fill.start_color.rgb}")
        
        return ";".join(styles) if styles else ""
    
    def _extract_cell_alignment(self, cell):
        """提取单元格对齐方式"""
        if cell.alignment:
            if cell.alignment.horizontal == 'left':
                return 'left'
            elif cell.alignment.horizontal == 'right':
                return 'right'
        return 'center'
    
    def _extract_merged_cells(self, ws):
        """提取合并单元格信息"""
        merged = []
        for merged_range in ws.merged_cells.ranges:
            merged.append({
                "起始": f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}",
                "结束": f"{get_column_letter(merged_range.max_col)}{merged_range.max_row}",
                "起始行": merged_range.min_row,
                "起始列": merged_range.min_col,
                "结束行": merged_range.max_row,
                "结束列": merged_range.max_col
            })
        return merged
    
    def preview_template(self, config, query_params=None, excel_path=None):
        """
        预览模板 - 直接从Excel文件读取单元格布局，确保与原表100%一致
        
        参数:
            config: JSON配置字典（含填充后的单元格数据）
            query_params: 查询参数
            excel_path: Excel文件路径（用于读取原始布局）
        
        返回:
            HTML字符串
        """
        if excel_path and os.path.exists(excel_path):
            return self._preview_from_excel(excel_path, config, query_params)
        return self._preview_from_config(config, query_params)
    
    def _preview_from_excel(self, excel_path, config, query_params=None):
        """直接从Excel文件生成HTML，应用填充值"""
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        wb = load_workbook(excel_path, read_only=False, keep_links=False)
        ws = wb.active
        
        max_row = ws.max_row
        max_col = ws.max_column
        
        # 列宽（Excel单位）
        col_widths = []
        for c in range(1, max_col + 1):
            cl = get_column_letter(c)
            cd = ws.column_dimensions.get(cl)
            w = (cd.width or 8.43) if cd else 8.43
            col_widths.append(float(w))
        
        # 计算总宽度和百分比
        total_width = sum(col_widths)
        
        # 合并单元格
        merge_map = {}
        for mc in ws.merged_cells.ranges:
            for r in range(mc.min_row, mc.max_row + 1):
                for c in range(mc.min_col, mc.max_col + 1):
                    key = f"{r}-{c}"
                    if r == mc.min_row and c == mc.min_col:
                        merge_map[key] = {
                            'rs': mc.max_row - mc.min_row + 1,
                            'cs': mc.max_col - mc.min_col + 1,
                            'master': True
                        }
                    else:
                        merge_map[key] = {'master': False}
        
        # 填充值映射
        cells_data = config.get('单元格数据', [])
        filled_map = {}
        for cell in cells_data:
            filled_map[f"{cell['行号']}-{cell['列号']}"] = cell
        
        # 页面设置信息
        page_setup = config.get('页面设置', {})
        orientation = page_setup.get('方向', '纵向')
        paper_type = page_setup.get('纸张类型', 9)
        margins = page_setup.get('边距', {})
        left_mm = round(float(margins.get('左', 0.354)) * 25.4)
        right_mm = round(float(margins.get('右', 0.354)) * 25.4)
        top_mm = round(float(margins.get('上', 0.984)) * 25.4)
        bottom_mm = round(float(margins.get('下', 0.984)) * 25.4)
        
        paper_dimensions = {
            1: (216, 279), 3: (279, 432), 4: (432, 279),
            5: (216, 356), 8: (297, 420), 9: (210, 297),
            11: (148, 210), 12: (250, 353), 13: (176, 250),
        }
        pw, ph = paper_dimensions.get(paper_type, (210, 297))
        if orientation == '横向':
            page_w_mm, page_h_mm = ph, pw
        else:
            page_w_mm, page_h_mm = pw, ph
        
        printable_w_mm = page_w_mm - left_mm - right_mm
        printable_h_mm = page_h_mm - top_mm - bottom_mm
        
        # 检测页面分页行（基于行高和可打印高度计算）
        page_break_rows = set()
        
        # 计算每行高度（像素），累加到超过可打印高度时分页
        accumulated_h = 0
        # 可打印高度转为像素：mm * 96dpi / 25.4
        printable_h_px = printable_h_mm * 96.0 / 25.4
        
        for r in range(1, max_row + 1):
            rd = ws.row_dimensions.get(r)
            rh = (rd.height or 20) if rd else 20
            accumulated_h += rh
            
            if accumulated_h > printable_h_px and r < max_row:
                page_break_rows.add(r)
                accumulated_h = rh  # 重置，新页从当前行开始
        
        # 生成HTML
        html = self._build_print_styles_full(config, page_w_mm, page_h_mm, left_mm, right_mm, top_mm, bottom_mm)
        
        # 表格容器：使用百分比宽度确保适配打印页面
        html += '<div class="table-wrapper" style="width:100%;overflow-x:auto;">'
        html += '<table style="border-collapse:collapse;font-family:SimSun,serif;width:100%;table-layout:fixed;">'
        html += '<colgroup>'
        for w in col_widths:
            pct = f"{(w / total_width * 100):.4f}%"
            # 设置最小列宽60px，防止中文短文本在窄列中异常换行
            html += f'<col style="width:{pct};min-width:60px">'
        html += '</colgroup>'
        
        for r in range(1, max_row + 1):
            rd = ws.row_dimensions.get(r)
            rh = (rd.height or 20) if rd else 20
            row_style = f'height:{rh:.0f}px'
            if r in page_break_rows:
                row_style += ';page-break-before:always'
            html += f'<tr style="{row_style}">'
            
            c = 1
            while c <= max_col:
                key = f"{r}-{c}"
                minfo = merge_map.get(key)
                if minfo and not minfo['master']:
                    c += 1
                    continue
                
                cell = ws.cell(r, c)
                style = self._build_cell_style_excel(cell)
                
                # 优先使用填充值
                filled = filled_map.get(key, {})
                cv = filled.get('显示值', '')
                if cv is None:
                    cv = ''
                cv = str(cv)
                
                raw_val = filled.get('值', '')
                if raw_val and isinstance(raw_val, str) and '{{' in raw_val:
                    cv = self._resolve_placeholders(raw_val, query_params)
                
                # 如果没有填充值，使用原始文本
                if not cv and cell.value is not None:
                    cv = str(cell.value)
                
                cv = cv.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                cv = cv.replace('\r\n', '\n').replace('\r', '\n')
                has_newline = '\n' in cv
                cv = cv.replace('\n', '<br>')
                
                if has_newline:
                    style = style.replace('text-align:center', 'text-align:left')
                    if 'white-space' not in style:
                        style += ';white-space:pre-wrap;word-break:break-all'
                
                # 短中文文本（≤6字）且无换行符，使用 nowrap 防止意外换行
                if not has_newline and len(cv) <= 6 and any('\u4e00' <= ch <= '\u9fff' for ch in cv):
                    if 'white-space' not in style:
                        style += ';white-space:nowrap'
                
                rowspan = minfo['rs'] if minfo else 1
                colspan = minfo['cs'] if minfo else 1
                
                rs_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
                cs_attr = f' colspan="{colspan}"' if colspan > 1 else ''
                
                html += f'<td data-row="{r}" data-col="{c}" style="{style}"{rs_attr}{cs_attr}>{cv or "&nbsp;"}</td>'
                
                c += colspan
            
            html += '</tr>'
        
        html += '</table></div>'
        wb.close()
        
        # 落款日期字间距调整
        def add_date_letter_spacing(match):
            s = match.group(1)
            rest = match.group(2)
            content = match.group(3)
            if 'letter-spacing' not in s:
                s += ';letter-spacing:0.5em'
            return f'<td style="{s}"{rest}>{content}</td>'
        
        html = re.sub(
            r'<td style="([^"]*)"([^>]*?)>(\s*年\s+月\s+日\s*)</td>',
            add_date_letter_spacing, html
        )
        
        return html
    
    def _build_cell_style_excel(self, cell):
        """从openpyxl单元格提取完整样式为CSS内联样式"""
        styles = ['padding:2px 4px']
        
        if cell.font:
            f = cell.font
            if f.name:
                styles.append(f"font-family:'{f.name}'")
            if f.size:
                styles.append(f"font-size:{f.size}pt")
            if f.bold:
                styles.append('font-weight:bold')
            if f.italic:
                styles.append('font-style:italic')
            if f.underline and f.underline != 'none':
                styles.append('text-decoration:underline')
            if f.color:
                try:
                    c = f.color.rgb
                    if c and c != '00000000' and len(c) >= 6:
                        styles.append(f"color:#{c[-6:]}")
                except:
                    pass
        
        if cell.fill and cell.fill.start_color:
            try:
                c = cell.fill.start_color.rgb
                if c and c != '00000000' and len(c) >= 6:
                    styles.append(f"background-color:#{c[-6:]}")
            except:
                pass
        
        # 边框
        border_map = {'thin':'1px','medium':'2px','thick':'3px','hair':'0.5px',
            'dashed':'1px','dotted':'1px','double':'3px','mediumDashed':'2px'}
        if cell.border:
            for side_name, css_prop in [('top','border-top'),('bottom','border-bottom'),('left','border-left'),('right','border-right')]:
                side = getattr(cell.border, side_name, None)
                if side and side.style and side.style != 'none':
                    w = border_map.get(side.style, '1px')
                    try:
                        sc = side.color.rgb if side.color else None
                        color = f"#{sc[-6:]}" if sc and sc != '00000000' else '#000'
                    except:
                        color = '#000'
                    styles.append(f"{css_prop}:{w} solid {color}")
        else:
            styles.append('border:none')
        
        # 对齐
        if cell.alignment:
            a = cell.alignment
            h = a.horizontal or 'center'
            v = a.vertical or 'center'
            h_map = {'left':'left','center':'center','right':'right','justify':'justify'}
            v_map = {'top':'top','center':'middle','bottom':'bottom','justify':'middle'}
            styles.append(f"text-align:{h_map.get(h,'center')}")
            styles.append(f"vertical-align:{v_map.get(v,'middle')}")
            if a.wrap_text:
                styles.append('white-space:pre-wrap;word-break:break-all')
        
        return ';'.join(styles)
    
    def _preview_from_config(self, config, query_params=None):
        """降级方案：使用配置JSON渲染（兼容旧数据）"""
        col_widths = config.get('列宽', [])
        row_heights = config.get('行高', {})
        cells_data = config.get('单元格数据', [])
        styles_data = config.get('样式数据', {})
        merged_cells = config.get('合并单元格', [])
        
        max_row = 0
        max_col = len(col_widths)
        for cell in cells_data:
            max_row = max(max_row, cell['行号'])
        
        merge_map = {}
        for mc in merged_cells:
            for r in range(mc['起始行'], mc['结束行'] + 1):
                for c in range(mc['起始列'], mc['结束列'] + 1):
                    key = f"{r}-{c}"
                    if r == mc['起始行'] and c == mc['起始列']:
                        merge_map[key] = {
                            'rs': mc['结束行'] - mc['起始行'] + 1,
                            'cs': mc['结束列'] - mc['起始列'] + 1,
                            'master': True
                        }
                    else:
                        merge_map[key] = {'master': False}
        
        cell_value_map = {}
        for cell in cells_data:
            cell_value_map[f"{cell['行号']}-{cell['列号']}"] = cell
        
        html = self._build_print_styles(config)
        html += '<table style="border-collapse:collapse;font-family:SimSun,serif;">'
        html += '<colgroup>'
        for w in col_widths:
            pixel_w = max(int(float(w) * 9), 30)
            html += f'<col style="width:{pixel_w}px;min-width:{pixel_w}px">'
        html += '</colgroup>'
        
        for r in range(1, max_row + 1):
            rh = row_heights.get(str(r), row_heights.get(r, None))
            height_style = f' style="height:{rh}px;"' if rh else ''
            html += f'<tr{height_style}>'
            
            c = 1
            while c <= max_col:
                key = f"{r}-{c}"
                minfo = merge_map.get(key)
                
                if minfo and not minfo['master']:
                    c += 1
                    continue
                
                coord = f"{get_column_letter(c)}{r}"
                style_info = styles_data.get(coord, {})
                
                cell = cell_value_map.get(key, {})
                cv = cell.get('显示值', '')
                if cv is None:
                    cv = ''
                cv = str(cv)
                
                raw_val = cell.get('值', '')
                if raw_val and isinstance(raw_val, str) and '{{' in raw_val:
                    cv = self._resolve_placeholders(raw_val, query_params)
                
                cv = cv.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                cv = cv.replace('\r\n', '\n').replace('\r', '\n')
                has_newline = '\n' in cv
                cv = cv.replace('\n', '<br>')
                
                rowspan = minfo['rs'] if minfo else 1
                colspan = minfo['cs'] if minfo else 1
                
                cell_style = self._build_cell_inline_style(style_info)
                if has_newline:
                    cell_style = cell_style.replace('text-align:center', 'text-align:left')
                    cell_style = (cell_style + ';text-align:left;vertical-align:top;white-space:pre-wrap;word-break:break-all') if cell_style else 'text-align:left;vertical-align:top;white-space:pre-wrap;word-break:break-all'
                
                rs_attr = f' rowspan="{rowspan}"' if rowspan > 1 else ''
                cs_attr = f' colspan="{colspan}"' if colspan > 1 else ''
                
                html += f'<td data-row="{r}" data-col="{c}" style="{cell_style}"{rs_attr}{cs_attr}>{cv or "&nbsp;"}</td>'
                
                c += colspan
            
            html += '</tr>'
        
        html += '</table>'
        
        def add_date_letter_spacing(match):
            style = match.group(1)
            rest = match.group(2)
            content = match.group(3)
            if 'letter-spacing' not in style:
                style += ';letter-spacing:0.5em'
            return f'<td style="{style}"{rest}>{content}</td>'
        
        html = re.sub(
            r'<td style="([^"]*)"([^>]*?)>(\s*年\s+月\s+日\s*)</td>',
            add_date_letter_spacing, html
        )
        
        return html
    
    def _build_cell_inline_style(self, style_info):
        style_parts = ['padding: 2px 4px']
        
        font_info = style_info.get('字体', {})
        if font_info.get('名称'):
            style_parts.append(f"font-family: '{font_info['名称']}'")
        if font_info.get('大小'):
            style_parts.append(f"font-size: {font_info['大小']}pt")
        if font_info.get('粗体'):
            style_parts.append('font-weight: bold')
        if font_info.get('斜体'):
            style_parts.append('font-style: italic')
        if font_info.get('颜色'):
            style_parts.append(f"color: #{font_info['颜色']}")
        if font_info.get('删除线'):
            style_parts.append('text-decoration: line-through')
        
        fill_info = style_info.get('填充', {})
        if fill_info.get('前景色'):
            style_parts.append(f"background-color: #{fill_info['前景色']}")
        
        border_info = style_info.get('边框', {})
        if border_info:
            border_width_map = {
                'thin': '1px', 'medium': '2px', 'thick': '3px',
                'hair': '0.5px', 'dotted': '1px', 'dashed': '1px',
                'dashDot': '1px', 'dashDotDot': '1px',
                'double': '3px', 'mediumDashed': '2px', 'mediumDashDot': '2px',
                'mediumDashDotDot': '2px', 'slantDashDot': '1px'
            }
            for side in ['top', 'bottom', 'left', 'right']:
                side_info = border_info.get(side)
                if side_info and side_info.get('样式'):
                    color = side_info.get('颜色', '000000')
                    width = border_width_map.get(side_info['样式'], '1px')
                    if color:
                        style_parts.append(f"border-{side}: {width} solid #{color}")
                    else:
                        style_parts.append(f"border-{side}: {width} solid #000")
        else:
            style_parts.append('border: none')
        
        align_info = style_info.get('对齐', {})
        h_align = align_info.get('水平')
        v_align = align_info.get('垂直')
        if h_align:
            style_parts.append(f"text-align: {h_align}")
        if v_align:
            style_parts.append(f"vertical-align: {'middle' if v_align == 'center' else v_align}")
        if align_info.get('自动换行'):
            style_parts.append('white-space: pre-wrap; word-wrap: break-word')
        
        result = '; '.join(style_parts)
        return result
    
    def _resolve_placeholders(self, text, query_params=None):
        import re
        from datetime import datetime as dt
        
        now = dt.now()
        year = now.year
        month = now.month
        
        if query_params and '年月' in query_params:
            ym = query_params['年月']
            parts = ym.split('-')
            if len(parts) == 2:
                year = int(parts[0])
                month = int(parts[1])
        
        def resolve_date(match):
            key = match.group(1).strip()
            
            if key == '年月+1':
                m = month + 1
                y = year
                if m > 12:
                    m = 1
                    y += 1
                return f'{y}年{m}月'
            elif key == '年月-1':
                m = month - 1
                y = year
                if m < 1:
                    m = 12
                    y -= 1
                return f'{y}年{m}月'
            elif key == '年月':
                return f'{year}年{month}月'
            elif key == '年月日':
                return f'{now.year}年{now.month}月{now.day}日'
            elif key == '年':
                return str(year)
            elif key == '月':
                return str(month)
            elif key == '日':
                return str(now.day)
            else:
                return match.group(0)
        
        result = re.sub(r'\{\{([^{}]+)\}\}', resolve_date, str(text))
        return result
    
    def generate_print_html(self, config, query_params=None, excel_path=None):
        """
        生成打印用的完整HTML文档
        
        参数:
            config: JSON配置字典
            query_params: 查询参数（如：{"年月": "2026-05"}），用于解析日期占位符
            excel_path: Excel文件路径（用于读取原始布局）

        返回:
            完整HTML文档字符串
        """
        table_html = self.preview_template(config, query_params, excel_path=excel_path)

        page_setup = config.get('页面设置', {})
        orientation = page_setup.get('方向', '纵向')
        css_orientation = 'landscape' if orientation == '横向' else 'portrait'

        paper_size_map = {
            1: 'letter', 3: 'tabloid', 4: 'ledger',
            5: 'legal', 8: 'A3', 9: 'A4', 11: 'A5',
            12: 'B4', 13: 'B5',
        }
        paper_dimensions = {
            1: (216, 279), 3: (279, 432), 4: (432, 279),
            5: (216, 356), 6: (140, 216), 7: (184, 267),
            8: (297, 420), 9: (210, 297), 10: (210, 297),
            11: (148, 210), 12: (250, 353), 13: (176, 250),
        }

        paper_type = page_setup.get('纸张类型')
        page_size_name = None
        page_width_mm = None
        page_height_mm = None

        if paper_type and paper_type in paper_size_map:
            page_size_name = paper_size_map[paper_type]
        elif paper_type and paper_type in paper_dimensions:
            w, h = paper_dimensions[paper_type]
            if orientation == '横向':
                page_width_mm, page_height_mm = h, w
            else:
                page_width_mm, page_height_mm = w, h

        if not page_size_name and not page_width_mm:
            paper_width = page_setup.get('纸张宽度')
            paper_height = page_setup.get('纸张高度')
            if paper_width and paper_height:
                if orientation == '横向':
                    page_width_mm = round(float(paper_height) * 0.3528)
                    page_height_mm = round(float(paper_width) * 0.3528)
                else:
                    page_width_mm = round(float(paper_width) * 0.3528)
                    page_height_mm = round(float(paper_height) * 0.3528)
            else:
                page_size_name = 'A4'

        margins = page_setup.get('边距', {})
        top_inch = float(margins.get('上', 0.984))
        right_inch = float(margins.get('右', 0.354))
        bottom_inch = float(margins.get('下', 0.984))
        left_inch = float(margins.get('左', 0.354))
        top_mm = round(top_inch * 25.4)
        right_mm = round(right_inch * 25.4)
        bottom_mm = round(bottom_inch * 25.4)
        left_mm = round(left_inch * 25.4)

        if page_size_name:
            page_rule = f'@page {{ size: {page_size_name} {css_orientation}; margin: {top_mm}mm {right_mm}mm {bottom_mm}mm {left_mm}mm; }}'
        else:
            page_rule = f'@page {{ size: {page_width_mm}mm {page_height_mm}mm; margin: {top_mm}mm {right_mm}mm {bottom_mm}mm {left_mm}mm; }}'

        template_name = config.get('模板名称', '文档')

        full_html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{template_name}</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{
    font-family: SimSun, serif;
    display: flex;
    justify-content: center;
    align-items: flex-start;
    padding: {top_mm}mm {right_mm}mm {bottom_mm}mm {left_mm}mm;
    min-height: 100vh;
}}
{page_rule}
@media print {{
    body {{
        margin: 0;
        padding: 0;
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
    }}
    @page {{
        margin: {top_mm}mm {right_mm}mm {bottom_mm}mm {left_mm}mm;
    }}
}}
table {{
    -webkit-print-color-adjust: exact;
    print-color-adjust: exact;
}}
</style>
</head>
<body>
{table_html}
</body>
</html>'''

        return full_html

    def _build_print_styles(self, config):
        page_setup = config.get('页面设置', {})
        orientation = page_setup.get('方向', '纵向')
        css_orientation = 'landscape' if orientation == '横向' else 'portrait'
        
        paper_size_map = {
            1: 'letter', 3: 'tabloid', 4: 'ledger',
            5: 'legal', 8: 'A3', 9: 'A4', 11: 'A5',
            12: 'B4', 13: 'B5',
        }
        
        paper_dimensions = {
            1: (216, 279), 3: (279, 432), 4: (432, 279),
            5: (216, 356), 6: (140, 216), 7: (184, 267),
            8: (297, 420), 9: (210, 297), 10: (210, 297),
            11: (148, 210), 12: (250, 353), 13: (176, 250),
        }
        
        paper_type = page_setup.get('纸张类型')
        
        page_size_name = None
        page_width_mm = None
        page_height_mm = None
        
        if paper_type and paper_type in paper_size_map:
            page_size_name = paper_size_map[paper_type]
        elif paper_type and paper_type in paper_dimensions:
            w, h = paper_dimensions[paper_type]
            if orientation == '横向':
                page_width_mm, page_height_mm = h, w
            else:
                page_width_mm, page_height_mm = w, h
        
        if not page_size_name and not page_width_mm:
            paper_width = page_setup.get('纸张宽度')
            paper_height = page_setup.get('纸张高度')
            if paper_width and paper_height:
                if orientation == '横向':
                    page_width_mm = round(float(paper_height) * 0.3528)
                    page_height_mm = round(float(paper_width) * 0.3528)
                else:
                    page_width_mm = round(float(paper_width) * 0.3528)
                    page_height_mm = round(float(paper_height) * 0.3528)
            else:
                page_size_name = 'A4'
        
        margins = page_setup.get('边距', {})
        top_inch = float(margins.get('上', 0.984))
        right_inch = float(margins.get('右', 0.354))
        bottom_inch = float(margins.get('下', 0.984))
        left_inch = float(margins.get('左', 0.354))
        
        top_mm = round(top_inch * 25.4)
        right_mm = round(right_inch * 25.4)
        bottom_mm = round(bottom_inch * 25.4)
        left_mm = round(left_inch * 25.4)
        
        styles = '<style>'
        
        if page_size_name:
            styles += f'@page {{ size: {page_size_name} {css_orientation}; margin: {top_mm}mm {right_mm}mm {bottom_mm}mm {left_mm}mm; }}'
        else:
            styles += f'@page {{ size: {page_width_mm}mm {page_height_mm}mm; margin: {top_mm}mm {right_mm}mm {bottom_mm}mm {left_mm}mm; }}'
        
        styles += '@media print { '
        styles += 'body { margin: 0; padding: 0; -webkit-print-color-adjust: exact; print-color-adjust: exact; }'
        styles += '}'
        styles += 'table { -webkit-print-color-adjust: exact; print-color-adjust: exact; }'
        styles += '</style>'
        return styles
    
    def _build_print_styles_full(self, config, page_w_mm, page_h_mm, left_mm, right_mm, top_mm, bottom_mm):
        """生成完整的打印样式CSS（含预览容器样式）"""
        styles = '<style>'
        styles += f'@page {{ size: {page_w_mm}mm {page_h_mm}mm; margin: {top_mm}mm {right_mm}mm {bottom_mm}mm {left_mm}mm; }}'
        styles += '@media print { '
        styles += 'body { margin: 0; padding: 0; -webkit-print-color-adjust: exact; print-color-adjust: exact; }'
        styles += '.table-wrapper { overflow: visible !important; width: 100% !important; }'
        styles += '}'
        styles += 'table { -webkit-print-color-adjust: exact; print-color-adjust: exact; }'
        styles += '@media screen { '
        styles += '.table-wrapper { max-width: 100%; overflow-x: auto; padding: 10px; }'
        styles += '}'
        styles += '</style>'
        return styles
    
    def fill_template_data(self, config, query_params):
        """
        自动填报数据（基于单元格数据直接修改值）
        
        两轮处理：
        1. 先处理所有数据库映射（含聚合查询），收集字段值
        2. 再处理所有公式映射，引用已填字段进行计算
        
        参数:
            config: JSON配置字典（包含字段映射）
            query_params: 查询参数（如：{"职工ID": "xxx", "年月": "2026-05"}）
        
        返回:
            填充后的JSON配置
        """
        filled_config = copy.deepcopy(config)
        
        for cell in filled_config.get('单元格数据', []):
            raw_val = cell.get('值', '')
            if raw_val and isinstance(raw_val, str) and '{{' in raw_val:
                resolved = self._resolve_placeholders(raw_val, query_params)
                cell['值'] = resolved
                cell['显示值'] = resolved
        
        if not config.get('字段映射'):
            return filled_config
        
        STAT_METHODS = ('计数', '求和', '平均值', '最大值', '最小值', '求积')
        
        db_mappings = {}
        formula_mappings = {}
        
        for field_name, mapping in config['字段映射'].items():
            data_source = mapping.get('数据源', '')
            transform_func = mapping.get('转换函数', '')
            
            if transform_func and not data_source:
                formula_mappings[field_name] = mapping
            elif data_source:
                db_mappings[field_name] = mapping
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # #region debug-point B:mappings-loaded
        import json as _dj, time as _dt, os as _dos
        _logp = _dos.path.join(_dos.path.dirname(_dos.path.dirname(__file__)), 'debug_fill_log.txt')
        with open(_logp, 'a', encoding='utf-8') as _df:
            _df.write(f"[{_dt.time()}] B:mappings db={len(db_mappings)} formula={len(formula_mappings)} detail={ {k:v.get('数据源','') for k,v in db_mappings.items()} }\n")
        # #endregion
        
        try:
            field_values = {}
            
            for field_name, mapping in db_mappings.items():
                data_source = mapping.get('数据源', '')
                if not data_source:
                    continue
                
                table, column = data_source.split('.')
                transform_func = mapping.get('转换函数', '')
                dict_values = mapping.get('字典值选择', [])
                
                if transform_func in STAT_METHODS:
                    row_filter = mapping.get('行筛选', {})
                    value = self._query_aggregated_value(cursor, table, column, transform_func, config, dict_values, row_filter)
                else:
                    value = self._query_field_value(cursor, table, column, query_params)
                
                if value is not None:
                    field_values[field_name] = value
                    # 支持同一字段多处位置：遍历位置列表填充所有位置
                    for pos in mapping.get('位置列表', [{'行': mapping.get('行'), '列': mapping.get('列')}]):
                        self._set_cell_value(filled_config, pos['行'], pos['列'], value)
            
            # #region debug-point E:values-collected
            import json as _dj, time as _dt, os as _dos
            _logp = _dos.path.join(_dos.path.dirname(_dos.path.dirname(__file__)), 'debug_fill_log.txt')
            with open(_logp, 'a', encoding='utf-8') as _df:
                _df.write(f"[{_dt.time()}] E:collected count={len(field_values)} values={ {k:str(v) for k,v in field_values.items()} }\n")
            # #endregion
            
            sorted_formula_mappings = sorted(formula_mappings.items(), key=lambda x: (x[1].get('行', 0), x[1].get('列', 0)))
            
            for field_name, mapping in sorted_formula_mappings:
                formula = mapping.get('转换函数', '')
                if not formula:
                    continue
                
                value = self._evaluate_formula(formula, field_values)
                
                if value is not None:
                    # 支持同一字段多处位置：遍历位置列表填充所有位置
                    for pos in mapping.get('位置列表', [{'行': mapping.get('行'), '列': mapping.get('列')}]):
                        self._set_cell_value(filled_config, pos['行'], pos['列'], value)
                    field_values[field_name] = value
            
            count_standard_pairs = []
            field_to_cells = {}  # 改为支持多位置
            for field_name, mapping in config['字段映射'].items():
                field_to_cells[field_name] = mapping.get('位置列表', [{'行': mapping.get('行'), '列': mapping.get('列')}])
            
            for field_name, mapping in formula_mappings.items():
                formula = mapping.get('转换函数', '')
                import re as _re
                refs = _re.findall(r'\{(.+?)\}', formula)
                if len(refs) == 2 and '*' in formula:
                    count_field = None
                    standard_field = None
                    if '人数' in refs[0] and '标准' in refs[1]:
                        count_field, standard_field = refs[0], refs[1]
                    elif '人数' in refs[1] and '标准' in refs[0]:
                        count_field, standard_field = refs[1], refs[0]
                    if count_field and standard_field:
                        count_standard_pairs.append((count_field, standard_field))
            
            for cell in filled_config.get('单元格数据', []):
                val = cell.get('显示值')
                if val == 0 or val == '0':
                    cell['显示值'] = ''
                    cell['值'] = ''
            
            for count_field, standard_field in count_standard_pairs:
                count_val = field_values.get(count_field, None)
                if count_val == 0 or count_val == '0':
                    if standard_field in field_to_cells:
                        for pos in field_to_cells[standard_field]:
                            sr, sc = pos['行'], pos['列']
                            for cell in filled_config.get('单元格数据', []):
                                if cell.get('行号') == sr and cell.get('列号') == sc:
                                    cell['显示值'] = ''
                                    cell['值'] = ''
                                    break
        finally:
            cursor.close()
            conn.close()
        
        模板名称 = config.get('模板名称', '')
        if '绩效工资审批' in 模板名称:
            try:
                year_month = query_params.get('年月', '')
                if not year_month:
                    year_month = query_params.get('年月日', '')
                if year_month and '-' in year_month:
                    parts = year_month.split('-')
                    if len(parts) == 2:
                        year = int(parts[0])
                        month = int(parts[1])
                        remarks = self.get_performance_remarks(year, month)
                        if remarks:
                            for cell in filled_config.get('单元格数据', []):
                                if cell.get('行号') == 28 and cell.get('列号') == 1:
                                    现有文本 = str(cell.get('显示值', '') or '')
                                    cell['显示值'] = f"备注：\n{remarks}"
                                    cell['值'] = cell['显示值']
                                    break
                            备注合并 = {
                                "结束": "E28",
                                "起始": "A28",
                                "结束列": 5,
                                "结束行": 28,
                                "起始列": 1,
                                "起始行": 28
                            }
                            现有合并列表 = filled_config.get('合并单元格', [])
                            if not isinstance(现有合并列表, list):
                                现有合并列表 = []
                            已有 = any(
                                m.get('起始行') == 28 and m.get('起始列') == 1
                                for m in 现有合并列表
                            )
                            if not 已有:
                                现有合并列表.append(备注合并)
                                filled_config['合并单元格'] = 现有合并列表
            except Exception as e:
                print(f"备注信息填充失败: {e}")
            
            # 人事部门意见栏填充
            try:
                绩效人数 = None
                绩效金额 = None
                乡镇补贴人数 = None
                乡镇补贴金额 = None
                遗留人数 = None
                遗留金额 = None
                
                for cell in filled_config.get('单元格数据', []):
                    r, c, v = cell.get('行号'), cell.get('列号'), cell.get('显示值')
                    if r == 20 and c == 2:
                        绩效人数 = int(float(v)) if v and str(v).strip() else 0
                    elif r == 20 and c == 4:
                        绩效金额 = float(v) if v and str(v).strip() else 0.0
                    elif r == 21 and c == 2:
                        乡镇补贴人数 = int(float(v)) if v and str(v).strip() else 0
                    elif r == 21 and c == 4:
                        乡镇补贴金额 = float(v) if v and str(v).strip() else 0.0
                    elif r == 24 and c == 2:
                        遗留人数 = int(float(v)) if v and str(v).strip() else 0
                    elif r == 24 and c == 3:
                        遗留金额 = float(v) if v and str(v).strip() else 0.0
                
                if 绩效人数 is not None and 绩效金额 is not None:
                    合计金额 = 绩效金额 + (乡镇补贴金额 or 0.0)
                    总计金额 = 合计金额 + (遗留金额 or 0.0)
                    
                    无乡镇补贴名单 = ''
                    无乡镇补贴人数 = 0
                    try:
                        conn2 = get_db_connection()
                        cur2 = conn2.cursor()
                        cur2.execute("""
                            SELECT t.name FROM teacher_basic_info t
                            JOIN employee_tag_relations etr1 ON t.id = etr1.employee_id
                            JOIN personal_dict_dictionary pdd1 ON etr1.tag_id = pdd1.id
                            WHERE pdd1.biao_qian = '绩效工资'
                            AND NOT EXISTS (
                                SELECT 1 FROM employee_tag_relations etr2
                                JOIN personal_dict_dictionary pdd2 ON etr2.tag_id = pdd2.id
                                WHERE etr2.employee_id = t.id AND pdd2.biao_qian = '乡镇补贴'
                            )
                            ORDER BY t.name
                        """)
                        names = [r[0] for r in cur2.fetchall() if r[0]]
                        无乡镇补贴名单 = '、'.join(names) if names else '无'
                        无乡镇补贴人数 = len(names)
                        cur2.close()
                        conn2.close()
                    except Exception as e:
                        print(f"无乡镇补贴查询失败: {e}")
                    无乡镇补贴名单 = 无乡镇补贴名单 or '无'
                    
                    人事部门意见映射 = {
                        18: f"基础性绩效工资{绩效人数}人，{绩效金额:.0f}元；",
                        19: f"生活补贴{乡镇补贴人数 or 0}人，{乡镇补贴金额 or 0:.0f}元；",
                        20: f"合计{合计金额:.0f}元；",
                        21: f"岗位设置遗留{遗留人数 or 0}人，{遗留金额 or 0:.2f}元；",
                        22: f"总计{总计金额:.2f}元。",
                        23: f"无乡镇补贴{无乡镇补贴人数}人，{无乡镇补贴名单}。"
                    }
                    
                    for cell in filled_config.get('单元格数据', []):
                        row = cell.get('行号')
                        col = cell.get('列号')
                        if col == 5 and row in 人事部门意见映射:
                            cell['显示值'] = 人事部门意见映射[row]
                            cell['值'] = 人事部门意见映射[row]
            except Exception as e:
                print(f"人事部门意见填充失败: {e}")
        
        # ========== 字段值自动扩散 ==========
        # 已填充的字段值自动匹配模板中所有同名标签并填充，无需逐个配置
        try:
            _auto_spread_field_values(filled_config, field_values, config)
        except Exception as e:
            print(f"字段值自动扩散失败: {e}")
        
        return filled_config
    
    def _set_cell_value(self, config, target_row, target_col, value):
        """设置配置中指定行列的值"""
        for cell in config.get('单元格数据', []):
            if cell['行号'] == target_row and cell['列号'] == target_col:
                cell['值'] = value
                cell['显示值'] = str(value)
                break
        
        if '行数据' in config:
            for row_data in config['行数据']:
                if row_data['行号'] == target_row:
                    for cell in row_data['单元格']:
                        if cell['列号'] == target_col:
                            cell['文本'] = str(value)
                            break
    
    def _evaluate_formula(self, formula, field_values):
        """
        计算公式表达式
        
        支持：
        - 基本运算符: + - * / % ^ ()
        - 字段引用: {字段名称}
        - 聚合函数: SUM(a,b,...) AVG(a,b,...) MAX(a,b,...) MIN(a,b,...)
        - 条件函数: IF(条件, 真值, 假值)
        - 数学函数: ROUND(x,n) ABS(x) SQRT(x) POW(x,y) CEIL(x) FLOOR(x)
        
        参数:
            formula: 公式表达式字符串
            field_values: {字段名称: 值} 字典
        
        返回:
            计算结果
        """
        import math
        import re
        
        expr = formula
        
        expr = expr.replace('｛', '{').replace('｝', '}')
        
        for field_name, value in field_values.items():
            if value is None:
                value = 0
            placeholder = '{' + field_name + '}'
            expr = expr.replace(placeholder, str(value))
        
        expr = re.sub(r'\{[^}]+\}', '0', expr)
        
        def _sum(*args):
            return sum(args)
        
        def _avg(*args):
            nums = [a for a in args if a is not None]
            return sum(nums) / len(nums) if nums else 0
        
        def _if(cond, tv, fv):
            return tv if cond else fv
        
        safe_dict = {
            'SUM': _sum,
            'AVG': _avg,
            'MAX': max,
            'MIN': min,
            'IF': _if,
            'ROUND': round,
            'ABS': abs,
            'SQRT': math.sqrt,
            'POW': pow,
            'CEIL': math.ceil,
            'FLOOR': math.floor,
            '__builtins__': {}
        }
        
        try:
            result = eval(expr, safe_dict)
            if isinstance(result, float):
                if result == int(result):
                    result = int(result)
                elif abs(result) < 0.001:
                    result = 0
            return result
        except Exception as e:
            print(f"公式计算失败 [{formula}]: {e}")
            return None
    
    def _query_field_value(self, cursor, table, column, query_params):
        """从数据库查询单个字段值"""
        try:
            身份证号 = None
            if '职工ID' in query_params:
                emp_id = query_params['职工ID']
                if str(emp_id).isdigit():
                    cursor.execute(
                        "SELECT id_card FROM teacher_basic_info WHERE id = %s LIMIT 1",
                        (int(emp_id),)
                    )
                    id_row = cursor.fetchone()
                    if id_row:
                        身份证号 = id_row[0]
                else:
                    身份证号 = emp_id
            elif '身份证号' in query_params:
                身份证号 = query_params['身份证号']

            if 身份证号:
                sql = f"SELECT {column} FROM {table} WHERE id_card = %s LIMIT 1"
                cursor.execute(sql, (身份证号,))
                row = cursor.fetchone()
                result = row[0] if row else None
                # #region debug-point C:query-field-value
                import json as _dj, time as _dt, os as _dos
                _logp = _dos.path.join(_dos.path.dirname(_dos.path.dirname(__file__)), 'debug_fill_log.txt')
                with open(_logp, 'a', encoding='utf-8') as _df:
                    _df.write(f"[{_dt.time()}] C:query_field table={table} col={column} id_card={身份证号} sql={sql} has_result={row is not None} val={str(result)}\n")
                # #endregion
                return result
            return None
        except Exception as e:
            print(f"查询字段值失败: {e}")
            return None
    
    def _get_scope_id_cards(self, cursor, config):
        """
        根据统计范围和填报口径获取符合条件的id_card列表
        
        返回:
            (id_cards_list, None): 有过滤条件且有匹配结果
            (None, None): 没有设置过滤条件
            (None, []): 有过滤条件但没有匹配结果
        """
        统计范围 = config.get('统计范围', {})
        填报口径 = config.get('填报口径', {})
        
        unit_scope = 统计范围.get('单位范围', {})
        tag_ids = 填报口径.get('标签ID列表', []) if isinstance(填报口径, dict) else []
        
        has_unit_filter = unit_scope and any(v.get('unit_id') for v in unit_scope.values())
        
        if not has_unit_filter and not tag_ids:
            return None, None
        
        conditions = []
        params = []
        
        if has_unit_filter:
            all_descendant_ids = set()
            for level_name, level_data in unit_scope.items():
                uid = level_data.get('unit_id')
                if uid:
                    cursor.execute("""
                        WITH RECURSIVE unit_tree AS (
                            SELECT id FROM unit_hierarchy WHERE id = %s
                            UNION ALL
                            SELECT uh.id FROM unit_hierarchy uh
                            JOIN unit_tree ut ON uh.parent_id = ut.id
                        )
                        SELECT id FROM unit_tree
                    """, (int(uid),))
                    all_descendant_ids.update(r[0] for r in cursor.fetchall())
            
            if all_descendant_ids:
                conditions.append(
                    "t.id IN (SELECT DISTINCT t2.id FROM teacher_basic_info t2 "
                    "JOIN teacher_unit tu ON t2.id_card = tu.id_card "
                    "WHERE CAST(tu.unit_1 AS integer) = ANY(%s))"
                )
                params.append(list(all_descendant_ids))
            else:
                return None, []
        
        if tag_ids:
            conditions.append(
                "t.id IN (SELECT employee_id FROM employee_tag_relations WHERE tag_id = ANY(%s))"
            )
            params.append(list(tag_ids))
        
        if not conditions:
            return None, None
        
        where_clause = " AND ".join(conditions)
        sql = f"SELECT DISTINCT t.id_card FROM teacher_basic_info t WHERE {where_clause}"
        
        try:
            cursor.execute(sql, params)
            id_cards = [r[0] for r in cursor.fetchall()]
            if not id_cards:
                return None, []
            return id_cards, None
        except Exception as e:
            print(f"获取范围过滤id_card失败: {e}")
            return None, []
    
    def _table_has_column(self, cursor, table, column_name):
        try:
            cursor.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.columns
                    WHERE table_name = %s AND column_name = %s
                )
            """, (table, column_name))
            return cursor.fetchone()[0]
        except Exception:
            return False
    
    def _get_first_business_column(self, cursor, table):
        """获取表中第一个非系统业务字段名（用于字典表筛选键）"""
        try:
            cursor.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = %s
                  AND column_name NOT IN ('id', 'id_card', 'created_at', 'updated_at')
                ORDER BY ordinal_position
                LIMIT 1
            """, (table,))
            row = cursor.fetchone()
            return row[0] if row else None
        except Exception:
            return None
    
    def _get_dict_value_column(self, cursor, table, key_column):
        """获取字典表中键字段之后的下一个业务字段作为取值字段"""
        try:
            cursor.execute("""
                SELECT column_name FROM information_schema.columns
                WHERE table_name = %s
                  AND column_name NOT IN ('id', 'id_card', 'created_at', 'updated_at')
                  AND column_name != %s
                ORDER BY ordinal_position
                LIMIT 1
            """, (table, key_column))
            row = cursor.fetchone()
            return row[0] if row else None
        except Exception:
            return None
    
    def _query_aggregated_value(self, cursor, table, column, transform_func, config, dict_values=None, row_filter=None):
        """
        执行聚合查询（支持计数/求和/平均值/最大值/最小值/求积/取值）
        row_filter: 行筛选条件，如 {"offset": 1} 用于取值时偏移
        """
        agg_map = {
            '计数': 'COUNT({col})',
            '求和': 'COALESCE(SUM(NULLIF("{col}", \'\')::numeric), 0)',
            '平均值': 'AVG(NULLIF("{col}", \'\')::numeric)',
            '最大值': 'MAX(NULLIF("{col}", \'\')::numeric)',
            '最小值': 'MIN(NULLIF("{col}", \'\')::numeric)',
            '求积': 'EXP(SUM(LN(NULLIF(NULLIF("{col}", \'\')::numeric, 0))))',
            '取值': '{col}',
        }
        
        if transform_func not in agg_map:
            return None
        
        agg_expr = agg_map[transform_func]
        
        id_cards, status = self._get_scope_id_cards(cursor, config)
        
        _GLOBAL_SCOPE_TABLES = {'v_teacher_retirement_category', 'personal_statistics'}
        if table in _GLOBAL_SCOPE_TABLES:
            id_cards = None
        
        has_id_card = self._table_has_column(cursor, table, 'id_card')
        
        str_dict = None
        if dict_values and len(dict_values) > 0:
            str_dict = [str(v) for v in dict_values]
        
        if transform_func == '取值' and str_dict:
            filter_col = self._get_first_business_column(cursor, table) or column
            value_col = column
            dict_placeholders = ','.join(['%s'] * len(str_dict))
            sql = f'SELECT {agg_expr.format(col=value_col)} FROM {table} WHERE "{filter_col}" IN ({dict_placeholders})'
            params = tuple(str_dict)
        elif id_cards is not None and status is None and has_id_card:
            placeholders = ','.join(['%s'] * len(id_cards))
            sql = f"SELECT {agg_expr.format(col=column)} FROM {table} WHERE id_card IN ({placeholders})"
            params = list(id_cards)
            if str_dict:
                dict_placeholders = ','.join(['%s'] * len(str_dict))
                sql += f' AND "{column}" IN ({dict_placeholders})'
                params.extend(str_dict)
            params = tuple(params)
        elif id_cards is not None and status is None and not has_id_card:
            filter_col = self._get_first_business_column(cursor, table) or column
            value_col = self._get_dict_value_column(cursor, table, filter_col) if filter_col == column else column
            if value_col is None:
                value_col = column
            if str_dict:
                dict_placeholders = ','.join(['%s'] * len(str_dict))
                sql = f'SELECT {agg_expr.format(col=value_col)} FROM {table} WHERE "{filter_col}" IN ({dict_placeholders})'
                params = tuple(str_dict)
            else:
                sql = f"SELECT {agg_expr.format(col=value_col)} FROM {table}"
                params = None
        elif status is not None and not status:
            return 0 if transform_func == '计数' else None
        elif str_dict:
            filter_col = column
            if not has_id_card:
                filter_col = self._get_first_business_column(cursor, table) or column
            value_col = column
            if not has_id_card:
                value_col = self._get_dict_value_column(cursor, table, filter_col) or column
            dict_placeholders = ','.join(['%s'] * len(str_dict))
            sql = f'SELECT {agg_expr.format(col=value_col)} FROM {table} WHERE "{filter_col}" IN ({dict_placeholders})'
            params = tuple(str_dict)
        else:
            sql = f"SELECT {agg_expr.format(col=column)} FROM {table}"
            params = None
        
        if transform_func == '取值':
            if row_filter and 'offset' in row_filter:
                sql += f' ORDER BY name LIMIT 1 OFFSET {row_filter["offset"]}'
            else:
                sql += ' LIMIT 1'
        
        try:
            cursor.execute(sql, params or ())
            row = cursor.fetchone()
            result = row[0] if row else None
            if result is not None and isinstance(result, float):
                if result == int(result):
                    result = int(result)
                elif abs(result) < 0.001:
                    result = 0
            # #region debug-point D:query-aggregated
            import json as _dj, time as _dt, os as _dos
            _logp = _dos.path.join(_dos.path.dirname(_dos.path.dirname(__file__)), 'debug_fill_log.txt')
            with open(_logp, 'a', encoding='utf-8') as _df:
                _df.write(f"[{_dt.time()}] D:aggregated table={table} col={column} agg={transform_func} sql={sql} params={str(params)} has_result={row is not None} val={str(result)}\n")
            # #endregion
            return result
        except Exception as e:
            cursor.connection.rollback()
            print(f"聚合查询失败 [{transform_func}] {table}.{column}: {e}")
            return None
    
    def fill_and_export_from_original(self, original_file_path, mappings, query_params, output_path):
        """
        基于原始Excel文件填充数据并导出（保证100%格式一致）
        
        参数:
            original_file_path: 原始Excel文件路径
            mappings: 字段映射字典
            query_params: 查询参数
            output_path: 输出文件路径
        
        返回:
            输出文件路径
        """
        STAT_METHODS = ('计数', '求和', '平均值', '最大值', '最小值', '求积')
        
        wb = load_workbook(original_file_path)
        ws = wb.active
        
        db_mappings = {}
        formula_mappings = {}
        
        for field_name, mapping in mappings.items():
            data_source = mapping.get('数据源', '')
            transform_func = mapping.get('转换函数', '')
            
            if transform_func and not data_source:
                formula_mappings[field_name] = mapping
            elif data_source:
                db_mappings[field_name] = mapping
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            field_values = {}
            
            for field_name, mapping in db_mappings.items():
                data_source = mapping.get('数据源', '')
                if not data_source:
                    continue
                
                row_idx = mapping['行']
                col_idx = mapping['列']
                
                table, column = data_source.split('.')
                
                transform_func = mapping.get('转换函数', '')
                
                if transform_func in STAT_METHODS:
                    value = self._query_aggregated_value(cursor, table, column, transform_func, {})
                else:
                    value = self._query_field_value(cursor, table, column, query_params)
                
                if value is not None:
                    field_values[field_name] = value
                    ws.cell(row=row_idx, column=col_idx, value=str(value))
            
            for field_name, mapping in formula_mappings.items():
                formula = mapping.get('转换函数', '')
                if not formula:
                    continue
                
                value = self._evaluate_formula(formula, field_values)
                
                if value is not None:
                    ws.cell(row=mapping['行'], column=mapping['列'], value=str(value))
        finally:
            cursor.close()
            conn.close()
        
        wb.save(output_path)
        wb.close()
        
        return output_path
    
    def export_to_excel(self, config, output_path):
        """
        导出为Excel文件（基于完整元数据实现100%格式复制）
        
        参数:
            config: JSON配置字典（包含完整元数据）
            output_path: 输出文件路径
        """
        wb = Workbook()
        ws = wb.active
        
        if '页面设置' in config:
            page_setup = config['页面设置']
            ws.page_setup.orientation = 'portrait' if page_setup.get('方向') == '纵向' else 'landscape'
            if page_setup.get('缩放比例'):
                ws.page_setup.scale = page_setup['缩放比例']
            if page_setup.get('纸张类型') is not None:
                ws.page_setup.paperSize = page_setup['纸张类型']
        
        if '页边距' in config:
            margins = config['页边距']
            ws.page_margins.left = margins.get('左', 0.71)
            ws.page_margins.right = margins.get('右', 0.71)
            ws.page_margins.top = margins.get('上', 0.98)
            ws.page_margins.bottom = margins.get('下', 0.98)
            ws.page_margins.header = margins.get('页眉', 0.5)
            ws.page_margins.footer = margins.get('页脚', 0.5)
        
        if '列宽' in config:
            for col_idx, width_val in enumerate(config['列宽'], 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = float(width_val)
        
        if '行高' in config:
            for row_idx, height in config['行高'].items():
                ws.row_dimensions[int(row_idx)].height = height
        
        if '单元格数据' in config and '样式数据' in config:
            for cell_data in config['单元格数据']:
                row_idx = cell_data['行号']
                col_idx = cell_data['列号']
                ws_cell = ws.cell(row_idx, col_idx)
                ws_cell.value = cell_data['值']
                
                cell_key = cell_data['坐标']
                if cell_key in config['样式数据']:
                    self._apply_full_style(ws_cell, config['样式数据'][cell_key])
        
        elif '行数据' in config:
            for row_data in config['行数据']:
                row_idx = row_data['行号']
                if '高度' in row_data:
                    ws.row_dimensions[row_idx].height = row_data['高度']
                
                for cell_data in row_data['单元格']:
                    col_idx = cell_data['列号']
                    ws_cell = ws.cell(row_idx, col_idx)
                    ws_cell.value = cell_data.get('文本', '')
                    
                    if '样式' in cell_data or '对齐' in cell_data:
                        self._apply_cell_style(ws_cell, cell_data)
        
        if '合并单元格' in config:
            for merged in config['合并单元格']:
                if '起始' in merged and '结束' in merged:
                    ws.merge_cells(f"{merged['起始']}:{merged['结束']}")
                elif '起始行' in merged and '起始列' in merged and '结束行' in merged and '结束列' in merged:
                    ws.merge_cells(
                        start_row=merged['起始行'],
                        start_column=merged['起始列'],
                        end_row=merged['结束行'],
                        end_column=merged['结束列']
                    )
        
        wb.save(output_path)
        wb.close()
        
        return output_path
    
    def _apply_cell_style(self, ws_cell, cell_data):
        """应用单元格样式到Excel单元格（简化版）"""
        style = cell_data.get('样式', '')
        
        if 'font-size:' in style:
            import re
            match = re.search(r'font-size:(\d+)pt', style)
            if match:
                ws_cell.font = Font(size=int(match.group(1)))
        
        if 'font-weight:bold' in style:
            ws_cell.font = Font(bold=True)
        
        align = cell_data.get('对齐', 'center')
        if align == 'left':
            ws_cell.alignment = Alignment(horizontal='left', vertical='center')
        elif align == 'right':
            ws_cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            ws_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def _apply_full_style(self, ws_cell, style_data):
        """应用完整样式到Excel单元格（实现100%格式复制）"""
        font_props = {}
        if '字体' in style_data:
            font_info = style_data['字体']
            if font_info.get('大小'):
                font_props['size'] = font_info['大小']
            if font_info.get('粗体'):
                font_props['bold'] = font_info['粗体']
            if font_info.get('斜体'):
                font_props['italic'] = font_info['斜体']
            if font_info.get('名称'):
                font_props['name'] = font_info['名称']
            if font_info.get('颜色'):
                from openpyxl.styles.colors import RGB
                font_props['color'] = RGB(font_info['颜色'])
            if font_info.get('下划线'):
                font_props['underline'] = font_info['下划线']
            if font_info.get('删除线'):
                font_props['strike'] = font_info['删除线']
        
        if font_props:
            ws_cell.font = Font(**font_props)
        
        if '填充' in style_data:
            fill_info = style_data['填充']
            if fill_info.get('前景色'):
                from openpyxl.styles.colors import RGB
                fill = PatternFill(
                    patternType=fill_info.get('图案类型', 'solid'),
                    fgColor=RGB(fill_info['前景色'])
                )
                ws_cell.fill = fill
        
        if '边框' in style_data:
            border_info = style_data['边框']
            sides = {}
            for side_name in ['left', 'right', 'top', 'bottom']:
                if side_name in border_info:
                    side_data = border_info[side_name]
                    side_color = None
                    if side_data.get('颜色'):
                        from openpyxl.styles.colors import RGB
                        side_color = RGB(side_data['颜色'])
                    sides[side_name] = Side(
                        style=side_data['样式'],
                        color=side_color
                    )
            if sides:
                ws_cell.border = Border(**sides)
        
        if '对齐' in style_data:
            align_info = style_data['对齐']
            ws_cell.alignment = Alignment(
                horizontal=align_info.get('水平', 'center'),
                vertical=align_info.get('垂直', 'center'),
                wrapText=align_info.get('自动换行', False),
                textRotation=align_info.get('旋转角度', 0)
            )
        
        if '数字格式' in style_data:
            ws_cell.number_format = style_data['数字格式']
    
    def save_template_config(self, config):
        """
        保存模板配置到数据库
        
        参数:
            config: JSON配置字典
        """
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("""
                INSERT INTO template_configs (模板ID, 模板名称, 模板类型, 配置JSON, 原始文件路径, 创建时间, 更新时间)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (模板ID) DO UPDATE SET
                    模板名称 = EXCLUDED.模板名称,
                    模板类型 = EXCLUDED.模板类型,
                    配置JSON = EXCLUDED.配置JSON,
                    原始文件路径 = EXCLUDED.原始文件路径,
                    更新时间 = EXCLUDED.更新时间
            """, (
                config['模板ID'],
                config['模板名称'],
                config['模板类型'],
                json.dumps(config, ensure_ascii=False),
                config.get('原始文件', ''),
                config['导入时间'],
                config['导入时间']
            ))
            
            conn.commit()
        finally:
            cursor.close()
            conn.close()
    
    def _resolve_template_id(self, cursor, template_id):
        """
        将模板标识（可能是模板id或模板名称）解析为实际的模板id
        
        参数:
            cursor: 数据库游标
            template_id: 模板标识（可能是 tpl_xxx 或模板名称）
        
        返回:
            实际的模板id，如果未找到返回None
        """
        # 先直接按模板id查
        cursor.execute("SELECT 模板id FROM template_configs WHERE 模板id = %s", (template_id,))
        row = cursor.fetchone()
        if row:
            return row[0]
        # 再按模板名称查
        cursor.execute("SELECT 模板id FROM template_configs WHERE 模板名称 = %s", (template_id,))
        row = cursor.fetchone()
        if row:
            return row[0]
        return None
    
    def load_template_config(self, template_id):
        """
        从数据库加载模板配置
        
        参数:
            template_id: 模板ID（支持 tpl_xxx 或模板名称）
        
        返回:
            JSON配置字典
        """
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            actual_id = self._resolve_template_id(cursor, template_id)
            if not actual_id:
                return None
            
            cursor.execute("""
                SELECT 配置json FROM template_configs WHERE 模板id = %s
            """, (actual_id,))
            
            row = cursor.fetchone()
            if row:
                config_str = row[0]
                if isinstance(config_str, str):
                    return json.loads(config_str)
                return config_str
            return None
        finally:
            cursor.close()
            conn.close()
    
    def list_templates(self):
        """
        列出所有模板
        
        返回:
            模板列表
        """
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("""
                SELECT 模板ID, 模板名称, 模板类型, 原始文件路径, 创建时间, 更新时间, 配置JSON
                FROM template_configs
                ORDER BY 创建时间 DESC
            """)
            
            templates = []
            for row in cursor.fetchall():
                config_json = row[6] or {}
                if isinstance(config_json, str):
                    import json
                    config_json = json.loads(config_json)
                模板分类 = config_json.get('模板分类', '个人表')
                templates.append({
                    "模板ID": row[0],
                    "模板名称": row[1],
                    "模板类型": row[2],
                    "原始文件": row[3],
                    "创建时间": row[4],
                    "更新时间": row[5],
                    "模板分类": 模板分类
                })
            
            return templates
        finally:
            cursor.close()
            conn.close()
    
    def save_field_mapping(self, template_id, field_name, row, col, data_source, transform_func=None, default_value=None, dict_values=None):
        """
        保存字段映射
        
        参数:
            template_id: 模板ID
            field_name: 字段名称
            row: 行号
            col: 列号
            data_source: 数据源（表名.字段名）
            transform_func: 转换函数
            default_value: 默认值
            dict_values: 字典值选择（筛选值列表），以JSON存入默认值
        """
        import json
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            table, column = data_source.split('.') if '.' in data_source else (data_source, '')
            
            if dict_values and len(dict_values) > 0:
                default_value = json.dumps(dict_values, ensure_ascii=False)
            
            cursor.execute("""
                INSERT INTO template_field_mappings 
                (模板ID, 字段名称, 行号, 列号, 数据源表, 数据源字段, 转换函数, 默认值)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (模板ID, 字段名称) DO UPDATE SET
                    行号 = EXCLUDED.行号,
                    列号 = EXCLUDED.列号,
                    数据源表 = EXCLUDED.数据源表,
                    数据源字段 = EXCLUDED.数据源字段,
                    转换函数 = EXCLUDED.转换函数,
                    默认值 = EXCLUDED.默认值
            """, (
                template_id, field_name, row, col, table, column, transform_func, default_value
            ))
            
            conn.commit()
        finally:
            cursor.close()
            conn.close()
    
    def load_field_mappings(self, template_id):
        """
        加载字段映射
        
        参数:
            template_id: 模板ID（支持 tpl_xxx 或模板名称）
        
        返回:
            字段映射字典
            每个字段支持多个位置，格式：
            {
                "字段名": {
                    "数据源": "表名.字段名",
                    "转换函数": "...",
                    "位置列表": [{"行": 1, "列": 5}, {"行": 3, "列": 7}, ...],
                    "行": 1,   # 兼容旧代码：第一个位置的行
                    "列": 5,   # 兼容旧代码：第一个位置的列
                    ...
                }
            }
        """
        conn = get_db_connection()
        cursor = conn.cursor()
        
        import json as _json
        
        try:
            actual_id = self._resolve_template_id(cursor, template_id)
            if not actual_id:
                return {}
            
            cursor.execute("""
                SELECT 字段名称, 行号, 列号, 数据源表, 数据源字段, 转换函数, 默认值
                FROM template_field_mappings
                WHERE 模板id = %s
                ORDER BY 行号, 列号
            """, (actual_id,))
            
            mappings = {}
            for row in cursor.fetchall():
                field_name = row[0]
                raw_default = row[6]
                字典值选择 = []
                行筛选 = {}
                if raw_default and raw_default.strip():
                    if raw_default.strip().startswith('['):
                        try:
                            parsed = _json.loads(raw_default)
                            if isinstance(parsed, list):
                                字典值选择 = parsed
                        except Exception:
                            pass
                    elif raw_default.strip().startswith('{'):
                        try:
                            parsed = _json.loads(raw_default)
                            if isinstance(parsed, dict):
                                行筛选 = parsed
                        except Exception:
                            pass
                
                new_pos = {"行": row[1], "列": row[2]}
                data_source = f"{row[3]}.{row[4]}" if row[3] and row[4] else None
                transform_func = row[5]
                
                if field_name in mappings:
                    # 同一字段名，追加位置
                    mappings[field_name]["位置列表"].append(new_pos)
                else:
                    mappings[field_name] = {
                        "行": row[1],        # 兼容旧代码
                        "列": row[2],        # 兼容旧代码
                        "数据源": data_source,
                        "转换函数": transform_func,
                        "默认值": raw_default,
                        "字典值选择": 字典值选择,
                        "行筛选": 行筛选,
                        "位置列表": [new_pos]  # 新增：支持多位置
                    }
            
            return mappings
        finally:
            cursor.close()
            conn.close()

    def delete_field_mapping(self, template_id, field_name):
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                DELETE FROM template_field_mappings
                WHERE 模板ID = %s AND 字段名称 = %s
            """, (template_id, field_name))
            conn.commit()
            return cursor.rowcount > 0
        finally:
            cursor.close()
            conn.close()

    def delete_template(self, template_id):
        conn = get_db_connection()
        cursor = conn.cursor()

        try:
            cursor.execute("SELECT 原始文件路径 FROM template_configs WHERE 模板ID = %s", (template_id,))
            row = cursor.fetchone()
            if not row:
                return {"成功": False, "消息": "模板不存在", "已删除文件": False}

            original_filename = row[0] or ''

            cursor.execute("DELETE FROM template_field_mappings WHERE 模板ID = %s", (template_id,))
            cursor.execute("DELETE FROM template_configs WHERE 模板ID = %s", (template_id,))
            conn.commit()

            file_deleted = False
            if original_filename:
                file_path = os.path.join(
                    os.path.dirname(os.path.dirname(__file__)),
                    'uploads', 'templates',
                    original_filename
                )
                if os.path.exists(file_path):
                    os.remove(file_path)
                    file_deleted = True

            return {"成功": True, "消息": "模板删除成功", "已删除文件": file_deleted}

        except Exception as e:
            conn.rollback()
            return {"成功": False, "消息": str(e), "已删除文件": False}
        finally:
            cursor.close()
            conn.close()

    def get_performance_remarks(self, year, month):
        """
        获取绩效工资审批表的备注信息
        
        参数:
            year: 年份
            month: 月份
        
        返回:
            格式化的备注文本
        """
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            report_month = f"{year}-{month:02d}"
            
            cursor.execute("""
                SELECT remark_type, teacher_name, original_status, new_status, 
                       original_post, new_post, change_category, change_detail
                FROM performance_pay_remarks
                WHERE report_period = %s
                ORDER BY id
            """, (report_month,))
            
            remarks_records = cursor.fetchall()
            
            notes_groups = {}
            
            for row in remarks_records:
                remark_type = row[0]
                teacher_name = row[1]
                original_status = row[2] if row[2] else ''
                new_status = row[3] if row[3] else ''
                original_post = row[4] if row[4] else ''
                new_post = row[5] if row[5] else ''
                change_category = row[6] if row[6] else ''
                change_detail = row[7] if row[7] else ''
                
                key = None
                group_label = ''
                
                if change_category == 'status_change':
                    if new_status in ['调离']:
                        level = original_post if original_post else '教师'
                        key = f'调离_{level}'
                        group_label = f'{level}调离'
                    elif new_status in ['调出']:
                        level = original_post if original_post else '教师'
                        key = f'调出_{level}'
                        group_label = f'{level}调出'
                    elif new_status in ['离职']:
                        level = original_post if original_post else '教师'
                        key = f'离职_{level}'
                        group_label = f'{level}离职'
                    elif new_status in ['辞职']:
                        level = original_post if original_post else '教师'
                        key = f'辞职_{level}'
                        group_label = f'{level}辞职'
                    elif new_status in ['去世', '死亡']:
                        if original_status == '退休':
                            key = '去世_退休'
                            group_label = '退休教师死亡'
                        else:
                            level = original_post if original_post else '教师'
                            key = f'死亡_{level}'
                            if level == '教师':
                                group_label = '教师死亡'
                            else:
                                group_label = f'{level}死亡'
                    elif new_status in ['退休'] and original_status == '在职':
                        level = original_post if original_post else '教师'
                        key = f'退休_{level}'
                        if level == '教师':
                            group_label = '教师退休'
                        else:
                            group_label = f'{level}退休'
                    elif new_status in ['病休']:
                        level = original_post if original_post else '教师'
                        key = f'病休_{level}'
                        group_label = f'{level}病休'
                    elif new_status in ['延迟退休']:
                        level = original_post if original_post else '教师'
                        key = f'延迟退休_{level}'
                        group_label = f'{level}延迟退休'
                    elif new_status in ['在职'] and original_status == '退休':
                        level = original_post if original_post else '教师'
                        key = f'返聘_{level}'
                        group_label = f'{level}返聘'
                    elif new_status in ['挂职锻炼']:
                        level = original_post if original_post else '教师'
                        key = f'挂职_{level}'
                        group_label = f'{level}挂职锻炼'
                    elif new_status in ['待岗']:
                        level = original_post if original_post else '教师'
                        key = f'待岗_{level}'
                        group_label = f'{level}待岗'
                    elif new_status in ['停薪留职']:
                        level = original_post if original_post else '教师'
                        key = f'停薪留职_{level}'
                        group_label = f'{level}停薪留职'
                
                elif change_category == 'position_change':
                    if original_post == '一级教师' and new_post == '高级教师':
                        key = '晋升_一级_高级'
                        group_label = '一级教师晋升高级教师'
                    elif original_post == '二级教师' and new_post == '一级教师':
                        key = '晋升_二级_一级'
                        group_label = '二级教师晋升一级教师'
                    elif original_post == '三级教师' and new_post == '二级教师':
                        key = '晋升_三级_二级'
                        group_label = '三级教师晋升二级教师'
                    elif original_post == '三级教师' and new_post == '一级教师':
                        key = '晋升_三级_一级'
                        group_label = '三级教师晋升一级教师'
                    elif original_post == '二级教师' and new_post == '高级教师':
                        key = '晋升_二级_高级'
                        group_label = '二级教师晋升高级教师'
                    elif '高级工' in original_post and '技师' in new_post:
                        key = f'晋升_{original_post}_{new_post}'
                        group_label = f'{original_post}晋升{new_post}'
                    elif '技师' in original_post and '高级技师' in new_post:
                        key = f'晋升_{original_post}_{new_post}'
                        group_label = f'{original_post}晋升{new_post}'
                    elif '初级工' in original_post and '中级工' in new_post:
                        key = f'晋升_{original_post}_{new_post}'
                        group_label = f'{original_post}晋升{new_post}'
                    elif '中级工' in original_post and '高级工' in new_post:
                        key = f'晋升_{original_post}_{new_post}'
                        group_label = f'{original_post}晋升{new_post}'
                    elif '九级管理' in original_post and '八级管理' in new_post:
                        key = f'晋升_{original_post}_{new_post}'
                        group_label = f'{original_post}晋升{new_post}'
                    elif '八级管理' in original_post and '七级管理' in new_post:
                        key = f'晋升_{original_post}_{new_post}'
                        group_label = f'{original_post}晋升{new_post}'
                    else:
                        key = f'岗位变更_{original_post}_{new_post}'
                        group_label = f'{original_post}变更为{new_post}'
                
                elif change_category == 'new_add':
                    if change_detail == 'transfer_in' or '调入' in change_detail:
                        level = new_post if new_post else (original_post if original_post else '教师')
                        key = f'调入_{level}'
                        group_label = f'{level}调入'
                    elif change_detail == 'new_hire' or '新录聘' in change_detail:
                        level = new_post if new_post else (original_post if original_post else '教师')
                        key = f'新录聘_{level}'
                        group_label = f'{level}新录聘'
                    elif change_detail == 'management':
                        level = new_post if new_post else (original_post if original_post else '九级管理')
                        key = f'管理新增_{level}'
                        group_label = f'{level}新增'
                    elif change_detail == 'graduate' or '应届' in change_detail or '毕业生' in change_detail:
                        level = new_post if new_post else (original_post if original_post else '教师')
                        key = f'应届_{level}'
                        group_label = f'{level}应届毕业生'
                    elif change_detail == 'talent' or '引进' in change_detail:
                        level = new_post if new_post else (original_post if original_post else '教师')
                        key = f'引进_{level}'
                        group_label = f'{level}人才引进'
                    elif change_detail == 'intern' or '见习' in change_detail:
                        level = new_post if new_post else (original_post if original_post else '教师')
                        key = f'见习_{level}'
                        group_label = f'{level}见习期'
                    else:
                        level = new_post if new_post else '教师'
                        key = f'新增_{level}'
                        group_label = f'{level}调入'
                
                if key and teacher_name:
                    if key not in notes_groups:
                        notes_groups[key] = {'label': group_label, 'names': []}
                    notes_groups[key]['names'].append(teacher_name)
            
            notes_parts = []
            seq = 1
            for key, group in notes_groups.items():
                count = len(group['names'])
                names_str = '、'.join(group['names'])
                if count > 0:
                    notes_parts.append(f"{seq}.{group['label']}{count}人：{names_str}")
                    seq += 1
            
            result = '\n'.join(notes_parts) if notes_parts else ''
            return result
        except Exception as e:
            print(f"获取备注信息失败: {e}")
            return ''
        finally:
            cursor.close()
            conn.close()


template_engine = UniversalTemplateEngine()
