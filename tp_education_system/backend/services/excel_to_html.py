"""
Excel转HTML服务 - 保留完整格式
"""
from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Font, Alignment
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List, Optional
import html


def get_color_value(color) -> Optional[str]:
    """
    安全地获取颜色值，处理不同类型的颜色对象
    """
    if color is None:
        return None
    
    try:
        # 如果是RGB类型
        if hasattr(color, 'rgb') and color.rgb:
            rgb_value = color.rgb
            # 可能是字符串或RGB对象
            if isinstance(rgb_value, str):
                return rgb_value[:6] if len(rgb_value) >= 6 else rgb_value
            else:
                # 可能是RGB对象，尝试转换为字符串
                return str(rgb_value)[:6]
        
        # 如果是主题颜色
        if hasattr(color, 'theme') and color.theme is not None:
            # 主题颜色，返回默认黑色
            return '000000'
        
        # 如果是索引颜色
        if hasattr(color, 'indexed') and color.indexed is not None:
            # 索引颜色，返回默认黑色
            return '000000'
            
    except Exception:
        pass
    
    return None


class ExcelToHtmlConverter:
    """Excel转HTML转换器"""

    @staticmethod
    def convert(excel_path: str, marked_cells: List[Dict] = None, selected_cell: Dict = None) -> str:
        """
        将Excel转换为HTML，保留所有格式
        
        Args:
            excel_path: Excel文件路径
            marked_cells: 已标记的单元格列表 [{row, col, label}]
            selected_cell: 当前选中的单元格 {row, col}
        """
        if marked_cells is None:
            marked_cells = []
        
        # 加载工作簿
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # 获取所有合并单元格
        merged_cells = {}
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            merged_cells[(min_row, min_col)] = {
                'min_row': min_row,
                'min_col': min_col,
                'max_row': max_row,
                'max_col': max_col,
                'rowspan': max_row - min_row + 1,
                'colspan': max_col - min_col + 1
            }
            # 标记被合并的单元格
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if (r, c) != (min_row, min_col):
                        merged_cells[(r, c)] = None  # None表示被合并，不显示
        
        # 计算表格总宽度
        total_width = 0
        col_widths = {}
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            col_dim = ws.column_dimensions.get(col_letter)
            if col_dim and col_dim.width:
                # Excel列宽转换为像素 (1个单位约等于8像素)
                width_px = int(col_dim.width * 8)
            else:
                width_px = 80  # 默认宽度
            col_widths[col_idx] = width_px
            total_width += width_px
        
        # 开始生成HTML
        html_parts = []
        
        # HTML头部
        html_parts.append(f'''<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
* {{
    margin: 0;
    padding: 0;
    box-sizing: border-box;
}}
body {{
    font-family: "Microsoft YaHei", "SimSun", Arial, sans-serif;
    font-size: 12px;
    background: #f0f0f0;
    padding: 20px;
}}
.excel-container {{
    background: white;
    display: inline-block;
    box-shadow: 0 2px 8px rgba(0,0,0,0.1);
}}
table {{
    border-collapse: collapse;
    table-layout: fixed;
    width: {total_width}px;
}}
td {{
    border: 1px solid #000;
    padding: 2px 4px;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
    cursor: pointer;
    position: relative;
}}
td:hover {{
    outline: 2px solid #409eff;
    outline-offset: -2px;
    z-index: 10;
}}
td.marked {{
    outline: 3px solid #1890ff !important;
    outline-offset: -3px;
    background-color: #e6f7ff !important;
    z-index: 20;
}}
td.marked::after {{
    content: attr(data-label);
    position: absolute;
    top: -20px;
    left: 0;
    background: #1890ff;
    color: white;
    padding: 2px 6px;
    font-size: 10px;
    border-radius: 2px;
    white-space: nowrap;
    z-index: 100;
}}
td.selected {{
    outline: 3px solid #f56c6c !important;
    outline-offset: -3px;
    background-color: #fef0f0 !important;
    z-index: 30;
}}
</style>
</head>
<body>
<div class="excel-container">
<table>
''')
        
        # 生成表格内容
        for row_idx in range(1, ws.max_row + 1):
            html_parts.append('<tr>')
            
            # 获取行高
            row_dim = ws.row_dimensions.get(row_idx)
            if row_dim and row_dim.height:
                height_px = int(row_dim.height * 1.33)  # Excel行高转像素
            else:
                height_px = 20
            
            for col_idx in range(1, ws.max_column + 1):
                # 检查是否是被合并的单元格
                if merged_cells.get((row_idx, col_idx)) is None:
                    continue  # 跳过被合并的单元格
                
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # 单元格样式
                styles = []
                
                # 宽度
                if col_idx in col_widths:
                    styles.append(f'width: {col_widths[col_idx]}px')
                
                # 高度
                styles.append(f'height: {height_px}px')
                
                # 获取单元格样式
                if cell.has_style:
                    # 字体
                    if cell.font:
                        if cell.font.bold:
                            styles.append('font-weight: bold')
                        if cell.font.size:
                            styles.append(f'font-size: {cell.font.size}pt')
                        if cell.font.color:
                            font_color = get_color_value(cell.font.color)
                            if font_color:
                                styles.append(f'color: #{font_color}')
                    
                    # 背景色
                    if cell.fill and cell.fill.fill_type == 'solid':
                        bg_color = get_color_value(cell.fill.fgColor)
                        if bg_color:
                            styles.append(f'background-color: #{bg_color}')
                    
                    # 对齐
                    if cell.alignment:
                        if cell.alignment.horizontal:
                            h_align = cell.alignment.horizontal
                            if h_align == 'center':
                                styles.append('text-align: center')
                            elif h_align == 'right':
                                styles.append('text-align: right')
                            elif h_align == 'left':
                                styles.append('text-align: left')
                        
                        if cell.alignment.vertical:
                            v_align = cell.alignment.vertical
                            if v_align == 'center':
                                styles.append('vertical-align: middle')
                            elif v_align == 'top':
                                styles.append('vertical-align: top')
                            elif v_align == 'bottom':
                                styles.append('vertical-align: bottom')
                    
                    # 边框
                    if cell.border:
                        border_styles = []
                        if cell.border.left and cell.border.left.style:
                            left_color = get_color_value(cell.border.left.color) or '000000'
                            border_styles.append(f"border-left: 1px solid #{left_color}")
                        if cell.border.right and cell.border.right.style:
                            right_color = get_color_value(cell.border.right.color) or '000000'
                            border_styles.append(f"border-right: 1px solid #{right_color}")
                        if cell.border.top and cell.border.top.style:
                            top_color = get_color_value(cell.border.top.color) or '000000'
                            border_styles.append(f"border-top: 1px solid #{top_color}")
                        if cell.border.bottom and cell.border.bottom.style:
                            bottom_color = get_color_value(cell.border.bottom.color) or '000000'
                            border_styles.append(f"border-bottom: 1px solid #{bottom_color}")
                        styles.extend(border_styles)
                
                # 检查是否是合并单元格的主单元格
                merge_info = merged_cells.get((row_idx, col_idx))
                
                # 构建class
                classes = []
                
                # 检查是否被标记
                marked_info = None
                for marked in marked_cells:
                    if marked.get('row') == row_idx - 1 and marked.get('col') == col_idx - 1:
                        marked_info = marked
                        classes.append('marked')
                        break
                
                # 检查是否被选中
                if selected_cell and selected_cell.get('row') == row_idx - 1 and selected_cell.get('col') == col_idx - 1:
                    classes.append('selected')
                
                class_attr = f' class="{" ".join(classes)}"' if classes else ''
                
                # 构建属性
                attrs = [f'style="{"; ".join(styles)}"']
                
                if marked_info and marked_info.get('label'):
                    attrs.append(f'data-label="{html.escape(marked_info["label"])}"')
                
                # 合并单元格属性
                if merge_info:
                    if merge_info['rowspan'] > 1:
                        attrs.append(f'rowspan="{merge_info["rowspan"]}"')
                    if merge_info['colspan'] > 1:
                        attrs.append(f'colspan="{merge_info["colspan"]}"')
                
                # 单元格内容
                value = cell.value if cell.value is not None else ''
                if isinstance(value, str):
                    value = html.escape(value)
                else:
                    value = str(value)
                
                # 点击事件
                attrs.append(f'onclick="parent.postMessage({{type: \"cellClick\", row: {row_idx - 1}, col: {col_idx - 1}}}, \"*\")"')
                
                html_parts.append(f'<td{class_attr} {" ".join(attrs)}>{value}</td>')
            
            html_parts.append('</tr>')
        
        # HTML尾部
        html_parts.append('''
</table>
</div>
</body>
</html>''')
        
        return ''.join(html_parts)


# 便捷函数
def excel_to_html(excel_path: str, **kwargs) -> str:
    """Excel转HTML便捷函数"""
    return ExcelToHtmlConverter.convert(excel_path, **kwargs)
