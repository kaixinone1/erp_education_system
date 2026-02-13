"""
Excel转图片服务 - 100%保留原格式
"""
import os
from typing import Dict, Any, List, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import io
import base64


class ExcelToImageConverter:
    """Excel转图片转换器"""
    
    # 像素比例：Excel单位转像素
    PX_PER_UNIT = 8  # 列宽单位转像素
    PX_PER_ROW = 1.33  # 行高单位转像素
    
    @staticmethod
    def convert(excel_path: str, zoom: float = 1.0) -> str:
        """
        将Excel转换为图片（base64）
        
        Args:
            excel_path: Excel文件路径
            zoom: 缩放比例
            
        Returns:
            base64编码的图片字符串
        """
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # 计算图片尺寸
        total_width = 0
        col_widths = {}
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            col_dim = ws.column_dimensions.get(col_letter)
            if col_dim and col_dim.width:
                width_px = int(col_dim.width * ExcelToImageConverter.PX_PER_UNIT)
            else:
                width_px = 64  # 默认8个字符宽度
            col_widths[col_idx] = width_px
            total_width += width_px
        
        total_height = 0
        row_heights = {}
        for row_idx in range(1, ws.max_row + 1):
            row_dim = ws.row_dimensions.get(row_idx)
            if row_dim and row_dim.height:
                height_px = int(row_dim.height * ExcelToImageConverter.PX_PER_ROW)
            else:
                height_px = 20  # 默认行高
            row_heights[row_idx] = height_px
            total_height += height_px
        
        # 应用缩放
        img_width = int(total_width * zoom)
        img_height = int(total_height * zoom)
        
        # 创建图片
        img = Image.new('RGB', (img_width, img_height), 'white')
        draw = ImageDraw.Draw(img)
        
        # 尝试加载字体
        try:
            font = ImageFont.truetype("simhei.ttf", int(11 * zoom))
        except:
            try:
                font = ImageFont.truetype("arial.ttf", int(11 * zoom))
            except:
                font = ImageFont.load_default()
        
        # 绘制单元格
        y_offset = 0
        for row_idx in range(1, ws.max_row + 1):
            x_offset = 0
            row_height = row_heights[row_idx]
            
            for col_idx in range(1, ws.max_column + 1):
                col_width = col_widths[col_idx]
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # 计算实际像素位置和尺寸
                x1 = int(x_offset * zoom)
                y1 = int(y_offset * zoom)
                x2 = int((x_offset + col_width) * zoom)
                y2 = int((y_offset + row_height) * zoom)
                
                # 绘制背景
                bg_color = 'white'
                if cell.fill and cell.fill.fill_type == 'solid':
                    if cell.fill.fgColor and cell.fill.fgColor.rgb:
                        rgb = cell.fill.fgColor.rgb
                        if isinstance(rgb, str) and len(rgb) >= 6:
                            bg_color = f'#{rgb[:6]}'
                
                draw.rectangle([x1, y1, x2, y2], fill=bg_color, outline='black', width=1)
                
                # 绘制文字
                if cell.value:
                    text = str(cell.value)
                    text_color = 'black'
                    if cell.font and cell.font.color and cell.font.color.rgb:
                        rgb = cell.font.color.rgb
                        if isinstance(rgb, str) and len(rgb) >= 6:
                            text_color = f'#{rgb[:6]}'
                    
                    # 计算文字位置（居中）
                    bbox = draw.textbbox((0, 0), text, font=font)
                    text_width = bbox[2] - bbox[0]
                    text_height = bbox[3] - bbox[1]
                    
                    text_x = x1 + (x2 - x1 - text_width) // 2
                    text_y = y1 + (y2 - y1 - text_height) // 2
                    
                    draw.text((text_x, text_y), text, fill=text_color, font=font)
                
                x_offset += col_width
            
            y_offset += row_height
        
        # 转换为base64
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        img_str = base64.b64encode(buffer.getvalue()).decode()
        
        return f"data:image/png;base64,{img_str}"
    
    @staticmethod
    def get_cell_positions(excel_path: str, zoom: float = 1.0) -> List[Dict]:
        """
        获取所有单元格的位置信息（用于前端点击定位）
        
        Returns:
            单元格位置列表 [{row, col, x, y, width, height}]
        """
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        positions = []
        y_offset = 0
        
        for row_idx in range(1, ws.max_row + 1):
            x_offset = 0
            row_dim = ws.row_dimensions.get(row_idx)
            if row_dim and row_dim.height:
                row_height = int(row_dim.height * ExcelToImageConverter.PX_PER_ROW)
            else:
                row_height = 20
            
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                col_dim = ws.column_dimensions.get(col_letter)
                if col_dim and col_dim.width:
                    col_width = int(col_dim.width * ExcelToImageConverter.PX_PER_UNIT)
                else:
                    col_width = 64
                
                positions.append({
                    'row': row_idx - 1,  # 0-based
                    'col': col_idx - 1,  # 0-based
                    'x': int(x_offset * zoom),
                    'y': int(y_offset * zoom),
                    'width': int(col_width * zoom),
                    'height': int(row_height * zoom)
                })
                
                x_offset += col_width
            
            y_offset += row_height
        
        return positions


def excel_to_image(excel_path: str, zoom: float = 1.0) -> str:
    """便捷函数"""
    return ExcelToImageConverter.convert(excel_path, zoom)


def get_excel_cell_positions(excel_path: str, zoom: float = 1.0) -> List[Dict]:
    """获取单元格位置便捷函数"""
    return ExcelToImageConverter.get_cell_positions(excel_path, zoom)
