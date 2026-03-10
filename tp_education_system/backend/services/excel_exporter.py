"""
Excel导出服务
将报表数据导出为Excel，保留原始模板的表格结构
"""
import os
import tempfile
from typing import Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
    
    def export_report(self, template_id: str, data: Dict, teacher_id: str,
                     page_settings: Dict = None) -> Optional[str]:
        """
        导出报表为Excel
        按照原始模板的表格结构布局
        
        Args:
            template_id: 模板ID
            data: 填充数据
            teacher_id: 教师ID
            page_settings: 页面设置
            
        Returns:
            Excel文件路径
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "退休呈报表"
            
            # 设置页面
            paper_size = page_settings.get('paper_size', 'A4') if page_settings else 'A4'
            is_landscape = page_settings.get('is_landscape', False) if page_settings else False
            
            if paper_size == 'A3':
                ws.page_setup.paperSize = ws.PAPERSIZE_A3
            else:
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
            
            if is_landscape:
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            else:
                ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            
            # 设置边距
            ws.page_margins.left = 0.3
            ws.page_margins.right = 0.3
            ws.page_margins.top = 0.3
            ws.page_margins.bottom = 0.3
            
            # 定义样式
            title_font = Font(name='宋体', size=16, bold=True)
            header_font = Font(name='宋体', size=10, bold=True)
            normal_font = Font(name='宋体', size=10)
            small_font = Font(name='宋体', size=9)
            
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置列宽 - 根据原始模板的复杂表格结构
            col_widths = [8, 6, 8, 6, 8, 6, 8, 6, 8, 6, 8, 6, 8, 6, 8, 6, 8, 6, 8, 6]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            current_row = 1
            
            # 标题行
            ws.merge_cells(f'A{current_row}:T{current_row}')
            title_cell = ws[f'A{current_row}']
            title_cell.value = "职工退休呈报表"
            title_cell.font = title_font
            title_cell.alignment = center_align
            ws.row_dimensions[current_row].height = 30
            current_row += 1
            
            # 按照原始模板的表格结构创建单元格
            # 第1行：呈报单位意见头部
            self._create_merged_cell(ws, current_row, 1, 2, 1, "呈\n报\n单\n位\n意\n见", header_font, center_align, thin_border)
            self._create_merged_cell(ws, current_row, 3, 10, 1, f"经研究，同意 {data.get('姓名', '')} 同志按以下第（    ）条办理", normal_font, left_align, thin_border)
            current_row += 1
            
            # 第2行：退休类型选择
            self._create_merged_cell(ws, current_row, 3, 10, 1, "（一）弹性提前退休", normal_font, left_align, thin_border)
            current_row += 1
            
            # 第3行
            self._create_merged_cell(ws, current_row, 3, 10, 1, "（二）法定退休年龄退休", normal_font, left_align, thin_align, thin_border)
            current_row += 1
            
            # 第4行：日期
            self._create_merged_cell(ws, current_row, 3, 10, 1, "", normal_font, left_align, thin_border)
            current_row += 1
            
            # 第5行：主管部门审核意见
            self._create_merged_cell(ws, current_row, 1, 2, 1, "主\n管\n部\n门\n审\n核\n意\n见", header_font, center_align, thin_border)
            self._create_merged_cell(ws, current_row, 3, 10, 1, "同意呈报。", normal_font, center_align, thin_border)
            current_row += 1
            
            # 第6行：日期
            self._create_merged_cell(ws, current_row, 3, 10, 1, "", normal_font, left_align, thin_border)
            current_row += 1
            
            # 第7行：退休一次性补贴审批
            self._create_merged_cell(ws, current_row, 1, 2, 1, "退\n休\n一\n次\n性\n补\n贴\n审\n批", header_font, center_align, thin_border)
            subsidy_text = f"根据鄂人社发【2017】8号文件规定，同意 {data.get('姓名', '')} 同志发放独生子女费        元，教育特殊贡献奖        元，"
            self._create_merged_cell(ws, current_row, 3, 10, 1, subsidy_text, normal_font, left_align, thin_border)
            current_row += 1
            
            # 第8行：执行时间
            self._create_merged_cell(ws, current_row, 3, 10, 1, "从      年      月执行。", normal_font, left_align, thin_border)
            current_row += 1
            
            # 第9行：基本信息表格标题
            ws.merge_cells(f'A{current_row}:T{current_row}')
            info_title = ws[f'A{current_row}']
            info_title.value = "退休人员基本信息及历次工资变动情况"
            info_title.font = header_font
            info_title.alignment = center_align
            ws.row_dimensions[current_row].height = 25
            current_row += 1
            
            # 第10-15行：基本信息
            # 姓名、性别、出生日期
            self._create_cell(ws, current_row, 1, "姓名", header_font, center_align, thin_border)
            self._create_cell(ws, current_row, 2, data.get('姓名', ''), normal_font, center_align, thin_border)
            self._create_cell(ws, current_row, 3, "性别", header_font, center_align, thin_border)
            self._create_cell(ws, current_row, 4, data.get('性别', ''), normal_font, center_align, thin_border)
            self._create_cell(ws, current_row, 5, "出生日期", header_font, center_align, thin_border)
            self._create_cell(ws, current_row, 6, data.get('出生日期', ''), normal_font, center_align, thin_border)
            self._create_cell(ws, current_row, 7, "民族", header_font, center_align, thin_border)
            self._create_cell(ws, current_row, 8, data.get('民族', ''), normal_font, center_align, thin_border)
            self._create_cell(ws, current_row, 9, "是否独生子女", header_font, center_align, thin_border)
            self._create_cell(ws, current_row, 10, data.get('是否独生子女', ''), normal_font, center_align, thin_border)
            current_row += 1
            
            # 身份证号码
            self._create_cell(ws, current_row, 1, "身份证号码", header_font, center_align, thin_border)
            ws.merge_cells(f'B{current_row}:J{current_row}')
            id_card_cell = ws[f'B{current_row}']
            id_card_cell.value = data.get('身份证号码', '')
            id_card_cell.font = normal_font
            id_card_cell.alignment = center_align
            id_card_cell.border = thin_border
            current_row += 1
            
            # 参加工作时间、工作年限
            self._create_cell(ws, current_row, 1, "参加工作时间", header_font, center_align, thin_border)
            self._create_cell(ws, current_row, 2, data.get('参加工作时间', ''), normal_font, center_align, thin_border)
            self._create_cell(ws, current_row, 3, "工作年限", header_font, center_align, thin_border)
            self._create_cell(ws, current_row, 4, str(data.get('工作年限', '')), normal_font, center_align, thin_border)
            self._create_cell(ws, current_row, 5, "退休时间", header_font, center_align, thin_border)
            self._create_cell(ws, current_row, 6, data.get('退休时间', ''), normal_font, center_align, thin_border)
            current_row += 1
            
            # 职务、技术职称
            self._create_cell(ws, current_row, 1, "职务", header_font, center_align, thin_border)
            self._create_cell(ws, current_row, 2, data.get('职务', ''), normal_font, center_align, thin_border)
            self._create_cell(ws, current_row, 3, "技术职称", header_font, center_align, thin_border)
            self._create_cell(ws, current_row, 4, data.get('技术职称', ''), normal_font, center_align, thin_border)
            current_row += 1
            
            # 岗位工资、薪级工资
            self._create_cell(ws, current_row, 1, "岗位工资", header_font, center_align, thin_border)
            self._create_cell(ws, current_row, 2, str(data.get('岗位工资', '')), normal_font, center_align, thin_border)
            self._create_cell(ws, current_row, 3, "薪级工资", header_font, center_align, thin_border)
            self._create_cell(ws, current_row, 4, str(data.get('薪级工资', '')), normal_font, center_align, thin_border)
            current_row += 1
            
            # 第16-25行：岗位信息表格
            # 表头
            headers = ["事业管理岗位", "对应原职务", "薪级", "事业专技岗位", "对应原职务", "薪级", 
                      "事业工勤岗位", "对应技术等级", "薪级"]
            for i, header in enumerate(headers, 1):
                self._create_cell(ws, current_row, i, header, header_font, center_align, thin_border)
            current_row += 1
            
            # 岗位1-3
            for i in range(1, 4):
                self._create_cell(ws, current_row, 1, data.get(f'事业管理岗位{i}', ''), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 2, data.get(f'对应原职务{i}', ''), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 3, str(data.get(f'薪级{i}', '')), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 4, data.get(f'事业专技岗位{i}', ''), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 5, data.get(f'对应原职务{i+1}', ''), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 6, str(data.get(f'薪级{i+1}', '')), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 7, data.get(f'事业工勤岗位{i}', ''), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 8, data.get(f'对应技术等级{i}', ''), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 9, str(data.get(f'薪级{i+2}', '')), normal_font, center_align, thin_border)
                current_row += 1
            
            # 第26-35行：退休时岗位信息
            ws.merge_cells(f'A{current_row}:I{current_row}')
            retire_title = ws[f'A{current_row}']
            retire_title.value = "退休时岗位信息"
            retire_title.font = header_font
            retire_title.alignment = center_align
            current_row += 1
            
            # 退休时岗位表头
            for i, header in enumerate(headers[:6], 1):
                self._create_cell(ws, current_row, i, header, header_font, center_align, thin_border)
            current_row += 1
            
            # 退休时岗位7-9
            for i in range(7, 10):
                self._create_cell(ws, current_row, 1, data.get(f'退休时事业管理岗位{i}', ''), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 2, data.get(f'对应原职务{i}', ''), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 3, str(data.get(f'薪级{i}', '')), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 4, data.get(f'退休时事业专技岗位{i}', ''), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 5, data.get(f'对应原职务{i+1}', ''), normal_font, center_align, thin_border)
                self._create_cell(ws, current_row, 6, str(data.get(f'薪级{i+1}', '')), normal_font, center_align, thin_border)
                current_row += 1
            
            # 第36-40行：工作简历
            ws.merge_cells(f'A{current_row}:I{current_row}')
            resume_title = ws[f'A{current_row}']
            resume_title.value = "工作简历"
            resume_title.font = header_font
            resume_title.alignment = center_align
            current_row += 1
            
            # 工作简历表头
            resume_headers = ["自何年何月", "至何年何月", "所在单位及职务", "证明人及住址"]
            for i, header in enumerate(resume_headers, 1):
                self._create_cell(ws, current_row, i, header, header_font, center_align, thin_border)
            current_row += 1
            
            # 工作简历内容
            self._create_cell(ws, current_row, 1, data.get('自何年何月', ''), normal_font, center_align, thin_border)
            self._create_cell(ws, current_row, 2, data.get('至何年何月', ''), normal_font, center_align, thin_border)
            self._create_cell(ws, current_row, 3, data.get('所在单位及职务', ''), normal_font, center_align, thin_border)
            self._create_cell(ws, current_row, 4, data.get('证明人及住址', ''), normal_font, center_align, thin_border)
            current_row += 1
            
            # 直系亲属供养情况
            ws.merge_cells(f'A{current_row}:I{current_row}')
            family_title = ws[f'A{current_row}']
            family_title.value = f"直系亲属供养情况：{data.get('直系亲属供养情况', '')}"
            family_title.font = normal_font
            family_title.alignment = left_align
            current_row += 1
            
            # 备注
            ws.merge_cells(f'A{current_row}:I{current_row}')
            remark_cell = ws[f'A{current_row}']
            remark_cell.value = f"备注：{data.get('备注', '')}"
            remark_cell.font = normal_font
            remark_cell.alignment = left_align
            
            # 保存文件
            output_filename = f"{template_id}_{teacher_id}_完整报表.xlsx"
            output_path = os.path.join(self.temp_dir, output_filename)
            wb.save(output_path)
            
            return output_path
            
        except Exception as e:
            print(f"导出Excel失败: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _create_cell(self, ws, row, col, value, font, alignment, border):
        """创建单元格"""
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.font = font
        cell.alignment = alignment
        cell.border = border
    
    def _create_merged_cell(self, ws, row, start_col, end_col, num_rows, value, font, alignment, border):
        """创建合并单元格"""
        if num_rows > 1:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row + num_rows - 1, end_column=end_col)
        else:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        
        cell = ws.cell(row=row, column=start_col)
        cell.value = value
        cell.font = font
        cell.alignment = alignment
        cell.border = border
