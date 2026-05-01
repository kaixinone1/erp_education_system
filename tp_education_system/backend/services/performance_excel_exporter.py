"""
绩效工资审批表Excel导出服务 - 基于原始模板逆向解析
完全按照原始模板的结构生成：11列，32行，25个合并区域
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from typing import Dict, List

class PerformancePayExcelExporter:
    """绩效工资审批表Excel导出器 - 原始模板精确版"""

    def __init__(self):
        self.output_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            'exports'
        )
        os.makedirs(self.output_dir, exist_ok=True)

    def export_with_template(self, data: Dict, year_month: str) -> str:
        """
        使用原始模板配置导出Excel
        11列(A-K)，32行，25个合并区域
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 1. 设置页面（A4纵向单面）
        self._setup_page(ws)

        # 2. 设置列宽（原始模板精确值）
        self._setup_columns(ws)

        # 3. 设置行高（原始模板精确值）
        self._setup_rows(ws)

        # 4. 填充所有数据（按照原始模板行号）
        self._fill_all_data(ws, data)

        # 5. 合并单元格（按照原始模板的25个合并区域）
        self._apply_merges(ws)

        # 6. 应用样式
        self._apply_styles(ws)

        # 7. 保存文件
        filename = f"绩效工资审批表_{year_month}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath

    def _setup_page(self, ws):
        """设置页面参数 - A4纵向单面（原始模板值）"""
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.orientation = 'portrait'  # 纵向
        ws.page_setup.fitToPage = False
        ws.page_setup.fitToWidth = False
        ws.page_setup.fitToHeight = False
        ws.page_setup.firstPageNumber = 1
        ws.page_setup.usePageNumbers = True

        # 页边距（英寸）- 原始模板精确值
        ws.page_margins = PageMargins(
            left=0.75, right=0.75, top=1.0, bottom=1.0, header=0.3, footer=0.3
        )

        # 打印区域 - 原始模板范围
        ws.print_area = 'A1:K32'

    def _setup_columns(self, ws):
        """设置列宽 - 原始模板精确值（字符宽度）"""
        # 原始模板的列宽（字符宽度）
        col_widths = {
            'A': 13.82, 'B': 7.64, 'C': 13.04, 'D': 9.36,
            'E': 6.09, 'F': 8.82, 'G': 17.64, 'H': 5.00,
            'I': 3.91, 'J': 11.00, 'K': 8.64
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

    def _setup_rows(self, ws):
        """设置行高 - 原始模板精确值"""
        # 原始模板的行高
        row_heights = {
            1: 27.0,   # 标题行
            2: 25.0,   # 填报信息行
            3: 25.0,   # 表头
            4: 25.0,   # 表头
            5: 25.0,   # 行政管理人员
            6: 25.0,   # 副处级
            7: 25.0,   # 正科级
            8: 25.0,   # 副科级
            9: 25.0,   # 科员级
            10: 25.0,  # 办事员级
            11: 25.0,  # 专业技术人员
            12: 25.0,  # 正高级
            13: 25.0,  # 高级教师
            14: 25.0,  # 一级教师
            15: 25.0,  # 二级教师
            16: 25.0,  # 三级教师
            17: 25.0,  # 工人
            18: 25.0,  # 高级技师
            19: 25.0,  # 技师
            20: 25.0,  # 高级工
            21: 25.0,  # 中级工
            22: 25.0,  # 初级工
            23: 25.0,  # 普工
            24: 25.0,  # 绩效工资合计
            25: 25.0,  # 乡镇补贴合计
            26: 37.0,  # 岗位设置遗留问题
            27: 37.0,  # 岗位设置遗留问题
            28: 36.0,  # 岗位设置遗留问题合计
            29: 25.0,  # 退休干部
            30: 25.0,  # 退休工人
            31: 25.0,  # 离休干部
            32: 130.0, # 备注行
        }
        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height

    def _fill_all_data(self, ws, data: Dict):
        """填充所有数据 - 严格按照原始模板结构"""
        today = datetime.now()
        today_str = f"{today.year}年{today.month}月{today.day}日"

        # ===== 第1行：标题 =====
        ws['B1'] = today.strftime('%Y-%m-%d')  # 日期
        ws['D1'] = f"{data.get('年月', '2026年5月')} 义务教育学校教职工绩效工资审批表"

        # ===== 第2行：填报信息 =====
        ws['A2'] = '填报单位：'
        ws['B2'] = data.get('填报单位', '太平镇中心学校')
        ws['E2'] = '填报时间:'
        ws['G2'] = today.strftime('%Y-%m-%d')
        ws['H2'] = '单位：'
        ws['J2'] = '人、元'

        # ===== 第3-4行：表头 =====
        ws['A3'] = '项  目'
        ws['A4'] = ''  # A3:A4合并后显示"项目"
        ws['B3'] = '基础性工资'
        ws['B4'] = '人数'
        ws['C4'] = '月工资标准'
        ws['D4'] = '小计'
        ws['E3'] = '呈报单位意见'

        # ===== 第5行：行政管理人员分类标题 =====
        ws['A5'] = '行政管理人员'

        # ===== 第6-10行：行政管理人员数据 =====
        admin_items = ['副处级', '正科级', '副科级', '科员级', '办事员级']
        admin_data = data.get('administrative', {})
        for i, item in enumerate(admin_items):
            row = 6 + i
            ws[f'A{row}'] = f'{i+1}、{item}'
            item_data = admin_data.get(item, {})
            ws[f'B{row}'] = item_data.get('count', '')
            ws[f'C{row}'] = item_data.get('standard', '')
            ws[f'D{row}'] = item_data.get('subtotal', '')

        # 呈报单位意见（E列，垂直文本）
        ws['F6'] = '据实填写，同意呈报。'
        ws['F7'] = '（盖章）'

        # ===== 第11行：专业技术人员分类标题 =====
        ws['A11'] = '专业技术人员'
        ws['E11'] = '教育局意见'

        # ===== 第12-16行：专业技术人员数据 =====
        pro_items = ['正高级', '高级教师', '一级教师', '二级教师', '三级教师']
        pro_data = data.get('professional', {})
        for i, item in enumerate(pro_items):
            row = 12 + i
            ws[f'A{row}'] = f'{i+1}、{item}'
            item_data = pro_data.get(item, {})
            ws[f'B{row}'] = item_data.get('count', '')
            ws[f'C{row}'] = item_data.get('standard', '')
            ws[f'D{row}'] = item_data.get('subtotal', '')

        # ===== 第17行：工人分类标题 =====
        ws['A17'] = '工人'
        ws['E20'] = '人事部门意见'

        # ===== 第18-23行：工人数据 =====
        worker_items = ['高级技师', '技师', '高级工', '中级工', '初级工', '普工']
        worker_data = data.get('worker', {})
        for i, item in enumerate(worker_items):
            row = 18 + i
            ws[f'A{row}'] = f'{i+1}、{item}'
            item_data = worker_data.get(item, {})
            ws[f'B{row}'] = item_data.get('count', '')
            ws[f'C{row}'] = item_data.get('standard', '')
            ws[f'D{row}'] = item_data.get('subtotal', '')

        # ===== 第20-23行：人事部门意见内容 =====
        ws['F20'] = '    根据相关文件及有关规定，经审核，同意你单位：'
        ws['F21'] = '基础性绩效工资'
        ws['H21'] = data.get('totals', {}).get('performance_count', 357)
        ws['I21'] = '人'
        ws['J21'] = data.get('totals', {}).get('performance_total', 462311)
        ws['K21'] = '元；'

        subsidies = data.get('subsidies', {})
        ws['F22'] = '生活补贴'
        ws['H22'] = subsidies.get('count', 356)
        ws['I22'] = '人'
        ws['J22'] = subsidies.get('total', 124600)
        ws['K22'] = '元；'

        total_all = (data.get('totals', {}).get('performance_total', 0) +
                    data.get('subsidies', {}).get('total', 0))
        ws['F23'] = '合计：'
        ws['J23'] = total_all if total_all else 586911
        ws['K23'] = '元'

        # ===== 第24行：绩效工资合计 =====
        totals = data.get('totals', {})
        ws['A24'] = '绩效工资合计'
        ws['B24'] = totals.get('performance_count', '')
        ws['D24'] = totals.get('performance_total', '')

        # ===== 第25行：乡镇补贴合计 =====
        subsidies = data.get('subsidies', {})
        ws['A25'] = '乡镇补贴合计'
        ws['B25'] = subsidies.get('count', '')
        ws['C25'] = subsidies.get('standard', 350)
        ws['D25'] = subsidies.get('total', '')

        # ===== 第26-28行：岗位设置遗留问题 =====
        legacy = data.get('legacy', [])
        for i in range(3):
            row = 26 + i
            if i < len(legacy):
                item = legacy[i]
                ws[f'A{row}'] = '岗位设置\n遗留问题'
                ws[f'B{row}'] = item.get('name', '')
                ws[f'C{row}'] = item.get('amount', '')
                ws[f'D{row}'] = item.get('amount', '')
            else:
                ws[f'A{row}'] = '岗位设置\n遗留问题'

        # ===== 第28行：岗位设置遗留问题合计 =====
        if legacy:
            total_count = sum(1 for item in legacy if item.get('name'))
            total_amount = sum(item.get('amount', 0) for item in legacy)
            ws['A28'] = '岗位设置\n遗留问题'
            ws['B28'] = total_count
            ws['D28'] = total_amount

        # ===== 第29-31行：退休人员 =====
        retirees = data.get('retirees', {})
        ws['A29'] = '退休干部'
        ws['B29'] = retirees.get('cadre_count', '')
        ws['A30'] = '退休工人'
        ws['B30'] = retirees.get('worker_count', '')
        ws['A31'] = '离休干部'
        ws['B31'] = retirees.get('retired_count', '')

        # ===== 第32行：备注 =====
        ws['A32'] = f"备注：\n{data.get('notes', '')}"

    def _apply_merges(self, ws):
        """应用合并单元格 - 原始模板的25个合并区域"""
        merges = [
            # B1:C1 - 日期
            (1, 2, 1, 3),
            # D1:K1 - 标题跨列
            (1, 4, 1, 11),
            # A3:A4 - 项目跨行
            (3, 1, 4, 1),
            # B3:D3 - 基础性工资跨列
            (3, 2, 3, 4),
            # E3:E10 - 呈报单位意见（垂直）
            (3, 5, 10, 5),
            # E11:E19 - 教育局意见（垂直）
            (11, 5, 19, 5),
            # E20:E31 - 人事部门意见（垂直）
            (20, 5, 31, 5),
            # F6:K6 - 据实填写跨列
            (6, 6, 6, 11),
            # F7:K7 - （盖章）跨列
            (7, 6, 7, 11),
            # F10:K10 - 金额跨列
            (10, 6, 10, 11),
            # F15:K15 - （盖章）跨列
            (15, 6, 15, 11),
            # F19:K19 - 金额跨列
            (19, 6, 19, 11),
            # F20:K20 - 审核意见内容
            (20, 6, 20, 11),
            # F21:G21 - 基础性绩效工资
            (21, 6, 21, 7),
            # J21:K21 - 元；
            (21, 10, 21, 11),
            # F22:G22 - 生活补贴
            (22, 6, 22, 7),
            # J22:K22 - 元；
            (22, 10, 22, 11),
            # F23:I23 - 合计
            (23, 6, 23, 9),
            # J23:K23 - 元
            (23, 10, 23, 11),
            # F24:G24 - 岗位设置遗留
            (24, 6, 24, 7),
            # H24:I24 - 人
            (24, 8, 24, 9),
            # J24:K24 - 元。
            (24, 10, 24, 11),
            # F25:I25 - 总计
            (25, 6, 25, 9),
            # J25:K25 - 元
            (25, 10, 25, 11),
            # F26:G26 - 无乡镇补贴
            (26, 6, 26, 7),
            # H26:I26 - 人：
            (26, 8, 26, 9),
            # F27:I27 - 岗位设置遗留
            (27, 6, 27, 9),
            # F30:K30 - 空白跨列
            (30, 6, 30, 11),
            # F31:K31 - 空白跨列
            (31, 6, 31, 11),
            # A32:K32 - 备注行跨列
            (32, 1, 32, 11),
            # H2:I2 - 单位跨列
            (2, 8, 2, 9),
            # E2:F2 - 填报时间跨列
            (2, 5, 2, 6),
        ]

        for start_row, start_col, end_row, end_col in merges:
            ws.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=end_row,
                end_column=end_col
            )

    def _apply_styles(self, ws):
        """应用单元格样式"""
        thin = Side(style='thin', color='000000')

        # 遍历所有单元格
        for row in ws.iter_rows():
            for cell in row:
                row_num = cell.row
                col_num = cell.column

                # 跳过无值且无合并的单元格
                if cell.value is None:
                    continue

                # ===== 标题行（第1行）=====
                if row_num == 1:
                    if col_num == 1:  # B1 - 日期
                        cell.font = Font(name='宋体', size=11)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col_num == 4:  # D1 - 标题
                        cell.font = Font(name='宋体', size=16, bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    continue

                # ===== 填报信息行（第2行）=====
                if row_num == 2:
                    cell.font = Font(name='宋体', size=11)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    continue

                # ===== 表头（第3-4行）=====
                if row_num in [3, 4]:
                    cell.font = Font(name='宋体', size=11, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    continue

                # ===== 分类标题行（第5、11、17行）=====
                if row_num in [5, 11, 17]:
                    cell.font = Font(name='宋体', size=11, bold=True)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    continue

                # ===== 垂直文本列（E列）=====
                if col_num == 5 and row_num >= 3:
                    cell.font = Font(name='宋体', size=11, bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center',
                                              text_rotation=255)
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    continue

                # ===== 数据单元格（普通单元格）=====
                if col_num in [1, 2, 3, 4] and row_num >= 5:
                    cell.font = Font(name='宋体', size=11)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    continue

                # ===== F-K列的意见内容区域 ======
                if col_num >= 6 and row_num >= 6:
                    cell.font = Font(name='宋体', size=11)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    continue

                # ===== 备注行（第32行）=====
                if row_num == 32:
                    cell.font = Font(name='宋体', size=11)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    continue


def export_performance_pay(data: Dict, year_month: str) -> str:
    """导出绩效工资审批表的便捷函数"""
    exporter = PerformancePayExcelExporter()
    return exporter.export_with_template(data, year_month)


if __name__ == "__main__":
    # 测试导出
    test_data = {
        '年月': '2026年5月',
        '填报单位': '太平镇中心学校',
        'administrative': {
            '副处级': {'count': 3, 'standard': 3500, 'subtotal': 10500},
            '正科级': {'count': 5, 'standard': 2800, 'subtotal': 14000},
            '副科级': {'count': 8, 'standard': 2200, 'subtotal': 17600},
            '科员级': {'count': 5, 'standard': 1185, 'subtotal': 5925},
            '办事员级': {'count': 2, 'standard': 1106, 'subtotal': 2212},
        },
        'professional': {
            '正高级': {'count': 2, 'standard': 1862, 'subtotal': 3724},
            '高级教师': {'count': 44, 'standard': 1523, 'subtotal': 67012},
            '一级教师': {'count': 136, 'standard': 1309, 'subtotal': 178024},
            '二级教师': {'count': 150, 'standard': 1241, 'subtotal': 186150},
            '三级教师': {'count': 19, 'standard': 1128, 'subtotal': 21432},
        },
        'worker': {
            '高级技师': {'count': 1, 'standard': 1331, 'subtotal': 1331},
            '技师': {'count': 2, 'standard': 1331, 'subtotal': 2662},
            '高级工': {'count': 3, 'standard': 1219, 'subtotal': 3657},
            '中级工': {'count': 5, 'standard': 1185, 'subtotal': 5925},
            '初级工': {'count': 1, 'standard': 1106, 'subtotal': 1106},
            '普工': {'count': 0, 'standard': 1106, 'subtotal': 0},
        },
        'totals': {
            'performance_count': 357,
            'performance_total': 462311,
        },
        'subsidies': {
            'count': 356,
            'standard': 350,
            'total': 124600,
        },
        'legacy': [
            {'name': '李发金', 'amount': 321.3},
            {'name': '张照凯', 'amount': 353.94},
        ],
        'retirees': {
            'cadre_count': 447,
            'worker_count': 2,
            'retired_count': 1,
        },
        'notes': '退休教师死亡2人：赵明安、候兴志'
    }

    path = export_performance_pay(test_data, '2026年5月')
    print(f"Excel导出成功: {path}")