"""
Excel导出工具 - 绩效工资审批表 (模板2 - 6列版)
100%复制原始模板格式，处理合并单元格
"""
import os
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime, timedelta

TEMPLATE_PATH = r"D:\erp_thirteen\数据库信息\模板\义务教育学校教职工绩效工资审批表（2）.xlsx"
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'exports')


def get_current_date():
    """获取当前日期，格式：2026年5月1日"""
    today = datetime.now()
    return f"{today.year}年{today.month}月{today.day}日"


def get_next_date():
    """获取明天日期，格式：2026年5月2日"""
    tomorrow = datetime.now() + timedelta(days=1)
    return f"{tomorrow.year}年{tomorrow.month}月{tomorrow.day}日"


def unmerge_and_write(ws, cell_range, value):
    """取消合并单元格并写入值"""
    try:
        ws.unmerge_cells(cell_range)
    except:
        pass
    ws[cell_range.split(':')[0]] = value


def export_performance_pay_approval_v2(data: dict, output_dir=None) -> str:
    """
    导出绩效工资审批表Excel（100%复制模板格式）
    """
    if output_dir is None:
        output_dir = OUTPUT_DIR

    os.makedirs(output_dir, exist_ok=True)

    year_month = data.get('年月', '未知')
    filename = f"绩效工资审批表_{year_month}.xlsx"
    filepath = os.path.join(output_dir, filename)

    shutil.copy2(TEMPLATE_PATH, filepath)

    wb = load_workbook(filepath)
    ws = wb.active

    current_date = get_current_date()
    next_date = get_next_date()

    report_unit = data.get('填报单位', '太平镇中心学校')

    # 取消需要写入的合并单元格
    merged_to_restore = []

    # B1:C1 合并 - 年月（取消合并，B1留空，日期放C1，样式与D1一致）
    try:
        ws.unmerge_cells('B1:C1')
    except:
        pass
    ws['B1'] = None
    ws['C1'] = year_month
    ws['C1'].font = Font(name='宋体', size=16, bold=True)
    ws['C1'].alignment = Alignment(horizontal='left', vertical='center')

    # B2:D2 合并 - 填报单位（取消合并后只写B2，C2和D2自动为空）
    try:
        ws.unmerge_cells('B2:D2')
    except:
        pass
    ws['B2'] = report_unit

    # E2:F2 合并 - 填报时间（取消合并，写入日期）
    try:
        ws.unmerge_cells('E2:F2')
    except:
        pass
    ws['E2'] = f"填报时间: {current_date}                    单位：人/元"
    ws['F2'] = f"填报时间: {current_date}                    单位：人/元"

    # ===== 行政管理人员 =====
    admin = data.get('administrative', {})
    # 第5行：副处级
    ws['B5'] = admin.get('副处级', {}).get('count', '')
    ws['C5'] = admin.get('副处级', {}).get('standard', '')
    ws['D5'] = admin.get('副处级', {}).get('subtotal', '')
    # 第6行：正科级
    ws['B6'] = admin.get('正科级', {}).get('count', '')
    ws['C6'] = admin.get('正科级', {}).get('standard', '')
    ws['D6'] = admin.get('正科级', {}).get('subtotal', '')
    # 第7行：副科级
    ws['B7'] = admin.get('副科级', {}).get('count', '')
    ws['C7'] = admin.get('副科级', {}).get('standard', '')
    ws['D7'] = admin.get('副科级', {}).get('subtotal', '')
    # 第8行：科员级
    ws['B8'] = admin.get('科员级', {}).get('count', '')
    ws['C8'] = admin.get('科员级', {}).get('standard', '')
    ws['D8'] = admin.get('科员级', {}).get('subtotal', '')
    # 第9行：办事员级
    ws['B9'] = admin.get('办事员级', {}).get('count', '')
    ws['C9'] = admin.get('办事员级', {}).get('standard', '')
    ws['D9'] = admin.get('办事员级', {}).get('subtotal', '')

    # ===== 呈报单位意见（F3:F9区域 - 添加外边框）=====
    ws.unmerge_cells('F3:F9')

    thin = Side(border_style='medium', color='000000')

    ws['F6'] = "据实填写，同意呈报。"
    ws['F6'].font = Font(name='宋体', size=11)
    ws['F6'].alignment = Alignment(horizontal='center', vertical='center')

    ws['F7'] = "（盖章）"
    ws['F7'].font = Font(name='宋体', size=11)
    ws['F7'].alignment = Alignment(horizontal='center', vertical='center')

    ws['F9'] = current_date
    ws['F9'].font = Font(name='宋体', size=11)
    ws['F9'].alignment = Alignment(horizontal='left', vertical='bottom', indent=21)

    # F3: 上边框 + 左边框 + 右边框
    ws['F3'].border = Border(top=thin, left=thin, right=thin)
    # F4-F8: 左边框 + 右边框（无上下边框）
    for row in [4, 5, 6, 7, 8]:
        ws[f'F{row}'].border = Border(left=thin, right=thin)
    # F9: 下边框 + 右边框 + 左边框
    ws['F9'].border = Border(bottom=thin, right=thin, left=thin)

    # ===== 专业技术人员 =====
    pro = data.get('professional', {})
    # 第11行：正高级
    ws['B11'] = pro.get('正高级', {}).get('count', '')
    ws['C11'] = pro.get('正高级', {}).get('standard', '')
    ws['D11'] = pro.get('正高级', {}).get('subtotal', '')
    # 第12行：高级教师
    ws['B12'] = pro.get('高级教师', {}).get('count', '')
    ws['C12'] = pro.get('高级教师', {}).get('standard', '')
    ws['D12'] = pro.get('高级教师', {}).get('subtotal', '')
    # 第13行：一级教师
    ws['B13'] = pro.get('一级教师', {}).get('count', '')
    ws['C13'] = pro.get('一级教师', {}).get('standard', '')
    ws['D13'] = pro.get('一级教师', {}).get('subtotal', '')
    # 第14行：二级教师
    ws['B14'] = pro.get('二级教师', {}).get('count', '')
    ws['C14'] = pro.get('二级教师', {}).get('standard', '')
    ws['D14'] = pro.get('二级教师', {}).get('subtotal', '')
    # 第15行：三级教师
    ws['B15'] = pro.get('三级教师', {}).get('count', '')
    ws['C15'] = pro.get('三级教师', {}).get('standard', '')
    ws['D15'] = pro.get('三级教师', {}).get('subtotal', '')

    # ===== 教育局意见（F10:F18区域 - 添加外边框）=====
    ws.unmerge_cells('F10:F18')
    ws['F13'] = "（盖章）"
    ws['F13'].font = Font(name='宋体', size=11)
    ws['F13'].alignment = Alignment(horizontal='center', vertical='center')

    ws['F18'] = next_date
    ws['F18'].font = Font(name='宋体', size=11)
    ws['F18'].alignment = Alignment(horizontal='left', vertical='bottom', indent=21)

    ws['F10'].border = Border(top=thin, left=thin, right=thin)
    for row in range(11, 18):
        ws[f'F{row}'].border = Border(left=thin, right=thin)
    ws['F18'].border = Border(bottom=thin, right=thin, left=thin)

    # ===== 工人 =====
    worker = data.get('worker', {})
    # 第17行：高级技师
    ws['B17'] = worker.get('高级技师', {}).get('count', '')
    ws['C17'] = worker.get('高级技师', {}).get('standard', '')
    ws['D17'] = worker.get('高级技师', {}).get('subtotal', '')
    # 第18行：技师
    ws['B18'] = worker.get('技师', {}).get('count', '')
    ws['C18'] = worker.get('技师', {}).get('standard', '')
    ws['D18'] = worker.get('技师', {}).get('subtotal', '')

    # ===== 教育局意见（F10:F18合并单元格）- 已在上面处理 =====

    # 第19行：高级工
    ws['B19'] = worker.get('高级工', {}).get('count', '')
    ws['C19'] = worker.get('高级工', {}).get('standard', '')
    ws['D19'] = worker.get('高级工', {}).get('subtotal', '')
    # 第20行：中级工
    ws['B20'] = worker.get('中级工', {}).get('count', '')
    ws['C20'] = worker.get('中级工', {}).get('standard', '')
    ws['D20'] = worker.get('中级工', {}).get('subtotal', '')
    # 第21行：初级工
    ws['B21'] = worker.get('初级工', {}).get('count', '')
    ws['C21'] = worker.get('初级工', {}).get('standard', '')
    ws['D21'] = worker.get('初级工', {}).get('subtotal', '')
    # 第22行：普工
    ws['B22'] = worker.get('普工', {}).get('count', '')
    ws['C22'] = worker.get('普工', {}).get('standard', '')
    ws['D22'] = worker.get('普工', {}).get('subtotal', '')

    # ===== 人事部门意见（F19:F30区域 - 添加外边框）=====
    ws.unmerge_cells('F19:F30')

    totals = data.get('totals', {})
    subsidies = data.get('subsidies', {})

    perfCount = totals.get('performance_count', 0)
    perfTotal = totals.get('performance_total', 0)
    subCount = subsidies.get('count', 0)
    subTotal = subsidies.get('total', 0)
    legacyCount = totals.get('legacy_count', 0)
    legacyTotal = totals.get('legacy_total', 0)
    totalAll = perfTotal + subTotal + legacyTotal
    noSubsidyCount = data.get('no_subsidy_count', 0)
    noSubsidyNames = data.get('no_subsidy_names', '')

    ws['F20'] = "根据相关文件及有关规定，经审核，同意你单位："
    ws['F20'].font = Font(name='宋体', size=10)
    ws['F20'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=2)

    ws['F21'] = f"基础性绩效工资{perfCount}人，{perfTotal}元；"
    ws['F22'] = f"生活补贴{subCount}人，{subTotal}元；"
    ws['F23'] = f"合计{perfTotal + subTotal}元；"
    ws['F24'] = f"岗位设置遗留{legacyCount}人，{legacyTotal}元；"
    ws['F25'] = f"总计{totalAll}元。"
    ws['F26'] = f"无乡镇补贴{noSubsidyCount}人，{noSubsidyNames}。"
    for row in range(21, 27):
        ws[f'F{row}'].font = Font(name='宋体', size=10)
        ws[f'F{row}'].alignment = Alignment(horizontal='right', vertical='center', wrap_text=True, indent=8)
        ws.row_dimensions[row].height = 25

    ws['F30'] = next_date
    ws['F30'].font = Font(name='宋体', size=11)
    ws['F30'].alignment = Alignment(horizontal='left', vertical='bottom', indent=21)

    ws['F19'].border = Border(top=thin, left=thin, right=thin)
    for row in range(20, 30):
        ws[f'F{row}'].border = Border(left=thin, right=thin)
    ws['F30'].border = Border(bottom=thin, right=thin, left=thin)

    # ===== 绩效工资合计（第23行）=====
    ws['B23'] = totals.get('performance_count', '')
    ws['C23'] = totals.get('performance_count', '')
    ws['D23'] = totals.get('performance_total', '')

    # ===== 乡镇补贴合计（第24行）=====
    ws['B24'] = subsidies.get('count', '')
    ws['C24'] = subsidies.get('standard', 350)
    ws['D24'] = subsidies.get('total', '')

    # ===== 岗位设置遗留问题（第25-26行）=====
    legacy = data.get('legacy', [])
    if len(legacy) > 0:
        ws['B25'] = legacy[0].get('name', '')
        ws['C25'] = legacy[0].get('amount', '')
        ws['D25'] = legacy[0].get('amount', '')
    if len(legacy) > 1:
        ws['B26'] = legacy[1].get('name', '')
        ws['C26'] = legacy[1].get('amount', '')
        ws['D26'] = legacy[1].get('amount', '')

    # ===== 岗位设置遗留问题合计（第27行）=====
    ws['B27'] = totals.get('legacy_count', '')
    ws['C27'] = totals.get('legacy_count', '')
    ws['D27'] = totals.get('legacy_total', '')

    # ===== 退休人员 =====
    retirees = data.get('retirees', {})
    # 第28行：退休干部
    ws['B28'] = retirees.get('cadre_count', '')
    ws['C28'] = retirees.get('cadre_count', '')
    # 第29行：退休工人
    ws['B29'] = retirees.get('worker_count', '')
    ws['C29'] = retirees.get('worker_count', '')
    # 第30行：离休干部
    ws['B30'] = retirees.get('retired_count', '')
    ws['C30'] = retirees.get('retired_count', '')

    # ===== 备注（第31行 A31:F31合并单元格 - 保留模板备注：，内容左对齐无空格）=====
    notes = data.get('备注', '') or data.get('notes', '')
    if notes:
        lines = notes.split('\n')
        clean_lines = [line.lstrip() for line in lines if line.strip()]
        clean_notes = '\n'.join(clean_lines)
        ws['A31'] = f"备注：\n{clean_notes}"
    else:
        ws['A31'] = "备注："

    wb.save(filepath)
    return filepath