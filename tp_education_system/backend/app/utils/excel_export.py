"""
Excel导出工具 - 绩效工资审批表
严格按照原始HTML格式导出
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def export_performance_pay_approval(data, output_dir=None):
    """
    导出绩效工资审批表为Excel
    100%还原原始HTML格式
    """
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'exports')
    
    os.makedirs(output_dir, exist_ok=True)
    
    # 创建文件名
    year_month = data.get('年月', '未知')
    filename = f"绩效工资审批表_{year_month}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "绩效工资审批表"
    
    # 设置页面（A4纸张，页边距）
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    
    # 页边距（单位：英寸）
    ws.page_margins.left = 0.75
    ws.page_margins.right = 0.75
    ws.page_margins.top = 1.0
    ws.page_margins.bottom = 1.0
    ws.page_margins.header = 0.5
    ws.page_margins.footer = 0.5
    
    # 定义样式
    # 边框样式 - 0.5pt实线
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 字体样式
    font_title = Font(name='宋体', size=18, bold=True)
    font_header = Font(name='宋体', size=12, bold=True)
    font_normal = Font(name='宋体', size=10)
    font_month = Font(name='宋体', size=16, bold=True)
    
    # 对齐方式
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    
    # 列宽设置（单位：字符宽度）
    column_widths = {
        'A': 12,   # 项目列
        'B': 10,   # 人数
        'C': 12,   # 月工资标准
        'D': 10,   # 小计
        'E': 8,    # 呈报单位意见
        'F': 12,   # 教育局意见/人事部门意见内容
        'G': 12,
        'H': 12,
        'I': 12,
        'J': 12,
        'K': 12,
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 行高设置
    row_height = 25  # 25pt
    
    current_row = 1
    
    # 第1行：标题
    ws.merge_cells('A1:A1')
    ws['A1'] = ''
    ws['A1'].font = font_normal
    
    ws.merge_cells('B1:C1')
    ws['B1'] = year_month
    ws['B1'].font = font_month
    ws['B1'].alignment = align_right
    
    ws.merge_cells('D1:I1')
    ws['D1'] = '义务教育学校教职工绩效工资审批表'
    ws['D1'].font = font_title
    ws['D1'].alignment = align_center
    
    ws.row_dimensions[1].height = row_height
    current_row = 2
    
    # 第2行：填报信息
    ws['A2'] = '填报单位：'
    ws['A2'].font = font_normal
    ws['A2'].alignment = align_left
    
    ws.merge_cells('B2:D2')
    ws['B2'] = data.get('填报单位', '太平中心学校')
    ws['B2'].font = font_normal
    ws['B2'].alignment = align_left
    ws['B2'].border = Border(bottom=Side(style='thin'))
    
    ws['E2'] = '填报时间:'
    ws['E2'].font = font_normal
    ws['E2'].alignment = align_right
    
    ws['F2'] = data.get('填报时间', '')
    ws['F2'].font = font_normal
    ws['F2'].alignment = align_left
    ws['F2'].border = Border(bottom=Side(style='thin'))
    
    ws['G2'] = '单位：'
    ws['G2'].font = font_normal
    ws['G2'].alignment = align_right
    
    ws['H2'] = '人、元'
    ws['H2'].font = font_normal
    ws['H2'].alignment = align_left
    
    ws.row_dimensions[2].height = row_height
    current_row = 3
    
    # 第3-4行：表头
    ws.merge_cells(f'A{current_row}:A{current_row+1}')
    ws[f'A{current_row}'] = '项  目'
    ws[f'A{current_row}'].font = font_header
    ws[f'A{current_row}'].alignment = align_center
    ws[f'A{current_row}'].border = thin_border
    
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = '基础性工资'
    ws[f'B{current_row}'].font = font_header
    ws[f'B{current_row}'].alignment = align_center
    ws[f'B{current_row}'].border = thin_border
    
    ws.merge_cells(f'E{current_row}:E{current_row+4}')
    ws[f'E{current_row}'] = '呈报单位意见'
    ws[f'E{current_row}'].font = font_header
    ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', text_rotation=255)
    ws[f'E{current_row}'].border = thin_border
    
    ws.merge_cells(f'F{current_row}:K{current_row+1}')
    ws[f'F{current_row}'] = '据实填写，同意呈报。'
    ws[f'F{current_row}'].font = font_normal
    ws[f'F{current_row}'].alignment = align_left
    ws[f'F{current_row}'].border = thin_border
    
    ws.row_dimensions[current_row].height = row_height
    current_row += 1
    
    # 第4行：表头第二行
    ws[f'B{current_row}'] = '人数'
    ws[f'B{current_row}'].font = font_normal
    ws[f'B{current_row}'].alignment = align_center
    ws[f'B{current_row}'].border = thin_border
    
    ws[f'C{current_row}'] = '月工资标准'
    ws[f'C{current_row}'].font = font_normal
    ws[f'C{current_row}'].alignment = align_center
    ws[f'C{current_row}'].border = thin_border
    
    ws[f'D{current_row}'] = '小计'
    ws[f'D{current_row}'].font = font_normal
    ws[f'D{current_row}'].alignment = align_center
    ws[f'D{current_row}'].border = thin_border
    
    ws.row_dimensions[current_row].height = row_height
    current_row += 1
    
    # 行政管理人员
    ws[f'A{current_row}'] = '行政管理人员'
    ws[f'A{current_row}'].font = font_header
    ws[f'A{current_row}'].alignment = align_center
    ws[f'A{current_row}'].border = thin_border
    
    ws.merge_cells(f'F{current_row}:K{current_row+2}')
    ws[f'F{current_row}'] = '（盖章）\n\n' + data.get('填报时间', '')
    ws[f'F{current_row}'].font = font_normal
    ws[f'F{current_row}'].alignment = align_center
    ws[f'F{current_row}'].border = thin_border
    
    ws.row_dimensions[current_row].height = row_height
    current_row += 1
    
    # 各级别数据行
    admin_levels = [
        ('1、副处级', '副处级人数', '副处级标准'),
        ('2、正科级', '正科级人数', '正科级标准'),
        ('3、副科级', '副科级人数', '副科级标准'),
    ]
    
    for label, count_key, std_key in admin_levels:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = align_left
        ws[f'A{current_row}'].border = thin_border
        
        count = data.get(count_key, 0)
        std = data.get(std_key, 0)
        
        ws[f'B{current_row}'] = count if count else ''
        ws[f'B{current_row}'].font = font_normal
        ws[f'B{current_row}'].alignment = align_center
        ws[f'B{current_row}'].border = thin_border
        
        ws[f'C{current_row}'] = std if std else ''
        ws[f'C{current_row}'].font = font_normal
        ws[f'C{current_row}'].alignment = align_center
        ws[f'C{current_row}'].border = thin_border
        
        ws[f'D{current_row}'] = count * std if count and std else ''
        ws[f'D{current_row}'].font = font_normal
        ws[f'D{current_row}'].alignment = align_center
        ws[f'D{current_row}'].border = thin_border
        
        ws.row_dimensions[current_row].height = row_height
        current_row += 1
    
    # 科员级和办事员级（带教育局意见）
    ws[f'A{current_row}'] = '4、科员级'
    ws[f'A{current_row}'].font = font_normal
    ws[f'A{current_row}'].alignment = align_left
    ws[f'A{current_row}'].border = thin_border
    
    count = data.get('科员级人数', 0)
    std = data.get('科员级标准', 1185)
    
    ws[f'B{current_row}'] = count if count else ''
    ws[f'B{current_row}'].font = font_normal
    ws[f'B{current_row}'].alignment = align_center
    ws[f'B{current_row}'].border = thin_border
    
    ws[f'C{current_row}'] = std
    ws[f'C{current_row}'].font = font_normal
    ws[f'C{current_row}'].alignment = align_center
    ws[f'C{current_row}'].border = thin_border
    
    ws[f'D{current_row}'] = count * std if count else ''
    ws[f'D{current_row}'].font = font_normal
    ws[f'D{current_row}'].alignment = align_center
    ws[f'D{current_row}'].border = thin_border
    
    ws.merge_cells(f'E{current_row}:E{current_row+4}')
    ws[f'E{current_row}'] = '教育局意见'
    ws[f'E{current_row}'].font = font_header
    ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', text_rotation=255)
    ws[f'E{current_row}'].border = thin_border
    
    ws.merge_cells(f'F{current_row}:K{current_row+4}')
    ws[f'F{current_row}'] = '（盖章）\n\n' + data.get('填报时间', '')
    ws[f'F{current_row}'].font = font_normal
    ws[f'F{current_row}'].alignment = align_center
    ws[f'F{current_row}'].border = thin_border
    
    ws.row_dimensions[current_row].height = row_height
    current_row += 1
    
    # 办事员级
    ws[f'A{current_row}'] = '5、办事员级'
    ws[f'A{current_row}'].font = font_normal
    ws[f'A{current_row}'].alignment = align_left
    ws[f'A{current_row}'].border = thin_border
    
    count = data.get('办事员级人数', 0)
    std = data.get('办事员级标准', 0)
    
    ws[f'B{current_row}'] = count if count else ''
    ws[f'B{current_row}'].font = font_normal
    ws[f'B{current_row}'].alignment = align_center
    ws[f'B{current_row}'].border = thin_border
    
    ws[f'C{current_row}'] = std if std else ''
    ws[f'C{current_row}'].font = font_normal
    ws[f'C{current_row}'].alignment = align_center
    ws[f'C{current_row}'].border = thin_border
    
    ws[f'D{current_row}'] = count * std if count and std else ''
    ws[f'D{current_row}'].font = font_normal
    ws[f'D{current_row}'].alignment = align_center
    ws[f'D{current_row}'].border = thin_border
    
    ws.row_dimensions[current_row].height = row_height
    current_row += 1
    
    # 专业技术人员
    ws[f'A{current_row}'] = '专业技术人员'
    ws[f'A{current_row}'].font = font_header
    ws[f'A{current_row}'].alignment = align_center
    ws[f'A{current_row}'].border = thin_border
    
    ws.row_dimensions[current_row].height = row_height
    current_row += 1
    
    # 教师级别数据
    teacher_levels = [
        ('1、正高级教师', '正高级教师人数', '正高级教师标准', 1862),
        ('2、高级教师', '高级教师人数', '高级教师标准', 1523),
        ('3、一级教师', '一级教师人数', '一级教师标准', 1309),
        ('4、二级教师', '二级教师人数', '二级教师标准', 1241),
        ('5、三级教师', '三级教师人数', '三级教师标准', 1128),
    ]
    
    for i, (label, count_key, std_key, default_std) in enumerate(teacher_levels):
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = align_left
        ws[f'A{current_row}'].border = thin_border
        
        count = data.get(count_key, 0)
        std = data.get(std_key, default_std)
        
        ws[f'B{current_row}'] = count if count else ''
        ws[f'B{current_row}'].font = font_normal
        ws[f'B{current_row}'].alignment = align_center
        ws[f'B{current_row}'].border = thin_border
        
        ws[f'C{current_row}'] = std
        ws[f'C{current_row}'].font = font_normal
        ws[f'C{current_row}'].alignment = align_center
        ws[f'C{current_row}'].border = thin_border
        
        ws[f'D{current_row}'] = count * std if count else ''
        ws[f'D{current_row}'].font = font_normal
        ws[f'D{current_row}'].alignment = align_center
        ws[f'D{current_row}'].border = thin_border
        
        # 第一行教师添加人事部门意见
        if i == 0:
            ws.merge_cells(f'E{current_row}:E{current_row+11}')
            ws[f'E{current_row}'] = '人事部门意见'
            ws[f'E{current_row}'].font = font_header
            ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center', text_rotation=255)
            ws[f'E{current_row}'].border = thin_border
            
            # 构建审批内容
            approval_text = f"""    根据相关文件及有关规定，经审核，同意你单位：

基础性绩效工资    {data.get('绩效人数合计', 0)}人：    {data.get('绩效工资合计', 0)}元；
生活补贴    {data.get('在职人数', 0)}人：    {data.get('乡镇补贴合计', 0)}元；
岗位设置遗留问题    {data.get('遗留问题人数', 0)}人：    {data.get('遗留问题金额', 0)}元；

合计：    {data.get('绩效工资合计', 0) + data.get('乡镇补贴合计', 0) + data.get('遗留问题金额', 0)}元

注：无生活补贴    {data.get('无补贴人数', 0)}人：    {data.get('无补贴名单', '')}


{data.get('填报时间', '')}"""
            
            ws.merge_cells(f'F{current_row}:K{current_row+11}')
            ws[f'F{current_row}'] = approval_text
            ws[f'F{current_row}'].font = font_normal
            ws[f'F{current_row}'].alignment = align_left
            ws[f'F{current_row}'].border = thin_border
        
        ws.row_dimensions[current_row].height = row_height
        current_row += 1
    
    # 工人
    ws[f'A{current_row}'] = '工人'
    ws[f'A{current_row}'].font = font_header
    ws[f'A{current_row}'].alignment = align_center
    ws[f'A{current_row}'].border = thin_border
    
    ws.row_dimensions[current_row].height = row_height
    current_row += 1
    
    # 工人级别数据
    worker_levels = [
        ('1、高级技师', '高级技师人数', '高级技师标准', 0),
        ('2、技师', '技师人数', '技师标准', 1331),
        ('3、高级工', '高级工人数', '高级工标准', 1219),
        ('4、中级工', '中级工人数', '中级工标准', 1185),
        ('5、初级工', '初级工人数', '初级工标准', 1106),
        ('6、普工', '普工人数', '普工标准', 1106),
    ]
    
    for label, count_key, std_key, default_std in worker_levels:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = font_normal
        ws[f'A{current_row}'].alignment = align_left
        ws[f'A{current_row}'].border = thin_border
        
        count = data.get(count_key, 0)
        std = data.get(std_key, default_std)
        
        ws[f'B{current_row}'] = count if count else ''
        ws[f'B{current_row}'].font = font_normal
        ws[f'B{current_row}'].alignment = align_center
        ws[f'B{current_row}'].border = thin_border
        
        ws[f'C{current_row}'] = std if std else ''
        ws[f'C{current_row}'].font = font_normal
        ws[f'C{current_row}'].alignment = align_center
        ws[f'C{current_row}'].border = thin_border
        
        ws[f'D{current_row}'] = count * std if count and std else ''
        ws[f'D{current_row}'].font = font_normal
        ws[f'D{current_row}'].alignment = align_center
        ws[f'D{current_row}'].border = thin_border
        
        ws.row_dimensions[current_row].height = row_height
        current_row += 1
    
    # 绩效汇总
    ws[f'A{current_row}'] = '绩效人数'
    ws[f'A{current_row}'].font = font_header
    ws[f'A{current_row}'].alignment = align_center
    ws[f'A{current_row}'].border = thin_border
    
    ws[f'B{current_row}'] = data.get('绩效人数合计', 0)
    ws[f'B{current_row}'].font = font_normal
    ws[f'B{current_row}'].alignment = align_center
    ws[f'B{current_row}'].border = thin_border
    
    ws[f'C{current_row}'] = '绩效合计'
    ws[f'C{current_row}'].font = font_header
    ws[f'C{current_row}'].alignment = align_center
    ws[f'C{current_row}'].border = thin_border
    
    ws[f'D{current_row}'] = data.get('绩效工资合计', 0)
    ws[f'D{current_row}'].font = font_normal
    ws[f'D{current_row}'].alignment = align_center
    ws[f'D{current_row}'].border = thin_border
    
    ws.row_dimensions[current_row].height = row_height
    current_row += 1
    
    # 乡镇补贴
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = '乡镇工作补贴人数'
    ws[f'A{current_row}'].font = font_header
    ws[f'A{current_row}'].alignment = align_center
    ws[f'A{current_row}'].border = thin_border
    
    ws[f'C{current_row}'] = data.get('在职人数', 0)
    ws[f'C{current_row}'].font = font_normal
    ws[f'C{current_row}'].alignment = align_center
    ws[f'C{current_row}'].border = thin_border
    
    ws[f'D{current_row}'] = '标准'
    ws[f'D{current_row}'].font = font_header
    ws[f'D{current_row}'].alignment = align_center
    ws[f'D{current_row}'].border = thin_border
    
    ws.merge_cells(f'E{current_row}:F{current_row}')
    ws[f'E{current_row}'] = data.get('乡镇补贴标准', 350)
    ws[f'E{current_row}'].font = font_normal
    ws[f'E{current_row}'].alignment = align_center
    ws[f'E{current_row}'].border = thin_border
    
    ws[f'G{current_row}'] = '金额'
    ws[f'G{current_row}'].font = font_header
    ws[f'G{current_row}'].alignment = align_center
    ws[f'G{current_row}'].border = thin_border
    
    ws.merge_cells(f'H{current_row}:K{current_row}')
    ws[f'H{current_row}'] = data.get('乡镇补贴合计', 0)
    ws[f'H{current_row}'].font = font_normal
    ws[f'H{current_row}'].alignment = align_center
    ws[f'H{current_row}'].border = thin_border
    
    ws.row_dimensions[current_row].height = row_height
    current_row += 1
    
    # 退休人员
    ws[f'A{current_row}'] = '退休干部人数'
    ws[f'A{current_row}'].font = font_header
    ws[f'A{current_row}'].alignment = align_center
    ws[f'A{current_row}'].border = thin_border
    
    ws[f'B{current_row}'] = data.get('退休干部', 0)
    ws[f'B{current_row}'].font = font_normal
    ws[f'B{current_row}'].alignment = align_center
    ws[f'B{current_row}'].border = thin_border
    
    ws.merge_cells(f'C{current_row}:D{current_row}')
    ws[f'C{current_row}'] = '退休工人人数'
    ws[f'C{current_row}'].font = font_header
    ws[f'C{current_row}'].alignment = align_center
    ws[f'C{current_row}'].border = thin_border
    
    ws.merge_cells(f'E{current_row}:F{current_row}')
    ws[f'E{current_row}'] = data.get('退休职工', 0)
    ws[f'E{current_row}'].font = font_normal
    ws[f'E{current_row}'].alignment = align_center
    ws[f'E{current_row}'].border = thin_border
    
    ws[f'G{current_row}'] = '离休干部人数'
    ws[f'G{current_row}'].font = font_header
    ws[f'G{current_row}'].alignment = align_center
    ws[f'G{current_row}'].border = thin_border
    
    ws.merge_cells(f'H{current_row}:K{current_row}')
    ws[f'H{current_row}'] = data.get('离休干部人数', 0)
    ws[f'H{current_row}'].font = font_normal
    ws[f'H{current_row}'].alignment = align_center
    ws[f'H{current_row}'].border = thin_border
    
    ws.row_dimensions[current_row].height = row_height
    current_row += 1
    
    # 岗位设置遗留问题
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = '岗位设置遗留问题'
    ws[f'A{current_row}'].font = font_header
    ws[f'A{current_row}'].alignment = align_center
    ws[f'A{current_row}'].border = thin_border
    
    ws.merge_cells(f'C{current_row}:D{current_row}')
    ws[f'C{current_row}'] = data.get('遗留问题详情', '')
    ws[f'C{current_row}'].font = font_normal
    ws[f'C{current_row}'].alignment = align_left
    ws[f'C{current_row}'].border = thin_border
    
    ws[f'E{current_row}'] = '人数'
    ws[f'E{current_row}'].font = font_header
    ws[f'E{current_row}'].alignment = align_center
    ws[f'E{current_row}'].border = thin_border
    
    ws.merge_cells(f'F{current_row}:G{current_row}')
    ws[f'F{current_row}'] = data.get('遗留问题人数', 0)
    ws[f'F{current_row}'].font = font_normal
    ws[f'F{current_row}'].alignment = align_center
    ws[f'F{current_row}'].border = thin_border
    
    ws[f'H{current_row}'] = '金额'
    ws[f'H{current_row}'].font = font_header
    ws[f'H{current_row}'].alignment = align_center
    ws[f'H{current_row}'].border = thin_border
    
    ws.merge_cells(f'I{current_row}:K{current_row}')
    ws[f'I{current_row}'] = data.get('遗留问题金额', 0)
    ws[f'I{current_row}'].font = font_normal
    ws[f'I{current_row}'].alignment = align_center
    ws[f'I{current_row}'].border = thin_border
    
    ws.row_dimensions[current_row].height = 37  # 37pt
    current_row += 1
    
    # 备注
    ws.merge_cells(f'A{current_row}:K{current_row}')
    ws[f'A{current_row}'] = f"备注: {data.get('备注', '')}"
    ws[f'A{current_row}'].font = font_normal
    ws[f'A{current_row}'].alignment = align_left
    ws[f'A{current_row}'].border = thin_border
    
    ws.row_dimensions[current_row].height = 21  # 21pt
    
    # 保存文件
    wb.save(filepath)
    
    return filepath
